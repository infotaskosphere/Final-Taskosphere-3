import os, io, base64, asyncio, time, logging
from fastapi import APIRouter, UploadFile, File, HTTPException, Depends
from backend.dependencies import get_current_user

router = APIRouter(prefix="/api/ai", tags=["AI Document Reader"])

_batch_logger = logging.getLogger("ai_document_reader.batching")

# ═══════════════════════════════════════════════════════════════════════════
# GROQ AUTO PAGE BATCHING (max 3 images/request) — additive, backward compatible.
# Used by the scanned-PDF branch below and can be reused by the OCR pipeline.
# ═══════════════════════════════════════════════════════════════════════════

def _batch_size() -> int:
    try:
        n = int(os.environ.get("GROQ_MAX_IMAGES_PER_REQUEST", "3"))
    except ValueError:
        n = 3
    return max(1, min(n, 3))  # Groq hard cap = 3


def _parallel_batches() -> int:
    try:
        n = int(os.environ.get("GROQ_MAX_PARALLEL_BATCHES", "3"))
    except ValueError:
        n = 3
    return max(1, n)


async def _groq_ocr_single_page(img_b64: str, mime: str, prompt: str) -> str:
    """OCR one page via Groq with 1 retry."""
    for attempt in (1, 2):
        try:
            return await _groq_vision_raw(img_b64, mime, prompt)
        except Exception as e:
            if attempt == 2:
                _batch_logger.warning(f"Single-page OCR failed after retry: {e}")
                return ""
            await asyncio.sleep(0.4)
    return ""


async def _groq_ocr_batch(pages: list, prompt: str) -> str:
    """OCR a batch (<=3 images). Retry once, then fall back to per-page OCR
    so a single bad page never fails the whole batch."""
    if not pages:
        return ""
    for attempt in (1, 2):
        try:
            return await _groq_vision_multipage_raw(pages, prompt)
        except Exception as e:
            if attempt == 1:
                _batch_logger.warning(f"Batch OCR attempt 1 failed, retrying: {e}")
                await asyncio.sleep(0.6)
                continue
            _batch_logger.warning(f"Batch OCR retry failed, splitting into single pages: {e}")
    # Fallback: OCR each page individually and stitch
    results = []
    for i, (b64, mime) in enumerate(pages, 1):
        results.append(await _groq_ocr_single_page(b64, mime, prompt))
    return "\n".join(results)


async def _groq_vision_batched_pages(
    page_images_b64: list,
    prompt: str,
    progress_cb=None,
) -> str:
    """
    Auto-batch pages into groups of <= GROQ_MAX_IMAGES_PER_REQUEST (default 3),
    run up to GROQ_MAX_PARALLEL_BATCHES batches concurrently, merge results
    strictly in page order. Retries once per batch, then splits to single-page
    OCR on failure. Releases each page's base64 payload after its batch runs
    to keep memory usage bounded for large PDFs.
    """
    bsize = _batch_size()
    pcount = _parallel_batches()
    batches = [page_images_b64[i:i + bsize] for i in range(0, len(page_images_b64), bsize)]
    total = len(batches)
    results: list = [None] * total

    _batch_logger.info(
        f"Groq batched OCR: {len(page_images_b64)} pages -> {total} batches "
        f"(batch_size={bsize}, parallel={pcount})"
    )

    sem = asyncio.Semaphore(pcount)
    done_counter = {"n": 0}
    lock = asyncio.Lock()

    async def _run(idx: int, batch):
        async with sem:
            t0 = time.time()
            text = await _groq_ocr_batch(batch, prompt)
            _batch_logger.info(f"Batch {idx + 1}/{total} done in {time.time() - t0:.2f}s")
            # Release b64 payloads for this batch immediately
            batch.clear()
            results[idx] = text or ""
            async with lock:
                done_counter["n"] += 1
                if progress_cb:
                    try:
                        progress_cb(done_counter["n"], total)
                    except Exception:
                        pass

    await asyncio.gather(*[_run(i, b) for i, b in enumerate(batches)])
    # Merge preserving original page order
    return "\n\n".join(r for r in results if r)


# ── Gemini client (PDF text, Excel, CSV) ─────────────────────────────────────
def _get_gemini_model():
    key = os.environ.get("GEMINI_API_KEY", "")
    if not key:
        raise HTTPException(
            status_code=500,
            detail="GEMINI_API_KEY is not configured on the server."
        )
    try:
        import google.generativeai as genai
        genai.configure(api_key=key)
        return genai.GenerativeModel("gemini-2.0-flash")
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="google-generativeai package not installed."
        )


# ── Provider selection ────────────────────────────────────────────────────────
def _provider() -> str:
    p = (os.environ.get("AI_PROVIDER") or "").strip().lower()
    if p in ("gemini", "google", "google-ai"):
        return "gemini"
    if p == "groq":
        return "groq"
    if os.environ.get("GEMINI_API_KEY") or os.environ.get("GOOGLE_API_KEY"):
        return "gemini"
    return "groq"


def _gemini_key() -> str:
    return (os.environ.get("GEMINI_API_KEY")
            or os.environ.get("GOOGLE_API_KEY")
            or os.environ.get("GOOGLE_AI_STUDIO_API_KEY")
            or "").strip()


# ── Gemini vision — single image ─────────────────────────────────────────────
async def _gemini_vision(image_b64: str, mime_type: str, prompt: str) -> str:
    import httpx
    key = _gemini_key()
    if not key:
        raise HTTPException(status_code=500, detail="GEMINI_API_KEY is not configured on the server.")
    model = (os.environ.get("GEMINI_VISION_MODEL") or "gemini-2.5-flash").strip()
    url = f"https://generativelanguage.googleapis.com/v1beta/models/{model}:generateContent"
    body = {
        "contents": [{
            "role": "user",
            "parts": [
                {"text": prompt},
                {"inline_data": {"mime_type": mime_type, "data": image_b64}},
            ],
        }],
        "generationConfig": {"temperature": 0.2, "maxOutputTokens": 2048},
    }
    async with httpx.AsyncClient(timeout=90) as client:
        resp = await client.post(url, params={"key": key},
                                 headers={"Content-Type": "application/json"}, json=body)
    if resp.status_code == 429:
        raise HTTPException(status_code=429, detail="Gemini quota exceeded. Please wait a moment and try again.")
    if resp.status_code != 200:
        raise HTTPException(status_code=422, detail=f"Gemini API error {resp.status_code}: {resp.text[:300]}")
    data = resp.json()
    try:
        parts = data["candidates"][0]["content"]["parts"]
        return "".join(p.get("text", "") for p in parts if isinstance(p, dict))
    except (KeyError, IndexError, TypeError):
        raise HTTPException(status_code=422, detail="Gemini returned an empty response.")


async def _gemini_vision_multipage(page_images_b64: list, prompt: str) -> str:
    import httpx
    key = _gemini_key()
    if not key:
        raise HTTPException(status_code=500, detail="GEMINI_API_KEY is not configured on the server.")
    model = (os.environ.get("GEMINI_VISION_MODEL") or "gemini-2.5-flash").strip()
    url = f"https://generativelanguage.googleapis.com/v1beta/models/{model}:generateContent"
    parts = [{"text": prompt}]
    for img_b64, mime in page_images_b64:
        parts.append({"inline_data": {"mime_type": mime, "data": img_b64}})
    body = {
        "contents": [{"role": "user", "parts": parts}],
        "generationConfig": {"temperature": 0.2, "maxOutputTokens": 2048},
    }
    async with httpx.AsyncClient(timeout=120) as client:
        resp = await client.post(url, params={"key": key},
                                 headers={"Content-Type": "application/json"}, json=body)
    if resp.status_code == 429:
        raise HTTPException(status_code=429, detail="Gemini quota exceeded. Please wait a moment and try again.")
    if resp.status_code != 200:
        raise HTTPException(status_code=422, detail=f"Gemini API error {resp.status_code}: {resp.text[:300]}")
    data = resp.json()
    try:
        rparts = data["candidates"][0]["content"]["parts"]
        return "".join(p.get("text", "") for p in rparts if isinstance(p, dict))
    except (KeyError, IndexError, TypeError):
        raise HTTPException(status_code=422, detail="Gemini returned an empty response.")


# ── Groq vision — single image ────────────────────────────────────────────────
async def _groq_vision_raw(image_b64: str, mime_type: str, prompt: str) -> str:
    import httpx
    groq_key = os.environ.get("GROQ_API_KEY", "")
    if not groq_key:
        raise HTTPException(status_code=500, detail="GROQ_API_KEY is not configured on the server.")

    payload = {
        "model": os.environ.get("GROQ_VISION_MODEL", "meta-llama/llama-4-maverick-17b-128e-instruct"),
        "messages": [{
            "role": "user",
            "content": [
                {"type": "image_url", "image_url": {"url": f"data:{mime_type};base64,{image_b64}"}},
                {"type": "text", "text": prompt},
            ],
        }],
        "max_tokens": 2048,
    }
    async with httpx.AsyncClient(timeout=60) as client:
        resp = await client.post(
            "https://api.groq.com/openai/v1/chat/completions",
            headers={"Authorization": f"Bearer {groq_key}", "Content-Type": "application/json"},
            json=payload,
        )
    if resp.status_code == 429:
        raise HTTPException(status_code=429, detail="Groq quota exceeded. Please wait a moment and try again.")
    if resp.status_code != 200:
        raise HTTPException(status_code=422, detail=f"Groq API error {resp.status_code}: {resp.text[:300]}")
    return resp.json()["choices"][0]["message"]["content"]


async def _groq_vision_multipage_raw(page_images_b64: list, prompt: str) -> str:
    import httpx
    groq_key = os.environ.get("GROQ_API_KEY", "")
    if not groq_key:
        raise HTTPException(status_code=500, detail="GROQ_API_KEY is not configured on the server.")

    content = []
    for img_b64, mime in page_images_b64:
        content.append({"type": "image_url", "image_url": {"url": f"data:{mime};base64,{img_b64}"}})
    content.append({"type": "text", "text": prompt})

    payload = {
        "model": os.environ.get("GROQ_VISION_MODEL", "meta-llama/llama-4-maverick-17b-128e-instruct"),
        "messages": [{"role": "user", "content": content}],
        "max_tokens": 2048,
    }
    async with httpx.AsyncClient(timeout=90) as client:
        resp = await client.post(
            "https://api.groq.com/openai/v1/chat/completions",
            headers={"Authorization": f"Bearer {groq_key}", "Content-Type": "application/json"},
            json=payload,
        )
    if resp.status_code == 429:
        raise HTTPException(status_code=429, detail="Groq quota exceeded. Please wait a moment and try again.")
    if resp.status_code != 200:
        raise HTTPException(status_code=422, detail=f"Groq API error {resp.status_code}: {resp.text[:300]}")
    return resp.json()["choices"][0]["message"]["content"]


# ── Provider-agnostic vision wrappers ────────────────────────────────────────
async def _groq_vision(image_b64: str, mime_type: str, prompt: str) -> str:
    """Vision call — prefers Gemini when configured, falls back to Groq.
    Named `_groq_vision` for backwards compatibility with existing callers."""
    if _provider() == "gemini":
        try:
            return await _gemini_vision(image_b64, mime_type, prompt)
        except HTTPException as e:
            if os.environ.get("GROQ_API_KEY") and e.status_code in (422, 429, 500):
                return await _groq_vision_raw(image_b64, mime_type, prompt)
            raise
    return await _groq_vision_raw(image_b64, mime_type, prompt)


async def _groq_vision_multipage(page_images_b64: list, prompt: str) -> str:
    if _provider() == "gemini":
        try:
            return await _gemini_vision_multipage(page_images_b64, prompt)
        except HTTPException as e:
            if os.environ.get("GROQ_API_KEY") and e.status_code in (422, 429, 500):
                return await _groq_vision_multipage_raw(page_images_b64, prompt)
            raise
    return await _groq_vision_multipage_raw(page_images_b64, prompt)


# ── Main route ────────────────────────────────────────────────────────────────
@router.post("/analyze-document")
async def analyze_document(
    file: UploadFile = File(...),
    current_user=Depends(get_current_user)
):
    contents = await file.read()
    filename = file.filename or "uploaded_file"

    async def _execute_existing_reader():
        ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""

        # ── Excel (.xlsx / .xlsm) → Gemini ───────────────────────────────────────
        if ext in ("xlsx", "xlsm"):
            try:
                import openpyxl
                wb = openpyxl.load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
                lines = []
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    lines.append(f"\n--- Sheet: {sheet_name} ---")
                    for row in ws.iter_rows(values_only=True):
                        row_text = "\t".join("" if v is None else str(v) for v in row)
                        if row_text.strip():
                            lines.append(row_text)
                text_content = "\n".join(lines)
            except Exception as e:
                raise HTTPException(status_code=422, detail=f"Could not parse Excel file: {e}")

            model = _get_gemini_model()
            prompt = (
                "You are a financial document analyst. Below is spreadsheet data.\n"
                "Please provide:\n"
                "1. A clear summary of what this spreadsheet contains\n"
                "2. All key figures, totals, and important numbers\n"
                "3. Any trends or patterns you notice\n"
                "4. Actionable insights or observations\n\n"
                f"{text_content[:40000]}"
            )
            response = await model.generate_content_async(prompt)
            return {"filename": filename, "analysis": response.text}

        # ── Excel (.xls) → Gemini ─────────────────────────────────────────────────
        if ext == "xls":
            try:
                import pandas as pd
                df = pd.read_excel(io.BytesIO(contents), engine="xlrd")
                text_content = df.to_string(index=False)
            except Exception as e:
                raise HTTPException(status_code=422, detail=f"Could not parse .xls file: {e}")

            model = _get_gemini_model()
            prompt = (
                "You are a data analyst. Below is spreadsheet data.\n"
                "Summarise it clearly: key columns, totals, patterns, and insights.\n\n"
                f"{text_content[:40000]}"
            )
            response = await model.generate_content_async(prompt)
            return {"filename": filename, "analysis": response.text}

        # ── CSV → Gemini ──────────────────────────────────────────────────────────
        if ext == "csv":
            try:
                text_content = contents.decode("utf-8", errors="replace")
            except Exception as e:
                raise HTTPException(status_code=422, detail=f"Could not read CSV: {e}")

            model = _get_gemini_model()
            prompt = (
                "You are a data analyst. Below is CSV data.\n"
                "Summarise it: key columns, totals, patterns, and insights.\n\n"
                f"{text_content[:40000]}"
            )
            response = await model.generate_content_async(prompt)
            return {"filename": filename, "analysis": response.text}

        # ── PDF → Gemini (text PDF) or Groq (scanned PDF) ────────────────────────
        if ext == "pdf":
            try:
                import pdfplumber
                extracted_pages = []
                with pdfplumber.open(io.BytesIO(contents)) as pdf:
                    for i, page in enumerate(pdf.pages[:30]):
                        text = page.extract_text()
                        if text and text.strip():
                            extracted_pages.append(f"--- Page {i+1} ---\n{text.strip()}")
                text_content = "\n\n".join(extracted_pages)
            except Exception as e:
                raise HTTPException(status_code=422, detail=f"Could not read PDF: {e}")

            if text_content.strip():
                # Normal text-based PDF → Gemini
                model = _get_gemini_model()
                prompt = (
                    "You are a document analyst. Below is extracted text from a PDF document.\n"
                    "Please provide:\n"
                    "1. Document type and purpose\n"
                    "2. All key data, figures, totals, dates, and party names\n"
                    "3. A clear structured summary\n"
                    "4. Any important observations or anomalies\n\n"
                    f"{text_content[:40000]}"
                )
                response = await model.generate_content_async(prompt)
                return {"filename": filename, "analysis": response.text}
            else:
                # Scanned PDF (no text layer) → render pages as images → Vision.
                # Uses auto-batching (Groq max 3 images/request), parallel batches,
                # retry+split fallback, ordered merge, streaming memory release.
                try:
                    from PIL import Image as PILImage
                    page_images_b64 = []
                    with pdfplumber.open(io.BytesIO(contents)) as pdf:
                        for page in pdf.pages:  # unlimited — batching handles it
                            pil_img = page.to_image(resolution=150).original
                            if pil_img.mode not in ("RGB", "L"):
                                pil_img = pil_img.convert("RGB")
                            buf = io.BytesIO()
                            pil_img.save(buf, format="JPEG", quality=85)
                            img_b64 = base64.b64encode(buf.getvalue()).decode()
                            page_images_b64.append((img_b64, "image/jpeg"))
                            # Release PIL/buffer refs promptly
                            del pil_img, buf

                    if not page_images_b64:
                        raise HTTPException(status_code=422, detail="No pages could be rendered from this PDF.")

                    prompt = (
                        "You are a document analyst. These are pages from a scanned PDF document.\n"
                        "Please read carefully and provide:\n"
                        "1. Document type and purpose\n"
                        "2. All key data, figures, dates, and names\n"
                        "3. A structured summary\n"
                        "4. Important observations"
                    )
                    # Prefer Gemini multi-page if configured (no 3-image cap);
                    # otherwise use Groq auto-batched pipeline.
                    if _provider() == "gemini":
                        try:
                            analysis = await _gemini_vision_multipage(page_images_b64, prompt)
                            return {"filename": filename, "analysis": analysis}
                        except HTTPException as e:
                            if not os.environ.get("GROQ_API_KEY") or e.status_code not in (422, 429, 500):
                                raise
                    analysis = await _groq_vision_batched_pages(page_images_b64, prompt)
                    return {"filename": filename, "analysis": analysis}
                except HTTPException:
                    raise
                except Exception as e:
                    raise HTTPException(status_code=422, detail=f"Scanned PDF could not be processed: {e}")

        # ── Images (JPG, PNG, WEBP) → Groq vision ────────────────────────────────
        if ext in ("jpg", "jpeg", "png", "webp", "gif"):
            try:
                from PIL import Image as PILImage
                img = PILImage.open(io.BytesIO(contents))
                if img.mode not in ("RGB", "L"):
                    img = img.convert("RGB")
                buf = io.BytesIO()
                img.save(buf, format="JPEG", quality=85)
                img_b64 = base64.b64encode(buf.getvalue()).decode()
            except Exception as e:
                raise HTTPException(status_code=422, detail=f"Could not open image: {e}")

            prompt = (
                "You are a document analyst. Look at this image carefully.\n"
                "Please provide:\n"
                "1. What type of document or image this is\n"
                "2. All key data, numbers, dates, names, and amounts visible\n"
                "3. A structured summary of all important information\n"
                "4. Any observations or red flags"
            )
            analysis = await _groq_vision(img_b64, "image/jpeg", prompt)
            return {"filename": filename, "analysis": analysis}

        raise HTTPException(
            status_code=415,
            detail=f"Unsupported file type '.{ext}'. Supported: PDF, Excel (.xlsx/.xls), CSV, JPG, PNG, WEBP."
        )

    from backend.ai import ai_router
    # Extract raw OCR text via the enterprise OCR pipeline
    raw_ocr_text = await ai_router.process_ocr(contents, filename)
    # Classify document
    classification = await ai_router.classify_document(contents, filename, raw_ocr_text)
    # Process document which runs classification, saves classification results, and orchestrates the AI extraction
    return await ai_router.process_document(contents, filename, current_user, _execute_existing_reader)
