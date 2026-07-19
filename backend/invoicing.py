import uuid
import sqlite3
import logging
import re
import asyncio
import time
import base64
import tempfile
import os
import smtplib
import json
import csv

from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication

from datetime import datetime, timezone, date, timedelta
from io import BytesIO, StringIO
from typing import List, Optional, Literal, Any, Dict

from fastapi import (
    APIRouter,
    Depends,
    HTTPException,
    Query,
    status,
    BackgroundTasks,
    UploadFile,
    File,
    Form,
)
from fastapi.responses import StreamingResponse

from pydantic import BaseModel, Field, field_validator

from backend.dependencies import db, get_current_user, check_module_permission
from backend.models import User

# ✅ Google imports (clean)
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseUpload


# ═══════════════════════════════════════════════════════════
# OPTIONAL LIBRARIES (AUTO-INSTALL)
# ═══════════════════════════════════════════════════════════

try:
    from fpdf import FPDF
except ImportError:
    import subprocess, sys
    subprocess.check_call([sys.executable, "-m", "pip", "install", "fpdf2"])
    from fpdf import FPDF

try:
    import openpyxl
except ImportError:
    import subprocess, sys
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl

try:
    import xml.etree.ElementTree as ET
except ImportError:
    ET = None


# ═══════════════════════════════════════════════════════════
# LOGGER + ROUTER
# ═══════════════════════════════════════════════════════════

logger = logging.getLogger(__name__)
router = APIRouter(tags=["Invoicing"])


# ═══════════════════════════════════════════════════════════
# DRIVE FOLDERS (MAKE SURE THESE ARE ACCESSIBLE)
# ═══════════════════════════════════════════════════════════

DRIVE_FOLDERS = {
    "invoices":     "1NhadvUmWtZ8x37FrJ2oeKTCOvHVyCyPv",
    "payments":     "1VPtuX6u_L-WPfLk0ZTawHrsyrXSMBfGu",
    "credit_notes": "1vY1mJexT-NJso6U1HLBeKaOgI6IFw9nc",
    "backups":      "1pWNDV2Yym3mvWYDQ9WmUiqrmqndT-Z9q",
}


# ═══════════════════════════════════════════════════════════
# DRIVE SCOPES
# ═══════════════════════════════════════════════════════════

DRIVE_SCOPES = [
    "https://www.googleapis.com/auth/drive",
    "https://www.googleapis.com/auth/drive.file",
    "https://www.googleapis.com/auth/drive.metadata.readonly",
]


def _get_drive_refresh_token() -> str | None:
    """
    Return the Google Drive refresh token from:
      1. GOOGLE_REFRESH_TOKEN env var (fastest — checked first)
      2. MongoDB app_settings collection (set by the frontend Connect flow)
    Caches the DB token into os.environ so subsequent calls in the same
    process skip the DB round-trip.
    """
    env_token = os.getenv("GOOGLE_REFRESH_TOKEN")
    if env_token:
        return env_token

    # Lazy import to avoid circular dependency at module load time
    try:
        import asyncio
        from backend.dependencies import db as _db

        async def _fetch():
            doc = await _db["app_settings"].find_one({"_id": "google_drive"})
            if doc and doc.get("connected") and doc.get("refresh_token"):
                return doc["refresh_token"]
            return None

        # Run synchronously — safe to call from sync helper functions
        # NOTE: pool.submit must receive a callable, not a coroutine.
        # We pass a lambda so asyncio.run() is called inside the new thread.
        try:
            import concurrent.futures
            with concurrent.futures.ThreadPoolExecutor(max_workers=1) as pool:
                token = pool.submit(lambda: asyncio.run(_fetch())).result(timeout=5)
        except Exception:
            token = None

        if token:
            # Cache in env so the next call in this process is instant
            os.environ["GOOGLE_REFRESH_TOKEN"] = token
            return token
    except Exception as e:
        logger.warning(f"Could not read Drive token from DB: {e}")

    return None


def _drive_configured() -> bool:
    return bool(
        _get_drive_refresh_token() and
        os.getenv("GOOGLE_CLIENT_ID") and
        os.getenv("GOOGLE_CLIENT_SECRET")
    )


# ═══════════════════════════════════════════════════════════
# AUTH — OAuth (REFRESH TOKEN BASED)
# ═══════════════════════════════════════════════════════════


def _get_drive_service():
    """
    Build and return an authenticated Google Drive v3 service.
    Handles stale/cached env tokens by falling back to DB and
    auto-clears invalid tokens so the UI shows the correct status.
    """
    import google.auth.exceptions

    client_id     = os.getenv("GOOGLE_CLIENT_ID")
    client_secret = os.getenv("GOOGLE_CLIENT_SECRET")

    if not client_id or not client_secret:
        raise HTTPException(500, "Missing GOOGLE_CLIENT_ID or GOOGLE_CLIENT_SECRET")

    def _build_from_token(token: str):
        creds = Credentials(
            None,
            refresh_token=token,
            token_uri="https://oauth2.googleapis.com/token",
            client_id=client_id,
            client_secret=client_secret,
            scopes=DRIVE_SCOPES,
        )
        creds.refresh(Request())
        return build("drive", "v3", credentials=creds, cache_discovery=False)

    def _fetch_db_token_sync():
        import asyncio
        from backend.dependencies import db as _db

        async def _fetch():
            doc = await _db["app_settings"].find_one({"_id": "google_drive"})
            return doc.get("refresh_token") if (doc and doc.get("connected")) else None

        try:
            import concurrent.futures
            with concurrent.futures.ThreadPoolExecutor(max_workers=1) as pool:
                return pool.submit(lambda: asyncio.run(_fetch())).result(timeout=5)
        except Exception:
            return None

    def _mark_drive_disconnected_sync():
        import asyncio
        from backend.dependencies import db as _db

        async def _update():
            await _db["app_settings"].update_one(
                {"_id": "google_drive"},
                {"$set": {"connected": False, "refresh_token": None, "access_token": None}},
            )

        try:
            import concurrent.futures
            with concurrent.futures.ThreadPoolExecutor(max_workers=1) as pool:
                pool.submit(lambda: asyncio.run(_update())).result(timeout=5)
        except Exception:
            pass

    # Step 1: Try token from env-var cache or DB
    refresh_token = _get_drive_refresh_token()

    if not refresh_token:
        raise HTTPException(
            500,
            "Google Drive is not connected. Go to Settings -> Integrations -> Connect Google Drive.",
        )

    try:
        return _build_from_token(refresh_token)

    except google.auth.exceptions.RefreshError as e:
        err_msg = str(e).lower()

        # Both mean the same thing from the admin's point of view: the
        # cached refresh token no longer works with the currently configured
        # GOOGLE_CLIENT_ID/SECRET (invalid_grant = token was revoked/expired;
        # unauthorized_client = token was issued for a *different* OAuth
        # client than the one currently configured, e.g. after rotating
        # credentials in Google Cloud Console or Render's env vars).
        # Either way the fix is the same: reconnect Drive.
        is_stale_credential = "invalid_grant" in err_msg or "unauthorized_client" in err_msg

        if not is_stale_credential:
            logger.error(f"DRIVE REFRESH ERROR: {e}", exc_info=True)
            raise HTTPException(500, f"Google Drive auth failed: {str(e)}")

        # --- stale/mismatched credentials: cached token is unusable ---
        logger.warning(
            f"Google Drive refresh token rejected ({err_msg}). "
            "Clearing env cache and checking DB for a newer token."
        )

        # Clear the stale env-var cache
        os.environ.pop("GOOGLE_REFRESH_TOKEN", None)

        # Try fetching a fresh token directly from DB
        # (user may have reconnected via UI after the env var was cached)
        db_token = _fetch_db_token_sync()

        if db_token and db_token != refresh_token:
            try:
                svc = _build_from_token(db_token)
                # Cache the working token
                os.environ["GOOGLE_REFRESH_TOKEN"] = db_token
                return svc
            except Exception as db_err:
                logger.warning(f"DB token also failed: {db_err}")

        # Both tokens are invalid: mark Drive as disconnected in DB
        # so Settings page shows the correct disconnected state
        _mark_drive_disconnected_sync()

        raise HTTPException(
            500,
            "Google Drive's connection has stopped working (the saved token no "
            "longer matches this app's Google credentials). Please go to "
            "Settings -> General Settings -> Google Drive and click Reconnect "
            "to re-authorize.",
        )

    except HTTPException:
        raise

    except Exception as e:
        logger.error(f"DRIVE AUTH ERROR: {e}", exc_info=True)
        raise HTTPException(500, f"Google Drive auth failed: {str(e)}")

# ═══════════════════════════════════════════════════════════
# UPLOAD
# ═══════════════════════════════════════════════════════════

async def _upload_to_drive(
    content_bytes: bytes,
    filename: str,
    folder_key: str,
    mime_type: str,
    custom_parent_id: str = None
):
    """Upload bytes to Drive using OAuth"""

    try:
        service = _get_drive_service()

        parent_id = custom_parent_id if custom_parent_id else DRIVE_FOLDERS.get(folder_key)

        if not parent_id:
            raise HTTPException(500, f"Invalid folder key: {folder_key}")
        file_metadata = {
            "name": filename,
            "parents": [parent_id]
        }

        media = MediaIoBaseUpload(
            BytesIO(content_bytes),
            mimetype=mime_type,
            resumable=True
        )

        file = service.files().create(
            body=file_metadata,
            media_body=media,
            fields="id, webViewLink, name"
        ).execute()

        logger.info(f"Uploaded to Drive → {filename}")

        return file.get("webViewLink")

    except Exception as e:
        logger.error(f"❌ DRIVE UPLOAD FAILED: {e}", exc_info=True)
        raise HTTPException(
            status_code=500,
            detail=f"Drive upload failed: {str(e)}"
        )

# ═══════════════════════════════════════════════════════════
# CLIENT FOLDER (DYNAMIC)
# ═══════════════════════════════════════════════════════════

def _get_or_create_client_folder(
    client_name: str
) -> str:
    if not client_name or not client_name.strip():
        return DRIVE_FOLDERS["invoices"]

    try:
        service = _get_drive_service()

        cn = client_name.strip()

        query = (
            f"'{DRIVE_FOLDERS['invoices']}' in parents and name='{cn}' "
            f"and mimeType='application/vnd.google-apps.folder' and trashed=false"
        )

        results = service.files().list(
            q=query,
            fields="files(id,name)"
        ).execute()

        existing = results.get("files", [])

        if existing:
            return existing[0]["id"]

        folder_metadata = {
            "name": cn,
            "mimeType": "application/vnd.google-apps.folder",
            "parents": [DRIVE_FOLDERS["invoices"]],
        }

        folder = service.files().create(
            body=folder_metadata,
            fields="id"
        ).execute()

        return folder["id"]

    except Exception as e:
        logger.warning(f"Could not create Drive folder for '{client_name}': {e}")
        return DRIVE_FOLDERS["invoices"]

# ═══════════════════════════════════════════════════════════
# CONSTANTS
# ═══════════════════════════════════════════════════════════

GST_RATES = [0.0, 5.0, 12.0, 18.0, 28.0]
UNITS = ["service", "nos", "kg", "ltr", "mtr", "sqft", "hr", "day", "month", "year", "set", "lot", "pcs", "box"]
PAYMENT_MODES = ["cash", "cheque", "neft", "rtgs", "imps", "upi", "card", "other"]
INV_STATUS = ["draft", "sent", "partially_paid", "paid", "overdue", "cancelled", "credit_note"]
PURCHASE_STATUS = ["outstanding", "partially_paid", "paid", "cancelled"]

KB_TXN_TYPES = {
    1: "tax_invoice", 2: "purchase", 3: "payment_received", 4: "payment_made",
    7: "credit_note", 21: "estimate", 27: "delivery_challan", 65: "proforma",
}
KB_PAY_STATUS = {1: "sent", 2: "partially_paid", 3: "paid"}

# ═══════════════════════════════════════════════════════════
# AMOUNT IN WORDS
# ═══════════════════════════════════════════════════════════

_ONES = ["", "One", "Two", "Three", "Four", "Five", "Six", "Seven", "Eight", "Nine", "Ten",
         "Eleven", "Twelve", "Thirteen", "Fourteen", "Fifteen", "Sixteen", "Seventeen",
         "Eighteen", "Nineteen"]
_TENS = ["", "", "Twenty", "Thirty", "Forty", "Fifty", "Sixty", "Seventy", "Eighty", "Ninety"]


def _amount_in_words(n: float) -> str:
    try:
        rupees = int(n)
        paise = round((n - rupees) * 100)

        def _grp(num):
            if num == 0: return ""
            if num < 20: return _ONES[num] + " "
            if num < 100: return _TENS[num // 10] + (" " + _ONES[num % 10] if num % 10 else "") + " "
            return _ONES[num // 100] + " Hundred " + _grp(num % 100)

        def _convert(num):
            if num == 0: return "Zero "
            r = ""
            cr = num // 10_000_000; num %= 10_000_000
            lk = num // 100_000; num %= 100_000
            th = num // 1000; num %= 1000
            if cr: r += _grp(cr) + "Crore "
            if lk: r += _grp(lk) + "Lakh "
            if th: r += _grp(th) + "Thousand "
            r += _grp(num)
            return r

        r = _convert(rupees).strip()
        p = f" and {_convert(paise).strip()} Paise" if paise else ""
        return f"Rupees {r}{p} Only"
    except Exception:
        return f"Rupees {n:.2f} Only"


# ═══════════════════════════════════════════════════════════
# PERMISSION HELPER
# ═══════════════════════════════════════════════════════════

def _perm(user: User) -> bool:
    if user.role == "admin": return True
    perms = user.permissions if isinstance(user.permissions, dict) else (
        user.permissions.model_dump() if user.permissions else {})
    return bool(perms.get("can_create_quotations") or perms.get("can_manage_invoices"))


# ═══════════════════════════════════════════════════════════
# NEXT INVOICE NUMBER
# ═══════════════════════════════════════════════════════════

async def _next_invoice_no(
    prefix: str = "INV",
    company_id: str = None,
    separator: str = "/",
    include_fy: bool = True,
    fy_format: str = "short",
    include_month: bool = False,
    number_padding: int = 3,
    invoice_type: str = None,
) -> str:
    """
    Generate the next available invoice number using the exact format from
    Invoice Settings (prefix, separator, FY label, month, padding).
    MAX-based scan so deletions/renames never cause collisions.

    FIX: also scoped by ``invoice_type`` (when provided), not just by the
    text prefix. Previously the scan matched ANY document whose invoice_no
    happened to start with the given prefix, regardless of its actual
    invoice_type. Since the invoice_no field is user-editable, a Proforma
    or Estimate that was manually saved with a custom "INV/..." style
    number (instead of its own PRO/EST prefix) would silently inflate the
    Tax Invoice sequence for that company — the numbers would "exist" in
    the scan but never show up when filtering the list to Tax Invoice
    only. Scoping by invoice_type as well closes that gap.
    """
    today = date.today()
    fy_start = today.year if today.month >= 4 else today.year - 1

    if fy_format == "long":
        fy_label = f"{fy_start}-{fy_start + 1}"
    else:
        fy_label = f"{fy_start % 100:02d}-{(fy_start + 1) % 100:02d}"

    month_str = f"{today.month:02d}"
    sep = separator if separator and separator.lower() != "none" else ""

    # Build regex to scan existing invoices with this exact format
    # Pattern matches: prefix [sep fy]? [sep month]? sep NUMBER
    scan_parts = [re.escape(prefix)]
    if include_fy:
        scan_parts.append(re.escape(fy_label))
    if include_month:
        scan_parts.append(re.escape(month_str))
    scan_parts.append(r"(\d+)")
    pattern = re.escape(sep).join(scan_parts)
    pattern = f"^{pattern}$"

    query: dict = {"invoice_no": {"$regex": f"^{re.escape(prefix)}"}}
    if company_id:
        query["company_id"] = company_id
    if invoice_type:
        query["invoice_type"] = invoice_type

    cursor = db.invoices.find(query, {"_id": 0, "invoice_no": 1})
    max_seq = 0
    async for doc in cursor:
        m = re.match(pattern, doc.get("invoice_no", ""))
        if m:
            seq = int(m.group(1))
            if seq > max_seq:
                max_seq = seq

    def _build(seq: int) -> str:
        parts = [prefix]
        if include_fy:
            parts.append(fy_label)
        if include_month:
            parts.append(month_str)
        parts.append(str(seq).zfill(number_padding))
        return sep.join(parts)

    candidate_seq = max_seq + 1
    for _ in range(50):
        candidate = _build(candidate_seq)
        dup_filter: dict = {"invoice_no": candidate}
        if company_id:
            dup_filter["company_id"] = company_id
        taken = await db.invoices.find_one(dup_filter)
        if not taken:
            return candidate
        candidate_seq += 1

    return _build(candidate_seq)


# ═══════════════════════════════════════════════════════════
# EMAIL SENDER
# ═══════════════════════════════════════════════════════════

def _send_email(to_email: str, subject: str, html_body: str, pdf_bytes: bytes, filename: str, company_email: str):
    smtp_server = os.getenv("SMTP_SERVER", "smtp.gmail.com")
    smtp_port = int(os.getenv("SMTP_PORT", 587))
    smtp_user = os.getenv("SMTP_USER")
    smtp_password = os.getenv("SMTP_PASSWORD")
    from_email = company_email or os.getenv("DEFAULT_FROM_EMAIL", "noreply@yourcompany.com")
    if not smtp_user or not smtp_password:
        raise HTTPException(500, "Email service not configured")
    msg = MIMEMultipart()
    msg["From"] = from_email
    msg["To"] = to_email
    msg["Subject"] = subject
    msg.attach(MIMEText(html_body, "html"))
    attachment = MIMEApplication(pdf_bytes, _subtype="pdf")
    attachment.add_header("Content-Disposition", "attachment", filename=filename)
    msg.attach(attachment)
    try:
        server = smtplib.SMTP(smtp_server, smtp_port)
        server.starttls()
        server.login(smtp_user, smtp_password)
        server.sendmail(from_email, to_email, msg.as_string())
        server.quit()
    except Exception as e:
        raise HTTPException(500, f"Email sending failed: {e}")


# ═══════════════════════════════════════════════════════════
# SAFE PDF HELPERS
# ═══════════════════════════════════════════════════════════

def _s(v, maxl=0):
    if v is None: return ""
    t = str(v).encode("latin-1", errors="replace").decode("latin-1")
    return t[:maxl] if maxl and len(t) > maxl else t


def _cell(pdf, w, h, txt="", border=0, align="L", fill=False, nl=False):
    pdf.cell(w, h, txt, border, 1 if nl else 0, align, fill)


def _mcell(pdf, w, h, txt, border=0, align="L", fill=False):
    pdf.multi_cell(w, h, txt, border, align, fill)


def _embed_logo(pdf, logo_b64, x, y, h):
    if not logo_b64: return
    tmp_path = None
    try:
        raw = re.sub(r"^data:image/[^;]+;base64,", "", logo_b64)
        img_bytes = base64.b64decode(raw)
        suffix = ".png" if "png" in logo_b64[:30] else ".jpg"
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=suffix)
        tmp.write(img_bytes); tmp.close(); tmp_path = tmp.name
        pdf.image(tmp_path, x=x, y=y, h=h)
    except Exception as e:
        logger.warning(f"Logo embed failed: {e}")
    finally:
        if tmp_path and os.path.exists(tmp_path):
            try: os.unlink(tmp_path)
            except: pass


def _lighten(c, f=0.88): return tuple(int(x + (255 - x) * f) for x in c)
def _darken(c, f=0.65): return tuple(int(x * f) for x in c)


# FIX: helper to parse hex color strings like "#0D3B66" → (13, 59, 102)
def _hex_to_rgb(hex_color: str) -> tuple:
    """Convert a CSS hex color string to an (R, G, B) tuple."""
    try:
        h = hex_color.strip().lstrip("#")
        if len(h) == 3:
            h = "".join(c * 2 for c in h)
        return (int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16))
    except Exception:
        return (13, 59, 102)  # fallback to default navy


# ═══════════════════════════════════════════════════════════
# UNIVERSAL BACKUP PARSERS
# ═══════════════════════════════════════════════════════════

def _safe_float(val, default=0.0):
    if val is None: return default
    try: return float(val)
    except: return default


def _safe_str(val, default=""):
    if val is None: return default
    return str(val).strip()


def _safe_date(val, default=None):
    if not val: return default or date.today().isoformat()
    s = str(val).strip()
    if " " in s: s = s.split(" ")[0]
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try: return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except: pass
    return default or date.today().isoformat()


def _parse_vyp_file(file_path: str) -> dict:
    """Parse KhataBook .vyp / .vyb SQLite backup.

    .vyb files are ZIP archives that contain a single .vyp SQLite file inside.
    We detect this and extract the inner database before connecting.
    """
    import zipfile as _zipfile
    _tmp_vyp_path = None
    try:
        # .vyb = ZIP wrapper around a .vyp SQLite file
        if _zipfile.is_zipfile(file_path):
            with _zipfile.ZipFile(file_path, "r") as _zf:
                _inner = next(
                    (n for n in _zf.namelist() if n.lower().endswith((".vyp", ".db"))),
                    _zf.namelist()[0] if _zf.namelist() else None,
                )
                if not _inner:
                    raise HTTPException(400, "Empty .vyb archive — no database found inside.")
                import tempfile as _tempfile
                _tmp = _tempfile.NamedTemporaryFile(delete=False, suffix=".vyp")
                _tmp.write(_zf.read(_inner))
                _tmp.close()
                _tmp_vyp_path = _tmp.name
            file_path = _tmp_vyp_path
    except HTTPException:
        raise
    except Exception as _e:
        raise HTTPException(400, f"Failed to unpack .vyb archive: {_e}")

    try:
        conn = sqlite3.connect(file_path)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        result = {"source": "khatabook", "source_label": "KhataBook (.vyp / .vyb)",
                  "firms": [], "clients": [], "items": [], "invoices": [], "payments": [], "stats": {}}

        try:
            cursor.execute("SELECT * FROM kb_firms")
            for row in cursor.fetchall():
                result["firms"].append({
                    "firm_id": row["firm_id"], "firm_name": _safe_str(row["firm_name"]),
                    "firm_email": _safe_str(row["firm_email"]), "firm_phone": _safe_str(row["firm_phone"]),
                    "firm_address": _safe_str(row["firm_address"]), "firm_gstin": _safe_str(row["firm_gstin_number"]),
                    "firm_state": _safe_str(row["firm_state"]),
                })
        except Exception as e:
            logger.warning(f"VYP firms: {e}")

        tax_map = {}
        try:
            cursor.execute("SELECT tax_code_id, tax_rate, tax_code_type FROM kb_tax_code")
            for row in cursor.fetchall():
                tax_map[row["tax_code_id"]] = {"rate": _safe_float(row["tax_rate"]), "type": row["tax_code_type"]}
        except Exception as e:
            logger.warning(f"VYP tax: {e}")

        name_map = {}
        try:
            cursor.execute("""SELECT name_id, full_name, phone_number, email, address,
                              name_gstin_number, name_state FROM kb_names WHERE name_type IN (1, 2)""")
            for row in cursor.fetchall():
                c = {"name_id": row["name_id"], "full_name": _safe_str(row["full_name"]),
                     "phone_number": _safe_str(row["phone_number"]), "email": _safe_str(row["email"]),
                     "address": _safe_str(row["address"]), "name_gstin_number": _safe_str(row["name_gstin_number"]),
                     "name_state": _safe_str(row["name_state"])}
                name_map[row["name_id"]] = c
                result["clients"].append(c)
        except Exception as e:
            logger.warning(f"VYP clients: {e}")

        item_map = {}
        try:
            cursor.execute("""SELECT item_id, item_name, item_sale_unit_price, item_purchase_unit_price,
                              item_hsn_sac_code, item_tax_id, item_code, item_description, item_stock_quantity
                              FROM kb_items WHERE item_is_active = 1 OR item_is_active IS NULL""")
            for row in cursor.fetchall():
                gst_rate = 0.0
                tid = row["item_tax_id"]
                if tid and tid in tax_map:
                    tc = tax_map[tid]
                    gst_rate = tc["rate"] if tc["type"] == 1 else tc["rate"] * 2
                item = {"item_id": row["item_id"], "name": _safe_str(row["item_name"]),
                        "sale_price": _safe_float(row["item_sale_unit_price"]),
                        "hsn_sac": _safe_str(row["item_hsn_sac_code"]), "gst_rate": gst_rate}
                item_map[row["item_id"]] = item
                result["items"].append(item)
        except Exception as e:
            logger.warning(f"VYP items: {e}")

        lineitem_map = {}
        try:
            cursor.execute("""SELECT lineitem_id, lineitem_txn_id, item_id, quantity,
                              priceperunit, total_amount, lineitem_tax_amount,
                              lineitem_discount_amount, lineitem_description, lineitem_discount_percent
                              FROM kb_lineitems""")
            for row in cursor.fetchall():
                txn_id = row["lineitem_txn_id"]
                if txn_id not in lineitem_map: lineitem_map[txn_id] = []
                item_data = item_map.get(row["item_id"], {})
                lineitem_map[txn_id].append({
                    "description": _safe_str(row["lineitem_description"]) or item_data.get("name", "Item"),
                    "hsn_sac": item_data.get("hsn_sac", ""),
                    "quantity": _safe_float(row["quantity"], 1),
                    "unit_price": _safe_float(row["priceperunit"]),
                    "total_amount": _safe_float(row["total_amount"]),
                    "tax_amount": _safe_float(row["lineitem_tax_amount"]),
                    "discount_amount": _safe_float(row["lineitem_discount_amount"]),
                    "discount_pct": _safe_float(row["lineitem_discount_percent"]),
                    "gst_rate": item_data.get("gst_rate", 18.0),
                })
        except Exception as e:
            logger.warning(f"VYP lineitems: {e}")

        _seen_invoice_nos: set = set()  # guards against same-firm/year/type/num collisions
        try:
            cursor.execute("""SELECT txn_id, txn_date, txn_name_id, txn_cash_amount, txn_balance_amount,
                              txn_type, txn_discount_percent, txn_tax_percent, txn_discount_amount,
                              txn_tax_amount, txn_due_date, txn_description, txn_ref_number_char,
                              txn_payment_status, txn_firm_id, txn_invoice_prefix, txn_billing_address,
                              txn_round_off_amount, txn_po_ref_number
                              FROM kb_transactions WHERE txn_status != 0 ORDER BY txn_date ASC""")
            for row in cursor.fetchall():
                txn_type_code = row["txn_type"]
                txn_type = KB_TXN_TYPES.get(txn_type_code)
                if not txn_type: continue
                client = name_map.get(row["txn_name_id"], {})
                txn_date = _safe_date(row["txn_date"])
                due_date = _safe_date(row["txn_due_date"], txn_date)
                prefix = _safe_str(row["txn_invoice_prefix"]) or "KB"
                ref_num = _safe_str(row["txn_ref_number_char"])
                # Build a collision-proof invoice number:
                #   {prefix}-{firm_id}-{fiscal_year}-{doc_type}-{ref_num}
                # This ensures the same sequential number issued by different
                # firms, in different years, or as different document types
                # (invoice vs estimate vs delivery challan) never collides.
                # e.g.  KB-1-2021-INV-52  ≠  KB-4-2022-INV-52  ≠  KB-1-2021-EST-52
                _fy   = txn_date[:4] if txn_date and len(txn_date) >= 4 else "00"
                _firm = str(row["txn_firm_id"]) if row["txn_firm_id"] else "0"
                _type_code = {
                    "tax_invoice":       "INV",
                    "credit_note":       "CN",
                    "estimate":          "EST",
                    "proforma":          "PRO",
                    "delivery_challan":  "DC",
                }.get(txn_type, "TXN")
                invoice_no = (
                    f"{prefix}-{_firm}-{_fy}-{_type_code}-{ref_num}"
                    if ref_num
                    else f"KB-{row['txn_id']}"
                )
                # Last-resort tie-breaker: if two records in the SAME firm/year/type
                # genuinely share a reference number (data-entry error in Vyapar),
                # append the txn_id so neither record is lost during import.
                if invoice_no in _seen_invoice_nos:
                    invoice_no = f"{invoice_no}-{row['txn_id']}"
                _seen_invoice_nos.add(invoice_no)
                items_list = lineitem_map.get(row["txn_id"], [])
                subtotal = sum(li["total_amount"] for li in items_list)
                total_tax = sum(li["tax_amount"] for li in items_list)
                total_discount = sum(li["discount_amount"] for li in items_list)
                if not items_list and _safe_float(row["txn_cash_amount"]) > 0:
                    cash_amt = _safe_float(row["txn_cash_amount"])
                    tax_amt = _safe_float(row["txn_tax_amount"])
                    subtotal = cash_amt
                    total_tax = tax_amt
                    items_list = [{"description": _safe_str(row["txn_description"]) or "Imported",
                                   "hsn_sac": "", "quantity": 1, "unit_price": cash_amt,
                                   "total_amount": cash_amt + tax_amt, "tax_amount": tax_amt,
                                   "discount_amount": _safe_float(row["txn_discount_amount"]),
                                   "discount_pct": _safe_float(row["txn_discount_percent"]),
                                   "gst_rate": _safe_float(row["txn_tax_percent"]) or 18.0}]
                grand_total = subtotal + total_tax - total_discount + _safe_float(row["txn_round_off_amount"])
                if txn_type == "payment_received":
                    result["payments"].append({"_kb_id": row["txn_id"],
                        "client_name": client.get("full_name", "Unknown"),
                        "amount": _safe_float(row["txn_cash_amount"]), "payment_date": txn_date,
                        "payment_mode": "other", "company_id": row["txn_firm_id"]})
                    continue
                if txn_type in ("payment_made", "purchase"): continue
                inv_type = {"credit_note": "credit_note", "estimate": "estimate", "proforma": "proforma"}.get(txn_type, "tax_invoice")
                pay_status = KB_PAY_STATUS.get(row["txn_payment_status"], "draft")
                firm = next((f for f in result["firms"] if f["firm_id"] == row["txn_firm_id"]), None)
                is_interstate = False
                if firm and client.get("name_state"):
                    is_interstate = firm.get("firm_state", "").lower() != client.get("name_state", "").lower()
                formatted_items = []
                for li in items_list:
                    gst_rate = li.get("gst_rate", 18.0) or 18.0
                    taxable = li["total_amount"] - li.get("tax_amount", 0)
                    if taxable <= 0: taxable = li["total_amount"]
                    half = gst_rate / 2
                    if is_interstate:
                        igst = round(taxable * gst_rate / 100, 2) if li.get("tax_amount", 0) == 0 else li["tax_amount"]
                        formatted_items.append({"description": li["description"], "hsn_sac": li.get("hsn_sac", ""),
                            "quantity": li["quantity"], "unit": "service", "unit_price": li["unit_price"],
                            "discount_pct": li.get("discount_pct", 0), "gst_rate": gst_rate,
                            "taxable_value": taxable, "cgst_rate": 0, "sgst_rate": 0, "igst_rate": gst_rate,
                            "cgst_amount": 0, "sgst_amount": 0, "igst_amount": igst, "total_amount": taxable + igst})
                    else:
                        tax_half = round(li["tax_amount"] / 2, 2) if li.get("tax_amount", 0) else round(taxable * half / 100, 2)
                        formatted_items.append({"description": li["description"], "hsn_sac": li.get("hsn_sac", ""),
                            "quantity": li["quantity"], "unit": "service", "unit_price": li["unit_price"],
                            "discount_pct": li.get("discount_pct", 0), "gst_rate": gst_rate,
                            "taxable_value": taxable, "cgst_rate": half, "sgst_rate": half, "igst_rate": 0,
                            "cgst_amount": tax_half, "sgst_amount": tax_half, "igst_amount": 0,
                            "total_amount": taxable + tax_half * 2})
                total_taxable = sum(i["taxable_value"] for i in formatted_items)
                total_cgst = sum(i["cgst_amount"] for i in formatted_items)
                total_sgst = sum(i["sgst_amount"] for i in formatted_items)
                total_igst = sum(i["igst_amount"] for i in formatted_items)
                total_gst = total_cgst + total_sgst + total_igst
                computed_grand = round(total_taxable + total_gst, 2)
                balance_amt = _safe_float(row["txn_balance_amount"])
                amount_paid = max(0, computed_grand - balance_amt) if balance_amt >= 0 else computed_grand
                result["invoices"].append({
                    "_kb_id": row["txn_id"], "invoice_type": inv_type, "invoice_no": invoice_no,
                    "invoice_date": txn_date, "due_date": due_date,
                    "client_name": client.get("full_name", "Unknown"),
                    "client_email": client.get("email", ""), "client_phone": client.get("phone_number", ""),
                    "client_gstin": client.get("name_gstin_number", ""),
                    "client_address": _safe_str(row["txn_billing_address"]) or client.get("address", ""),
                    "client_state": client.get("name_state", ""), "is_interstate": is_interstate,
                    "items": formatted_items,
                    "subtotal": round(subtotal, 2), "total_discount": round(total_discount, 2),
                    "total_taxable": round(total_taxable, 2), "total_cgst": round(total_cgst, 2),
                    "total_sgst": round(total_sgst, 2), "total_igst": round(total_igst, 2),
                    "total_gst": round(total_gst, 2),
                    "grand_total": computed_grand if computed_grand > 0 else round(grand_total, 2),
                    "amount_paid": round(amount_paid, 2),
                    "amount_due": round(balance_amt if balance_amt >= 0 else 0, 2),
                    "status": pay_status, "payment_terms": "Imported from KhataBook",
                    "notes": _safe_str(row["txn_description"]),
                    "reference_no": _safe_str(row["txn_po_ref_number"]),
                    "company_id": row["txn_firm_id"],
                })
        except Exception as e:
            logger.error(f"VYP transactions: {e}", exc_info=True)

        result["stats"] = {k: len(result[k]) for k in ["firms", "clients", "items", "invoices", "payments"]}
        conn.close()
        return result
    except Exception as e:
        raise HTTPException(400, f"Failed to parse .vyp/.vyb file: {e}")
    finally:
        # Clean up the temp .vyp extracted from a .vyb ZIP, if any
        if _tmp_vyp_path:
            try:
                import os as _os
                _os.unlink(_tmp_vyp_path)
            except Exception:
                pass


def _parse_excel_file(file_path: str, filename: str) -> dict:
    """
    Multi-format Excel/CSV parser.  Auto-detects and handles:
      1. Vyapar / KhataBook / Billing-software Sale Report  (Generated On header)
      2. Tally DayBook Export  (Date | Particulars | Voucher Type | Debit | Credit)
      3. GSTR B2B Invoice Report / B2BInvoices CSV  (Supplier GSTIN, Party Name, Inv. No.)
      4. Invoices Excel List  (Party Name, Invoice Number, GSTIN, CGST, SGST, IGST …)
      5. Marg / Busy / Zoho-style sales reports  (flexible column mapping)
      6. Taskosphere / standard invoice template  (Client Name, Rate, GST Rate …)
    """
    result = {"source": "excel", "source_label": f"Excel/CSV ({filename})",
              "firms": [], "clients": [], "items": [], "invoices": [], "payments": [], "stats": {}}

    fn_lower = filename.lower()

    # ── helpers ──────────────────────────────────────────────────────────────
    def _sf(v, default=0.0):
        return _safe_float(str(v).replace(",", "") if v is not None else "", default)

    def _col_idx(headers_list, *keys):
        """Return first header index matching any key (case-insensitive, partial)."""
        hl = [str(h or "").lower().strip() for h in headers_list]
        for key in keys:
            kl = key.lower()
            for i, h in enumerate(hl):
                if kl in h:
                    return i
        return -1

    def _get_dict(row_dict, *keys, default=""):
        """Fuzzy-match key in a dict row."""
        for k in keys:
            for rk in row_dict.keys():
                if k.lower().replace(" ", "") in rk.lower().replace(" ", ""):
                    v = row_dict[rk]
                    if v is not None and str(v).strip():
                        return str(v).strip()
        return default

    def _build_invoice(client_name, inv_no, inv_date, taxable, cgst, sgst, igst,
                       gst_rate=18, client_gstin="", client_phone="", client_email="",
                       client_addr="", client_state="", reference_no="", notes="",
                       payment_terms="Due on receipt", inv_type="tax_invoice",
                       amount_paid=0.0, status="draft"):
        gst_total = round(cgst + sgst + igst, 2)
        grand     = round(taxable + gst_total, 2)
        due_amt   = round(grand - amount_paid, 2)
        is_inter  = igst > 0
        half      = gst_rate / 2
        due_dt    = (_safe_date(None, (date.today() + timedelta(days=30)).isoformat()))
        try:
            due_dt = (datetime.strptime(inv_date, "%Y-%m-%d") + timedelta(days=30)).strftime("%Y-%m-%d")
        except Exception:
            pass
        return {
            "invoice_type": inv_type, "invoice_no": inv_no, "invoice_date": inv_date,
            "due_date": due_dt, "client_name": client_name,
            "client_email": client_email, "client_phone": client_phone,
            "client_gstin": client_gstin, "client_address": client_addr,
            "client_state": client_state, "is_interstate": is_inter,
            "reference_no": reference_no, "notes": notes, "payment_terms": payment_terms,
            "items": [{
                "description": notes or f"Sale - {inv_no}",
                "hsn_sac": "", "quantity": 1, "unit": "service",
                "unit_price": taxable, "discount_pct": 0, "gst_rate": gst_rate,
                "taxable_value": round(taxable, 2),
                "cgst_rate": 0 if is_inter else half,
                "sgst_rate": 0 if is_inter else half,
                "igst_rate": gst_rate if is_inter else 0,
                "cgst_amount": 0 if is_inter else cgst,
                "sgst_amount": 0 if is_inter else sgst,
                "igst_amount": igst if is_inter else 0,
                "total_amount": grand,
            }],
            "subtotal": round(taxable, 2), "total_taxable": round(taxable, 2),
            "total_cgst": round(0 if is_inter else cgst, 2),
            "total_sgst": round(0 if is_inter else sgst, 2),
            "total_igst": round(igst if is_inter else 0, 2),
            "total_gst": gst_total, "grand_total": grand,
            "amount_paid": round(amount_paid, 2), "amount_due": due_amt, "status": status,
        }

    # ── load raw rows as a list-of-lists ────────────────────────────────────
    try:
        if fn_lower.endswith(".csv"):
            with open(file_path, "r", encoding="utf-8-sig", errors="replace") as f:
                raw_text = f.read()
            import io as _io
            reader = csv.reader(_io.StringIO(raw_text))
            all_rows = [r for r in reader]
        else:
            engine = "xlrd" if fn_lower.endswith(".xls") else "openpyxl"
            try:
                import pandas as pd
                xl = pd.ExcelFile(file_path, engine=engine)
                # Use first sheet that looks relevant
                sheet = next(
                    (s for s in xl.sheet_names if any(
                        kw in str(s).lower()
                        for kw in ["b2b", "invoice", "sale", "purchase", "daybook", "day book", "ledger"]
                    )),
                    xl.sheet_names[0],
                )
                df = pd.read_excel(xl, sheet_name=sheet, header=None, dtype=str)
                df = df.fillna("")
                all_rows = df.values.tolist()
            except Exception:
                wb = openpyxl.load_workbook(file_path, data_only=True)
                ws = wb.active
                all_rows = [[str(c if c is not None else "") for c in row] for row in ws.iter_rows(values_only=True)]
    except Exception as e:
        raise HTTPException(400, f"Cannot open file '{filename}': {e}")

    # Flatten to clean strings
    def _cell(row, idx, default=""):
        if idx < 0 or idx >= len(row): return default
        v = row[idx]
        return str(v).strip() if v is not None else default

    # ── FORMAT 1: Vyapar / KhataBook / Billing Sale Report ──────────────────
    # Signature: one of first 3 rows contains "Generated On" or "Sale Report"
    is_sale_report = any(
        ("generated on" in str(r[0]).lower() or "sale report" in " ".join(str(c) for c in r[:3]).lower())
        for r in all_rows[:5] if r and r[0]
    )
    if is_sale_report:
        hi = next((i for i, r in enumerate(all_rows) if r and str(r[0]).strip().lower() in ("date", "voucher date")), 2)
        for r in all_rows[hi + 1:]:
            if not r or not str(r[0]).strip(): continue
            raw_date = _cell(r, 0); order_no = _cell(r, 1); inv_no_raw = _cell(r, 2)
            party   = _cell(r, 3) or "Unknown"; gstin_v = _cell(r, 4); phone_v = _cell(r, 5)
            txn_type= _cell(r, 6).lower(); total = _sf(r[7] if len(r) > 7 else "")
            pay_type= _cell(r, 8); received = _sf(r[9] if len(r) > 9 else ""); balance = _sf(r[10] if len(r) > 10 else "")
            pay_sts = _cell(r, 11).lower(); desc_v = _cell(r, 12)
            if total <= 0: continue
            inv_date = date.today().isoformat()
            for sep in ("/", "-"):
                if sep in raw_date:
                    pts = raw_date.split(sep)
                    try:
                        yr = pts[2] if len(pts[2]) == 4 else "20" + pts[2]
                        inv_date = f"{yr}-{pts[1].zfill(2)}-{pts[0].zfill(2)}"
                        break
                    except Exception:
                        pass
            status = "paid" if pay_sts == "paid" else ("partially_paid" if received > 0 and balance > 0 else ("sent" if balance > 0 else "draft"))
            inv_type = "credit_note" if "credit" in txn_type else "tax_invoice"
            taxable  = round(total / 1.18, 2); gst_amt = round(total - taxable, 2)
            result["invoices"].append(_build_invoice(
                party, f"SR-{inv_no_raw or len(result['invoices'])+1:04d}",
                inv_date, taxable, round(gst_amt/2,2), round(gst_amt/2,2), 0,
                18, gstin_v, phone_v, "", "", "", order_no, desc_v or pay_type,
                pay_type or "Due on receipt", inv_type, received, status,
            ))
        result["stats"] = {"firms": 0, "clients": len({i["client_name"] for i in result["invoices"]}),
                           "items": 0, "invoices": len(result["invoices"]), "payments": 0}
        return result

    # ── Locate header row (first row with ≥3 non-empty cells) ────────────────
    def _find_header(rows, keywords=None):
        kw = keywords or ["gstin", "invoice", "party", "date", "amount", "debit", "credit", "particulars"]
        for i, r in enumerate(rows[:20]):
            rstr = " ".join(str(c) for c in (r or [])).lower()
            if sum(1 for k in kw if k in rstr) >= 2:
                return i
        # Fallback: first row with ≥ 3 filled cells
        for i, r in enumerate(rows[:15]):
            if sum(1 for c in (r or []) if str(c).strip()) >= 3:
                return i
        return 0

    # ── FORMAT 2: Tally DayBook Export ──────────────────────────────────────
    # Signature: header contains "Particulars" + "Debit"/"Credit" + "Voucher"
    hi = _find_header(all_rows, ["particulars", "narration", "debit", "credit", "voucher"])
    header_row = all_rows[hi] if hi < len(all_rows) else []
    header_str = " ".join(str(c) for c in header_row).lower()
    GSTIN_RE   = re.compile(r'\b\d{2}[A-Z]{5}\d{4}[A-Z][A-Z\d]Z[A-Z\d]\b')

    if ("particulars" in header_str or "narration" in header_str) and ("debit" in header_str or "credit" in header_str):
        date_idx  = _col_idx(header_row, "date")
        part_idx  = _col_idx(header_row, "particulars", "narration")
        vno_idx   = _col_idx(header_row, "voucher no", "vch no", "no.", "vno")
        vtype_idx = _col_idx(header_row, "voucher type", "vch type", "type")
        debit_idx = _col_idx(header_row, "debit", "dr")
        cred_idx  = _col_idx(header_row, "credit", "cr")
        last_date = date.today().isoformat()
        for r in all_rows[hi + 1:]:
            dval = _cell(r, date_idx)
            if dval and re.search(r'\d{1,2}[-/]\d{1,2}[-/]\d{2,4}', dval):
                last_date = _safe_date(dval.replace("-", "/"), last_date)
            vtype = _cell(r, vtype_idx).lower()
            if not vtype or "purchase" not in vtype:
                continue
            particulars = _cell(r, part_idx)
            vno   = _cell(r, vno_idx)
            debit = _sf(r[debit_idx] if debit_idx >= 0 and debit_idx < len(r) else "")
            cred  = _sf(r[cred_idx]  if cred_idx  >= 0 and cred_idx  < len(r) else "")
            amount = debit if debit > 0 else cred
            if not particulars or amount <= 0 or not vno:
                continue
            gm = GSTIN_RE.search(particulars.upper())
            if not gm:
                continue
            gstin   = gm.group(0)
            taxable = round(amount / 1.05, 2)
            gst     = round(amount - taxable, 2)
            result["invoices"].append(_build_invoice(
                particulars.split(chr(10))[0][:50],
                vno, last_date, taxable, round(gst/2,2), round(gst/2,2), 0,
                5, gstin, "", "", "", "", "", particulars[:80],
            ))
        if result["invoices"]:
            result["source"]       = "tally_daybook"
            result["source_label"] = f"Tally DayBook ({filename})"
            result["stats"]        = {"firms": 0, "clients": len({i["client_name"] for i in result["invoices"]}),
                                      "items": 0, "invoices": len(result["invoices"]), "payments": 0}
            return result

    # ── FORMAT 3: GSTR B2B Invoice Report / B2BInvoices CSV ─────────────────
    # Signature: header has "Supplier GSTIN" or "GSTIN of Supplier" + "Inv. No."
    is_gstr_b2b = (
        "supplier gstin" in header_str or
        "gstin of supplier" in header_str or
        ("inv. no." in header_str and "party name" in header_str) or
        "b2binvoices" in fn_lower or "as_per_books" in fn_lower
    )
    if is_gstr_b2b:
        # Re-find header: scan for the GSTIN row (may be preceded by company name rows)
        for i, r in enumerate(all_rows[:20]):
            rs = " ".join(str(c) for c in (r or [])).lower()
            if ("supplier gstin" in rs or "gstin of supplier" in rs) and ("inv" in rs or "party" in rs):
                hi = i; header_row = r; break

        gstin_i  = _col_idx(header_row, "supplier gstin", "gstin of supplier", "gstin")
        name_i   = _col_idx(header_row, "party name", "trade/legal", "supplier name")
        invno_i  = _col_idx(header_row, "inv. no.", "invoice number", "invoice no")
        date_i   = _col_idx(header_row, "inv. date", "invoice date", "date")
        val_i    = _col_idx(header_row, "inv. value", "invoice value", "total invoice value")
        tax_i    = _col_idx(header_row, "total taxable value", "taxable value")
        igst_i   = _col_idx(header_row, "integrated tax", "igst")
        cgst_i   = _col_idx(header_row, "central tax", "cgst")
        sgst_i   = _col_idx(header_row, "state / ut tax", "state/ut tax", "sgst")
        cess_i   = _col_idx(header_row, "cess")
        stype_i  = _col_idx(header_row, "supply type")
        rate_i   = _col_idx(header_row, "rate")

        for r in all_rows[hi + 1:]:
            gstin = _cell(r, gstin_i).upper().strip()
            inv_no= _cell(r, invno_i)
            if not gstin or len(gstin) != 15 or not re.match(r'^\d{2}[A-Z]', gstin) or not inv_no:
                continue
            stype   = _cell(r, stype_i).lower()
            is_inter= "inter" in stype and "intra" not in stype
            igst    = _sf(_cell(r, igst_i))
            cgst    = _sf(_cell(r, cgst_i))
            sgst    = _sf(_cell(r, sgst_i))
            cess    = _sf(_cell(r, cess_i))
            inv_val = _sf(_cell(r, val_i))
            tax_val = _sf(_cell(r, tax_i)) or round(inv_val / 1.05, 2)
            rate_raw= re.sub(r"[^0-9.]", "", _cell(r, rate_i) or "5")
            rate_v  = _safe_float(rate_raw, 5)

            # Parse date (DD-MM-YYYY or DD/MM/YYYY)
            raw_dt = _cell(r, date_i)
            inv_date = date.today().isoformat()
            for fmt in ("%d-%m-%Y", "%d/%m/%Y", "%d-%m-%y", "%d/%m/%y"):
                try: inv_date = datetime.strptime(raw_dt, fmt).strftime("%Y-%m-%d"); break
                except Exception: pass

            result["invoices"].append(_build_invoice(
                _cell(r, name_i) or gstin,
                inv_no, inv_date, tax_val,
                0 if is_inter else cgst, 0 if is_inter else sgst, igst if is_inter else 0,
                rate_v, gstin,
            ))
        if result["invoices"]:
            result["source"]       = "gstr_b2b"
            result["source_label"] = f"GSTR B2B Invoice Report ({filename})"
            result["stats"]        = {"firms": 0, "clients": len({i["client_name"] for i in result["invoices"]}),
                                      "items": 0, "invoices": len(result["invoices"]), "payments": 0}
            return result

    # ── FORMAT 4: Invoices Excel List (Party, Invoice No, Date, GST columns) ──
    # Signature: has "party name" or "invoice no" + tax columns
    is_invoice_list = (
        ("party name" in header_str or "invoice no" in header_str) and
        ("cgst" in header_str or "igst" in header_str or "gst" in header_str or "tax" in header_str)
    )
    if is_invoice_list:
        gstin_i  = _col_idx(header_row, "gstin", "gst no", "gst number")
        name_i   = _col_idx(header_row, "party name", "customer name", "client name", "vendor name")
        invno_i  = _col_idx(header_row, "invoice no", "invoice number", "bill no", "inv no", "voucher no")
        date_i   = _col_idx(header_row, "invoice date", "bill date", "date")
        val_i    = _col_idx(header_row, "invoice value", "invoice amount", "total amount", "net amount", "grand total")
        tax_i    = _col_idx(header_row, "taxable value", "taxable amount", "basic amount")
        igst_i   = _col_idx(header_row, "igst", "integrated tax")
        cgst_i   = _col_idx(header_row, "cgst", "central tax")
        sgst_i   = _col_idx(header_row, "sgst", "state tax")
        phone_i  = _col_idx(header_row, "phone", "mobile", "contact")
        addr_i   = _col_idx(header_row, "address", "billing address")
        state_i  = _col_idx(header_row, "state", "place of supply")
        hsn_i    = _col_idx(header_row, "hsn", "hsn/sac", "sac")
        notes_i  = _col_idx(header_row, "notes", "remarks", "narration", "description")

        for r in all_rows[hi + 1:]:
            client_name = _cell(r, name_i)
            if not client_name: continue
            gstin_v = _cell(r, gstin_i)
            if gstin_v and (len(gstin_v) != 15 or not re.match(r'^\d{2}', gstin_v)):
                gstin_v = ""
            inv_no  = _cell(r, invno_i) or f"IMP-{len(result['invoices'])+1:04d}"
            inv_date= _safe_date(_cell(r, date_i))
            inv_val = _sf(_cell(r, val_i))
            tax_val = _sf(_cell(r, tax_i))
            igst    = _sf(_cell(r, igst_i))
            cgst    = _sf(_cell(r, cgst_i))
            sgst    = _sf(_cell(r, sgst_i))
            is_inter= igst > 0
            if not tax_val:
                tax_val = round(inv_val / (1 + (igst+cgst+sgst)/max(inv_val,1)), 2) if inv_val else 0
            if inv_val <= 0 and tax_val <= 0: continue
            gst_total = igst + cgst + sgst
            gst_rate  = round(gst_total / tax_val * 100, 0) if tax_val > 0 else 18
            result["invoices"].append(_build_invoice(
                client_name, inv_no, inv_date, tax_val,
                cgst, sgst, igst, gst_rate,
                gstin_v, _cell(r, phone_i), "", _cell(r, addr_i), _cell(r, state_i),
                "", _cell(r, notes_i), "Due on receipt", "tax_invoice", 0,
            ))
        if result["invoices"]:
            result["source"]       = "invoice_list"
            result["source_label"] = f"Invoice List ({filename})"
            result["stats"]        = {"firms": 0, "clients": len({i["client_name"] for i in result["invoices"]}),
                                      "items": 0, "invoices": len(result["invoices"]), "payments": 0}
            return result

    # ── FORMAT 5 & 6: Marg / Busy / Zoho / Taskosphere flexible parser ───────
    # Convert rows to dict using best-guess header row
    hi2 = _find_header(all_rows)
    raw_headers = all_rows[hi2] if hi2 < len(all_rows) else []
    dict_rows: list[dict] = []
    for r in all_rows[hi2 + 1:]:
        rd: dict = {}
        for i, h in enumerate(raw_headers):
            if i < len(r) and h and str(h).strip():
                rd[str(h).strip()] = r[i] if i < len(r) else ""
        if any(str(v).strip() for v in rd.values()):
            dict_rows.append(rd)

    for row in dict_rows:
        client_name = _get_dict(row, "Client Name", "customer name", "party name", "party_name", "Name", "Buyer")
        if not client_name:
            continue
        desc       = _get_dict(row, "Description", "Item Description", "Particulars", "Item Name", default="Service")
        qty        = _sf(_get_dict(row, "Quantity", "Qty", "Units", default="1"), 1)
        rate       = _sf(_get_dict(row, "Rate", "Unit Price", "Price", "MRP", "Sale Rate", default="0"))
        gst_rate   = _sf(_get_dict(row, "GST Rate", "GST%", "Tax Rate", "GST", default="18"), 18)
        discount   = _sf(_get_dict(row, "Discount%", "Discount", "Disc%", default="0"))
        inv_date   = _safe_date(_get_dict(row, "Invoice Date", "Date", "Bill Date", "Inv Date"))
        due_date   = _safe_date(_get_dict(row, "Due Date", "Payment Due"),
                                (datetime.strptime(inv_date, "%Y-%m-%d") + timedelta(days=30)).strftime("%Y-%m-%d"))
        gstin_v    = _get_dict(row, "GSTIN", "GST No", "GST Number", "Supplier GSTIN")
        state_v    = _get_dict(row, "State", "Place of Supply")
        taxable    = round(qty * rate * (1 - discount / 100), 2)
        half       = gst_rate / 2
        cgst_v     = round(taxable * half / 100, 2)
        sgst_v     = round(taxable * half / 100, 2)
        igst_v     = _sf(_get_dict(row, "IGST", "Integrated Tax"))
        is_inter   = igst_v > 0 or ("inter" in state_v.lower() and "intra" not in state_v.lower())
        if is_inter:
            igst_v = igst_v or round(taxable * gst_rate / 100, 2)
            cgst_v = sgst_v = 0
        total = round(taxable + cgst_v + sgst_v + igst_v, 2)
        if total <= 0:
            continue
        result["invoices"].append({
            "invoice_type": "tax_invoice",
            "invoice_no": _get_dict(row, "Invoice No", "Invoice Number", "Bill No", "Ref No",
                                    default=f"IMP-{len(result['invoices'])+1:04d}"),
            "invoice_date": inv_date, "due_date": due_date,
            "client_name": client_name, "client_email": _get_dict(row, "Email", "E-Mail"),
            "client_phone": _get_dict(row, "Phone", "Mobile", "Contact"),
            "client_gstin": gstin_v,
            "client_address": _get_dict(row, "Address", "Billing Address"),
            "client_state": state_v, "is_interstate": is_inter,
            "reference_no": _get_dict(row, "Reference No", "Ref No", "Order No"),
            "notes": _get_dict(row, "Notes", "Remarks", "Narration"),
            "payment_terms": "Due on receipt",
            "items": [{
                "description": desc,
                "hsn_sac": _get_dict(row, "HSN/SAC", "HSN", "SAC"),
                "quantity": qty, "unit": _get_dict(row, "Unit", "UOM", default="nos"),
                "unit_price": rate, "discount_pct": discount, "gst_rate": gst_rate,
                "taxable_value": round(taxable, 2),
                "cgst_rate": 0 if is_inter else half,
                "sgst_rate": 0 if is_inter else half,
                "igst_rate": gst_rate if is_inter else 0,
                "cgst_amount": 0 if is_inter else cgst_v,
                "sgst_amount": 0 if is_inter else sgst_v,
                "igst_amount": igst_v if is_inter else 0,
                "total_amount": total,
            }],
            "subtotal": round(qty * rate, 2), "total_taxable": round(taxable, 2),
            "total_cgst": 0 if is_inter else round(cgst_v, 2),
            "total_sgst": 0 if is_inter else round(sgst_v, 2),
            "total_igst": round(igst_v, 2) if is_inter else 0,
            "total_gst": round(cgst_v + sgst_v + igst_v, 2),
            "grand_total": total, "amount_paid": 0, "amount_due": total,
            "status": "draft",
        })

    result["stats"] = {"firms": 0, "clients": len({i["client_name"] for i in result["invoices"]}),
                       "items": 0, "invoices": len(result["invoices"]), "payments": 0}
    return result


def _parse_tally_xml(file_path: str) -> dict:
    result = {"source": "tally", "source_label": "Tally XML",
              "firms": [], "clients": [], "items": [], "invoices": [], "payments": [], "stats": {}}
    try:
        tree = ET.parse(file_path)
        root = tree.getroot()
        vouchers = root.findall(".//VOUCHER") or root.findall(".//Voucher")
        for voucher in vouchers:
            v_type = (voucher.findtext("VOUCHERTYPENAME") or voucher.get("VCHTYPE", "")).lower()
            if v_type not in ("sales", "sale", "invoice", "credit note", "debit note"): continue
            v_date = voucher.findtext("DATE") or ""
            v_no = voucher.findtext("VOUCHERNUMBER") or ""
            party_name = voucher.findtext("PARTYLEDGERNAME") or ""
            items_list = []
            for entry in (voucher.findall(".//INVENTORYENTRIES.LIST") or voucher.findall(".//ALLINVENTORYENTRIES.LIST")):
                stock_name = entry.findtext("STOCKITEMNAME") or "Item"
                qty = abs(_safe_float(entry.findtext(".//ACTUALQTY") or "1", 1))
                rate = abs(_safe_float(entry.findtext(".//RATE") or "0"))
                amount = abs(_safe_float(entry.findtext(".//AMOUNT") or str(qty * rate)))
                items_list.append({"description": stock_name, "hsn_sac": "", "quantity": qty, "unit": "nos",
                    "unit_price": rate if rate > 0 else (amount / qty if qty > 0 else amount),
                    "discount_pct": 0, "gst_rate": 18, "taxable_value": amount,
                    "cgst_rate": 9, "sgst_rate": 9, "igst_rate": 0,
                    "cgst_amount": round(amount * 0.09, 2), "sgst_amount": round(amount * 0.09, 2),
                    "igst_amount": 0, "total_amount": round(amount * 1.18, 2)})
            if not items_list:
                for le in (voucher.findall(".//ALLLEDGERENTRIES.LIST") or []):
                    amt = _safe_float(le.findtext("AMOUNT") or "0")
                    if amt < 0:
                        ta = abs(amt)
                        items_list.append({"description": f"Tally - {v_no}", "hsn_sac": "",
                            "quantity": 1, "unit": "service", "unit_price": ta, "discount_pct": 0,
                            "gst_rate": 18, "taxable_value": ta,
                            "cgst_rate": 9, "sgst_rate": 9, "igst_rate": 0,
                            "cgst_amount": round(ta * 0.09, 2), "sgst_amount": round(ta * 0.09, 2),
                            "igst_amount": 0, "total_amount": round(ta * 1.18, 2)})
                        break
            if not items_list: continue
            tt = sum(i["taxable_value"] for i in items_list)
            tc = sum(i["cgst_amount"] for i in items_list)
            ts = sum(i["sgst_amount"] for i in items_list)
            result["invoices"].append({
                "invoice_type": "credit_note" if "credit" in v_type else "tax_invoice",
                "invoice_no": v_no or f"TALLY-{len(result['invoices'])+1:04d}",
                "invoice_date": _safe_date(v_date),
                "due_date": _safe_date(None, (date.today() + timedelta(days=30)).isoformat()),
                "client_name": party_name or "Unknown",
                "client_email": "", "client_phone": "", "client_gstin": "", "client_address": "",
                "client_state": "", "is_interstate": False, "items": items_list,
                "subtotal": round(tt, 2), "total_taxable": round(tt, 2),
                "total_cgst": round(tc, 2), "total_sgst": round(ts, 2), "total_igst": 0,
                "total_gst": round(tc + ts, 2), "grand_total": round(tt + tc + ts, 2),
                "amount_paid": 0, "amount_due": round(tt + tc + ts, 2), "status": "draft",
                "payment_terms": "Imported from Tally", "notes": "",
            })
    except ET.ParseError as e:
        raise HTTPException(400, f"Invalid XML: {e}")
    except Exception as e:
        raise HTTPException(400, f"Failed to parse Tally XML: {e}")
    result["stats"] = {"firms": 0, "clients": 0, "items": 0, "invoices": len(result["invoices"]), "payments": 0}
    return result


def _parse_json_file(file_path: str) -> dict:
    result = {"source": "json", "source_label": "JSON Backup",
              "firms": [], "clients": [], "items": [], "invoices": [], "payments": [], "stats": {}}
    try:
        with open(file_path, "r", encoding="utf-8") as f:
            data = json.load(f)
        invoices_data = data if isinstance(data, list) else (
            data.get("invoices") or data.get("bills") or data.get("transactions") or data.get("data") or [])
        for inv in invoices_data:
            if not isinstance(inv, dict): continue
            client_name = (inv.get("customer_name") or inv.get("party_name") or inv.get("client_name") or "Unknown")
            if client_name == "Unknown": continue
            items_list = []
            for item in (inv.get("items") or inv.get("line_items") or []):
                qty = _safe_float(item.get("quantity") or item.get("qty"), 1)
                price = _safe_float(item.get("price") or item.get("rate") or item.get("unit_price"))
                gst = _safe_float(item.get("gst_rate") or item.get("tax_rate"), 18)
                taxable = qty * price; half = gst / 2
                cgst = round(taxable * half / 100, 2); sgst = round(taxable * half / 100, 2)
                items_list.append({"description": item.get("name") or item.get("description") or "Item",
                    "hsn_sac": item.get("hsn") or "", "quantity": qty, "unit": item.get("unit") or "nos",
                    "unit_price": price, "discount_pct": _safe_float(item.get("discount")), "gst_rate": gst,
                    "taxable_value": round(taxable, 2), "cgst_rate": half, "sgst_rate": half, "igst_rate": 0,
                    "cgst_amount": cgst, "sgst_amount": sgst, "igst_amount": 0, "total_amount": round(taxable + cgst + sgst, 2)})
            if not items_list:
                total = _safe_float(inv.get("total") or inv.get("amount") or inv.get("grand_total"))
                if total > 0:
                    items_list.append({"description": "Imported item", "hsn_sac": "", "quantity": 1,
                        "unit": "service", "unit_price": total, "discount_pct": 0, "gst_rate": 18,
                        "taxable_value": total, "cgst_rate": 9, "sgst_rate": 9, "igst_rate": 0,
                        "cgst_amount": round(total * 0.09, 2), "sgst_amount": round(total * 0.09, 2),
                        "igst_amount": 0, "total_amount": round(total * 1.18, 2)})
            if not items_list: continue
            tt = sum(i["taxable_value"] for i in items_list)
            tc = sum(i["cgst_amount"] for i in items_list)
            ts = sum(i["sgst_amount"] for i in items_list)
            result["invoices"].append({
                "invoice_type": "tax_invoice",
                "invoice_no": inv.get("invoice_no") or f"IMP-{len(result['invoices'])+1:04d}",
                "invoice_date": _safe_date(inv.get("date") or inv.get("invoice_date")),
                "due_date": _safe_date(inv.get("due_date"), (date.today() + timedelta(days=30)).isoformat()),
                "client_name": client_name, "client_email": inv.get("email") or "",
                "client_phone": inv.get("phone") or "", "client_gstin": inv.get("gstin") or "",
                "client_address": inv.get("address") or "", "client_state": inv.get("state") or "",
                "is_interstate": False, "items": items_list,
                "subtotal": round(tt, 2), "total_taxable": round(tt, 2),
                "total_cgst": round(tc, 2), "total_sgst": round(ts, 2), "total_igst": 0,
                "total_gst": round(tc + ts, 2), "grand_total": round(tt + tc + ts, 2),
                "amount_paid": _safe_float(inv.get("amount_paid")),
                "amount_due": round(tt + tc + ts - _safe_float(inv.get("amount_paid")), 2),
                "status": "draft", "payment_terms": "Imported", "notes": inv.get("notes") or "",
            })
    except json.JSONDecodeError as e:
        raise HTTPException(400, f"Invalid JSON: {e}")
    except Exception as e:
        raise HTTPException(400, f"Failed to parse JSON: {e}")
    result["stats"] = {"firms": 0, "clients": len(result["clients"]), "items": 0,
                       "invoices": len(result["invoices"]), "payments": 0}
    return result


# ═══════════════════════════════════════════════════════════
# PYDANTIC MODELS
# ═══════════════════════════════════════════════════════════

class ProductCreate(BaseModel):
    name: str
    description: Optional[str] = None
    hsn_sac: Optional[str] = None
    unit: str = "service"
    unit_price: float = 0.0
    gst_rate: float = 18.0
    category: Optional[str] = None
    is_service: bool = True


class Product(ProductCreate):
    id: str
    created_by: str
    created_at: str


class InvoiceItem(BaseModel):
    product_id: Optional[str] = None
    description: str
    hsn_sac: Optional[str] = None
    quantity: float = 1.0
    unit: str = "service"
    unit_price: float = 0.0
    discount_pct: float = 0.0
    taxable_value: float = 0.0
    gst_rate: float = 18.0
    cgst_rate: float = 9.0
    sgst_rate: float = 9.0
    igst_rate: float = 0.0
    cgst_amount: float = 0.0
    sgst_amount: float = 0.0
    igst_amount: float = 0.0
    total_amount: float = 0.0
    item_details: Optional[str] = ""


class InvoiceCreate(BaseModel):
    invoice_type: Literal["tax_invoice", "proforma", "estimate", "credit_note", "debit_note"] = "tax_invoice"
    company_id: str
    client_id: Optional[str] = None
    lead_id: Optional[str] = None
    quotation_id: Optional[str] = None
    client_name: str
    client_address: str = ""
    client_email: str = ""
    client_phone: str = ""
    client_gstin: str = ""
    client_state: str = ""
    invoice_no: Optional[str] = None   # if provided by frontend, use it (after dup-check); else auto-generate
    invoice_date: str = ""
    due_date: str = ""
    supply_state: str = ""
    is_interstate: bool = False
    items: List[InvoiceItem] = []
    gst_rate: float = 18.0
    discount_amount: float = 0.0
    shipping_charges: float = 0.0
    other_charges: float = 0.0
    advance_received: float = 0.0
    payment_terms: str = "Due on receipt"
    notes: str = ""
    terms_conditions: str = ""
    reference_no: str = ""
    is_recurring: bool = False
    recurrence_pattern: Literal["monthly", "quarterly", "yearly"] = "monthly"
    recurrence_end: Optional[str] = None
    next_invoice_date: Optional[str] = None
    status: str = "draft"
    invoice_template: str = "prestige"
    invoice_theme: str = "classic_blue"
    invoice_custom_color: str = "#0D3B66"

    @field_validator("client_id", "lead_id", "quotation_id", mode="before")
    @classmethod
    def empty_to_none(cls, v):
        return None if v in ("", None) else v


class Invoice(InvoiceCreate):
    id: str
    invoice_no: str
    subtotal: float = 0.0
    total_discount: float = 0.0
    total_taxable: float = 0.0
    total_cgst: float = 0.0
    total_sgst: float = 0.0
    total_igst: float = 0.0
    total_gst: float = 0.0
    grand_total: float = 0.0
    amount_paid: float = 0.0
    amount_due: float = 0.0
    pdf_drive_link: str = ""
    created_by: str
    created_at: str
    updated_at: str


class PaymentCreate(BaseModel):
    invoice_id: str
    amount: float
    payment_date: str
    payment_mode: str = "neft"
    reference_no: str = ""
    notes: str = ""
    # Which specific bank account (from Accounts › Bank) the money landed in.
    # Only meaningful when payment_mode != "cash". Required at the API layer
    # (see _resolve_payment_bank_account) whenever the invoice's company has
    # at least one bank account on file, so a receipt always ties back to a
    # real account instead of only ever crediting the generic "Bank
    # Accounts" control total — which is what let the Bank Accounts page
    # balance drift out of step with what Accounting Reports showed.
    bank_account_id: Optional[str] = None


class Payment(PaymentCreate):
    id: str
    created_by: str
    created_at: str


class CreditNoteCreate(BaseModel):
    original_invoice_id: str
    company_id: str
    client_name: str
    reason: str
    items: List[InvoiceItem] = []
    notes: str = ""


# ═══════════════════════════════════════════════════════════
# CALCULATION ENGINE
# ═══════════════════════════════════════════════════════════

def _compute_item(item: InvoiceItem, is_interstate: bool, company_has_gst: bool = True) -> InvoiceItem:
    # Only a GST-registered entity may charge GST. Whether the invoice is a
    # "GST invoice" is decided entirely by the ISSUING company's own
    # registration status (Company Profile → "GST Registered" toggle /
    # `has_gst`) — never by whether the *buyer* happens to have a GSTIN.
    # If the issuing company isn't GST-registered, force the item's rate to
    # zero here so no CGST/SGST/IGST can be computed or stored downstream,
    # no matter what gst_rate the request payload contained.
    if not company_has_gst:
        item.gst_rate = 0.0
    discount = round(item.unit_price * item.quantity * item.discount_pct / 100, 2)
    taxable = round(item.unit_price * item.quantity - discount, 2)
    gst = item.gst_rate
    if is_interstate:
        igst_amt = round(taxable * gst / 100, 2)
        item.cgst_rate = 0.0; item.sgst_rate = 0.0; item.igst_rate = gst
        item.cgst_amount = 0.0; item.sgst_amount = 0.0; item.igst_amount = igst_amt
        item.total_amount = round(taxable + igst_amt, 2)
    else:
        half = gst / 2
        cgst_amt = round(taxable * half / 100, 2)
        sgst_amt = round(taxable * half / 100, 2)
        item.cgst_rate = half; item.sgst_rate = half; item.igst_rate = 0.0
        item.cgst_amount = cgst_amt; item.sgst_amount = sgst_amt; item.igst_amount = 0.0
        item.total_amount = round(taxable + cgst_amt + sgst_amt, 2)
    item.taxable_value = taxable
    return item


async def _company_has_gst(company_id: str) -> bool:
    """Whether the ISSUING company (the one raising the invoice/quotation,
    not the buyer) is GST-registered. Defaults to True (GST applies) when
    there's no company_id yet or no explicit `has_gst: false` on record, so
    existing companies that never touched the toggle keep behaving as
    before. Only an explicit `has_gst: false` on the company profile turns
    GST off."""
    if not company_id:
        return True
    company = await db.companies.find_one({"id": company_id}, {"_id": 0, "has_gst": 1})
    if not company:
        return True
    return company.get("has_gst") is not False


def _compute_invoice_totals(inv_data: dict, company_has_gst: bool = True) -> dict:
    items = inv_data.get("items", [])
    interstate = inv_data.get("is_interstate", False)
    computed = []
    for raw in items:
        it = InvoiceItem(**raw) if isinstance(raw, dict) else raw
        it = _compute_item(it, interstate, company_has_gst)
        computed.append(it.model_dump())
    subtotal = sum(i["unit_price"] * i["quantity"] for i in computed)
    total_discount = sum(i["unit_price"] * i["quantity"] * i["discount_pct"] / 100 for i in computed)
    total_taxable = sum(i["taxable_value"] for i in computed)
    total_cgst = sum(i["cgst_amount"] for i in computed)
    total_sgst = sum(i["sgst_amount"] for i in computed)
    total_igst = sum(i["igst_amount"] for i in computed)
    total_gst = round(total_cgst + total_sgst + total_igst, 2)
    shipping = float(inv_data.get("shipping_charges", 0))
    other = float(inv_data.get("other_charges", 0))
    discount_extra = float(inv_data.get("discount_amount", 0))
    grand_total = round(total_taxable + total_gst + shipping + other - discount_extra, 2)
    inv_data.update({
        "items": computed,
        "subtotal": round(subtotal, 2),
        "total_discount": round(total_discount + discount_extra, 2),
        "total_taxable": round(total_taxable, 2),
        "total_cgst": round(total_cgst, 2),
        "total_sgst": round(total_sgst, 2),
        "total_igst": round(total_igst, 2),
        "total_gst": total_gst,
        "grand_total": grand_total,
    })
    return inv_data


# ═══════════════════════════════════════════════════════════
# PDF BUILDER  — v6.1: matches Invoice Settings preview tab
#   • Reads invoice_custom_color / company color (no more hardcoded navy)
#   • Full-width dark header: logo + company left, doc-type right
#   • Bill To (left) / Due Date (right) two-column section
#   • Items table with Disc% column matching the UI preview
#   • Signature block bottom-right
# ═══════════════════════════════════════════════════════════

def _build_invoice_pdf(inv: dict, company: dict) -> BytesIO:
    # ── Resolve brand color from invoice settings → company settings → default
    raw_color = (
        inv.get("invoice_custom_color")
        or company.get("invoice_custom_color")
        or company.get("brand_color")
        or "#0D3B66"
    )
    BRAND = _hex_to_rgb(raw_color)
    BL    = _lighten(BRAND, 0.92)   # very light tint for alternating rows
    BD    = _darken(BRAND, 0.65)    # darker shade for accents
    DARK  = (30, 41, 59)
    MUTED = (100, 116, 139)
    WHITE = (255, 255, 255)
    GREEN = (22, 163, 74)
    RED   = (220, 38, 38)

    inv_type_labels = {
        "tax_invoice": "Tax Invoice",
        "proforma":    "Proforma Invoice",
        "estimate":    "Estimate",
        "credit_note": "Credit Note",
        "debit_note":  "Debit Note",
    }
    title_label = inv_type_labels.get(inv.get("invoice_type", "tax_invoice"), "Tax Invoice")
    is_cn = inv.get("invoice_type") == "credit_note"
    _inv  = inv
    _MUTED = MUTED

    class PDF(FPDF):
        def header(self): pass
        def footer(self):
            self.set_y(-12)
            self.set_font("Helvetica", "I", 7)
            self.set_text_color(*_MUTED)
            _cell(self, 0, 5,
                  _s(f"This is a computer-generated document.  ·  {_inv.get('invoice_no','')}  ·  Page {self.page_no()}"),
                  align="C", nl=True)

    pdf = PDF(orientation="P", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=18)
    pdf.add_page()

    MARGIN     = 14
    CONTENT_W  = pdf.w - MARGIN * 2   # ≈ 182 mm on A4

    # ══════════════════════════════════════════════════════
    # HEADER BAND  — full-width coloured rectangle
    # ══════════════════════════════════════════════════════
    HEADER_H = 44
    header_color = RED if is_cn else BRAND
    pdf.set_fill_color(*header_color)
    pdf.rect(0, 0, pdf.w, HEADER_H, "F")

    # Logo (top-left, inside margin)
    if company.get("logo_base64"):
        _embed_logo(pdf, company["logo_base64"], x=MARGIN, y=7, h=16)
        logo_offset = 20   # shift text right if logo present
    else:
        logo_offset = 0

    # Company info — left column
    pdf.set_xy(MARGIN + logo_offset, 7)
    pdf.set_font("Helvetica", "B", 12)
    pdf.set_text_color(*WHITE)
    _cell(pdf, CONTENT_W * 0.55, 6, _s(company.get("name", "")), nl=True)

    pdf.set_x(MARGIN + logo_offset)
    pdf.set_font("Helvetica", "", 7.5)
    pdf.set_text_color(210, 225, 245)

    addr_lines = _s(company.get("address", ""))
    if addr_lines:
        _mcell(pdf, CONTENT_W * 0.55, 4, addr_lines)
        pdf.set_x(MARGIN + logo_offset)

    contact_parts = []
    if company.get("phone"): contact_parts.append(f"Ph: {company['phone']}")
    if company.get("email"): contact_parts.append(company["email"])
    if contact_parts:
        _cell(pdf, CONTENT_W * 0.55, 4, _s("  ·  ".join(contact_parts)), nl=True)
        pdf.set_x(MARGIN + logo_offset)

    if company.get("gstin"):
        pdf.set_font("Helvetica", "B", 7.5)
        pdf.set_text_color(*WHITE)
        _cell(pdf, CONTENT_W * 0.55, 4, _s(f"GSTIN: {company['gstin']}"), nl=True)
    if company.get("pan"):
        pdf.set_font("Helvetica", "", 7.5)
        pdf.set_text_color(210, 225, 245)
        pdf.set_x(MARGIN + logo_offset)
        _cell(pdf, CONTENT_W * 0.55, 4, _s(f"PAN: {company['pan']}"), nl=True)

    # Document type — right column (aligned to right edge)
    right_col_x = MARGIN + CONTENT_W * 0.60
    right_col_w = CONTENT_W * 0.40

    pdf.set_xy(right_col_x, 7)
    pdf.set_font("Helvetica", "B", 20)
    pdf.set_text_color(*WHITE)
    _cell(pdf, right_col_w, 10, title_label, align="R", nl=True)

    pdf.set_x(right_col_x)
    pdf.set_font("Helvetica", "", 8)
    pdf.set_text_color(210, 225, 245)
    inv_no = inv.get("invoice_no", "")
    _cell(pdf, right_col_w, 5, _s(f"# {inv_no}"), align="R", nl=True)

    pdf.set_x(right_col_x)
    _cell(pdf, right_col_w, 5, _s(f"Date: {inv.get('invoice_date', '')}"), align="R", nl=True)

    # ══════════════════════════════════════════════════════
    # BILL TO  /  DUE DATE  — two-column section
    # ══════════════════════════════════════════════════════
    info_y = HEADER_H + 5
    pdf.set_xy(MARGIN, info_y)

    # Left: Bill To
    bill_w = CONTENT_W * 0.56
    pdf.set_font("Helvetica", "B", 7.5)
    pdf.set_text_color(*MUTED)
    _cell(pdf, bill_w, 5, "BILL TO", nl=True)

    pdf.set_x(MARGIN)
    pdf.set_font("Helvetica", "B", 11)
    pdf.set_text_color(*DARK)
    _cell(pdf, bill_w, 6, _s(inv.get("client_name", ""), 50), nl=True)

    pdf.set_x(MARGIN)
    pdf.set_font("Helvetica", "", 8)
    pdf.set_text_color(*MUTED)
    if inv.get("client_address"):
        _mcell(pdf, bill_w, 4, _s(inv["client_address"], 90))
        pdf.set_x(MARGIN)

    contact_c = []
    if inv.get("client_phone"): contact_c.append(inv["client_phone"])
    if inv.get("client_email"): contact_c.append(inv["client_email"])
    if contact_c:
        _cell(pdf, bill_w, 4, _s("  ·  ".join(contact_c)), nl=True)

    if inv.get("client_gstin"):
        pdf.set_x(MARGIN)
        pdf.set_font("Helvetica", "B", 8)
        pdf.set_text_color(*DARK)
        _cell(pdf, bill_w, 4, _s(f"GSTIN: {inv['client_gstin']}"), nl=True)

    # Right: Due Date block
    due_x = MARGIN + CONTENT_W * 0.62
    due_w = CONTENT_W * 0.38
    pdf.set_xy(due_x, info_y)

    pdf.set_font("Helvetica", "B", 7.5)
    pdf.set_text_color(*MUTED)
    _cell(pdf, due_w, 5, "DUE DATE", align="R", nl=True)

    pdf.set_x(due_x)
    pdf.set_font("Helvetica", "B", 13)
    pdf.set_text_color(*DARK)
    _cell(pdf, due_w, 7, _s(inv.get("due_date", "")), align="R", nl=True)

    pdf.set_x(due_x)
    pdf.set_font("Helvetica", "", 8)
    pdf.set_text_color(*MUTED)
    _cell(pdf, due_w, 5, _s(inv.get("payment_terms", "Due on receipt")), align="R", nl=True)

    if inv.get("reference_no"):
        pdf.set_x(due_x)
        _cell(pdf, due_w, 5, _s(f"PO: {inv['reference_no']}"), align="R", nl=True)

    supply_label = "Interstate (IGST)" if inv.get("is_interstate") else "Intrastate (CGST+SGST)"
    pdf.set_x(due_x)
    _cell(pdf, due_w, 5, _s(supply_label), align="R", nl=True)

    # Divider
    div_y = max(pdf.get_y(), info_y + 30) + 2
    pdf.set_draw_color(*_lighten(BRAND, 0.70))
    pdf.set_line_width(0.3)
    pdf.line(MARGIN, div_y, MARGIN + CONTENT_W, div_y)
    pdf.set_line_width(0.2)

    # ══════════════════════════════════════════════════════
    # ITEMS TABLE
    # ══════════════════════════════════════════════════════
    table_y = div_y + 3
    pdf.set_xy(MARGIN, table_y)

    is_inter = inv.get("is_interstate", False)

    # Column widths — match UI preview: Description | HSN/SAC | Qty | Disc% | Rate | (tax cols) | Amount
    sr_w   = 7
    hsn_w  = 20
    qty_w  = 13
    disc_w = 13
    rate_w = 24
    amt_w  = 26
    tax_w  = 18   # per CGST / SGST / IGST column
    n_tax_cols = 1 if is_inter else 2
    desc_w = CONTENT_W - sr_w - hsn_w - qty_w - disc_w - rate_w - tax_w * n_tax_cols - amt_w

    def _th(txt, w, a="C"):
        pdf.set_fill_color(*BRAND)
        pdf.set_text_color(*WHITE)
        pdf.set_font("Helvetica", "B", 7)
        _cell(pdf, w, 7, txt, align=a, fill=True, nl=False)

    _th("Sr",          sr_w)
    _th("Description", desc_w, "L")
    _th("HSN/SAC",     hsn_w)
    _th("Qty",         qty_w)
    _th("Disc%",       disc_w)
    _th("Rate",        rate_w, "R")
    if is_inter:
        _th("IGST%",   tax_w)
    else:
        _th("CGST%",   tax_w)
        _th("SGST%",   tax_w)
    _th("Amount",      amt_w, "R")
    _cell(pdf, 0, 0, "", nl=True)

    items = inv.get("items", [])
    for idx, it in enumerate(items, 1):
        row_bg = BL if idx % 2 == 0 else WHITE
        pdf.set_fill_color(*row_bg)
        pdf.set_text_color(*DARK)
        pdf.set_font("Helvetica", "", 7.5)

        _cell(pdf, sr_w,   7, str(idx),                                align="C", fill=True, nl=False)
        _cell(pdf, desc_w, 7, _s(it.get("description", ""), 42),      align="L", fill=True, nl=False)
        _cell(pdf, hsn_w,  7, _s(it.get("hsn_sac", "")[:12]),         align="C", fill=True, nl=False)
        _cell(pdf, qty_w,  7, f"{it.get('quantity', 1):.2f}",         align="C", fill=True, nl=False)
        _cell(pdf, disc_w, 7, f"{it.get('discount_pct', 0):.1f}%",    align="C", fill=True, nl=False)
        _cell(pdf, rate_w, 7, f"Rs.{it.get('unit_price', 0):,.2f}",   align="R", fill=True, nl=False)
        if is_inter:
            _cell(pdf, tax_w, 7, f"{it.get('igst_rate', 0):.1f}%",   align="C", fill=True, nl=False)
        else:
            cgst_r = it.get("cgst_rate", it.get("gst_rate", 18) / 2)
            sgst_r = it.get("sgst_rate", it.get("gst_rate", 18) / 2)
            _cell(pdf, tax_w, 7, f"{cgst_r:.1f}%",                    align="C", fill=True, nl=False)
            _cell(pdf, tax_w, 7, f"{sgst_r:.1f}%",                    align="C", fill=True, nl=False)
        _cell(pdf, amt_w,  7, f"Rs.{it.get('total_amount', 0):,.2f}", align="R", fill=True, nl=True)

        # Optional sub-detail line (unit / item_details)
        sub_parts = []
        if it.get("item_details"): sub_parts.append(_s(it["item_details"]))
        if it.get("unit"):         sub_parts.append(_s(it["unit"]))
        if sub_parts:
            pdf.set_x(MARGIN + sr_w)
            pdf.set_font("Helvetica", "I", 6.5)
            pdf.set_text_color(*MUTED)
            pdf.set_fill_color(*row_bg)
            sub_w = desc_w + hsn_w + qty_w + disc_w + rate_w + tax_w * n_tax_cols + amt_w
            _cell(pdf, sub_w, 4, _s("  ".join(sub_parts)), align="L", fill=True, nl=True)

    # ══════════════════════════════════════════════════════
    # TOTALS BLOCK
    # ══════════════════════════════════════════════════════
    def _trow(label, value, bold=False, strike=False):
        bg = BRAND if bold else BL
        tc = WHITE if bold else DARK
        pdf.set_fill_color(*bg)
        pdf.set_text_color(*tc)
        pdf.set_font("Helvetica", "B" if bold else "", 8 if not bold else 9)
        _cell(pdf, CONTENT_W - amt_w, 7, label, align="R", fill=True, nl=False)
        prefix = "-Rs." if strike else "Rs."
        val_str = f"{prefix}{abs(float(value)):,.2f}"
        _cell(pdf, amt_w, 7, val_str, align="R", fill=True, nl=True)

    pdf.set_x(MARGIN)
    _trow("Subtotal", inv.get("subtotal", 0))
    if float(inv.get("total_discount", 0)) > 0:
        _trow("Discount", inv.get("total_discount", 0), strike=True)
    _trow("Taxable Value", inv.get("total_taxable", 0))

    tt = float(inv.get("total_taxable", 1)) or 1
    if is_inter:
        igst_pct = round(float(inv.get("total_igst", 0)) / tt * 100, 1)
        _trow(f"IGST ({igst_pct:.1f}%)", inv.get("total_igst", 0))
    else:
        cgst_pct = round(float(inv.get("total_cgst", 0)) / tt * 100, 1)
        sgst_pct = round(float(inv.get("total_sgst", 0)) / tt * 100, 1)
        _trow(f"CGST ({cgst_pct:.1f}%)", inv.get("total_cgst", 0))
        _trow(f"SGST ({sgst_pct:.1f}%)", inv.get("total_sgst", 0))

    if float(inv.get("shipping_charges", 0)) > 0:
        _trow("Shipping Charges", inv.get("shipping_charges", 0))
    if float(inv.get("other_charges", 0)) > 0:
        _trow("Other Charges", inv.get("other_charges", 0))
    _trow("GRAND TOTAL", inv.get("grand_total", 0), bold=True)

    # Amount in words
    pdf.set_x(MARGIN)
    pdf.set_font("Helvetica", "I", 8)
    pdf.set_text_color(*MUTED)
    _cell(pdf, CONTENT_W, 5, _s(_amount_in_words(float(inv.get("grand_total", 0)))), nl=True)

    # PAID stamp (right side)
    if inv.get("status") == "paid":
        stamp_y = pdf.get_y() - 12
        pdf.set_xy(MARGIN + CONTENT_W - 56, stamp_y)
        pdf.set_draw_color(*GREEN)
        pdf.set_line_width(0.8)
        pdf.set_font("Helvetica", "B", 20)
        pdf.set_text_color(*GREEN)
        pdf.cell(50, 12, "PAID", border=1, align="C")
        pdf.set_line_width(0.2)

    # ══════════════════════════════════════════════════════
    # GST SUMMARY
    # ══════════════════════════════════════════════════════
    pdf.ln(5)
    pdf.set_font("Helvetica", "B", 9)
    pdf.set_text_color(*BRAND)
    _cell(pdf, 0, 5, "GST Summary", nl=True)
    pdf.set_draw_color(*_lighten(BRAND, 0.70))
    pdf.line(MARGIN, pdf.get_y(), MARGIN + CONTENT_W, pdf.get_y())
    pdf.ln(1)

    gst_sum: Dict[float, Dict[str, float]] = {}
    for it in items:
        r = float(it.get("gst_rate", 18))
        if r not in gst_sum:
            gst_sum[r] = {"taxable": 0.0, "cgst": 0.0, "sgst": 0.0, "igst": 0.0}
        gst_sum[r]["taxable"] += float(it.get("taxable_value", 0))
        gst_sum[r]["cgst"]    += float(it.get("cgst_amount",   0))
        gst_sum[r]["sgst"]    += float(it.get("sgst_amount",   0))
        gst_sum[r]["igst"]    += float(it.get("igst_amount",   0))

    g_w = CONTENT_W / 5
    pdf.set_fill_color(*BRAND)
    pdf.set_text_color(*WHITE)
    pdf.set_font("Helvetica", "B", 7.5)
    for col_hdr in ["GST Rate", "Taxable Amt", "CGST", "SGST / IGST", "Total GST"]:
        _cell(pdf, g_w, 6, col_hdr, align="C", fill=True, nl=False)
    _cell(pdf, 0, 0, "", nl=True)

    pdf.set_font("Helvetica", "", 7.5)
    for i, (rate, row) in enumerate(sorted(gst_sum.items())):
        pdf.set_fill_color(*(BL if i % 2 == 0 else WHITE))
        pdf.set_text_color(*DARK)
        gst_tot = row["cgst"] + row["sgst"] + row["igst"]
        _cell(pdf, g_w, 6, f"{rate:.1f}%",              align="C", fill=True, nl=False)
        _cell(pdf, g_w, 6, f"Rs.{row['taxable']:,.2f}", align="C", fill=True, nl=False)
        _cell(pdf, g_w, 6, f"Rs.{row['cgst']:,.2f}",   align="C", fill=True, nl=False)
        _cell(pdf, g_w, 6, f"Rs.{row['sgst'] or row['igst']:,.2f}", align="C", fill=True, nl=False)
        _cell(pdf, g_w, 6, f"Rs.{gst_tot:,.2f}",       align="C", fill=True, nl=True)

    # ══════════════════════════════════════════════════════
    # BANK DETAILS
    # ══════════════════════════════════════════════════════
    if company.get("bank_account_no") or company.get("bank_name"):
        pdf.ln(5)
        pdf.set_font("Helvetica", "B", 9)
        pdf.set_text_color(*BRAND)
        _cell(pdf, 0, 5, "Bank Details for Payment", nl=True)
        pdf.set_draw_color(*_lighten(BRAND, 0.70))
        pdf.line(MARGIN, pdf.get_y(), MARGIN + CONTENT_W, pdf.get_y())
        pdf.ln(1)
        half_w = CONTENT_W / 2
        for lbl, val in [
            ("Account Name", company.get("bank_account_name", "")),
            ("Bank Name",    company.get("bank_name",          "")),
            ("Account No",   company.get("bank_account_no",    "")),
            ("IFSC Code",    company.get("bank_ifsc",          "")),
        ]:
            pdf.set_font("Helvetica", "B", 8); pdf.set_text_color(*DARK)
            _cell(pdf, half_w * 0.42, 5, _s(f"{lbl}:"), nl=False)
            pdf.set_font("Helvetica", "", 8)
            _cell(pdf, half_w * 0.58, 5, _s(val), nl=True)

    # ══════════════════════════════════════════════════════
    # TERMS & NOTES
    # ══════════════════════════════════════════════════════
    if inv.get("terms_conditions") or inv.get("notes"):
        pdf.ln(5)
        pdf.set_font("Helvetica", "B", 9)
        pdf.set_text_color(*BRAND)
        _cell(pdf, 0, 5, "Terms & Notes", nl=True)
        pdf.set_draw_color(*_lighten(BRAND, 0.70))
        pdf.line(MARGIN, pdf.get_y(), MARGIN + CONTENT_W, pdf.get_y())
        pdf.ln(1)
        pdf.set_font("Helvetica", "", 8)
        pdf.set_text_color(*DARK)
        if inv.get("terms_conditions"):
            _mcell(pdf, CONTENT_W, 4, _s(inv["terms_conditions"]))
        if inv.get("notes"):
            _mcell(pdf, CONTENT_W, 4, _s(f"Note: {inv['notes']}"))

    # ══════════════════════════════════════════════════════
    # SIGNATURE — bottom-right
    # ══════════════════════════════════════════════════════
    pdf.ln(8)
    sig_y   = pdf.get_y()
    sig_x   = MARGIN + CONTENT_W - 58
    sig_w   = 55

    sig_b64 = company.get("signature_base64", "")
    if sig_b64:
        _embed_logo(pdf, sig_b64, x=sig_x, y=sig_y, h=14)
        sig_y += 16

    pdf.set_draw_color(*BRAND)
    pdf.set_line_width(0.4)
    pdf.line(sig_x, sig_y, sig_x + sig_w, sig_y)
    pdf.set_line_width(0.2)

    pdf.set_xy(sig_x, sig_y + 1)
    pdf.set_font("Helvetica", "B", 8)
    pdf.set_text_color(*DARK)
    _cell(pdf, sig_w, 5, _s(f"For {company.get('name', '')}"), align="C", nl=True)

    pdf.set_x(sig_x)
    pdf.set_font("Helvetica", "", 7)
    pdf.set_text_color(*MUTED)
    _cell(pdf, sig_w, 4, "Authorised Signatory", align="C", nl=True)

    # ── Output
    buf = BytesIO()
    buf.write(pdf.output())
    buf.seek(0)
    return buf


# ═══════════════════════════════════════════════════════════
# BACKUP IMPORT ENDPOINTS
# ═══════════════════════════════════════════════════════════

@router.post("/invoices/parse-vyp")
async def parse_vyp_file(
    file: UploadFile = File(...),
    current_user: User = Depends(check_module_permission("invoicing", "create"))
):
    if not _perm(current_user):
        raise HTTPException(403, "Access denied")

    tmp_path = None
    try:
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".vyp")
        tmp.write(await file.read())
        tmp.close()
        tmp_path = tmp.name

        return _parse_vyp_file(tmp_path)

    finally:
        if tmp_path and os.path.exists(tmp_path):
            try:
                os.unlink(tmp_path)
            except:
                pass


@router.post("/invoices/parse-backup")
async def parse_backup_file(
    file: UploadFile = File(...),
    current_user: User = Depends(check_module_permission("invoicing", "create"))
):
    if not _perm(current_user):
        raise HTTPException(403, "Access denied")

    filename = file.filename or "unknown"
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""

    supported = {
        "vyp": "vyp",
        "vyb": "vyp",
        "db": "vyp",
        "xml": "xml",
        "tbk": "xml",
        "xlsx": "excel",
        "xls": "excel",
        "csv": "excel",
        "json": "json",
    }

    parser_type = supported.get(ext)
    if not parser_type:
        raise HTTPException(
            400,
            f"Unsupported file format: .{ext}. Supported: {', '.join(f'.{k}' for k in supported)}"
        )

    tmp_path = None
    try:
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=f".{ext}")
        content = await file.read()
        tmp.write(content)
        tmp.close()
        tmp_path = tmp.name

        if parser_type == "vyp":
            result = _parse_vyp_file(tmp_path)
        elif parser_type == "xml":
            result = _parse_tally_xml(tmp_path)
        elif parser_type == "excel":
            result = _parse_excel_file(tmp_path, filename)
        elif parser_type == "json":
            result = _parse_json_file(tmp_path)
        else:
            raise HTTPException(400, "No parser available")

        # ✅ upload backup (already correct)
        if _drive_configured():
            await _upload_to_drive(
                content,
                f"Backup_{filename}",
                "backups",
                file.content_type or "application/octet-stream"
            )

        return result

    finally:
        if tmp_path and os.path.exists(tmp_path):
            try:
                os.unlink(tmp_path)
            except:
                pass


# ═══════════════════════════════════════════════════════════
# IMPORT BACKUP → SAVE TO MONGODB  (v6.3 addition)
# ═══════════════════════════════════════════════════════════

class ImportBackupRequest(BaseModel):
    company_id: str
    source: str = "unknown"
    invoices: List[dict] = []
    clients: List[dict] = []
    items: List[dict] = []
    payments: List[dict] = []
    skip_duplicates: bool = True


class ImportBackupResult(BaseModel):
    invoices_imported: int = 0
    invoices_skipped: int = 0
    clients_imported: int = 0
    clients_updated: int = 0
    items_imported: int = 0
    payments_imported: int = 0
    errors: List[str] = []


@router.post("/invoices/import-backup", response_model=ImportBackupResult)
async def import_backup(
    data: ImportBackupRequest,
    current_user: User = Depends(check_module_permission("invoicing", "create")),
):
    """
    Bulk-saves the JSON output of /invoices/parse-backup into MongoDB.

    v7 — rewrote all four sections to use bulk DB operations instead of
    one-by-one insert_one / find_one loops.  This drops ~4 800 individual
    round-trips down to ~10 queries total, keeping well under the 60 s
    Nginx / Render timeout even for large backups (1 000+ invoices).

    Frontend flow:
      1. POST /invoices/parse-backup  → preview JSON
      2. User reviews
      3. POST /invoices/import-backup → save to DB  ← this endpoint
    """
    if not _perm(current_user):
        raise HTTPException(403, "Access denied")

    now = datetime.now(timezone.utc).isoformat()
    result = ImportBackupResult()

    # ══════════════════════════════════════════════════════════════════════
    # 1.  CLIENTS  — one pre-fetch → batch insert + bulk_write updates
    # ══════════════════════════════════════════════════════════════════════
    from pymongo import UpdateOne, InsertOne

    client_name_to_id: dict = {}

    if data.clients:
        # Fetch ALL existing clients in one round-trip
        existing_clients = await db.clients.find(
            {}, {"_id": 0, "id": 1, "company_name": 1,
                 "phone": 1, "email": 1, "address": 1,
                 "client_gstin": 1, "state": 1}
        ).to_list(100_000)
        existing_map = {c["company_name"].strip().lower(): c for c in existing_clients}

        bulk_inserts = []
        bulk_updates = []   # list of UpdateOne operations

        for c in data.clients:
            name = (c.get("full_name") or c.get("client_name") or "").strip()
            if not name:
                continue
            key = name.lower()
            if key in existing_map:
                ex = existing_map[key]
                client_name_to_id[name] = ex["id"]
                update_fields = {}
                for field, src_key in [
                    ("phone",        "phone_number"),
                    ("email",        "email"),
                    ("address",      "address"),
                    ("client_gstin", "name_gstin_number"),
                    ("state",        "name_state"),
                ]:
                    bv = (c.get(src_key) or "").strip()
                    if bv and not (ex.get(field) or "").strip():
                        update_fields[field] = bv
                if update_fields:
                    update_fields["updated_at"] = now
                    bulk_updates.append(
                        UpdateOne({"id": ex["id"]}, {"$set": update_fields})
                    )
                    result.clients_updated += 1
            else:
                cid = str(uuid.uuid4())
                doc = {
                    "id":            cid,
                    "company_name":  name,
                    "phone":         (c.get("phone_number") or "").strip(),
                    "email":         (c.get("email") or "").strip(),
                    "address":       (c.get("address") or "").strip(),
                    "client_gstin":  (c.get("name_gstin_number") or "").strip(),
                    "state":         (c.get("name_state") or "").strip(),
                    "client_type":   "other",
                    "status":        "active",
                    "imported_from": data.source,
                    "created_by":    current_user.id,
                    "created_at":    now,
                    "updated_at":    now,
                }
                bulk_inserts.append(doc)
                client_name_to_id[name] = cid
                # also add to existing_map so duplicate names in same import batch
                # don't each get inserted separately
                existing_map[key] = {"id": cid, "company_name": name,
                                     "phone": doc["phone"], "email": doc["email"],
                                     "address": doc["address"],
                                     "client_gstin": doc["client_gstin"],
                                     "state": doc["state"]}

        try:
            if bulk_inserts:
                await db.clients.insert_many(bulk_inserts, ordered=False)
                result.clients_imported += len(bulk_inserts)
            if bulk_updates:
                await db.clients.bulk_write(bulk_updates, ordered=False)
        except Exception as e:
            result.errors.append(f"Clients bulk write: {e}")

    # ══════════════════════════════════════════════════════════════════════
    # 2.  PRODUCTS / ITEMS  — one pre-fetch → single insert_many
    # ══════════════════════════════════════════════════════════════════════
    if data.items:
        existing_items = await db.products.find(
            {"company_id": data.company_id}, {"_id": 0, "name": 1}
        ).to_list(100_000)
        existing_item_names = {i["name"].strip().lower() for i in existing_items}

        new_items = []
        for item in data.items:
            item_name = (item.get("name") or "").strip()
            if not item_name or item_name.lower() in existing_item_names:
                continue
            existing_item_names.add(item_name.lower())   # guard same-batch dups
            new_items.append({
                "id":            str(uuid.uuid4()),
                "company_id":    data.company_id,
                "name":          item_name,
                "description":   item.get("description", ""),
                "hsn_sac":       item.get("hsn_sac", ""),
                "unit":          "service",
                "unit_price":    float(item.get("sale_price", 0)),
                "gst_rate":      float(item.get("gst_rate", 18)),
                "is_service":    False,
                "imported_from": data.source,
                "created_by":    current_user.id,
                "created_at":    now,
            })
        try:
            if new_items:
                await db.products.insert_many(new_items, ordered=False)
                result.items_imported += len(new_items)
        except Exception as e:
            result.errors.append(f"Items bulk insert: {e}")

    # ══════════════════════════════════════════════════════════════════════
    # 3.  INVOICES  — one pre-fetch of existing numbers → single insert_many
    # ══════════════════════════════════════════════════════════════════════
    if data.invoices:
        # Pull every existing invoice_no for this company in one query
        if data.skip_duplicates:
            existing_inv_docs = await db.invoices.find(
                {"company_id": data.company_id},
                {"_id": 0, "invoice_no": 1}
            ).to_list(200_000)
            existing_inv_nos = {d["invoice_no"] for d in existing_inv_docs if d.get("invoice_no")}
        else:
            existing_inv_nos = set()

        new_invoices = []
        for inv in data.invoices:
            inv_no = (inv.get("invoice_no") or "").strip()
            if data.skip_duplicates and inv_no and inv_no in existing_inv_nos:
                result.invoices_skipped += 1
                continue

            # Guard same-batch duplicates (two invoices with the same number sent
            # in the same payload should not both be inserted)
            if inv_no:
                existing_inv_nos.add(inv_no)

            client_name = (inv.get("client_name") or "").strip()
            new_invoices.append({
                "id":                   str(uuid.uuid4()),
                "company_id":           data.company_id,
                "client_id":            client_name_to_id.get(client_name),
                "invoice_no":           inv_no or f"IMP-{str(uuid.uuid4())[:8].upper()}",
                "invoice_type":         inv.get("invoice_type", "tax_invoice"),
                "invoice_date":         inv.get("invoice_date", date.today().isoformat()),
                "due_date":             inv.get("due_date", date.today().isoformat()),
                "client_name":          client_name,
                "client_email":         inv.get("client_email", ""),
                "client_phone":         inv.get("client_phone", ""),
                "client_gstin":         inv.get("client_gstin", ""),
                "client_address":       inv.get("client_address", ""),
                "client_state":         inv.get("client_state", ""),
                "is_interstate":        inv.get("is_interstate", False),
                "items":                inv.get("items", []),
                "subtotal":             float(inv.get("subtotal", 0)),
                "total_discount":       float(inv.get("total_discount", 0)),
                "total_taxable":        float(inv.get("total_taxable", 0)),
                "total_cgst":           float(inv.get("total_cgst", 0)),
                "total_sgst":           float(inv.get("total_sgst", 0)),
                "total_igst":           float(inv.get("total_igst", 0)),
                "total_gst":            float(inv.get("total_gst", 0)),
                "grand_total":          float(inv.get("grand_total", 0)),
                "amount_paid":          float(inv.get("amount_paid", 0)),
                "amount_due":           float(inv.get("amount_due", 0)),
                "status":               inv.get("status", "draft"),
                "payment_terms":        inv.get("payment_terms", "Imported"),
                "notes":                inv.get("notes", ""),
                "reference_no":         inv.get("reference_no", ""),
                "supply_state":         inv.get("client_state", ""),
                "discount_amount":      0.0,
                "shipping_charges":     0.0,
                "other_charges":        0.0,
                "terms_conditions":     "",
                "is_recurring":         False,
                "recurrence_pattern":   "monthly",
                "recurrence_end":       None,
                "next_invoice_date":    None,
                "invoice_template":     "prestige",
                "invoice_theme":        "classic_blue",
                "invoice_custom_color": "#0D3B66",
                "pdf_drive_link":       "",
                "imported_from":        data.source,
                "created_by":           current_user.id,
                "created_at":           now,
                "updated_at":           now,
            })

        try:
            if new_invoices:
                await db.invoices.insert_many(new_invoices, ordered=False)
                result.invoices_imported += len(new_invoices)
        except Exception as e:
            result.errors.append(f"Invoices bulk insert: {e}")

    # ══════════════════════════════════════════════════════════════════════
    # 4.  PAYMENTS  — one pre-fetch of kb_source_ids → single insert_many
    # ══════════════════════════════════════════════════════════════════════
    if data.payments:
        # Fetch all already-imported source IDs for this company in one query
        existing_pay_docs = await db.payments.find(
            {"company_id": data.company_id, "kb_source_id": {"$exists": True, "$ne": ""}},
            {"_id": 0, "kb_source_id": 1}
        ).to_list(500_000)
        existing_kb_ids = {d["kb_source_id"] for d in existing_pay_docs if d.get("kb_source_id")}

        new_payments = []
        for pay in data.payments:
            amount = float(pay.get("amount", 0))
            if amount <= 0:
                continue
            kb_id = pay.get("_kb_id")
            kb_id_str = str(kb_id) if kb_id else ""
            if kb_id_str and kb_id_str in existing_kb_ids:
                continue
            if kb_id_str:
                existing_kb_ids.add(kb_id_str)   # guard same-batch dups
            new_payments.append({
                "id":            str(uuid.uuid4()),
                "invoice_id":    "",
                "company_id":    data.company_id,
                "kb_source_id":  kb_id_str,
                "client_name":   (pay.get("client_name") or "").strip(),
                "amount":        amount,
                "payment_date":  pay.get("payment_date", date.today().isoformat()),
                "payment_mode":  pay.get("payment_mode", "other"),
                "reference_no":  "",
                "notes":         f"Imported from {data.source}",
                "imported_from": data.source,
                "created_by":    current_user.id,
                "created_at":    now,
            })
        try:
            if new_payments:
                await db.payments.insert_many(new_payments, ordered=False)
                result.payments_imported += len(new_payments)
        except Exception as e:
            result.errors.append(f"Payments bulk insert: {e}")

    result.errors = result.errors[:50]
    logger.info(
        f"Import [{data.source}]: {result.invoices_imported} invoices, "
        f"{result.clients_imported} clients new, {result.clients_updated} clients updated, "
        f"{result.items_imported} items, {result.invoices_skipped} skipped, {len(result.errors)} errors"
    )
    return result



# ─────────────────────────────────────────────
# GET /invoices/backup-sources  — list distinct import sources with counts
# DELETE /invoices/remove-backup  — delete all data from a given import source
# ─────────────────────────────────────────────

@router.get("/invoices/backup-sources")
async def get_backup_sources(
    current_user: User = Depends(check_module_permission("invoicing", "view")),
):
    """
    Return all distinct import sources that exist in the database,
    together with the count of invoices, clients and products from each source.
    """
    pipeline = [
        {"$group": {"_id": "$imported_from", "invoice_count": {"$sum": 1}}},
        {"$match": {"_id": {"$ne": None}}},
    ]
    invoice_docs = await db.invoices.aggregate(pipeline).to_list(100)

    # Index by source for quick merging
    source_map: dict = {}
    for doc in invoice_docs:
        src = doc["_id"] or "unknown"
        source_map[src] = {
            "source": src,
            "invoice_count": doc["invoice_count"],
            "client_count": 0,
            "product_count": 0,
        }

    # Count clients per source
    async for doc in db.clients.aggregate([
        {"$group": {"_id": "$imported_from", "count": {"$sum": 1}}},
        {"$match": {"_id": {"$ne": None}}},
    ]):
        src = doc["_id"] or "unknown"
        if src in source_map:
            source_map[src]["client_count"] = doc["count"]

    # Count products per source
    async for doc in db.products.aggregate([
        {"$group": {"_id": "$imported_from", "count": {"$sum": 1}}},
        {"$match": {"_id": {"$ne": None}}},
    ]):
        src = doc["_id"] or "unknown"
        if src in source_map:
            source_map[src]["product_count"] = doc["count"]

    return list(source_map.values())


@router.delete("/invoices/remove-backup")
async def remove_backup(
    source: str,
    current_user: User = Depends(check_module_permission("invoicing", "delete")),
):
    """
    Delete ALL invoices, clients and products that were imported from *source*.
    Requires DELETE permission on the invoicing module.
    """
    if not source:
        raise HTTPException(status_code=400, detail="source query param is required")

    inv_res    = await db.invoices.delete_many({"imported_from": source})
    client_res = await db.clients.delete_many({"imported_from": source})
    prod_res   = await db.products.delete_many({"imported_from": source})

    return {
        "source": source,
        "invoices_deleted": inv_res.deleted_count,
        "clients_deleted":  client_res.deleted_count,
        "products_deleted": prod_res.deleted_count,
    }



# ─────────────────────────────────────────────
# ✅ NEW ENDPOINT (CHANGE 3 — ADD THIS HERE)
# ─────────────────────────────────────────────

@router.patch("/invoices/sync-client/{client_id}")
async def sync_invoices_for_client(
    client_id: str,
    data: dict,
    current_user: User = Depends(check_module_permission("invoicing", "create")),
):
    """
    Bulk-update client details across all invoices that reference this client_id.
    Called automatically when a client record is saved/updated.
    """
    if not _perm(current_user):
        raise HTTPException(403, "Access denied")
    allowed = {
        "client_name", "client_gstin", "client_phone",
        "client_email", "client_address", "client_state",
    }
    update_fields = {k: v for k, v in data.items() if k in allowed and v is not None and v != ""}
    if update_fields:
        update_fields["updated_at"] = datetime.now(timezone.utc).isoformat()
        await db.invoices.update_many({"client_id": client_id}, {"$set": update_fields})
    count = await db.invoices.count_documents({"client_id": client_id})
    return {"synced_invoices": count, "fields_updated": list(update_fields.keys())}

@router.post("/invoices/{inv_id}/upload-pdf-to-drive")
async def upload_pdf_bytes_to_drive(
    inv_id: str,
    payload: dict,
    current_user: User = Depends(check_module_permission("invoicing", "create"))
):
    if not _perm(current_user):
        raise HTTPException(403, "Access denied")

    if not _drive_configured():
        raise HTTPException(503, "Google Drive not configured")

    inv = await db.invoices.find_one({"id": inv_id}, {"_id": 0})
    if not inv:
        raise HTTPException(404, "Invoice not found")

    try:
        pdf_bytes = base64.b64decode(payload.get("pdf_base64"))
    except Exception:
        raise HTTPException(400, "Invalid PDF data")

    filename = payload.get("filename") or f"Invoice_{inv.get('invoice_no','')}.pdf"

    # optional: client folder
    try:
        client_folder_id = _get_or_create_client_folder(inv.get("client_name"))
    except Exception:
        client_folder_id = None

    drive_link = await _upload_to_drive(
        content_bytes=pdf_bytes,
        filename=filename,
        folder_key="invoices",
        mime_type="application/pdf",
        custom_parent_id=client_folder_id
    )

    if drive_link:
        await db.invoices.update_one(
            {"id": inv_id},
            {"$set": {"pdf_drive_link": drive_link}}
        )

    return {
        "drive_link": drive_link or ""
    }

# ═══════════════════════════════════════════════════════════
# PRODUCT CATALOG
# ═══════════════════════════════════════════════════════════

@router.post("/products", response_model=Product)
async def create_product(data: ProductCreate, current_user: User = Depends(check_module_permission("invoicing", "create"))):
    if not _perm(current_user): raise HTTPException(403, "Access denied")
    now = datetime.now(timezone.utc).isoformat()
    doc = {"id": str(uuid.uuid4()), **data.model_dump(), "created_by": current_user.id, "created_at": now}
    await db.products.insert_one(doc); doc.pop("_id", None)
    return doc


@router.get("/products")
async def list_products(search: Optional[str] = None, current_user: User = Depends(check_module_permission("invoicing", "view"))):
    if not _perm(current_user): raise HTTPException(403, "Access denied")
    q: dict = {}
    if current_user.role != "admin": q["created_by"] = current_user.id
    if search:
        q["$or"] = [{"name": {"$regex": search, "$options": "i"}},
                    {"description": {"$regex": search, "$options": "i"}}]
    return await db.products.find(q, {"_id": 0}).sort("name", 1).to_list(500)


@router.put("/products/{pid}")
async def update_product(pid: str, data: ProductCreate, current_user: User = Depends(check_module_permission("invoicing", "create"))):
    if not _perm(current_user): raise HTTPException(403, "Access denied")
    ex = await db.products.find_one({"id": pid})
    if not ex: raise HTTPException(404, "Product not found")
    if current_user.role != "admin" and ex.get("created_by") != current_user.id:
        raise HTTPException(403, "Not authorized")
    await db.products.update_one({"id": pid}, {"$set": data.model_dump()})
    return await db.products.find_one({"id": pid}, {"_id": 0})


@router.delete("/products/{pid}")
async def delete_product(pid: str, current_user: User = Depends(check_module_permission("invoicing", "delete"))):
    if not _perm(current_user): raise HTTPException(403, "Access denied")
    await db.products.delete_one({"id": pid})
    return {"message": "Product deleted"}



# ═══════════════════════════════════════════════════════════
# PURCHASE INVOICE UPLOAD + CLIENT LINKING
# ═══════════════════════════════════════════════════════════

_GSTIN_RE = re.compile(r"\b[0-9]{2}[A-Z]{5}[0-9]{4}[A-Z][1-9A-Z]Z[0-9A-Z]\b", re.I)
_MONEY_RE = re.compile(r"(?:₹|Rs\.?|INR)?\s*([0-9]{1,3}(?:,[0-9]{2,3})+(?:\.\d{1,2})?|[0-9]+(?:\.\d{1,2})?)", re.I)
_DATE_RE = re.compile(r"\b(\d{1,2}[/-]\d{1,2}[/-]\d{2,4}|\d{4}[/-]\d{1,2}[/-]\d{1,2}|\d{1,2}\s+[A-Za-z]{3,9}\s+\d{2,4})\b")

def _money_to_float(value: Any, default: float = 0.0) -> float:
    if value is None:
        return default
    try:
        txt = str(value).replace(",", "").replace("₹", "").replace("Rs.", "").replace("Rs", "").replace("INR", "").strip()
        return round(float(txt), 2)
    except Exception:
        return default

def _normalise_invoice_date(value: str | None) -> str:
    if not value:
        return ""
    raw = str(value).strip().replace(".", "-")
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%d/%m/%y", "%d-%m-%y", "%Y-%m-%d", "%Y/%m/%d", "%d %b %Y", "%d %B %Y", "%d %b %y", "%d %B %y"):
        try:
            return datetime.strptime(raw, fmt).date().isoformat()
        except Exception:
            pass
    return raw

def _clean_party_line(line: str) -> str:
    line = re.sub(r"^[\s:;\-–—]+", "", line or "")
    line = re.sub(r"\s{2,}", " ", line).strip(" :-\t")
    return line[:160]

def _find_value_near_labels(text: str, labels: list[str], max_len: int = 80) -> str:
    for label in labels:
        m = re.search(rf"{label}\s*[:#\-]?\s*([^\n]{{1,{max_len}}})", text, re.I)
        if m:
            val = _clean_party_line(m.group(1))
            val = re.split(r"\s{3,}|\t|\|", val)[0].strip()
            if val:
                return val
    return ""

def _find_amount_by_labels(text: str, labels: list[str]) -> float:
    for label in labels:
        m = re.search(label, text, re.I)
        if not m:
            continue
        window = text[m.end():m.end() + 180]
        amounts = [_money_to_float(x) for x in _MONEY_RE.findall(window)]
        amounts = [a for a in amounts if a > 0]
        if amounts:
            return max(amounts)
    return 0.0

def _guess_party_after_label(lines: list[str], labels: tuple[str, ...]) -> str:
    for i, line in enumerate(lines):
        low = line.lower()
        if any(lbl in low for lbl in labels):
            if ":" in line:
                val = _clean_party_line(line.split(":", 1)[1])
                if val and not _GSTIN_RE.search(val) and len(val) > 2:
                    return val
            for nxt in lines[i + 1:i + 5]:
                val = _clean_party_line(nxt)
                if not val or _GSTIN_RE.search(val):
                    continue
                if re.search(r"address|gstin|state|phone|email|invoice|date", val, re.I):
                    continue
                return val
    return ""

def _parse_purchase_invoice_text(text: str, filename: str = "") -> dict:
    compact = re.sub(r"[ \t]+", " ", text or "")
    lines = [_clean_party_line(x) for x in compact.splitlines() if _clean_party_line(x)]
    gstins = []
    for g in _GSTIN_RE.findall(compact):
        g = g.upper()
        if g not in gstins:
            gstins.append(g)
    invoice_no = _find_value_near_labels(compact, [r"invoice\s*(?:no|number|#)", r"bill\s*(?:no|number)", r"tax\s*invoice\s*(?:no|number)", r"document\s*(?:no|number)", r"voucher\s*(?:no|number)"], 50)
    if invoice_no:
        invoice_no = re.split(r"\s+(?:date|dt|dated|gstin|place)\b", invoice_no, flags=re.I)[0].strip(" :-#")
    date_value = _find_value_near_labels(compact, [r"invoice\s*date", r"bill\s*date", r"dated", r"date"], 40)
    dm = _DATE_RE.search(date_value or "") or _DATE_RE.search(compact)
    invoice_date = _normalise_invoice_date(dm.group(1) if dm else "")
    grand_total = _find_amount_by_labels(compact, [r"grand\s+total", r"invoice\s+total", r"total\s+amount", r"amount\s+payable", r"net\s+amount", r"total\s+value", r"bill\s+amount"])
    taxable = _find_amount_by_labels(compact, [r"taxable\s+(?:amount|value)", r"sub\s*total", r"subtotal"])
    cgst = _find_amount_by_labels(compact, [r"cgst"])
    sgst = _find_amount_by_labels(compact, [r"sgst"])
    igst = _find_amount_by_labels(compact, [r"igst"])
    total_gst = round(cgst + sgst + igst, 2) if any([cgst, sgst, igst]) else _find_amount_by_labels(compact, [r"total\s+gst", r"tax\s+amount", r"gst\s+amount"])
    if not grand_total:
        amounts = [_money_to_float(x) for x in _MONEY_RE.findall(compact)]
        amounts = [a for a in amounts if 1 <= a <= 10_00_00_000]
        if amounts:
            grand_total = max(amounts)
    buyer_name = _guess_party_after_label(lines, ("bill to", "billed to", "buyer", "customer", "recipient", "consignee"))
    supplier_name = _guess_party_after_label(lines, ("supplier", "seller", "from", "vendor"))
    if not supplier_name:
        for line in lines[:12]:
            if re.search(r"tax invoice|invoice|original|duplicate|gstin|phone|email|date|bill to|buyer", line, re.I):
                continue
            if len(line) >= 4:
                supplier_name = line
                break
    return {
        "file_name": filename,
        "invoice_no": invoice_no or "",
        "invoice_date": invoice_date,
        "supplier_name": supplier_name or "",
        "supplier_gstin": gstins[0] if gstins else "",
        "buyer_name": buyer_name or "",
        "buyer_gstin": gstins[1] if len(gstins) > 1 else "",
        "all_gstins": gstins,
        "taxable_amount": taxable,
        "total_gst": total_gst,
        "grand_total": grand_total,
        "currency": "INR",
        "parse_confidence": 0.68 if (invoice_no or grand_total or gstins) else 0.25,
        "raw_text_excerpt": compact[:5000],
    }

def _extract_json_object(text: str) -> dict:
    if not text:
        return {}
    cleaned = re.sub(r"^```(?:json)?\s*|\s*```$", "", text.strip(), flags=re.I | re.S).strip()
    m = re.search(r"\{.*\}", cleaned, re.S)
    if m:
        cleaned = m.group(0)
    try:
        return json.loads(cleaned)
    except Exception:
        return {}

async def _ai_extract_purchase_invoice(contents: bytes, filename: str, ext: str, mime_type: str) -> dict:
    prompt = ("Read this supplier purchase invoice and return ONLY valid JSON with these keys: supplier_name, supplier_gstin, buyer_name, buyer_gstin, invoice_no, invoice_date, taxable_amount, total_gst, grand_total, currency, parse_confidence. Use ISO yyyy-mm-dd for invoice_date when possible. Use numbers only for amounts.")
    try:
        from backend.ai_document_reader import _groq_vision, _groq_vision_multipage
        if ext == "pdf":
            import pdfplumber
            page_images_b64 = []
            with pdfplumber.open(BytesIO(contents)) as pdf:
                for page in pdf.pages[:4]:
                    pil_img = page.to_image(resolution=150).original
                    if pil_img.mode not in ("RGB", "L"):
                        pil_img = pil_img.convert("RGB")
                    buf = BytesIO()
                    pil_img.save(buf, format="JPEG", quality=85)
                    page_images_b64.append((base64.b64encode(buf.getvalue()).decode(), "image/jpeg"))
            if not page_images_b64:
                return {}
            answer = await _groq_vision_multipage(page_images_b64, prompt)
        else:
            from PIL import Image as PILImage
            img = PILImage.open(BytesIO(contents))
            if img.mode not in ("RGB", "L"):
                img = img.convert("RGB")
            buf = BytesIO()
            img.save(buf, format="JPEG", quality=85)
            answer = await _groq_vision(base64.b64encode(buf.getvalue()).decode(), "image/jpeg", prompt)
        parsed = _extract_json_object(answer)
        if parsed:
            parsed["invoice_date"] = _normalise_invoice_date(parsed.get("invoice_date"))
            for key in ("taxable_amount", "total_gst", "grand_total"):
                parsed[key] = _money_to_float(parsed.get(key))
            parsed["all_gstins"] = [g.upper() for g in _GSTIN_RE.findall(" ".join(str(v) for v in parsed.values()))]
            parsed["raw_text_excerpt"] = ""
            return parsed
    except Exception as e:
        logger.warning(f"Purchase invoice AI extraction failed for {filename}: {e}")
    return {}

async def _extract_purchase_invoice(contents: bytes, filename: str, content_type: str = "") -> dict:
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    text = ""
    if ext == "pdf":
        try:
            import pdfplumber
            pages = []
            with pdfplumber.open(BytesIO(contents)) as pdf:
                for i, page in enumerate(pdf.pages[:12]):
                    page_text = page.extract_text() or ""
                    if page_text.strip():
                        pages.append(f"--- Page {i + 1} ---\n{page_text.strip()}")
            text = "\n\n".join(pages)
        except Exception as e:
            logger.warning(f"Purchase invoice PDF text extraction failed for {filename}: {e}")
    elif ext in {"txt", "csv"}:
        text = contents.decode("utf-8", errors="replace")
    parsed = _parse_purchase_invoice_text(text, filename) if text.strip() else {}
    needs_ocr = not parsed or (not parsed.get("invoice_no") and not parsed.get("grand_total") and not parsed.get("all_gstins"))
    if needs_ocr and ext in {"pdf", "png", "jpg", "jpeg", "webp"}:
        ai_parsed = await _ai_extract_purchase_invoice(contents, filename, ext, content_type or "application/octet-stream")
        parsed = {**parsed, **{k: v for k, v in ai_parsed.items() if v not in (None, "", [], 0)}} if parsed else ai_parsed
    if not parsed:
        parsed = {"file_name": filename, "raw_text_excerpt": text[:5000]}
    parsed.setdefault("file_name", filename)
    parsed.setdefault("all_gstins", [g.upper() for g in _GSTIN_RE.findall(text or "")])
    parsed.setdefault("parse_confidence", 0.3)
    return parsed

def _client_access_query_for_purchase(current_user: User) -> dict:
    if getattr(current_user, "role", None) == "admin":
        return {}
    assigned_clients = []
    perms = getattr(current_user, "permissions", None)
    try:
        assigned_clients = getattr(perms, "assigned_clients", []) or []
    except Exception:
        if isinstance(perms, dict):
            assigned_clients = perms.get("assigned_clients", []) or []
    clauses = [{"created_by": current_user.id}, {"assigned_to": current_user.id}, {"assignments.user_id": current_user.id}]
    if assigned_clients:
        clauses.append({"id": {"$in": assigned_clients}})
    return {"$or": clauses}

async def _match_purchase_client(parsed: dict, current_user: User, client_id: str | None = None) -> dict | None:
    access = _client_access_query_for_purchase(current_user)
    if client_id:
        q = {"id": client_id}
        if access:
            q = {"$and": [access, q]}
        return await db.clients.find_one(q, {"_id": 0})
    gstins = []
    for g in [parsed.get("buyer_gstin"), *(parsed.get("all_gstins") or [])]:
        if g and str(g).upper() not in gstins:
            gstins.append(str(g).upper())
    for gst in gstins:
        gst_q = {"$or": [{"gstin": {"$regex": f"^{re.escape(gst)}$", "$options": "i"}}, {"client_gstin": {"$regex": f"^{re.escape(gst)}$", "$options": "i"}}]}
        q = {"$and": [access, gst_q]} if access else gst_q
        client = await db.clients.find_one(q, {"_id": 0})
        if client:
            return client
    haystack = (parsed.get("raw_text_excerpt") or "").lower()
    buyer_name = (parsed.get("buyer_name") or "").lower()
    if not haystack and not buyer_name:
        return None
    clients = await db.clients.find(access, {"_id": 0, "id": 1, "company_name": 1, "gstin": 1, "client_gstin": 1}).to_list(5000)
    for c in clients:
        name = (c.get("company_name") or "").strip().lower()
        if len(name) >= 4 and (name in haystack or name in buyer_name or buyer_name in name):
            return c
    return None

def _purchase_scope_query(current_user: User) -> dict:
    return {} if getattr(current_user, "role", None) == "admin" else {"created_by": current_user.id}

@router.get("/purchase-invoices")
async def list_purchase_invoices(client_id: Optional[str] = Query(None), search: Optional[str] = Query(None), page: int = Query(1, ge=1), page_size: int = Query(100, ge=1, le=1000), current_user: User = Depends(get_current_user)):
    if not _perm(current_user):
        raise HTTPException(403, "Access denied")
    q = _purchase_scope_query(current_user)
    if client_id:
        q["client_id"] = client_id
    if search:
        q["$or"] = [{"client_name": {"$regex": search, "$options": "i"}}, {"supplier_name": {"$regex": search, "$options": "i"}}, {"invoice_no": {"$regex": search, "$options": "i"}}, {"supplier_gstin": {"$regex": search, "$options": "i"}}]
    total = await db.purchase_invoices.count_documents(q)
    skip = (page - 1) * page_size
    items = await db.purchase_invoices.find(q, {"_id": 0}).sort("created_at", -1).skip(skip).limit(page_size).to_list(page_size)
    return {"purchase_invoices": items, "total": total, "page": page, "page_size": page_size}

@router.post("/purchase-invoices/upload")
async def upload_purchase_invoice(file: UploadFile = File(...), client_id: Optional[str] = Form(None), company_id: Optional[str] = Form(None), current_user: User = Depends(get_current_user)):
    if not _perm(current_user):
        raise HTTPException(403, "Access denied")
    filename = file.filename or "purchase-invoice"
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    if ext not in {"pdf", "png", "jpg", "jpeg", "webp", "txt", "csv"}:
        raise HTTPException(400, "Unsupported file type. Upload PDF, image, TXT, or CSV invoices.")
    contents = await file.read()
    if not contents:
        raise HTTPException(400, "Uploaded file is empty")
    parsed = await _extract_purchase_invoice(contents, filename, file.content_type or "")
    matched_client = await _match_purchase_client(parsed, current_user, client_id)
    gstins = [g.upper() for g in (parsed.get("all_gstins") or []) if g]
    client_gstin = ((matched_client or {}).get("gstin") or (matched_client or {}).get("client_gstin") or "").upper()
    supplier_gstin = (parsed.get("supplier_gstin") or "").upper()
    buyer_gstin = (parsed.get("buyer_gstin") or "").upper()
    if matched_client and client_gstin:
        buyer_gstin = client_gstin
        supplier_gstin = next((g for g in gstins if g != client_gstin), supplier_gstin)
    elif len(gstins) >= 2 and not buyer_gstin:
        buyer_gstin = gstins[1]
    now = datetime.now(timezone.utc).isoformat()
    invoice_no = (parsed.get("invoice_no") or f"PUR-{str(uuid.uuid4())[:8].upper()}").strip()
    duplicate_q = {"invoice_no": invoice_no, "created_by": current_user.id}
    if matched_client:
        duplicate_q["client_id"] = matched_client.get("id")
    if supplier_gstin:
        duplicate_q["supplier_gstin"] = supplier_gstin
    existing = await db.purchase_invoices.find_one(duplicate_q, {"_id": 0}) if invoice_no else None
    if existing:
        return {"purchase_invoice": existing, "matched_client": matched_client, "duplicate": True}
    grand_total_val = _money_to_float(parsed.get("grand_total"))
    doc = {
        "id": str(uuid.uuid4()), "company_id": company_id or "", "client_id": (matched_client or {}).get("id", client_id or ""),
        "client_name": (matched_client or {}).get("company_name", parsed.get("buyer_name", "")), "supplier_name": parsed.get("supplier_name", ""),
        "supplier_gstin": supplier_gstin, "buyer_name": parsed.get("buyer_name", ""), "buyer_gstin": buyer_gstin, "invoice_no": invoice_no,
        "invoice_date": parsed.get("invoice_date") or date.today().isoformat(), "taxable_amount": _money_to_float(parsed.get("taxable_amount")),
        "total_gst": _money_to_float(parsed.get("total_gst")), "grand_total": grand_total_val, "currency": parsed.get("currency") or "INR",
        "file_name": filename, "file_size": len(contents), "content_type": file.content_type or "application/octet-stream", "parse_confidence": float(parsed.get("parse_confidence") or 0),
        "raw_text_excerpt": parsed.get("raw_text_excerpt", "")[:5000], "created_by": current_user.id, "created_at": now, "updated_at": now,
        # Ledger automation: every saved purchase invoice is a real bill the moment it's
        # confirmed, so it starts "outstanding" (unlike Sales, there's no draft step here)
        # and immediately posts to the ledger — mirroring the Dr Purchases/GST Input,
        # Cr Accounts Payable entry an accountant would make by hand.
        "status": "outstanding", "amount_paid": 0.0, "amount_due": round(grand_total_val, 2),
    }
    await db.purchase_invoices.insert_one(doc)
    if matched_client:
        await db.clients.update_one({"id": matched_client["id"]}, {"$set": {"last_purchase_invoice_at": now}, "$inc": {"purchase_invoice_count": 1}})
    await sync_purchase_journal_entry(doc["id"])
    doc.pop("_id", None)
    return {"purchase_invoice": doc, "matched_client": matched_client, "duplicate": False}


class PurchaseInvoiceUpdate(BaseModel):
    client_id: Optional[str] = None
    client_name: Optional[str] = None
    company_id: Optional[str] = None
    supplier_name: Optional[str] = None
    supplier_gstin: Optional[str] = None
    buyer_name: Optional[str] = None
    buyer_gstin: Optional[str] = None
    invoice_no: Optional[str] = None
    invoice_date: Optional[str] = None
    taxable_amount: Optional[float] = None
    total_gst: Optional[float] = None
    grand_total: Optional[float] = None
    currency: Optional[str] = None


@router.put("/purchase-invoices/{invoice_id}")
async def update_purchase_invoice(invoice_id: str, data: PurchaseInvoiceUpdate, current_user: User = Depends(get_current_user)):
    if not _perm(current_user):
        raise HTTPException(403, "Access denied")
    existing = await db.purchase_invoices.find_one({"id": invoice_id})
    if not existing:
        raise HTTPException(404, "Purchase invoice not found")
    
    update_data = {k: v for k, v in data.model_dump().items() if v is not None}
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()

    # If the grand total changed, amount_due should move with it (unless the
    # invoice was already fully paid — don't silently reopen a paid bill).
    if "grand_total" in update_data and existing.get("status") != "paid":
        new_total = float(update_data["grand_total"] or 0)
        paid_so_far = float(existing.get("amount_paid") or 0)
        update_data["amount_due"] = round(max(new_total - paid_so_far, 0), 2)

    # If client_id is changed, update client names/counters if needed
    if "client_id" in update_data and update_data["client_id"] != existing.get("client_id"):
        new_client_id = update_data["client_id"]
        if new_client_id:
            client = await db.clients.find_one({"id": new_client_id})
            if client:
                update_data["client_name"] = client.get("company_name", "")
                await db.clients.update_one({"id": new_client_id}, {"$set": {"last_purchase_invoice_at": update_data["updated_at"]}, "$inc": {"purchase_invoice_count": 1}})
        old_client_id = existing.get("client_id")
        if old_client_id:
            await db.clients.update_one({"id": old_client_id}, {"$inc": {"purchase_invoice_count": -1}})

    await db.purchase_invoices.update_one({"id": invoice_id}, {"$set": update_data})
    await sync_purchase_journal_entry(invoice_id)
    updated = await db.purchase_invoices.find_one({"id": invoice_id}, {"_id": 0})
    return updated


@router.delete("/purchase-invoices/{invoice_id}")
async def delete_purchase_invoice(invoice_id: str, current_user: User = Depends(get_current_user)):
    if not _perm(current_user):
        raise HTTPException(403, "Access denied")
    existing = await db.purchase_invoices.find_one({"id": invoice_id})
    if not existing:
        raise HTTPException(404, "Purchase invoice not found")

    # Clean up any payments recorded against this bill (and their journal entries)
    # before deleting the bill's own journal entry, so nothing dangles in the ledger.
    payments = await db.purchase_payments.find({"purchase_invoice_id": invoice_id}, {"_id": 0, "id": 1}).to_list(500)
    for p in payments:
        existing_pe = await db.journal_entries.find_one({"source": "purchase_payment", "source_id": p["id"]})
        if existing_pe:
            await db.journal_lines.delete_many({"entry_id": existing_pe["id"]})
            await db.journal_entries.delete_one({"id": existing_pe["id"]})
    await db.purchase_payments.delete_many({"purchase_invoice_id": invoice_id})

    _old_entries = await db.journal_entries.find({"source": "purchase", "source_id": invoice_id}, {"_id": 0, "id": 1}).to_list(50)
    if _old_entries:
        _old_ids = [e["id"] for e in _old_entries]
        await db.journal_lines.delete_many({"entry_id": {"$in": _old_ids}})
        await db.journal_entries.delete_many({"id": {"$in": _old_ids}})

    await db.purchase_invoices.delete_one({"id": invoice_id})
    client_id = existing.get("client_id")
    if client_id:
        await db.clients.update_one({"id": client_id}, {"$inc": {"purchase_invoice_count": -1}})
    return {"success": True}


@router.patch("/purchase-invoices/{invoice_id}/status")
async def update_purchase_invoice_status(
    invoice_id: str,
    payload: dict,
    current_user: User = Depends(get_current_user),
):
    """
    Change a purchase bill's payment status — the Purchase-page equivalent of
    the Sale page's status dropdown. Accepted: outstanding | partially_paid | paid | cancelled

    Unlike the Sale-side status endpoint (which only adjusts the amount fields
    and leaves recording the actual receipt to a separate /payments call),
    marking a purchase bill 'paid' or 'partially_paid' here *also* records the
    cash/bank payment itself and posts its journal entry — matching the
    one-step "flip the status, the payment entry is done" flow that was asked for.

    Optional payload fields (used only for paid/partially_paid):
      amount        — payment amount. Defaults to the full amount_due for 'paid'.
      payment_date  — defaults to today.
      payment_mode  — 'cash' | 'bank' (default 'bank').
      bank_account_id — which bank account paid this vendor, when payment_mode
                        isn't cash. Required whenever the company has more than
                        one bank account on file (see _resolve_payment_bank_account).
      reference_no, notes
    """
    if not _perm(current_user):
        raise HTTPException(403, "Access denied")

    new_status = (payload.get("status") or "").strip().lower()
    if new_status not in PURCHASE_STATUS:
        raise HTTPException(400, f"Invalid status. Must be one of: {PURCHASE_STATUS}")

    inv = await db.purchase_invoices.find_one({"id": invoice_id})
    if not inv:
        raise HTTPException(404, "Purchase invoice not found")

    now_iso = datetime.now(timezone.utc).isoformat()
    grand_total = float(inv.get("grand_total") or 0)
    already_paid = float(inv.get("amount_paid") or 0)

    if new_status in ("paid", "partially_paid"):
        amount = float(payload.get("amount") or 0)
        if new_status == "paid" and amount <= 0:
            amount = round(max(grand_total - already_paid, 0), 2)
        if amount <= 0:
            raise HTTPException(400, "Payment amount must be greater than zero.")

        payment_mode = (payload.get("payment_mode") or "bank").strip().lower()
        bank_account = await _resolve_payment_bank_account(
            inv.get("company_id") or "", payment_mode, payload.get("bank_account_id")
        )
        payment_doc = {
            "id": str(uuid.uuid4()), "purchase_invoice_id": invoice_id,
            "company_id": inv.get("company_id") or "", "amount": amount,
            "payment_date": (payload.get("payment_date") or "").strip() or date.today().isoformat(),
            "payment_mode": payment_mode,
            "bank_account_id": bank_account["id"] if bank_account else None,
            "reference_no": (payload.get("reference_no") or "").strip(),
            "notes": (payload.get("notes") or "").strip(),
            "created_by": current_user.id, "created_at": now_iso,
        }
        await db.purchase_payments.insert_one(dict(payment_doc))
        await sync_purchase_payment_journal_entry(payment_doc["id"])

        all_payments = await db.purchase_payments.find({"purchase_invoice_id": invoice_id}, {"_id": 0}).to_list(500)
        total_paid = round(sum(float(p.get("amount", 0)) for p in all_payments), 2)
        amount_due = round(max(grand_total - total_paid, 0), 2)
        final_status = "paid" if amount_due <= 0 else "partially_paid"
        await db.purchase_invoices.update_one({"id": invoice_id}, {"$set": {
            "amount_paid": total_paid, "amount_due": amount_due, "status": final_status, "updated_at": now_iso,
        }})
        return {"id": invoice_id, "status": final_status, "amount_paid": total_paid, "amount_due": amount_due,
                "payment_id": payment_doc["id"]}

    if new_status == "cancelled":
        await db.purchase_invoices.update_one({"id": invoice_id}, {"$set": {
            "status": "cancelled", "amount_due": 0.0, "updated_at": now_iso,
        }})
        await sync_purchase_journal_entry(invoice_id)  # removes the bill's journal entry
        return {"id": invoice_id, "status": "cancelled", "amount_paid": inv.get("amount_paid", 0), "amount_due": 0.0}

    # new_status == "outstanding" — manual revert, doesn't touch already-recorded payments
    await db.purchase_invoices.update_one({"id": invoice_id}, {"$set": {"status": "outstanding", "updated_at": now_iso}})
    await sync_purchase_journal_entry(invoice_id)
    return {"id": invoice_id, "status": "outstanding", "amount_paid": already_paid, "amount_due": inv.get("amount_due", grand_total)}


@router.get("/purchase-payments")
async def list_purchase_payments(purchase_invoice_id: Optional[str] = None, current_user: User = Depends(get_current_user)):
    if not _perm(current_user):
        raise HTTPException(403, "Access denied")
    q: dict = {}
    if purchase_invoice_id:
        q["purchase_invoice_id"] = purchase_invoice_id
    return await db.purchase_payments.find(q, {"_id": 0}).sort("created_at", -1).to_list(1000)


@router.delete("/purchase-payments/{payment_id}")
async def delete_purchase_payment(payment_id: str, current_user: User = Depends(get_current_user)):
    if not _perm(current_user):
        raise HTTPException(403, "Access denied")
    payment = await db.purchase_payments.find_one({"id": payment_id})
    if not payment:
        raise HTTPException(404, "Payment not found")
    invoice_id = payment.get("purchase_invoice_id")

    result = await db.purchase_payments.delete_one({"id": payment_id})
    if result.deleted_count == 0:
        raise HTTPException(404, f"Payment {payment_id} not found")
    await sync_purchase_payment_journal_entry(payment_id)  # cleans up its journal entry (payment doc is gone)

    if invoice_id:
        inv = await db.purchase_invoices.find_one({"id": invoice_id})
        if inv:
            all_payments = await db.purchase_payments.find({"purchase_invoice_id": invoice_id}, {"_id": 0}).to_list(500)
            total_paid = round(sum(float(p.get("amount", 0)) for p in all_payments), 2)
            grand_total = float(inv.get("grand_total") or 0)
            amount_due = round(max(grand_total - total_paid, 0), 2)
            new_status = "paid" if amount_due <= 0 else ("partially_paid" if total_paid > 0 else "outstanding")
            await db.purchase_invoices.update_one({"id": invoice_id}, {"$set": {
                "amount_paid": total_paid, "amount_due": amount_due, "status": new_status,
                "updated_at": datetime.now(timezone.utc).isoformat(),
            }})

    return {"message": f"Payment {payment_id} deleted"}


# ═══════════════════════════════════════════════════════════
# INVOICE CRUD
# ═══════════════════════════════════════════════════════════

@router.post("/invoices", response_model=Invoice)
async def create_invoice(data: InvoiceCreate, current_user: User = Depends(check_module_permission("invoicing", "create"))):
    if not _perm(current_user): raise HTTPException(403, "Access denied")
    now = datetime.now(timezone.utc).isoformat()
    prefix = {"proforma": "PRO", "estimate": "EST", "credit_note": "CN", "debit_note": "DN"}.get(data.invoice_type, "INV")

    # Use frontend-supplied number if non-empty; otherwise auto-generate
    requested_no = (data.invoice_no or "").strip()
    if requested_no:
        # Duplicate check: reject if already in use (scoped to same company)
        dup_filter = {"invoice_no": requested_no}
        if data.company_id:
            dup_filter["company_id"] = data.company_id
        conflict = await db.invoices.find_one(dup_filter)
        if conflict:
            raise HTTPException(400, f"Invoice number '{requested_no}' is already in use. Please choose a different number.")
        inv_no = requested_no
    else:
        # Fallback: auto-generate using the same settings the frontend would use.
        # The frontend always sends invoice_no pre-filled; this path is a safety net.
        inv_no = await _next_invoice_no(
            prefix=prefix,
            company_id=data.company_id,
            separator="/",
            include_fy=True,
            fy_format="short",
            include_month=False,
            number_padding=3,
            invoice_type=data.invoice_type,
        )

    inv_date = data.invoice_date or date.today().isoformat()
    due_date = data.due_date or (date.today() + timedelta(days=30)).isoformat()
    advance = max(0.0, float(data.advance_received or 0))
    raw = {"id": str(uuid.uuid4()), "invoice_no": inv_no, "invoice_date": inv_date, "due_date": due_date,
           **data.model_dump(exclude={"invoice_no"}), "amount_paid": advance, "amount_due": 0.0, "pdf_drive_link": "",
           "created_by": current_user.id, "created_at": now, "updated_at": now}
    raw = _compute_invoice_totals(raw, await _company_has_gst(data.company_id))
    if raw["amount_due"] <= 0:
        raw["status"] = "paid"
    elif advance > 0:
        raw["status"] = "partially_paid"
    await db.invoices.insert_one({**raw})
    await sync_invoice_journal_entry(raw["id"])
    # Record advance as a payment entry so history is tracked
    if advance > 0:
        pay_id = str(uuid.uuid4())
        await db.payments.insert_one({
            "id": pay_id, "invoice_id": raw["id"],
            "company_id": data.company_id,
            "amount": advance, "payment_date": inv_date,
            "payment_mode": "advance", "reference_no": "",
            "notes": "Advance received at invoice creation",
            "created_by": current_user.id, "created_at": now,
        })
        await sync_payment_journal_entry(pay_id)
    raw.pop("_id", None)
    return raw


@router.get("/invoices")
async def list_invoices(
    page: int = Query(1, ge=1, description="Page number (1-based)"),
    page_size: int = Query(20, ge=1, le=10000, description="Results per page"),
    status: Optional[str] = Query(None, description="Filter by status"),
    search: Optional[str] = Query(None, description="Search by client name or invoice number"),
    current_user: User = Depends(check_module_permission("invoicing", "view")),
):
    if not _perm(current_user): raise HTTPException(403, "Access denied")
    q: dict = {} if current_user.role == "admin" else {"created_by": current_user.id}
    if status and status in INV_STATUS:
        q["status"] = status
    if search:
        q["$or"] = [
            {"client_name": {"$regex": search, "$options": "i"}},
            {"invoice_no": {"$regex": search, "$options": "i"}},
        ]
    total = await db.invoices.count_documents(q)
    skip = (page - 1) * page_size
    invoices = await (
        db.invoices.find(q, {"_id": 0})
        .sort("created_at", -1)
        .skip(skip)
        .limit(page_size)
        .to_list(page_size)
    )
    return {
        "invoices": invoices,
        "total": total,
        "page": page,
        "page_size": page_size,
        "total_pages": max(1, -(-total // page_size)),
    }


@router.get("/invoices/stats")
async def invoice_stats(year: Optional[int] = None, month: Optional[int] = None,
                        current_user: User = Depends(check_module_permission("invoicing", "view"))):
    if not _perm(current_user): raise HTTPException(403, "Access denied")
    q: dict = {"invoice_type": "tax_invoice", "status": {"$ne": "cancelled"}}
    if current_user.role != "admin": q["created_by"] = current_user.id
    every_inv = await db.invoices.find(q, {"_id": 0, "grand_total": 1, "amount_paid": 1,
        "amount_due": 1, "status": 1, "invoice_date": 1, "client_name": 1, "total_gst": 1}).to_list(5000)
    # Revenue/outstanding/trend figures now include draft invoices as well —
    # per business rule, a draft (renamed "Invoiced" in the UI) is recognised
    # revenue immediately. sync_invoice_journal_entry() also posts drafts so
    # that Trial Balance / P&L / Balance Sheet remain consistent with the
    # figures reported here. Only cancelled invoices are excluded.
    all_inv = list(every_inv)
    today = date.today()
    cur_year = year or today.year
    cur_mon = month or today.month

    def _in_month(d, y, m):
        try: dt = date.fromisoformat(d[:10]); return dt.year == y and dt.month == m
        except: return False

    total_rev = sum(i.get("grand_total", 0) for i in all_inv)
    # Outstanding: exclude paid and cancelled — use grand_total minus amount_paid
    # so old records with stale amount_due field don't inflate the number.
    def _calc_due(i):
        if i.get("status") in ("paid", "cancelled"):
            return 0.0
        grand = float(i.get("grand_total") or 0)
        paid  = float(i.get("amount_paid") or i.get("advance_received") or 0)
        return max(0.0, grand - paid)

    def _calc_paid(i):
        if i.get("status") == "paid":
            return float(i.get("grand_total") or 0)
        return float(i.get("amount_paid") or i.get("advance_received") or 0)

    total_out = sum(_calc_due(i) for i in all_inv)
    overdue_c = sum(1 for i in all_inv if i.get("status") not in ("paid", "cancelled", "draft") and _calc_due(i) > 0)
    mon_inv = [i for i in all_inv if _in_month(i.get("invoice_date", ""), cur_year, cur_mon)]
    trend = []
    for offset in range(11, -1, -1):
        dt = (date(today.year, today.month, 1) - timedelta(days=offset * 28))
        y_, m_ = dt.year, dt.month
        mi = [i for i in all_inv if _in_month(i.get("invoice_date", ""), y_, m_)]
        trend.append({"year": y_, "month": m_, "label": date(y_, m_, 1).strftime("%b %y"),
                      "revenue": sum(i.get("grand_total", 0) for i in mi),
                      "collected": sum(_calc_paid(i) for i in mi), "count": len(mi)})
    from collections import defaultdict
    client_rev: dict = defaultdict(float)
    for i in all_inv: client_rev[i.get("client_name", "Unknown")] += i.get("grand_total", 0)
    top_clients = sorted(client_rev.items(), key=lambda x: -x[1])[:5]
    return {
        "total_revenue": round(total_rev, 2), "total_outstanding": round(total_out, 2),
        "overdue_count": overdue_c, "total_invoices": len(every_inv),
        "month_revenue": round(sum(i.get("grand_total", 0) for i in mon_inv), 2),
        "month_collected": round(sum(i.get("amount_paid", 0) for i in mon_inv), 2),
        "month_invoices": len(mon_inv), "monthly_trend": trend,
        "top_clients": [{"name": n, "revenue": round(v, 2)} for n, v in top_clients],
        "paid_count": sum(1 for i in all_inv if i.get("status") == "paid"),
        "draft_count": sum(1 for i in every_inv if i.get("status") == "draft"),
        "total_gst": round(sum(i.get("total_gst", 0) for i in all_inv), 2),
    }


@router.get("/invoices/referral-summary")
async def invoice_referral_summary(
    current_user: User = Depends(check_module_permission("invoicing", "view")),
):
    """
    Group clients — and every invoice raised against them — by the
    referrer recorded on the client record (`referred_by`): the person
    or company via whom that client was brought in, and who a referral
    payout may be owed to.

    This lets the Invoicing screen show, for each referrer: which
    clients they brought in, how many invoices were raised for those
    clients, and how much has been billed / collected / is still due —
    so referral commissions can be tracked without leaving Invoicing.
    """
    from collections import defaultdict

    if not _perm(current_user):
        raise HTTPException(403, "Access denied")

    client_q: dict = {} if current_user.role == "admin" else {"created_by": current_user.id}
    clients = await db.clients.find(
        client_q, {"_id": 0, "id": 1, "company_name": 1, "referred_by": 1}
    ).to_list(10000)

    client_by_id = {c["id"]: c for c in clients if c.get("id")}
    client_by_name = {c.get("company_name"): c for c in clients if c.get("company_name")}

    inv_q: dict = {"status": {"$ne": "cancelled"}}
    if current_user.role != "admin":
        inv_q["created_by"] = current_user.id
    invoices = await db.invoices.find(
        inv_q,
        {
            "_id": 0, "client_id": 1, "client_name": 1,
            "grand_total": 1, "amount_paid": 1, "status": 1,
        },
    ).to_list(20000)

    def _paid(i: dict) -> float:
        if i.get("status") == "paid":
            return float(i.get("grand_total") or 0)
        return float(i.get("amount_paid") or 0)

    def _due(i: dict) -> float:
        if i.get("status") in ("paid", "cancelled"):
            return 0.0
        return max(0.0, float(i.get("grand_total") or 0) - _paid(i))

    groups: Dict[str, dict] = defaultdict(lambda: {
        "referrer": None,
        "clients": {},
        "invoice_count": 0,
        "total_invoiced": 0.0,
        "total_collected": 0.0,
        "total_due": 0.0,
    })

    for inv in invoices:
        client = client_by_id.get(inv.get("client_id")) or client_by_name.get(inv.get("client_name"))
        referrer = (client.get("referred_by") if client else None) or None
        key = referrer or "__unreferred__"
        g = groups[key]
        g["referrer"] = referrer
        g["invoice_count"] += 1
        g["total_invoiced"] += float(inv.get("grand_total") or 0)
        g["total_collected"] += _paid(inv)
        g["total_due"] += _due(inv)
        cname = (client.get("company_name") if client else None) or inv.get("client_name") or "Unknown"
        cid = client.get("id") if client else None
        g["clients"][cid or cname] = {"id": cid, "company_name": cname}

    result = []
    for g in groups.values():
        result.append({
            "referrer": g["referrer"],
            "is_unreferred": g["referrer"] is None,
            "client_count": len(g["clients"]),
            "clients": sorted(g["clients"].values(), key=lambda c: c["company_name"] or ""),
            "invoice_count": g["invoice_count"],
            "total_invoiced": round(g["total_invoiced"], 2),
            "total_collected": round(g["total_collected"], 2),
            "total_due": round(g["total_due"], 2),
        })

    result.sort(key=lambda r: -r["total_invoiced"])
    return {"groups": result}


# ══════════════════════════════════════════════════════════════
# FIX: drive-status MUST be declared BEFORE /invoices/{inv_id}
# In v6.0 it was after, so FastAPI matched "drive-status" as an
# invoice ID and routed it to get_invoice() → always 404.
# ══════════════════════════════════════════════════════════════

@router.get("/invoices/next-number")
async def get_next_invoice_number(
    company_id: str = Query(..., description="Company ID to scope the numbering"),
    invoice_type: str = Query("tax_invoice", description="Invoice type"),
    prefix: Optional[str] = Query(None, description="Custom prefix from Invoice Settings"),
    separator: str = Query("/", description="Separator character"),
    include_fy: bool = Query(True, description="Include financial year in number"),
    fy_format: str = Query("short", description="FY format: short=25-26, long=2025-2026"),
    include_month: bool = Query(False, description="Include month in number"),
    number_padding: int = Query(3, description="Zero-pad width for sequential number"),
    current_user: User = Depends(check_module_permission("invoicing", "view")),
):
    """
    Returns the next available invoice number using the exact format settings
    configured in Invoice Settings (prefix, separator, FY, month, padding).
    Scans existing invoices in the DB (MAX-based) so it is always accurate
    regardless of deletions or renames.
    """
    if not _perm(current_user):
        raise HTTPException(403, "Access denied")
    if not prefix:
        prefix = {
            "proforma":    "PRO",
            "estimate":    "EST",
            "credit_note": "CN",
            "debit_note":  "DN",
        }.get(invoice_type, "INV")
    next_no = await _next_invoice_no(
        prefix=prefix,
        company_id=company_id,
        separator=separator,
        include_fy=include_fy,
        fy_format=fy_format,
        include_month=include_month,
        number_padding=number_padding,
        invoice_type=invoice_type,
    )
    return {"invoice_no": next_no}


@router.get("/invoices/drive-status")
async def check_drive_status(current_user: User = Depends(check_module_permission("invoicing", "view"))):
    if not _perm(current_user): raise HTTPException(403, "Access denied")
    configured = _drive_configured()
    accessible = False
    if configured:
        try:
            service = _get_drive_service()
            service.files().list(pageSize=1).execute()
            accessible = True
        except Exception:
            pass
    return {
        "configured": configured,
        "accessible": accessible,
        "message": (
            "Google Drive is connected and ready" if accessible
            else "Drive credentials found but connection failed" if configured
            else "Google Drive not configured (optional feature)"
        ),
    }



# ═══════════════════════════════════════════════════════════
# INLINE STATUS UPDATE (from invoice list page dropdown)
# ═══════════════════════════════════════════════════════════

@router.patch("/invoices/{inv_id}/status")
async def update_invoice_status(
    inv_id: str,
    payload: dict,
    current_user: User = Depends(check_module_permission("invoicing", "create")),
):
    """
    Update only the status of an invoice — used by the list page dropdown.
    Accepted: draft | sent | partially_paid | paid | overdue | cancelled | credit_note

    When status is set to 'paid', amount_paid is synced to grand_total and
    amount_due is set to 0 so due/collected summaries are always correct even
    if the user never went through the formal payment recording flow.

    IMPORTANT — this quick toggle used to only touch the invoice document.
    That made the Invoicing list show "Paid / ₹0 due" while Accounting
    Reports (Party Ledger, Trial Balance, Balance Sheet) still carried the
    full invoice amount as outstanding, because nothing had ever credited
    Accounts Receivable — there was no `payments` record and no journal
    entry behind the "payment". We now auto-record the missing receipt (and
    its journal entry) the same way the formal "Record Payment" flow does,
    so every report agrees with what the dropdown shows. The auto-created
    receipt is tagged `auto_generated: True` so it can be identified and
    cleanly removed if the status is later changed away from Paid.

    When status is set to 'cancelled', amount_due is zeroed out so cancelled
    invoices don't inflate the outstanding balance.
    """
    if not _perm(current_user):
        raise HTTPException(403, "Access denied")

    new_status = (payload.get("status") or "").strip().lower()
    if new_status not in INV_STATUS:
        raise HTTPException(400, f"Invalid status. Must be one of: {INV_STATUS}")

    now_iso = datetime.now(timezone.utc).isoformat()
    history_entry = {
        "status": new_status,
        "changed_at": now_iso,
        "changed_by": getattr(current_user, "username", getattr(current_user, "email", "system")),
    }

    # Fetch current invoice so we can read grand_total for the paid sync
    inv = await db.invoices.find_one({"id": inv_id}, {"_id": 0})
    if not inv:
        raise HTTPException(404, "Invoice not found")

    fields_to_set = {"status": new_status, "updated_at": now_iso}

    # Real receipts already on file for this invoice (formal "Record
    # Payment" flow, imports, bank matches, etc.) — these already have their
    # own journal entries, so we must not double-count them.
    existing_payments = await db.payments.find({"invoice_id": inv_id}, {"_id": 0}).to_list(500)
    recorded_total = round(sum(float(p.get("amount") or 0) for p in existing_payments), 2)
    auto_payment = next((p for p in existing_payments if p.get("auto_generated")), None)

    if new_status == "paid":
        grand = float(inv.get("grand_total") or 0)
        shortfall = round(grand - recorded_total, 2)
        if shortfall > 0.004:
            # Nothing (or not enough) has actually been received for this
            # invoice yet — post the missing receipt so AR really clears.
            # payment_mode / bank_account_id come from the "Where did this
            # payment land?" dialog on the frontend (MarkPaidReceiptDialog).
            # _resolve_payment_bank_account enforces that a specific bank
            # account is chosen whenever this company has more than one on
            # file, instead of every quick-paid receipt silently landing in
            # the generic "Bank Accounts" control total with no way to tell
            # which real account received it.
            payment_mode = (payload.get("payment_mode") or "bank").strip().lower()
            if payment_mode not in ("cash", "bank"):
                payment_mode = "bank"
            bank_account = await _resolve_payment_bank_account(
                inv.get("company_id") or "", payment_mode, payload.get("bank_account_id")
            )
            bank_account_id = bank_account["id"] if bank_account else None
            if auto_payment:
                new_amount = round(float(auto_payment.get("amount") or 0) + shortfall, 2)
                await db.payments.update_one(
                    {"id": auto_payment["id"]},
                    {"$set": {"amount": new_amount, "payment_mode": payment_mode, "bank_account_id": bank_account_id}},
                )
                await sync_payment_journal_entry(auto_payment["id"])
            else:
                payment_id = str(uuid.uuid4())
                await db.payments.insert_one({
                    "id": payment_id, "invoice_id": inv_id, "amount": shortfall,
                    "payment_date": date.today().isoformat(), "payment_mode": payment_mode,
                    "bank_account_id": bank_account_id,
                    "reference_no": "", "notes": "Auto-recorded receipt — invoice marked Paid",
                    "auto_generated": True,
                    "created_by": getattr(current_user, "id", "system"), "created_at": now_iso,
                })
                await sync_payment_journal_entry(payment_id)
            recorded_total = grand
        fields_to_set["amount_paid"] = round(recorded_total, 2)
        fields_to_set["amount_due"] = 0.0

    elif new_status == "cancelled":
        # An auto-generated "quick paid" receipt was never real money in the
        # bank — remove it along with its journal entry so a cancelled
        # invoice doesn't leave a phantom credit sitting on the books.
        if auto_payment:
            await db.payments.delete_one({"id": auto_payment["id"]})
            await sync_payment_journal_entry(auto_payment["id"])
            recorded_total = round(recorded_total - float(auto_payment.get("amount") or 0), 2)
            fields_to_set["amount_paid"] = round(max(recorded_total, 0), 2)
        fields_to_set["amount_due"] = 0.0

    else:
        # Moving back to draft/sent/partially_paid: same cleanup — an
        # auto-generated receipt no longer matches reality once the user
        # says the invoice isn't (fully) paid after all.
        if auto_payment:
            await db.payments.delete_one({"id": auto_payment["id"]})
            await sync_payment_journal_entry(auto_payment["id"])
            recorded_total = round(recorded_total - float(auto_payment.get("amount") or 0), 2)
            grand = float(inv.get("grand_total") or 0)
            fields_to_set["amount_paid"] = round(max(recorded_total, 0), 2)
            fields_to_set["amount_due"] = round(max(grand - recorded_total, 0), 2)

    result = await db.invoices.update_one(
        {"id": inv_id},
        {
            "$set": fields_to_set,
            "$push": {"status_history": history_entry},
        },
    )
    if result.matched_count == 0:
        raise HTTPException(404, "Invoice not found")

    await sync_invoice_journal_entry(inv_id)

    return {
        "id":          inv_id,
        "status":      new_status,
        "amount_paid": fields_to_set.get("amount_paid", inv.get("amount_paid", 0)),
        "amount_due":  fields_to_set.get("amount_due",  inv.get("amount_due",  0)),
    }


@router.post("/invoices/reconcile-paid-receipts")
async def reconcile_paid_receipts(
    company_id: str = Query(""),
    current_user: User = Depends(check_module_permission("invoicing", "create")),
):
    """
    One-time backfill for invoices that were marked 'Paid' via the quick
    status dropdown *before* this endpoint auto-recorded a receipt for that
    action (see update_invoice_status above). Those invoices show
    "Paid / ₹0 due" on the Invoicing page but still show the full amount
    outstanding on Party Ledger / Trial Balance / Balance Sheet, because no
    payment or journal entry was ever posted for them.

    Safe to run more than once — an invoice is only touched if its recorded
    payments (db.payments) don't yet add up to its grand_total.
    """
    if not _perm(current_user):
        raise HTTPException(403, "Access denied")

    q: dict = {"status": "paid"}
    if company_id:
        q["company_id"] = company_id
    invoices = await db.invoices.find(q, {"_id": 0}).to_list(20000)

    fixed, skipped = [], 0
    now_iso = datetime.now(timezone.utc).isoformat()
    for inv in invoices:
        inv_id = inv["id"]
        grand = float(inv.get("grand_total") or 0)
        if grand <= 0:
            skipped += 1
            continue
        existing_payments = await db.payments.find({"invoice_id": inv_id}, {"_id": 0}).to_list(500)
        recorded_total = round(sum(float(p.get("amount") or 0) for p in existing_payments), 2)
        shortfall = round(grand - recorded_total, 2)
        if shortfall <= 0.004:
            skipped += 1
            continue
        payment_id = str(uuid.uuid4())
        await db.payments.insert_one({
            "id": payment_id, "invoice_id": inv_id, "amount": shortfall,
            "payment_date": inv.get("invoice_date") or date.today().isoformat(), "payment_mode": "bank",
            "reference_no": "", "notes": "Auto-recorded receipt — backfilled for invoice already marked Paid",
            "auto_generated": True,
            "created_by": getattr(current_user, "id", "system"), "created_at": now_iso,
        })
        await sync_payment_journal_entry(payment_id)
        await db.invoices.update_one(
            {"id": inv_id},
            {"$set": {"amount_paid": round(recorded_total + shortfall, 2), "amount_due": 0.0, "updated_at": now_iso}},
        )
        await sync_invoice_journal_entry(inv_id)
        fixed.append({"invoice_id": inv_id, "invoice_no": inv.get("invoice_no"), "amount_recorded": shortfall})

    return {"fixed_count": len(fixed), "skipped_count": skipped, "fixed": fixed}


@router.post("/invoices/reconcile-gst-registration")
async def reconcile_gst_registration(
    company_id: str = Query(""),
    current_user: User = Depends(check_module_permission("invoicing", "create")),
):
    """
    One-time backfill for invoices/credit/debit notes that carry CGST/SGST/
    IGST even though the issuing company profile is *not* GST-registered
    (Company Profile → "GST Registered" toggle off / `has_gst: false`).

    Before this fix, GST was computed from whatever `gst_rate` the request
    payload happened to send (or the buyer's own GSTIN), never checked
    against the issuing company's own registration — so a non-GST company
    could still end up with GST Output Payable on Trial Balance / Balance
    Sheet and GST amounts on the GSTR-1/GSTR-3B report. New invoices are now
    forced to zero GST automatically (see _compute_invoice_totals /
    _company_has_gst); this endpoint corrects documents that were already
    saved before that fix, by zeroing their tax and re-syncing the journal
    entry so every downstream report updates in one pass.

    Safe to run more than once — a document is only touched if it still
    carries non-zero GST and its company is (still) not GST-registered.
    """
    if not _perm(current_user):
        raise HTTPException(403, "Access denied")

    comp_q: dict = {"has_gst": False}
    if company_id:
        comp_q["id"] = company_id
    non_gst_companies = await db.companies.find(comp_q, {"_id": 0, "id": 1, "name": 1}).to_list(2000)
    if not non_gst_companies:
        return {"fixed_count": 0, "skipped_count": 0, "fixed": [], "message": "No non-GST-registered companies found."}

    now_iso = datetime.now(timezone.utc).isoformat()
    fixed, skipped = [], 0

    for comp in non_gst_companies:
        cid = comp["id"]
        docs = await db.invoices.find({"company_id": cid, "status": {"$ne": "cancelled"}}, {"_id": 0}).to_list(20000)
        for inv in docs:
            if float(inv.get("total_gst") or 0) <= 0.004:
                skipped += 1
                continue
            recomputed = _compute_invoice_totals(dict(inv), company_has_gst=False)
            recomputed["updated_at"] = now_iso
            recomputed["amount_due"] = round(
                recomputed["grand_total"] - float(inv.get("amount_paid") or 0), 2
            )
            for f in ("id", "created_by", "created_at", "_id"):
                recomputed.pop(f, None)
            await db.invoices.update_one({"id": inv["id"]}, {"$set": recomputed})
            await sync_invoice_journal_entry(inv["id"])
            fixed.append({
                "invoice_id": inv["id"], "invoice_no": inv.get("invoice_no"),
                "company_name": comp.get("name"), "old_total_gst": inv.get("total_gst"),
                "new_grand_total": recomputed["grand_total"],
            })

        # Quotations for the same non-GST company — zero their stored gst_rate/gst_amount too.
        qtns = await db.quotations.find({"company_id": cid}, {"_id": 0}).to_list(20000)
        for q in qtns:
            if float(q.get("gst_amount") or 0) <= 0.004:
                continue
            new_subtotal = float(q.get("subtotal") or 0)
            await db.quotations.update_one(
                {"id": q["id"]},
                {"$set": {"gst_rate": 0.0, "gst_amount": 0.0, "total": round(new_subtotal, 2), "updated_at": now_iso}},
            )
            fixed.append({"quotation_id": q["id"], "quotation_no": q.get("quotation_no"), "company_name": comp.get("name")})

    return {"fixed_count": len(fixed), "skipped_count": skipped, "fixed": fixed}


@router.get("/invoices/{inv_id}")
async def get_invoice(inv_id: str, current_user: User = Depends(check_module_permission("invoicing", "view"))):
    if not _perm(current_user): raise HTTPException(403, "Access denied")
    inv = await db.invoices.find_one({"id": inv_id}, {"_id": 0})
    if not inv: raise HTTPException(404, "Invoice not found")
    return inv


# ═══════════════════════════════════════════════════════════
# AUTOMATIC GSTIN MIGRATION (Company Master -> Invoice) ON UPDATE
# ═══════════════════════════════════════════════════════════
#
# Field names anywhere in an invoice document that represent the SELLING
# COMPANY's GSTIN (as opposed to `client_gstin`, which belongs to the buyer
# and must never be touched by this logic). Any of these keys, if already
# present on the saved invoice document or in the incoming update payload,
# are always re-stamped with the latest Company Master GSTIN before saving.
_COMPANY_GST_FIELD_NAMES = (
    "company_gstin",
    "company_gst_number",
    "seller_gstin",
    "supplier_gstin",
    "business_gstin",
    "invoice_header_gstin",
    "pdf_gstin",
    "header_gstin",
    "firm_gstin",
    "gst_number",
    "gstin",
)


async def _sync_invoice_gst_from_company(inv_id: str, ex: dict, data: dict, company_id: str) -> dict:
    """
    Automatic GSTIN migration during invoice updates.

    Loads the latest Company/Business profile linked to the invoice (by
    `company_id`) and compares its current GST-registration status/GSTIN
    against every GST-related field used anywhere on the invoice (see
    `_COMPANY_GST_FIELD_NAMES`). Any field found to be outdated is
    automatically corrected before the invoice is saved:
      - If the company IS GST-registered, every such field is stamped with
        its current GSTIN.
      - If the company is NOT GST-registered (has_gst explicitly False),
        every such field is cleared instead -- this prevents a non-GST
        company's invoice from carrying a stray/legacy seller GSTIN (which
        would otherwise cause it to incorrectly show up as GST payable /
        in GST returns even though the selling company charges no GST).
    If a stored PDF/Drive snapshot exists, it is cleared so the next
    generated PDF/print reflects the corrected GSTIN.

    This function is strictly additive and read-only with respect to every
    other part of the invoice: it never touches invoice number, invoice
    date, customer information, items, HSN codes, quantities, taxable
    values, CGST/SGST/IGST calculations, totals, ledger entries, journal
    entries, payment status, or bank reconciliation data.
    """
    if not company_id:
        return data

    company = await db.companies.find_one({"id": company_id}, {"_id": 0, "gstin": 1, "has_gst": 1})
    if company is None:
        return data

    # Same semantics as `_company_has_gst`: GST-registered unless the
    # company profile explicitly sets has_gst = False.
    company_is_gst_registered = company.get("has_gst", True) is not False
    current_gstin = (company.get("gstin") or "").strip() if company_is_gst_registered else ""

    fields_synced = []
    for field in _COMPANY_GST_FIELD_NAMES:
        # Only touch a field if it is already used somewhere on this invoice
        # (either present in the incoming payload or already stored on the
        # existing invoice document) -- never invent new fields.
        if field not in data and field not in ex:
            continue
        existing_value = str(data.get(field, ex.get(field)) or "").strip()
        if existing_value != current_gstin:
            data[field] = current_gstin
            fields_synced.append(field)

    if fields_synced:
        # Force regeneration of any stored invoice print/PDF snapshot so the
        # next generated/uploaded copy reflects the corrected GSTIN.
        data["pdf_drive_link"] = ""

    logger.info(
        f"Invoice GST fields synchronized from Company Master "
        f"(invoice_id={inv_id}, company_id={company_id}, "
        f"company_gst_registered={company_is_gst_registered}, current_gstin={current_gstin!r}, "
        f"fields_updated={fields_synced})"
    )
    return data


@router.put("/invoices/{inv_id}")
async def update_invoice(inv_id: str, data: dict, current_user: User = Depends(check_module_permission("invoicing", "create"))):
    if not _perm(current_user): raise HTTPException(403, "Access denied")
    ex = await db.invoices.find_one({"id": inv_id})
    if not ex: raise HTTPException(404, "Invoice not found")

    old_company_id  = (ex.get("company_id") or "").strip()
    new_company_id  = (data.get("company_id") or old_company_id).strip()
    company_changed = bool(new_company_id and new_company_id != old_company_id)

    # -- Handle invoice_no when company changes ----------------------------------
    new_invoice_no = (data.get("invoice_no") or "").strip()
    old_invoice_no = ex.get("invoice_no", "")

    if company_changed:
        # Company changed: always assign next sequential number for the NEW company.
        # The old company's number must not be carried over because it would break
        # the sequence of BOTH companies.
        invoice_type = (data.get("invoice_type") or ex.get("invoice_type") or "tax_invoice")
        prefix_map = {"proforma": "PRO", "estimate": "EST", "credit_note": "CN", "debit_note": "DN"}
        prefix = prefix_map.get(invoice_type, "INV")
        auto_no = await _next_invoice_no(
            prefix=prefix,
            company_id=new_company_id,
            separator="/",
            include_fy=True,
            fy_format="short",
            include_month=False,
            number_padding=3,
            invoice_type=invoice_type,
        )

        if new_invoice_no and new_invoice_no != old_invoice_no:
            # Frontend sent a new explicit number (pre-fetched from /next-number).
            # Verify it is not already taken in the new company.
            dup_filter = {"invoice_no": new_invoice_no, "id": {"$ne": inv_id}, "company_id": new_company_id}
            conflict = await db.invoices.find_one(dup_filter)
            if conflict:
                new_invoice_no = auto_no  # fall back to server-generated number
        else:
            # Frontend sent old number or nothing -- override with next-in-sequence.
            new_invoice_no = auto_no

        data["invoice_no"] = new_invoice_no
        logger.info(
            f"UPDATE invoice {inv_id}: company changed {old_company_id!r} -> {new_company_id!r}, "
            f"invoice_no reassigned {old_invoice_no!r} -> {new_invoice_no!r}"
        )
    elif new_invoice_no and new_invoice_no != old_invoice_no:
        # Same company, user manually changed the number -- duplicate check within same company.
        dup_filter = {"invoice_no": new_invoice_no, "id": {"$ne": inv_id}}
        if new_company_id:
            dup_filter["company_id"] = new_company_id
        conflict = await db.invoices.find_one(dup_filter)
        if conflict:
            raise HTTPException(400, f"Invoice number '{new_invoice_no}' is already in use by another invoice")
        # Allow the update -- will be included in the $set below
    elif not new_invoice_no:
        # If empty string sent, keep the original number
        data["invoice_no"] = old_invoice_no

    # -- Strip truly immutable fields -------------------------------------------
    for f in ("id", "created_by", "created_at", "amount_paid", "pdf_drive_link"):
        data.pop(f, None)

    # -- Automatic GSTIN migration: sync all company/seller GST fields on the
    #    invoice with the latest Company Master GSTIN before saving. This never
    #    touches invoice number/date, customer info, items, HSN codes, quantities,
    #    taxable values, GST calculations, totals, or any ledger/journal/payment data.
    data = await _sync_invoice_gst_from_company(inv_id, ex, data, new_company_id)

    data["updated_at"] = datetime.now(timezone.utc).isoformat()
    data = _compute_invoice_totals(data, await _company_has_gst(new_company_id))
    data["amount_due"] = round(data["grand_total"] - ex.get("amount_paid", 0), 2)
    logger.info(
        f"UPDATE invoice {inv_id}: $set keys = {list(data.keys())}, "
        f"invoice_no = {data.get('invoice_no', 'NOT PRESENT')!r}, "
        f"company_id = {data.get('company_id', 'NOT PRESENT')!r}"
    )
    await db.invoices.update_one({"id": inv_id}, {"$set": data})
    await sync_invoice_journal_entry(inv_id)
    return await db.invoices.find_one({"id": inv_id}, {"_id": 0})

@router.delete("/invoices/bulk-delete")
async def bulk_delete_invoices(ids: List[str], current_user: User = Depends(check_module_permission("invoicing", "delete"))):
    if not _perm(current_user): raise HTTPException(403, "Access denied")
    if not ids: raise HTTPException(400, "No IDs provided")
    deleted, failed = 0, 0
    for inv_id in ids:
        try:
            # Delete journal entries for associated payments
            payments = await db.payments.find({"invoice_id": inv_id}).to_list(1000)
            for p in payments:
                existing_pe = await db.journal_entries.find_one({"source": "payment", "source_id": p["id"]})
                if existing_pe:
                    await db.journal_lines.delete_many({"entry_id": existing_pe["id"]})
                    await db.journal_entries.delete_one({"id": existing_pe["id"]})
            await db.payments.delete_many({"invoice_id": inv_id})

            r = await db.invoices.delete_one({"id": inv_id})
            if r.deleted_count:
                deleted += 1
                await sync_invoice_journal_entry(inv_id)
            else:
                failed += 1
        except Exception: failed += 1
    return {"deleted": deleted, "failed": failed, "total": len(ids)}


@router.delete("/invoices/{inv_id}")
async def delete_invoice(inv_id: str, current_user: User = Depends(check_module_permission("invoicing", "delete"))):
    if not _perm(current_user): raise HTTPException(403, "Access denied")
    
    # Delete journal entries for associated payments
    payments = await db.payments.find({"invoice_id": inv_id}).to_list(1000)
    for p in payments:
        existing_pe = await db.journal_entries.find_one({"source": "payment", "source_id": p["id"]})
        if existing_pe:
            await db.journal_lines.delete_many({"entry_id": existing_pe["id"]})
            await db.journal_entries.delete_one({"id": existing_pe["id"]})
    await db.payments.delete_many({"invoice_id": inv_id})

    result = await db.invoices.delete_one({"id": inv_id})
    if result.deleted_count == 0: raise HTTPException(404, "Invoice not found")
    await sync_invoice_journal_entry(inv_id)
    return {"message": f"Invoice {inv_id} deleted"}


# ═══════════════════════════════════════════════════════════
# PDF DOWNLOAD — ALWAYS STREAMS LOCALLY
# ═══════════════════════════════════════════════════════════

@router.get("/invoices/{inv_id}/pdf")
async def download_invoice_pdf(
    inv_id: str,
    background_tasks: BackgroundTasks,
    current_user: User = Depends(check_module_permission("invoicing", "view")),
):
    """
    Generates a fresh PDF, streams it as a file download, AND - when Google
    Drive is configured - automatically uploads it to Drive in the background.
    One click = browser download + Drive backup. No separate call required.
    """
    if not _perm(current_user):
        raise HTTPException(403, "Access denied")

    inv = await db.invoices.find_one({"id": inv_id}, {"_id": 0})
    if not inv:
        raise HTTPException(404, "Invoice not found")

    company = await db.companies.find_one({"id": inv.get("company_id")}, {"_id": 0}) or {}

    try:
        pdf_buf = _build_invoice_pdf(inv, company)
    except Exception as e:
        logger.error(f"PDF generation failed for {inv_id}: {e}", exc_info=True)
        raise HTTPException(500, f"PDF generation failed: {e}")

    safe_name = (inv.get("invoice_no") or inv_id).replace("/", "_").replace("\\", "_")
    company_prefix = (company.get("name", "") or "").strip().replace(" ", "_").replace("/", "_").replace("\\", "_")
    filename = f"{company_prefix}_Invoice_{safe_name}.pdf" if company_prefix else f"Invoice_{safe_name}.pdf"
    pdf_bytes = pdf_buf.getvalue()

    # Auto-upload to Google Drive in the background (if configured)
    if _drive_configured():
        async def _bg_upload(inv_data: dict, raw_bytes: bytes, fname: str):
            try:
                client_name = inv_data.get("client_name", "").strip()
                try:
                    folder_id = _get_or_create_client_folder(client_name)
                except Exception:
                    folder_id = DRIVE_FOLDERS["invoices"]
                link = await _upload_to_drive(
                    raw_bytes, fname, "invoices", "application/pdf",
                    custom_parent_id=folder_id,
                )
                if link:
                    await db.invoices.update_one(
                        {"id": inv_data["id"]},
                        {"$set": {"pdf_drive_link": link,
                                  "updated_at": datetime.now(timezone.utc).isoformat()}},
                    )
                    logger.info(f"Auto-uploaded to Drive: {fname} -> {link}")
            except Exception as exc:
                logger.warning(f"Background Drive upload failed for {fname}: {exc}")

        background_tasks.add_task(_bg_upload, dict(inv), pdf_bytes, filename)

    return StreamingResponse(
        BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Content-Type": "application/pdf",
            "Cache-Control": "no-cache, no-store, must-revalidate",
        },
    )


# ═══════════════════════════════════════════════════════════
# OPTIONAL: UPLOAD PDF TO GOOGLE DRIVE
# ═══════════════════════════════════════════════════════════

@router.post("/invoices/{inv_id}/upload-to-drive")
async def upload_invoice_to_drive(inv_id: str, current_user: User = Depends(check_module_permission("invoicing", "create"))):
    """
    Optional: generate PDF and upload to Google Drive.
    Only works if Drive credentials are configured.
    """
    if not _perm(current_user):
        raise HTTPException(403, "Access denied")

    if not _drive_configured():
        raise HTTPException(
            503,
            "Google Drive is not configured. Add GOOGLE_SERVICE_ACCOUNT_JSON as an environment variable.",
        )

    inv = await db.invoices.find_one({"id": inv_id}, {"_id": 0})
    if not inv: raise HTTPException(404, "Invoice not found")

    company = await db.companies.find_one({"id": inv.get("company_id")}, {"_id": 0}) or {}

    try:
        pdf_buf = _build_invoice_pdf(inv, company)
        pdf_bytes = pdf_buf.getvalue()
    except Exception as e:
        raise HTTPException(500, f"PDF generation failed: {e}")

    client_name = inv.get("client_name", "").strip()
    try:
        client_folder_id = _get_or_create_client_folder(client_name)
    except Exception:
        client_folder_id = DRIVE_FOLDERS["invoices"]

    safe_inv_no = (inv.get("invoice_no") or inv_id).replace("/", "_").replace("\\", "_")
    company_prefix = (company.get("name", "") or "").strip().replace(" ", "_").replace("/", "_").replace("\\", "_")

    # ✅ FIX v6.2: added `await` to both _upload_to_drive calls —
    #    previously the coroutine object itself was stored in pdf_link,
    #    causing bson.errors.InvalidDocument when saving to MongoDB.
    _inv_fname = f"{company_prefix}_Invoice_{safe_inv_no}.pdf" if company_prefix else f"Invoice_{safe_inv_no}.pdf"
    pdf_link = await _upload_to_drive(
        pdf_bytes, _inv_fname, "invoices", "application/pdf",
        custom_parent_id=client_folder_id
    )
    _json_fname = f"{company_prefix}_Invoice_{safe_inv_no}.json" if company_prefix else f"Invoice_{safe_inv_no}.json"
    await _upload_to_drive(
        json.dumps(inv, default=str).encode(), _json_fname,
        "invoices", "application/json", custom_parent_id=client_folder_id
    )

    if pdf_link:
        await db.invoices.update_one({"id": inv_id},
            {"$set": {"pdf_drive_link": pdf_link, "updated_at": datetime.now(timezone.utc).isoformat()}})

    return {
        "status": "success" if pdf_link else "warning",
        "drive_link": pdf_link or "",
        "message": "Invoice uploaded to Google Drive" if pdf_link else "Upload failed — check Drive credentials",
        "invoice_no": inv.get("invoice_no", ""),
    }


# ═══════════════════════════════════════════════════════════
# CONVERT QUOTATION → INVOICE
# ═══════════════════════════════════════════════════════════

@router.post("/invoices/from-quotation/{qtn_id}")
async def convert_quotation(qtn_id: str, current_user: User = Depends(check_module_permission("invoicing", "create"))):
    if not _perm(current_user): raise HTTPException(403, "Access denied")
    q = await db.quotations.find_one({"id": qtn_id}, {"_id": 0})
    if not q: raise HTTPException(404, "Quotation not found")

    # Prevent duplicate conversion — return the existing invoice if one is already linked.
    if q.get("invoice_id"):
        existing = await db.invoices.find_one({"id": q["invoice_id"]}, {"_id": 0})
        if existing:
            existing.pop("_id", None)
            return existing

    # Resolve template/theme: prefer values stored on the quotation, then fall
    # back to the company's defaults so the converted invoice keeps its identity.
    company = await db.companies.find_one({"id": q.get("company_id")}, {"_id": 0}) or {}
    inv_template = q.get("invoice_template") or company.get("invoice_template") or "prestige"
    inv_theme    = q.get("invoice_theme")    or company.get("invoice_theme")    or "classic_blue"
    inv_color    = q.get("invoice_custom_color") or company.get("invoice_custom_color") or "#0D3B66"

    inv_items = [InvoiceItem(description=it.get("description", ""),
        quantity=float(it.get("quantity", 1)), unit=it.get("unit", "service"),
        unit_price=float(it.get("unit_price", 0)), gst_rate=float(q.get("gst_rate", 18)))
        for it in q.get("items", [])]
    create_data = InvoiceCreate(
        invoice_type="tax_invoice", company_id=q.get("company_id", ""),
        quotation_id=qtn_id, lead_id=q.get("lead_id"), client_id=q.get("client_id"),
        client_name=q.get("client_name", ""), client_address=q.get("client_address", ""),
        client_email=q.get("client_email", ""), client_phone=q.get("client_phone", ""),
        client_gstin=q.get("client_gstin", "") or "",
        client_state=q.get("client_state", "") or "",
        items=inv_items, gst_rate=q.get("gst_rate", 18),
        invoice_template=inv_template, invoice_theme=inv_theme,
        invoice_custom_color=inv_color,
        payment_terms=q.get("payment_terms", ""), notes=q.get("notes", ""), status="draft")
    invoice = await create_invoice(create_data, current_user)
    # Normalise: create_invoice may return a plain dict or a Pydantic model
    # depending on FastAPI version / call-site.  Handle both.
    if isinstance(invoice, dict):
        inv_dict = invoice
    else:
        try:
            inv_dict = invoice.model_dump()
        except AttributeError:
            inv_dict = dict(invoice)
    inv_id = inv_dict.get("id")
    inv_no = inv_dict.get("invoice_no")
    # Backlink: store invoice_id + invoice_no + status on the source quotation
    # so the UI knows it's been converted and can prevent duplicate conversion.
    if inv_id:
        try:
            await db.quotations.update_one(
                {"id": qtn_id},
                {"$set": {
                    "status": "converted",
                    "invoice_id": inv_id,
                    "invoice_no": inv_no,
                    "converted_at": datetime.now(timezone.utc).isoformat(),
                    "updated_at": datetime.now(timezone.utc).isoformat(),
                }},
            )
        except Exception as backlink_err:
            logger.error(
                f"convert_quotation: failed to write backlink on qtn {qtn_id} "
                f"(invoice {inv_id}): {backlink_err}",
                exc_info=True,
            )
    else:
        logger.error(
            f"convert_quotation: create_invoice returned no 'id' for qtn {qtn_id}; "
            f"backlink NOT saved.  invoice payload: {inv_dict!r}"
        )
    return inv_dict


# ═══════════════════════════════════════════════════════════
# SEND EMAIL
# ═══════════════════════════════════════════════════════════

@router.post("/invoices/{inv_id}/send-email")
async def send_invoice_email(inv_id: str, background_tasks: BackgroundTasks,
                              current_user: User = Depends(check_module_permission("invoicing", "create"))):
    if not _perm(current_user): raise HTTPException(403, "Access denied")
    inv = await db.invoices.find_one({"id": inv_id}, {"_id": 0})
    if not inv: raise HTTPException(404, "Invoice not found")
    if not inv.get("client_email"): raise HTTPException(400, "Client email not set")
    company = await db.companies.find_one({"id": inv.get("company_id")}, {"_id": 0}) or {}
    try:
        pdf_bytes = _build_invoice_pdf(inv, company).getvalue()
    except Exception as e:
        raise HTTPException(500, f"PDF generation failed: {e}")
    inv_no = inv.get("invoice_no", inv_id)
    subject = f"Invoice {inv_no} from {company.get('name', 'Your Company')}"
    html_body = f"""<h2>Invoice {inv_no}</h2>
    <p>Dear {inv.get('client_name', 'Customer')},</p>
    <p>Please find attached invoice <strong>{inv_no}</strong> for <strong>Rs.{inv.get('grand_total', 0):,.2f}</strong>.</p>
    <p>Due Date: {inv.get('due_date', 'N/A')}</p><br>
    <p>Regards,<br>{company.get('name', 'Your Company')}</p>"""
    background_tasks.add_task(_send_email, inv["client_email"], subject, html_body,
                               pdf_bytes, f"Invoice_{inv_no}.pdf", company.get("email", ""))
    await db.invoices.update_one({"id": inv_id},
        {"$set": {"status": "sent", "updated_at": datetime.now(timezone.utc).isoformat()}})
    return {"message": f"Email queued for {inv['client_email']}", "invoice_no": inv_no}


# ═══════════════════════════════════════════════════════════
# MARK SENT
# ═══════════════════════════════════════════════════════════

@router.post("/invoices/{inv_id}/mark-sent")
async def mark_invoice_sent(inv_id: str, current_user: User = Depends(check_module_permission("invoicing", "create"))):
    if not _perm(current_user): raise HTTPException(403, "Access denied")
    result = await db.invoices.update_one({"id": inv_id},
        {"$set": {"status": "sent", "updated_at": datetime.now(timezone.utc).isoformat()}})
    if result.matched_count == 0: raise HTTPException(404, "Invoice not found")
    return {"message": f"Invoice {inv_id} marked as sent", "status": "sent"}


# ═══════════════════════════════════════════════════════════
# RECURRING INVOICE GENERATOR
# ═══════════════════════════════════════════════════════════

@router.post("/invoices/{inv_id}/generate-recurring")
async def generate_recurring(inv_id: str, current_user: User = Depends(check_module_permission("invoicing", "create"))):
    if not _perm(current_user): raise HTTPException(403, "Access denied")
    template = await db.invoices.find_one({"id": inv_id}, {"_id": 0})
    if not template: raise HTTPException(404, "Template invoice not found")
    now = datetime.now(timezone.utc).isoformat()
    _rec_type = template.get("invoice_type", "tax_invoice")
    _rec_prefix = {"proforma": "PRO", "estimate": "EST", "credit_note": "CN", "debit_note": "DN"}.get(_rec_type, "INV")
    new_inv = {**template, "id": str(uuid.uuid4()), "invoice_no": await _next_invoice_no(_rec_prefix, template.get("company_id"), invoice_type=_rec_type),
               "invoice_date": date.today().isoformat(),
               "due_date": (date.today() + timedelta(days=30)).isoformat(),
               "status": "draft", "amount_paid": 0.0, "amount_due": template.get("grand_total", 0),
               "is_recurring": False, "created_at": now, "updated_at": now, "pdf_drive_link": ""}
    new_inv.pop("_id", None)
    await db.invoices.insert_one({**new_inv})
    new_inv.pop("_id", None)
    return {"status": "success", "invoice_no": new_inv["invoice_no"], "id": new_inv["id"]}


# ═══════════════════════════════════════════════════════════
# PAYMENT ENDPOINTS
# ═══════════════════════════════════════════════════════════

@router.post("/payments")
async def record_payment(data: PaymentCreate, current_user: User = Depends(check_module_permission("invoicing", "create"))):
    if not _perm(current_user): raise HTTPException(403, "Access denied")
    inv = await db.invoices.find_one({"id": data.invoice_id})
    if not inv: raise HTTPException(404, "Invoice not found")
    bank_account = await _resolve_payment_bank_account(
        inv.get("company_id") or "", data.payment_mode, data.bank_account_id
    )
    payment_data = {**data.model_dump(), "id": str(uuid.uuid4()),
                    "bank_account_id": bank_account["id"] if bank_account else None,
                    "created_by": current_user.id, "created_at": datetime.now(timezone.utc).isoformat()}
    await db.payments.insert_one({**payment_data})
    await sync_payment_journal_entry(payment_data["id"])
    payment_data.pop("_id", None)
    all_payments = await db.payments.find({"invoice_id": data.invoice_id}, {"_id": 0}).to_list(500)
    total_paid = sum(float(p.get("amount", 0)) for p in all_payments)
    grand_total = float(inv.get("grand_total", 0))
    amount_due = round(max(grand_total - total_paid, 0), 2)
    new_status = "paid" if amount_due <= 0 else ("partially_paid" if total_paid > 0 else inv.get("status", "sent"))
    await db.invoices.update_one({"id": data.invoice_id},
        {"$set": {"amount_paid": round(total_paid, 2), "amount_due": amount_due, "status": new_status,
                  "updated_at": datetime.now(timezone.utc).isoformat()}})
    await sync_invoice_journal_entry(data.invoice_id)
    return payment_data


@router.get("/payments")
async def list_payments(invoice_id: Optional[str] = None, current_user: User = Depends(check_module_permission("invoicing", "view"))):
    if not _perm(current_user): raise HTTPException(403, "Access denied")
    q: dict = {}
    if invoice_id: q["invoice_id"] = invoice_id
    return await db.payments.find(q, {"_id": 0}).sort("created_at", -1).to_list(1000)


@router.delete("/payments/{pid}")
async def delete_payment(pid: str, current_user: User = Depends(check_module_permission("invoicing", "delete"))):
    if not _perm(current_user): raise HTTPException(403, "Access denied")
    payment = await db.payments.find_one({"id": pid})
    if not payment: raise HTTPException(404, "Payment not found")
    invoice_id = payment.get("invoice_id")

    # Delete from DB
    result = await db.payments.delete_one({"id": pid})
    if result.deleted_count == 0: raise HTTPException(404, f"Payment {pid} not found")

    # Sync payment to ledger (will delete its journal entry)
    await sync_payment_journal_entry(pid)

    # Recalculate invoice payments & status
    if invoice_id:
        inv = await db.invoices.find_one({"id": invoice_id})
        if inv:
            all_payments = await db.payments.find({"invoice_id": invoice_id}, {"_id": 0}).to_list(500)
            total_paid = sum(float(p.get("amount", 0)) for p in all_payments)
            grand_total = float(inv.get("grand_total", 0))
            amount_due = round(max(grand_total - total_paid, 0), 2)
            new_status = "paid" if amount_due <= 0 else ("partially_paid" if total_paid > 0 else "sent")
            await db.invoices.update_one({"id": invoice_id},
                {"$set": {"amount_paid": round(total_paid, 2), "amount_due": amount_due, "status": new_status,
                          "updated_at": datetime.now(timezone.utc).isoformat()}})
            await sync_invoice_journal_entry(invoice_id)

    return {"message": f"Payment {pid} deleted"}


# ═══════════════════════════════════════════════════════════
# CREDIT NOTES
# ═══════════════════════════════════════════════════════════

@router.post("/credit-notes")
async def create_credit_note(data: CreditNoteCreate, current_user: User = Depends(check_module_permission("invoicing", "create"))):
    if not _perm(current_user): raise HTTPException(403, "Access denied")
    inv_no = await _next_invoice_no("CN", data.company_id, invoice_type="credit_note")
    now = datetime.now(timezone.utc).isoformat()
    raw = {"id": str(uuid.uuid4()), "invoice_no": inv_no, "invoice_type": "credit_note",
           **data.model_dump(), "invoice_date": date.today().isoformat(), "due_date": date.today().isoformat(),
           "created_by": current_user.id, "created_at": now, "updated_at": now,
           "status": "credit_note", "amount_paid": 0, "amount_due": 0, "pdf_drive_link": ""}
    raw = _compute_invoice_totals(raw, await _company_has_gst(data.company_id))


# ═══════════════════════════════════════════════════════════
# LEDGER SYNC HELPERS (links invoicing and payments to ledger)
# ═══════════════════════════════════════════════════════════

async def _resolve_payment_bank_account(company_id: str, payment_mode: str, bank_account_id: Optional[str]) -> Optional[dict]:
    """Validate the bank account a receipt/payment says it landed in.

    - payment_mode == "cash": no bank account involved, always returns None.
    - payment_mode == "bank": if the company has one or more accounts on
      file in Accounts › Bank, a valid bank_account_id belonging to that
      company is required — this is what forces "which bank account" to be
      asked (and answered) whenever there's more than one on file, instead
      of every receipt silently landing in the generic "Bank Accounts"
      total with no way to tell which real account actually received it.
      A company with zero bank accounts on file is allowed to proceed
      without one (nothing to pick from yet); the receipt still posts to
      the control account so the books aren't blocked on account setup.
    """
    mode = (payment_mode or "bank").strip().lower()
    if mode == "cash":
        return None

    accounts = await db.bank_accounts.find({"company_id": company_id}, {"_id": 0}).to_list(500)
    if not accounts:
        return None  # nothing on file yet — fall back to the control total

    if bank_account_id:
        acct = next((a for a in accounts if a["id"] == bank_account_id), None)
        if not acct:
            raise HTTPException(400, "Selected bank account was not found for this company.")
        return acct

    if len(accounts) == 1:
        return accounts[0]

    raise HTTPException(
        400,
        "This company has more than one bank account on file — please choose which "
        "bank account received this payment.",
    )


async def _sync_bank_transaction_for_payment(
    *, payment_id: str, matched_type: str, bank_account: Optional[dict], company_id: str,
    amount: float, txn_date: str, description: str, journal_entry_id: Optional[str], direction: str,
):
    """Keep the specific bank account's transaction feed (and therefore its
    balance on the Bank Accounts page — see list_bank_accounts) in step with
    receipts/payments recorded from the Invoicing/Purchase pages, the same
    way an uploaded bank statement would. `direction` is "credit" (money in,
    e.g. a customer receipt) or "debit" (money out, e.g. paying a vendor).
    Always clears any previous transaction row for this payment first so
    edits/deletes/status-reverts never leave a stale balance behind.
    """
    await db.bank_transactions.delete_many({"matched_id": payment_id, "matched_type": matched_type})
    if not bank_account or amount <= 0:
        return
    now = datetime.now(timezone.utc).isoformat()
    await db.bank_transactions.insert_one({
        "id": str(uuid.uuid4()), "bank_account_id": bank_account["id"], "company_id": company_id,
        "date": txn_date, "description": description, "reference": "",
        "debit": amount if direction == "debit" else 0.0,
        "credit": amount if direction == "credit" else 0.0,
        "balance_after": None,
        "matched_type": matched_type, "matched_id": payment_id, "matched_label": description,
        "journal_entry_id": journal_entry_id, "source_file": "", "created_by": "system", "created_at": now,
    })


async def sync_invoice_journal_entry(invoice_id: str):
    from backend.accounting_core import get_default_account_id, post_journal_entry
    
    # 1. Clean up any existing journal entries for this invoice
    _old_entries = await db.journal_entries.find({"source": "sale", "source_id": invoice_id}, {"_id": 0, "id": 1}).to_list(50)
    if _old_entries:
        _old_ids = [e["id"] for e in _old_entries]
        await db.journal_lines.delete_many({"entry_id": {"$in": _old_ids}})
        await db.journal_entries.delete_many({"id": {"$in": _old_ids}})
        
    # 2. Fetch the current invoice document
    inv = await db.invoices.find_one({"id": invoice_id})
    if not inv:
        return
        
    # 3. If invoice is cancelled we don't post a journal entry. Draft
    # invoices ARE now posted to the books — per business rule, a raised
    # invoice (even in "draft"/"Invoiced" state) is recognised revenue so
    # that Accounting Reports (Trial Balance / P&L / Balance Sheet) reflect
    # the same figure as the Sales/Invoicing page.
    status = inv.get("status", "draft")
    if status == "cancelled":
        return

    # 3b. Proforma invoices / estimates are quotations, not real invoices —
    # no GST liability or revenue has actually been raised yet, so they
    # must never be posted to Accounts Receivable / Sales. Posting them
    # (as this used to, since only "credit_note" got special-cased below)
    # inflated Trial Balance / P&L / Balance Sheet figures above what the
    # Sales/Invoicing page reports as real revenue.
    invoice_type = inv.get("invoice_type") or "tax_invoice"
    if invoice_type in ("proforma", "estimate"):
        return

    # 4. Extract required fields
    company_id = inv.get("company_id") or ""
    invoice_no = inv.get("invoice_no") or ""
    client_name = inv.get("client_name") or "Client"
    invoice_date = inv.get("invoice_date") or date.today().isoformat()
    
    grand_total = float(inv.get("grand_total") or 0)
    total_gst = float(inv.get("total_gst") or 0)
    
    if grand_total <= 0:
        return
        
    # 5. Get default ledger account IDs
    ar_id = await get_default_account_id(company_id, "1100")  # Accounts Receivable
    sales_id = await get_default_account_id(company_id, "4000")  # Sales / Fee Income
    gst_pay_id = await get_default_account_id(company_id, "2100")  # GST Output Payable
    
    if not ar_id or not sales_id or not gst_pay_id:
        return
        
    sales_amount = round(grand_total - total_gst, 2)
    
    # Create the double-entry lines
    if invoice_type == "credit_note" or status == "credit_note":
        # Credit Note logic: reduces accounts receivable and sales revenue
        lines = [
            {"account_id": sales_id, "account_name": "Sales / Fee Income", "debit": sales_amount, "credit": 0.0, "memo": f"Reversal of sales for Credit Note {invoice_no}"},
            {"account_id": ar_id, "account_name": "Accounts Receivable", "debit": 0.0, "credit": grand_total, "memo": f"Credit Note {invoice_no} to {client_name}"}
        ]
        if total_gst > 0:
            lines.append({"account_id": gst_pay_id, "account_name": "GST Output Payable", "debit": total_gst, "credit": 0.0, "memo": f"GST Reversal on Credit Note {invoice_no}"})
        narration = f"Credit Note {invoice_no} issued to {client_name}"
    else:
        # Standard invoice logic: increases accounts receivable and sales revenue
        lines = [
            {"account_id": ar_id, "account_name": "Accounts Receivable", "debit": grand_total, "credit": 0.0, "memo": f"Invoice {invoice_no} to {client_name}"},
            {"account_id": sales_id, "account_name": "Sales / Fee Income", "debit": 0.0, "credit": sales_amount, "memo": f"Sales revenue from Invoice {invoice_no}"}
        ]
        if total_gst > 0:
            lines.append({"account_id": gst_pay_id, "account_name": "GST Output Payable", "debit": 0.0, "credit": total_gst, "memo": f"GST Output on Invoice {invoice_no}"})
        narration = f"Sales Invoice {invoice_no} to {client_name}"
        
    try:
        await post_journal_entry(
            company_id=company_id,
            entry_date=invoice_date,
            narration=narration,
            lines=lines,
            source="sale",
            source_id=invoice_id,
            created_by=inv.get("created_by", "system")
        )
    except Exception as e:
        logging.error(f"Error posting invoice journal entry: {e}")


async def sync_payment_journal_entry(payment_id: str):
    from backend.accounting_core import get_default_account_id, post_journal_entry
    
    # 1. Clean up any existing journal entries for this payment
    _old_entries = await db.journal_entries.find({"source": "payment", "source_id": payment_id}, {"_id": 0, "id": 1}).to_list(50)
    if _old_entries:
        _old_ids = [e["id"] for e in _old_entries]
        await db.journal_lines.delete_many({"entry_id": {"$in": _old_ids}})
        await db.journal_entries.delete_many({"id": {"$in": _old_ids}})
        
    # 2. Fetch the current payment document
    payment = await db.payments.find_one({"id": payment_id})
    if not payment:
        return
        
    invoice_id = payment.get("invoice_id")
    inv = await db.invoices.find_one({"id": invoice_id}) if invoice_id else None
    invoice_no = inv.get("invoice_no") if inv else None
    # Standalone/imported receipts (e.g. bank-matched or bulk-imported
    # payments) carry their own client_name and have no invoice_id at all —
    # fall back to that instead of the generic "Client" placeholder, and
    # only mention "Client" as a last resort.
    client_name = (inv.get("client_name") if inv else None) or payment.get("client_name") or "Client"

    company_id = payment.get("company_id") or (inv.get("company_id") if inv else "")
    amount = float(payment.get("amount") or 0)
    payment_date = payment.get("payment_date") or date.today().isoformat()
    payment_mode = str(payment.get("payment_mode") or "bank").strip().lower()
    bank_account_id = payment.get("bank_account_id")
    
    if amount <= 0:
        # Nothing recorded for this payment (e.g. it was deleted or zeroed
        # out) — make sure any bank-specific transaction row from a previous
        # sync is cleared too, or that account's balance would stay wrong.
        await db.bank_transactions.delete_many({"matched_id": payment_id, "matched_type": "sale_payment"})
        return
        
    # 3. Resolve accounts
    ar_id = await get_default_account_id(company_id, "1100")  # Accounts Receivable
    cash_id = await get_default_account_id(company_id, "1000")  # Cash in Hand
    bank_id = await get_default_account_id(company_id, "1010")  # Bank Accounts
    
    if not ar_id or not cash_id or not bank_id:
        return
        
    # Decide debit account based on payment mode
    debit_acct_id = cash_id if payment_mode == "cash" else bank_id
    debit_acct_name = "Cash in Hand" if payment_mode == "cash" else "Bank Accounts"
    
    invoice_ref = f"Inv {invoice_no}" if invoice_no else "(no invoice linked)"
    lines = [
        {"account_id": debit_acct_id, "account_name": debit_acct_name, "debit": amount, "credit": 0.0, "memo": f"Receipt from {client_name} via {payment_mode.upper()}"},
        {"account_id": ar_id, "account_name": "Accounts Receivable", "debit": 0.0, "credit": amount, "memo": f"Payment by {client_name} {invoice_ref}"}
    ]
    
    narration = (
        f"Receipt from {client_name} for Invoice {invoice_no} ({payment_mode.upper()})"
        if invoice_no else
        f"Receipt from {client_name} ({payment_mode.upper()})"
    )
    
    entry_id = None
    try:
        entry = await post_journal_entry(
            company_id=company_id,
            entry_date=payment_date,
            narration=narration,
            lines=lines,
            source="payment",
            source_id=payment_id,
            created_by=payment.get("created_by", "system")
        )
        entry_id = entry.get("id") if entry else None
    except Exception as e:
        logging.error(f"Error posting payment journal entry: {e}")

    # 4. Mirror the receipt into the specific bank account's own transaction
    # feed (if one was recorded) so its balance on the Bank Accounts page
    # includes money received through Invoicing — not just uploaded bank
    # statements — the same way this receipt already updates the generic
    # "Bank Accounts" control total above.
    bank_account = None
    if payment_mode != "cash" and bank_account_id:
        bank_account = await db.bank_accounts.find_one({"id": bank_account_id, "company_id": company_id}, {"_id": 0})
    await _sync_bank_transaction_for_payment(
        payment_id=payment_id, matched_type="sale_payment", bank_account=bank_account, company_id=company_id,
        amount=amount, txn_date=payment_date, direction="credit", journal_entry_id=entry_id,
        description=(f"Receipt from {client_name} for Invoice {invoice_no}" if invoice_no else f"Receipt from {client_name}"),
    )


async def sync_purchase_journal_entry(invoice_id: str):
    """Purchase-side mirror of sync_invoice_journal_entry: Dr Purchases (+ Dr GST
    Input Credit if any), Cr Accounts Payable. Posted the moment a purchase bill
    is saved/edited (no draft gate — a saved purchase bill is already a real
    liability), and removed again if the bill is cancelled or deleted."""
    from backend.accounting_core import get_default_account_id, post_journal_entry

    # 1. Clean up any existing journal entry for this purchase bill
    _old_entries = await db.journal_entries.find({"source": "purchase", "source_id": invoice_id}, {"_id": 0, "id": 1}).to_list(50)
    if _old_entries:
        _old_ids = [e["id"] for e in _old_entries]
        await db.journal_lines.delete_many({"entry_id": {"$in": _old_ids}})
        await db.journal_entries.delete_many({"id": {"$in": _old_ids}})

    # 2. Fetch the current purchase invoice document
    inv = await db.purchase_invoices.find_one({"id": invoice_id})
    if not inv:
        return

    # 3. Cancelled bills don't sit in the ledger
    if inv.get("status") == "cancelled":
        return

    company_id = inv.get("company_id") or ""
    invoice_no = inv.get("invoice_no") or ""
    supplier_name = inv.get("supplier_name") or "Supplier"
    invoice_date = inv.get("invoice_date") or date.today().isoformat()

    grand_total = round(float(inv.get("grand_total") or 0), 2)
    total_gst = round(float(inv.get("total_gst") or 0), 2)

    if grand_total <= 0:
        return

    # 4. Get default ledger account IDs
    payable_id = await get_default_account_id(company_id, "2000")   # Accounts Payable
    purchases_id = await get_default_account_id(company_id, "5000")  # Purchases (expense)
    gst_input_id = await get_default_account_id(company_id, "1200")  # GST Input Credit

    if not payable_id or not purchases_id or not gst_input_id:
        return

    purchase_amount = round(grand_total - total_gst, 2)

    lines = [
        {"account_id": purchases_id, "account_name": "Purchases", "debit": purchase_amount, "credit": 0.0, "memo": f"Purchase Bill {invoice_no} from {supplier_name}"},
        {"account_id": payable_id, "account_name": "Accounts Payable", "debit": 0.0, "credit": grand_total, "memo": f"Bill {invoice_no} payable to {supplier_name}"},
    ]
    if total_gst > 0:
        lines.append({"account_id": gst_input_id, "account_name": "GST Input Credit", "debit": total_gst, "credit": 0.0, "memo": f"GST Input on Purchase {invoice_no}"})
    narration = f"Purchase Bill {invoice_no} from {supplier_name}"

    try:
        await post_journal_entry(
            company_id=company_id,
            entry_date=invoice_date,
            narration=narration,
            lines=lines,
            source="purchase",
            source_id=invoice_id,
            created_by=inv.get("created_by", "system"),
        )
    except Exception as e:
        logging.error(f"Error posting purchase journal entry: {e}")


async def sync_purchase_payment_journal_entry(payment_id: str):
    """Purchase-side mirror of sync_payment_journal_entry: Dr Accounts Payable,
    Cr Cash/Bank — the entry for actually paying a vendor bill."""
    from backend.accounting_core import get_default_account_id, post_journal_entry

    # 1. Clean up any existing journal entry for this payment
    _old_entries = await db.journal_entries.find({"source": "purchase_payment", "source_id": payment_id}, {"_id": 0, "id": 1}).to_list(50)
    if _old_entries:
        _old_ids = [e["id"] for e in _old_entries]
        await db.journal_lines.delete_many({"entry_id": {"$in": _old_ids}})
        await db.journal_entries.delete_many({"id": {"$in": _old_ids}})

    # 2. Fetch the current payment document
    payment = await db.purchase_payments.find_one({"id": payment_id})
    if not payment:
        return

    invoice_id = payment.get("purchase_invoice_id")
    inv = await db.purchase_invoices.find_one({"id": invoice_id})
    invoice_no = inv.get("invoice_no") if inv else "Unknown"
    supplier_name = inv.get("supplier_name") if inv else "Supplier"

    company_id = payment.get("company_id") or (inv.get("company_id") if inv else "")
    amount = round(float(payment.get("amount") or 0), 2)
    payment_date = payment.get("payment_date") or date.today().isoformat()
    payment_mode = str(payment.get("payment_mode") or "bank").strip().lower()
    bank_account_id = payment.get("bank_account_id")

    if amount <= 0:
        await db.bank_transactions.delete_many({"matched_id": payment_id, "matched_type": "purchase_payment"})
        return

    payable_id = await get_default_account_id(company_id, "2000")  # Accounts Payable
    cash_id = await get_default_account_id(company_id, "1000")     # Cash in Hand
    bank_id = await get_default_account_id(company_id, "1010")     # Bank Accounts

    if not payable_id or not cash_id or not bank_id:
        return

    credit_acct_id = cash_id if payment_mode == "cash" else bank_id
    credit_acct_name = "Cash in Hand" if payment_mode == "cash" else "Bank Accounts"

    lines = [
        {"account_id": payable_id, "account_name": "Accounts Payable", "debit": amount, "credit": 0.0, "memo": f"Bill {invoice_no} paid to {supplier_name}"},
        {"account_id": credit_acct_id, "account_name": credit_acct_name, "debit": 0.0, "credit": amount, "memo": f"Payment for Bill {invoice_no} via {payment_mode.upper()}"},
    ]
    narration = f"Payment to {supplier_name} for Bill {invoice_no} ({payment_mode.upper()})"

    entry_id = None
    try:
        entry = await post_journal_entry(
            company_id=company_id,
            entry_date=payment_date,
            narration=narration,
            lines=lines,
            source="purchase_payment",
            source_id=payment_id,
            created_by=payment.get("created_by", "system"),
        )
        entry_id = entry.get("id") if entry else None
    except Exception as e:
        logging.error(f"Error posting purchase payment journal entry: {e}")

    bank_account = None
    if payment_mode != "cash" and bank_account_id:
        bank_account = await db.bank_accounts.find_one({"id": bank_account_id, "company_id": company_id}, {"_id": 0})
    await _sync_bank_transaction_for_payment(
        payment_id=payment_id, matched_type="purchase_payment", bank_account=bank_account, company_id=company_id,
        amount=amount, txn_date=payment_date, direction="debit", journal_entry_id=entry_id,
        description=f"Payment to {supplier_name} for Bill {invoice_no}",
    )


# ── Reconciliation debounce ──────────────────────────────────────────────
# Trial Balance, P&L, Balance Sheet, and Ledger each independently call the
# two reconcile_and_sync_all_* functions below, and the Accounting Reports
# page loads all three reports at once. That used to mean 6 full
# invoices/payments/journal_entries scans firing concurrently on every page
# load or company switch, which is what made the page slow to render.
#
# Invoice/bill/payment create-edit-delete flows already call the targeted
# sync_*_journal_entry() functions directly and immediately, so these
# all-company scans only exist as a backfill safety net (e.g. for bills that
# existed before auto-posting was added). It's safe to skip a re-scan if one
# already ran for this company very recently.
_RECONCILE_TTL_SECONDS = 20
_reconcile_last_run: Dict[str, float] = {}
_reconcile_locks: Dict[str, "asyncio.Lock"] = {}


async def _run_reconcile_debounced(cache_key: str, fn, *args):
    lock = _reconcile_locks.setdefault(cache_key, asyncio.Lock())
    async with lock:
        last = _reconcile_last_run.get(cache_key, 0.0)
        if time.monotonic() - last < _RECONCILE_TTL_SECONDS:
            return
        await fn(*args)
        _reconcile_last_run[cache_key] = time.monotonic()


async def reconcile_and_sync_all_sales_and_payments(company_id: str):
    await _run_reconcile_debounced(f"sales:{company_id}", _reconcile_and_sync_all_sales_and_payments_impl, company_id)


async def reconcile_and_sync_all_purchases_and_payments(company_id: str):
    await _run_reconcile_debounced(f"purchases:{company_id}", _reconcile_and_sync_all_purchases_and_payments_impl, company_id)


async def _dedupe_journal_entries(company_id: str, source: str):
    """Collapse duplicate journal entries that share the same source_id.
    Duplicates can accumulate if sync_*_journal_entry ran concurrently (e.g.
    two overlapping report loads racing the reconciler) — the cleanup pass
    that lives inside sync_* only removes ONE existing entry before posting
    a new one, so a race can leave two, and a subsequent race leaves three.
    Keep the newest posting per source_id and delete the rest so the
    Journal Entries page doesn't show the same invoice three times."""
    entries = await db.journal_entries.find(
        {"company_id": company_id, "source": source, "source_id": {"$ne": None}},
        {"_id": 0, "id": 1, "source_id": 1, "created_at": 1},
    ).to_list(20000)
    by_src: Dict[str, list] = {}
    for e in entries:
        sid = e.get("source_id")
        if not sid:
            continue
        by_src.setdefault(sid, []).append(e)
    dup_ids: list = []
    for sid, es in by_src.items():
        if len(es) <= 1:
            continue
        es.sort(key=lambda x: x.get("created_at") or "", reverse=True)
        dup_ids.extend(e["id"] for e in es[1:])
    if dup_ids:
        await db.journal_lines.delete_many({"entry_id": {"$in": dup_ids}})
        await db.journal_entries.delete_many({"id": {"$in": dup_ids}})


async def _reconcile_and_sync_all_sales_and_payments_impl(company_id: str):
    try:
        await _dedupe_journal_entries(company_id, "sale")
        await _dedupe_journal_entries(company_id, "payment")

        # 1. Fetch all invoices for this company from the database
        invoices = await db.invoices.find({"company_id": company_id}).to_list(10000)
        invoice_ids = {inv["id"] for inv in invoices}
        # Drafts (renamed "Invoiced" in the UI) are now considered active for
        # journal-entry purposes — only cancelled invoices are skipped.
        # Proforma invoices / estimates are quotations, not real invoices,
        # and are never posted to the books (see sync_invoice_journal_entry)
        # — excluding them here too means any journal entry that was
        # incorrectly posted for one before this fix gets cleaned up as a
        # stale entry below, instead of being left behind forever.
        active_invoices = [
            inv for inv in invoices
            if inv.get("status", "draft") != "cancelled"
            and inv.get("invoice_type", "tax_invoice") not in ("proforma", "estimate")
        ]
        active_invoice_ids = {inv["id"] for inv in active_invoices}
        
        # 2. Fetch all existing journal entries with source "sale" for this company
        existing_sale_entries = await db.journal_entries.find({"company_id": company_id, "source": "sale"}).to_list(10000)
        sale_entry_by_source_id = {se["source_id"]: se for se in existing_sale_entries if se.get("source_id")}
        
        # 3. Clean up stale sale entries (deleted, draft, or cancelled invoices)
        stale_sale_ids = []
        for source_id, se in sale_entry_by_source_id.items():
            if source_id not in active_invoice_ids:
                stale_sale_ids.append(se["id"])
                
        if stale_sale_ids:
            await db.journal_lines.delete_many({"entry_id": {"$in": stale_sale_ids}})
            await db.journal_entries.delete_many({"id": {"$in": stale_sale_ids}})
            
        # 4. Sync missing/outdated sale entries
        for inv in active_invoices:
            inv_id = inv["id"]
            se = sale_entry_by_source_id.get(inv_id)
            if not se or abs(float(se.get("total_debit", 0)) - float(inv.get("grand_total", 0))) > 0.01:
                await sync_invoice_journal_entry(inv_id)
                
        # 5. Fetch all payments for this company
        payments = await db.payments.find({"company_id": company_id}).to_list(10000)
        if invoice_ids:
            missing_payments = await db.payments.find({
                "invoice_id": {"$in": list(invoice_ids)},
                "company_id": {"$exists": False}
            }).to_list(10000)
            pay_id_set = {p["id"] for p in payments}
            for mp in missing_payments:
                if mp["id"] not in pay_id_set:
                    payments.append(mp)
                    
        payment_ids = {p["id"] for p in payments}
        
        # 6. Fetch existing payment journal entries
        existing_pay_entries = await db.journal_entries.find({"company_id": company_id, "source": "payment"}).to_list(10000)
        pay_entry_by_source_id = {pe["source_id"]: pe for pe in existing_pay_entries if pe.get("source_id")}
        
        # 7. Clean up stale payment entries (deleted payments)
        stale_pay_ids = []
        for source_id, pe in pay_entry_by_source_id.items():
            if source_id not in payment_ids:
                stale_pay_ids.append(pe["id"])
                
        if stale_pay_ids:
            await db.journal_lines.delete_many({"entry_id": {"$in": stale_pay_ids}})
            await db.journal_entries.delete_many({"id": {"$in": stale_pay_ids}})
            
        # 8. Sync missing/outdated payment entries. Also re-sync entries that
        # were posted before the client_name fallback fix, which show up as
        # "Receipt from Client for Invoice Unknown" even though the payment
        # itself has a real client_name (common for imported/bank-matched
        # receipts that were never linked to an invoice).
        for p in payments:
            p_id = p["id"]
            pe = pay_entry_by_source_id.get(p_id)
            stale_narration = pe and "for Invoice Unknown" in (pe.get("narration") or "") and (p.get("client_name") or "").strip()
            if not pe or stale_narration or abs(float(pe.get("total_debit", 0)) - float(p.get("amount", 0))) > 0.01:
                await sync_payment_journal_entry(p_id)
                
    except Exception as e:
        logging.error(f"Error in reconcile_and_sync_all_sales_and_payments: {e}")


async def _reconcile_and_sync_all_purchases_and_payments_impl(company_id: str):
    """Purchase-side mirror of reconcile_and_sync_all_sales_and_payments.
    Runs on every report view (Trial Balance, P&L, Balance Sheet, Ledger) so
    any purchase bill or vendor payment that isn't yet reflected in the
    ledger — including bills uploaded before this automation existed — gets
    backfilled automatically, no manual "resync" step required. Debounced by
    the public reconcile_and_sync_all_purchases_and_payments wrapper above so
    it doesn't re-run on every single tab/report load."""
    try:
        await _dedupe_journal_entries(company_id, "purchase")
        await _dedupe_journal_entries(company_id, "purchase_payment")
        invoices = await db.purchase_invoices.find({"company_id": company_id}).to_list(10000)
        invoice_ids = {inv["id"] for inv in invoices}
        active_invoices = [inv for inv in invoices if inv.get("status", "outstanding") != "cancelled"]
        active_invoice_ids = {inv["id"] for inv in active_invoices}

        existing_purchase_entries = await db.journal_entries.find({"company_id": company_id, "source": "purchase"}).to_list(10000)
        entry_by_source_id = {e["source_id"]: e for e in existing_purchase_entries if e.get("source_id")}

        stale_ids = [e["id"] for source_id, e in entry_by_source_id.items() if source_id not in active_invoice_ids]
        if stale_ids:
            await db.journal_lines.delete_many({"entry_id": {"$in": stale_ids}})
            await db.journal_entries.delete_many({"id": {"$in": stale_ids}})

        for inv in active_invoices:
            inv_id = inv["id"]
            e = entry_by_source_id.get(inv_id)
            if not e or abs(float(e.get("total_debit", 0)) - float(inv.get("grand_total", 0))) > 0.01:
                await sync_purchase_journal_entry(inv_id)

        payments = await db.purchase_payments.find({"company_id": company_id}).to_list(10000)
        if invoice_ids:
            missing_payments = await db.purchase_payments.find({
                "purchase_invoice_id": {"$in": list(invoice_ids)},
                "company_id": {"$exists": False},
            }).to_list(10000)
            pay_id_set = {p["id"] for p in payments}
            for mp in missing_payments:
                if mp["id"] not in pay_id_set:
                    payments.append(mp)
        payment_ids = {p["id"] for p in payments}

        existing_pay_entries = await db.journal_entries.find({"company_id": company_id, "source": "purchase_payment"}).to_list(10000)
        pay_entry_by_source_id = {pe["source_id"]: pe for pe in existing_pay_entries if pe.get("source_id")}

        stale_pay_ids = [pe["id"] for source_id, pe in pay_entry_by_source_id.items() if source_id not in payment_ids]
        if stale_pay_ids:
            await db.journal_lines.delete_many({"entry_id": {"$in": stale_pay_ids}})
            await db.journal_entries.delete_many({"id": {"$in": stale_pay_ids}})

        for p in payments:
            p_id = p["id"]
            pe = pay_entry_by_source_id.get(p_id)
            if not pe or abs(float(pe.get("total_debit", 0)) - float(p.get("amount", 0))) > 0.01:
                await sync_purchase_payment_journal_entry(p_id)

    except Exception as e:
        logging.error(f"Error in reconcile_and_sync_all_purchases_and_payments: {e}")
