"""
AI Accounting Module for Taskosphere
=====================================
• Bank statement parser  (PDF – SBI/HDFC/ICICI/Axis + Excel/CSV)
• AI-powered transaction categorisation (OpenAI / rule-based fallback)
• Double-entry bookkeeping engine
• Indian Chart of Accounts (Ind AS compliant)
• Ledger management
• Trial Balance, P&L, Trading Account, Balance Sheet
"""

import io
import re
import csv
import uuid
import json
import logging
import os
from datetime import datetime, timezone, date
from typing import Optional, List, Dict, Any, Tuple
from enum import Enum

import pandas as pd
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query, status
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field

from backend.dependencies import db, get_current_user
from backend.models import User

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/accounting", tags=["Accounting"])

# ═══════════════════════════════════════════════════════════════════════════
# OPTIONAL IMPORTS  (auto-install if missing)
# ═══════════════════════════════════════════════════════════════════════════
try:
    import pdfplumber
except ImportError:
    import subprocess, sys
    subprocess.check_call([sys.executable, "-m", "pip", "install", "pdfplumber"])
    import pdfplumber

try:
    import openpyxl
except ImportError:
    import subprocess, sys
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl

# ═══════════════════════════════════════════════════════════════════════════
# ENUMS
# ═══════════════════════════════════════════════════════════════════════════

class AccountType(str, Enum):
    asset       = "Asset"
    liability   = "Liability"
    capital     = "Capital"
    revenue     = "Revenue"
    expense     = "Expense"
    contra      = "Contra"

class AccountSubType(str, Enum):
    # Assets
    current_asset   = "Current Asset"
    fixed_asset     = "Fixed Asset"
    bank            = "Bank"
    cash            = "Cash"
    receivable      = "Receivable"
    prepaid         = "Prepaid"
    investment      = "Investment"
    tax_asset       = "Tax Asset"
    # Liabilities
    current_liability = "Current Liability"
    long_term_liability = "Long Term Liability"
    payable         = "Payable"
    tax_liability   = "Tax Liability"
    # Capital
    equity          = "Equity"
    retained        = "Retained Earnings"
    # Revenue
    operating_revenue = "Operating Revenue"
    other_income    = "Other Income"
    # Expense
    direct_expense  = "Direct Expense"
    indirect_expense = "Indirect Expense"
    depreciation    = "Depreciation"
    tax_expense     = "Tax Expense"

class EntryType(str, Enum):
    debit  = "Dr"
    credit = "Cr"

# ═══════════════════════════════════════════════════════════════════════════
# INDIAN CHART OF ACCOUNTS  (default seed)
# ═══════════════════════════════════════════════════════════════════════════

DEFAULT_ACCOUNTS = [
    # ─── ASSETS ───────────────────────────────────────────────────────────
    {"code": "1001", "name": "Cash in Hand",           "type": "Asset",     "sub_type": "Cash",              "normal_balance": "Dr"},
    {"code": "1002", "name": "Bank Account – SBI",     "type": "Asset",     "sub_type": "Bank",              "normal_balance": "Dr"},
    {"code": "1003", "name": "Bank Account – HDFC",    "type": "Asset",     "sub_type": "Bank",              "normal_balance": "Dr"},
    {"code": "1004", "name": "Bank Account – ICICI",   "type": "Asset",     "sub_type": "Bank",              "normal_balance": "Dr"},
    {"code": "1005", "name": "Bank Account – Axis",    "type": "Asset",     "sub_type": "Bank",              "normal_balance": "Dr"},
    {"code": "1006", "name": "Bank Account – Others",  "type": "Asset",     "sub_type": "Bank",              "normal_balance": "Dr"},
    {"code": "1010", "name": "Sundry Debtors",          "type": "Asset",     "sub_type": "Receivable",        "normal_balance": "Dr"},
    {"code": "1020", "name": "Loans & Advances Given",  "type": "Asset",     "sub_type": "Current Asset",     "normal_balance": "Dr"},
    {"code": "1030", "name": "Stock / Inventory",       "type": "Asset",     "sub_type": "Current Asset",     "normal_balance": "Dr"},
    {"code": "1040", "name": "Prepaid Expenses",        "type": "Asset",     "sub_type": "Prepaid",           "normal_balance": "Dr"},
    {"code": "1050", "name": "TDS Receivable",          "type": "Asset",     "sub_type": "Tax Asset",         "normal_balance": "Dr"},
    {"code": "1051", "name": "Input GST – CGST",        "type": "Asset",     "sub_type": "Tax Asset",         "normal_balance": "Dr"},
    {"code": "1052", "name": "Input GST – SGST",        "type": "Asset",     "sub_type": "Tax Asset",         "normal_balance": "Dr"},
    {"code": "1053", "name": "Input GST – IGST",        "type": "Asset",     "sub_type": "Tax Asset",         "normal_balance": "Dr"},
    {"code": "1060", "name": "Fixed Assets – Furniture","type": "Asset",     "sub_type": "Fixed Asset",       "normal_balance": "Dr"},
    {"code": "1061", "name": "Fixed Assets – Computers","type": "Asset",     "sub_type": "Fixed Asset",       "normal_balance": "Dr"},
    {"code": "1062", "name": "Fixed Assets – Vehicles", "type": "Asset",     "sub_type": "Fixed Asset",       "normal_balance": "Dr"},
    {"code": "1063", "name": "Fixed Assets – Office Equip","type":"Asset",   "sub_type": "Fixed Asset",       "normal_balance": "Dr"},
    {"code": "1070", "name": "Accum. Depreciation",     "type": "Asset",     "sub_type": "Fixed Asset",       "normal_balance": "Cr"},
    {"code": "1080", "name": "Investments",             "type": "Asset",     "sub_type": "Investment",        "normal_balance": "Dr"},
    {"code": "1090", "name": "Security Deposits",       "type": "Asset",     "sub_type": "Current Asset",     "normal_balance": "Dr"},
    # ─── LIABILITIES ──────────────────────────────────────────────────────
    {"code": "2001", "name": "Sundry Creditors",        "type": "Liability", "sub_type": "Payable",           "normal_balance": "Cr"},
    {"code": "2010", "name": "Output GST – CGST",       "type": "Liability", "sub_type": "Tax Liability",     "normal_balance": "Cr"},
    {"code": "2011", "name": "Output GST – SGST",       "type": "Liability", "sub_type": "Tax Liability",     "normal_balance": "Cr"},
    {"code": "2012", "name": "Output GST – IGST",       "type": "Liability", "sub_type": "Tax Liability",     "normal_balance": "Cr"},
    {"code": "2020", "name": "TDS Payable",             "type": "Liability", "sub_type": "Tax Liability",     "normal_balance": "Cr"},
    {"code": "2021", "name": "Professional Tax Payable","type": "Liability", "sub_type": "Tax Liability",     "normal_balance": "Cr"},
    {"code": "2030", "name": "Salary Payable",          "type": "Liability", "sub_type": "Current Liability", "normal_balance": "Cr"},
    {"code": "2040", "name": "Bank Overdraft",          "type": "Liability", "sub_type": "Current Liability", "normal_balance": "Cr"},
    {"code": "2050", "name": "Loans Payable",           "type": "Liability", "sub_type": "Long Term Liability","normal_balance": "Cr"},
    {"code": "2060", "name": "Advance from Customers",  "type": "Liability", "sub_type": "Current Liability", "normal_balance": "Cr"},
    {"code": "2070", "name": "Audit Fees Payable",      "type": "Liability", "sub_type": "Current Liability", "normal_balance": "Cr"},
    # ─── CAPITAL ──────────────────────────────────────────────────────────
    {"code": "3001", "name": "Capital Account",         "type": "Capital",   "sub_type": "Equity",            "normal_balance": "Cr"},
    {"code": "3002", "name": "Drawings",                "type": "Capital",   "sub_type": "Equity",            "normal_balance": "Dr"},
    {"code": "3003", "name": "Retained Earnings",       "type": "Capital",   "sub_type": "Retained Earnings", "normal_balance": "Cr"},
    {"code": "3004", "name": "General Reserve",         "type": "Capital",   "sub_type": "Retained Earnings", "normal_balance": "Cr"},
    # ─── REVENUE ──────────────────────────────────────────────────────────
    {"code": "4001", "name": "Professional Fees",       "type": "Revenue",   "sub_type": "Operating Revenue", "normal_balance": "Cr"},
    {"code": "4002", "name": "Service Revenue",         "type": "Revenue",   "sub_type": "Operating Revenue", "normal_balance": "Cr"},
    {"code": "4003", "name": "Sales Revenue",           "type": "Revenue",   "sub_type": "Operating Revenue", "normal_balance": "Cr"},
    {"code": "4004", "name": "Interest Received",       "type": "Revenue",   "sub_type": "Other Income",      "normal_balance": "Cr"},
    {"code": "4005", "name": "Rent Received",           "type": "Revenue",   "sub_type": "Other Income",      "normal_balance": "Cr"},
    {"code": "4006", "name": "Commission Received",     "type": "Revenue",   "sub_type": "Other Income",      "normal_balance": "Cr"},
    {"code": "4007", "name": "Discount Received",       "type": "Revenue",   "sub_type": "Other Income",      "normal_balance": "Cr"},
    {"code": "4008", "name": "IT Refund / Govt Refund", "type": "Revenue",   "sub_type": "Other Income",      "normal_balance": "Cr"},
    {"code": "4009", "name": "Miscellaneous Income",    "type": "Revenue",   "sub_type": "Other Income",      "normal_balance": "Cr"},
    # ─── DIRECT EXPENSES (Trading A/c) ────────────────────────────────────
    {"code": "5001", "name": "Purchase of Goods",       "type": "Expense",   "sub_type": "Direct Expense",    "normal_balance": "Dr"},
    {"code": "5002", "name": "Freight & Cartage",       "type": "Expense",   "sub_type": "Direct Expense",    "normal_balance": "Dr"},
    {"code": "5003", "name": "Direct Labour",           "type": "Expense",   "sub_type": "Direct Expense",    "normal_balance": "Dr"},
    # ─── INDIRECT EXPENSES (P&L) ──────────────────────────────────────────
    {"code": "6001", "name": "Salaries & Wages",        "type": "Expense",   "sub_type": "Indirect Expense",  "normal_balance": "Dr"},
    {"code": "6002", "name": "Rent Expense",            "type": "Expense",   "sub_type": "Indirect Expense",  "normal_balance": "Dr"},
    {"code": "6003", "name": "Electricity & Utilities", "type": "Expense",   "sub_type": "Indirect Expense",  "normal_balance": "Dr"},
    {"code": "6004", "name": "Internet & Telephone",    "type": "Expense",   "sub_type": "Indirect Expense",  "normal_balance": "Dr"},
    {"code": "6005", "name": "Office Supplies",         "type": "Expense",   "sub_type": "Indirect Expense",  "normal_balance": "Dr"},
    {"code": "6006", "name": "Professional Charges",    "type": "Expense",   "sub_type": "Indirect Expense",  "normal_balance": "Dr"},
    {"code": "6007", "name": "Bank Charges",            "type": "Expense",   "sub_type": "Indirect Expense",  "normal_balance": "Dr"},
    {"code": "6008", "name": "Travel & Conveyance",     "type": "Expense",   "sub_type": "Indirect Expense",  "normal_balance": "Dr"},
    {"code": "6009", "name": "Marketing & Advertising", "type": "Expense",   "sub_type": "Indirect Expense",  "normal_balance": "Dr"},
    {"code": "6010", "name": "Software Subscriptions",  "type": "Expense",   "sub_type": "Indirect Expense",  "normal_balance": "Dr"},
    {"code": "6011", "name": "Insurance Premium",       "type": "Expense",   "sub_type": "Indirect Expense",  "normal_balance": "Dr"},
    {"code": "6012", "name": "Repairs & Maintenance",   "type": "Expense",   "sub_type": "Indirect Expense",  "normal_balance": "Dr"},
    {"code": "6013", "name": "Printing & Stationery",   "type": "Expense",   "sub_type": "Indirect Expense",  "normal_balance": "Dr"},
    {"code": "6014", "name": "Audit & Legal Fees",      "type": "Expense",   "sub_type": "Indirect Expense",  "normal_balance": "Dr"},
    {"code": "6015", "name": "Miscellaneous Expenses",  "type": "Expense",   "sub_type": "Indirect Expense",  "normal_balance": "Dr"},
    {"code": "6016", "name": "ATM Cash Withdrawal",     "type": "Expense",   "sub_type": "Indirect Expense",  "normal_balance": "Dr"},
    {"code": "6017", "name": "Depreciation",            "type": "Expense",   "sub_type": "Depreciation",      "normal_balance": "Dr"},
    {"code": "6018", "name": "Bad Debts Written Off",   "type": "Expense",   "sub_type": "Indirect Expense",  "normal_balance": "Dr"},
    {"code": "6019", "name": "Postage & Courier",       "type": "Expense",   "sub_type": "Indirect Expense",  "normal_balance": "Dr"},
    {"code": "6020", "name": "Staff Welfare",           "type": "Expense",   "sub_type": "Indirect Expense",  "normal_balance": "Dr"},
    # ─── TAX ──────────────────────────────────────────────────────────────
    {"code": "7001", "name": "Income Tax",              "type": "Expense",   "sub_type": "Tax Expense",       "normal_balance": "Dr"},
    {"code": "7002", "name": "Provision for Tax",       "type": "Liability", "sub_type": "Tax Liability",     "normal_balance": "Cr"},
]

# ═══════════════════════════════════════════════════════════════════════════
# TRANSACTION CATEGORISATION RULES  (keyword → account code)
# ═══════════════════════════════════════════════════════════════════════════

# (pattern, debit_account, credit_account, narration)
# For CREDIT transactions (money IN): Dr Bank, Cr <credit_account>
# For DEBIT  transactions (money OUT): Dr <debit_account>, Cr Bank

CATEGORISATION_RULES: List[Dict] = [
    # ── INCOME ─────────────────────────────────────────────────────────────
    {"patterns": [r"UPI.*CR", r"NEFT.*CR", r"IMPS.*CR", r"BY TRANSFER"],
     "credit_ac": "4002", "narration": "Service Revenue / Receipt"},
    {"patterns": [r"ITDTAX REFUND", r"IT REFUND", r"INCOME TAX REFUND", r"GOV.*REFUND"],
     "credit_ac": "4008", "narration": "IT Refund"},
    {"patterns": [r"INTEREST CREDIT", r"INT CREDIT"],
     "credit_ac": "4004", "narration": "Interest Received"},
    {"patterns": [r"CHEQUE DEPOSIT", r"CHQ DEP"],
     "credit_ac": "4002", "narration": "Cheque Deposit – Service Receipt"},
    {"patterns": [r"CLEARING.*CHEQUE"],
     "credit_ac": "4002", "narration": "Clearing – Receipt"},

    # ── EXPENSES ───────────────────────────────────────────────────────────
    {"patterns": [r"ATM WDL", r"ATM CASH", r"ATM WD"],
     "debit_ac": "6016", "narration": "ATM Cash Withdrawal"},
    {"patterns": [r"SALARY", r"SAL CR", r"WAGES"],
     "debit_ac": "6001", "narration": "Salary & Wages"},
    {"patterns": [r"RENT", r"RENTAL"],
     "debit_ac": "6002", "narration": "Rent Expense"},
    {"patterns": [r"ELECTRICITY", r"BESCOM", r"TNEB", r"DGVCL", r"TORRENT POWER"],
     "debit_ac": "6003", "narration": "Electricity Charges"},
    {"patterns": [r"INTERNET", r"BROADBAND", r"AIRTEL", r"JIO", r"BSNL", r"TELEPHONE", r"MOBILE BILL"],
     "debit_ac": "6004", "narration": "Internet & Telephone"},
    {"patterns": [r"INSURANCE", r"LIC", r"NEW INDIA", r"BAJAJ ALLIANZ", r"HDFC LIFE", r"FIRE.*INS"],
     "debit_ac": "6011", "narration": "Insurance Premium"},
    {"patterns": [r"JUST DIAL", r"WORKINDIA", r"NAUKRI", r"GOOGLE ADS", r"FACEBOOK ADS"],
     "debit_ac": "6009", "narration": "Marketing & Advertising"},
    {"patterns": [r"TALLY", r"VYAPAR", r"ZOHO", r"FOURV", r"SOFTWARE"],
     "debit_ac": "6010", "narration": "Software Subscription"},
    {"patterns": [r"AUDIT", r"CA FEE", r"LEGAL"],
     "debit_ac": "6014", "narration": "Audit & Legal Fees"},
    {"patterns": [r"GAYA BUSINESS", r"GAYA.*SERVICE"],
     "debit_ac": "6014", "narration": "Audit / Professional Fees (Gaya Business Service)"},
    {"patterns": [r"ICI.*BUSINESS", r"ICICI.*BUSINESS"],
     "debit_ac": "6006", "narration": "Professional Charges"},
    {"patterns": [r"SN TRADELINK", r"TRADELINK"],
     "debit_ac": "5001", "narration": "Purchase of Goods"},
    {"patterns": [r"KALYAN RESORT"],
     "debit_ac": "6008", "narration": "Travel & Conveyance"},
    {"patterns": [r"A/C KEEPING", r"ACCOUNT KEEPING", r"A/C CHGS", r"SERVICE CHARGE"],
     "debit_ac": "6007", "narration": "Bank Charges"},
    {"patterns": [r"ATM.*AMC", r"DEBIT.*AMC", r"CARD AMC"],
     "debit_ac": "6007", "narration": "Bank Charges – Card AMC"},
    {"patterns": [r"HDFC AC PAYMENT", r"CREDIT CARD PAYMENT", r"CC PAYMENT"],
     "debit_ac": "2001", "narration": "Credit Card Payment"},
    {"patterns": [r"INSUFFICIENT BAL.*DECLINE"],
     "debit_ac": "6007", "narration": "Bank Charges – Decline Fee"},
    {"patterns": [r"MANTHAN DESAI", r"CA.*ASSOC", r"MANTHAN.*ASSO"],
     "debit_ac": "6014", "narration": "CA / Accountant Fees"},
    {"patterns": [r"TO TRANSFER", r"NEFT.*DR", r"IMPS.*DR", r"TO CLEARING"],
     "debit_ac": "6015", "narration": "Transfer / Payment"},
    {"patterns": [r"DEBIT CARD.*PG", r"by debit card"],
     "debit_ac": "6015", "narration": "Debit Card Payment"},
]

# ═══════════════════════════════════════════════════════════════════════════
# PYDANTIC MODELS
# ═══════════════════════════════════════════════════════════════════════════

class AccountCreate(BaseModel):
    code: str
    name: str
    type: str
    sub_type: str
    normal_balance: str = "Dr"
    opening_balance: float = 0.0
    description: Optional[str] = None

class JournalLineIn(BaseModel):
    account_code: str
    type: EntryType          # Dr / Cr
    amount: float
    narration: Optional[str] = None

class JournalEntryCreate(BaseModel):
    date: str                # ISO date  YYYY-MM-DD
    narration: str
    lines: List[JournalLineIn]
    bank_statement_id: Optional[str] = None
    ref_no: Optional[str] = None

class BankStatementUpload(BaseModel):
    bank_name: str
    account_number: str
    bank_account_code: str   # Chart-of-Accounts code, e.g. "1002"
    financial_year: str      # e.g. "2024-25"

# ═══════════════════════════════════════════════════════════════════════════
# BANK STATEMENT PARSERS
# ═══════════════════════════════════════════════════════════════════════════

def _clean_amount(val: Any) -> float:
    """Convert any amount representation to float."""
    if val is None:
        return 0.0
    s = str(val).replace(",", "").replace(" ", "").strip()
    if s in ("", "-", "–"):
        return 0.0
    try:
        return abs(float(s))
    except ValueError:
        return 0.0


def _parse_date(val: Any) -> Optional[str]:
    """Return ISO date string or None. Handles many Indian bank date formats."""
    if not val:
        return None
    s = str(val).strip()
    # Remove trailing time portion if present (e.g. "01/04/2024 00:00:00")
    s = re.split(r"\s+\d{1,2}:\d{2}", s)[0].strip()
    for fmt in (
        "%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%d %b %Y", "%d-%b-%Y",
        "%d/%m/%y", "%d-%m-%y", "%d %B %Y", "%d-%B-%Y",
        "%d %b %y", "%d-%b-%y", "%d.%m.%Y", "%d.%m.%y",
        "%Y/%m/%d", "%m/%d/%Y", "%b %d, %Y", "%B %d, %Y",
    ):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    # Try dateutil as last resort
    try:
        from dateutil import parser as du_parser
        return du_parser.parse(s, dayfirst=True).strftime("%Y-%m-%d")
    except Exception:
        pass
    return None


def parse_sbi_pdf(content: bytes) -> List[Dict]:
    """Parse SBI bank statement PDF → list of transaction dicts."""
    transactions = []
    try:
        with pdfplumber.open(io.BytesIO(content)) as pdf:
            for page in pdf.pages:
                tables = page.extract_tables()
                for table in tables:
                    for row in table:
                        if not row or len(row) < 6:
                            continue
                        # Detect header row
                        if any(str(c).strip().lower() in ("txn date", "value date", "date") for c in row if c):
                            continue
                        # Try to extract date from first or second column
                        date_val = _parse_date(row[0]) or _parse_date(row[1])
                        if not date_val:
                            continue
                        # Description is usually col 2
                        desc = str(row[2] or "").strip().replace("\n", " ")
                        # Ref no col 3, Branch col 4, Debit col 5, Credit col 6, Balance col 7
                        debit  = _clean_amount(row[5] if len(row) > 5 else None)
                        credit = _clean_amount(row[6] if len(row) > 6 else None)
                        balance= _clean_amount(row[7] if len(row) > 7 else None)
                        ref_no = str(row[3] or "").strip() if len(row) > 3 else ""
                        if debit == 0 and credit == 0:
                            continue
                        transactions.append({
                            "date":    date_val,
                            "description": desc,
                            "ref_no":  ref_no,
                            "debit":   debit,
                            "credit":  credit,
                            "balance": balance,
                        })
    except Exception as e:
        logger.error("SBI PDF parse error: %s", e)
    return transactions


def parse_generic_pdf(content: bytes) -> List[Dict]:
    """Generic PDF bank statement parser (HDFC/ICICI/Axis pattern)."""
    transactions = []
    date_pattern = re.compile(r"\b(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})\b")
    amount_pattern = re.compile(r"[\d,]+\.\d{2}")

    try:
        with pdfplumber.open(io.BytesIO(content)) as pdf:
            for page in pdf.pages:
                text = page.extract_text() or ""
                lines = text.split("\n")
                for line in lines:
                    dm = date_pattern.search(line)
                    if not dm:
                        continue
                    amounts = amount_pattern.findall(line)
                    if not amounts:
                        continue
                    amounts_f = [_clean_amount(a) for a in amounts]
                    date_val = _parse_date(dm.group(1))
                    if not date_val:
                        continue
                    # Heuristic: last amount = balance, second-last = debit or credit
                    debit = credit = balance = 0.0
                    if len(amounts_f) >= 3:
                        balance = amounts_f[-1]
                        # check Dr/Cr marker in line
                        if "Dr" in line or "debit" in line.lower():
                            debit = amounts_f[-2]
                        else:
                            credit = amounts_f[-2]
                    elif len(amounts_f) == 2:
                        balance = amounts_f[-1]
                        credit = amounts_f[0]
                    elif len(amounts_f) == 1:
                        credit = amounts_f[0]

                    if debit == 0 and credit == 0:
                        continue
                    # Extract description: everything between date match and first amount
                    desc_part = line[dm.end():].strip()
                    desc_part = amount_pattern.sub("", desc_part)
                    desc_part = re.sub(r"\s+", " ", desc_part).strip()[:120]
                    transactions.append({
                        "date":        date_val,
                        "description": desc_part or line[:80],
                        "ref_no":      "",
                        "debit":       debit,
                        "credit":      credit,
                        "balance":     balance,
                    })
    except Exception as e:
        logger.error("Generic PDF parse error: %s", e)
    return transactions


def _df_to_transactions(df):
    """Convert a normalised DataFrame to transaction list using flexible column matching."""
    import pandas as pd
    transactions = []
    df.columns = [str(c).strip().lower() for c in df.columns]
    col_map = {}
    for c in df.columns:
        if re.search(r"txn.*date|tran.*date|value date", c):
            col_map.setdefault("date", c)
        elif re.search(r"^date$", c) and "date" not in col_map:
            col_map.setdefault("date", c)
        elif re.search(r"desc|narr|particular|detail|remark", c):
            col_map.setdefault("desc", c)
        elif re.search(r"debit|withdraw|dr\b", c):
            col_map.setdefault("debit", c)
        elif re.search(r"credit|deposit|cr\b", c):
            col_map.setdefault("credit", c)
        elif re.search(r"balance|bal\b", c):
            col_map.setdefault("balance", c)
        elif re.search(r"ref|chq|cheque|utr", c):
            col_map.setdefault("ref", c)
    if "date" not in col_map:
        return transactions
    for _, row in df.iterrows():
        date_val = _parse_date(row.get(col_map.get("date", ""), ""))
        if not date_val:
            continue
        debit  = _clean_amount(row.get(col_map.get("debit",  ""), 0))
        credit = _clean_amount(row.get(col_map.get("credit", ""), 0))
        if debit == 0 and credit == 0:
            continue
        transactions.append({
            "date":        date_val,
            "description": str(row.get(col_map.get("desc",    ""), ""))[:200].strip(),
            "ref_no":      str(row.get(col_map.get("ref",     ""), "")).strip(),
            "debit":       debit,
            "credit":      credit,
            "balance":     _clean_amount(row.get(col_map.get("balance", ""), 0)),
        })
    return transactions


def _parse_sbi_tsv_xls(content: bytes) -> List[Dict]:
    """
    Parse SBI bank statements exported as .xls but actually tab-separated text.
    Header rows contain account metadata; transaction rows start after Txn Date header.
    """
    transactions = []
    try:
        text = content.decode("utf-8", errors="replace")
        lines = text.splitlines()
        header_idx = None
        for i, line in enumerate(lines):
            if re.search(r"txn\s*date", line, re.IGNORECASE):
                header_idx = i
                break
        if header_idx is None:
            return transactions
        rows = [line.split("	") for line in lines[header_idx:] if line.strip()]
        rows = [r for r in rows if r and not r[0].strip().startswith("*")]
        if len(rows) < 2:
            return transactions
        headers = [h.strip().lower() for h in rows[0]]
        data_rows = rows[1:]
        col_map = {}
        for i, h in enumerate(headers):
            if re.search(r"txn.*date|tran.*date", h):
                col_map.setdefault("date", i)
            elif re.search(r"desc|narr|particular|detail|remark", h):
                col_map.setdefault("desc", i)
            elif re.search(r"debit|withdraw|dr", h):
                col_map.setdefault("debit", i)
            elif re.search(r"credit|deposit|cr", h):
                col_map.setdefault("credit", i)
            elif re.search(r"balance|bal", h):
                col_map.setdefault("balance", i)
            elif re.search(r"ref|chq|cheque|utr", h):
                col_map.setdefault("ref", i)
        if "date" not in col_map:
            return transactions
        def _get(row, key):
            idx = col_map.get(key)
            if idx is None or idx >= len(row):
                return ""
            return row[idx].strip()
        for row in data_rows:
            date_val = _parse_date(_get(row, "date"))
            if not date_val:
                continue
            debit  = _clean_amount(_get(row, "debit"))
            credit = _clean_amount(_get(row, "credit"))
            if debit == 0 and credit == 0:
                continue
            transactions.append({
                "date":        date_val,
                "description": _get(row, "desc")[:200],
                "ref_no":      _get(row, "ref"),
                "debit":       debit,
                "credit":      credit,
                "balance":     _clean_amount(_get(row, "balance")),
            })
    except Exception as e:
        logger.error("SBI TSV-XLS parse error: %s", e)
    return transactions


def parse_excel_statement(content: bytes, filename: str = "") -> List[Dict]:
    """
    Parse Excel / CSV / TSV bank statement.
    Tries multiple strategies:
      1. SBI tab-delimited .xls (plain text disguised as Excel)
      2. Real Excel via openpyxl
      3. Real Excel via xlrd (older .xls binary)
      4. CSV (comma, semicolon, or tab separated)
    """
    import pandas as pd
    fname_lower = filename.lower()

    # Strategy 1: SBI tab-delimited text masquerading as .xls
    try:
        snippet = content[:500].decode("utf-8", errors="replace")
        if re.search(r"account\s*name|txn\s*date|account\s*number|sbin", snippet, re.IGNORECASE):
            txns = _parse_sbi_tsv_xls(content)
            if txns:
                logger.info("Parsed %d txns via SBI TSV-XLS strategy", len(txns))
                return txns
    except Exception as e:
        logger.warning("SBI TSV sniff failed: %s", e)

    # Strategy 2: Real Excel via openpyxl
    if not fname_lower.endswith(".csv"):
        try:
            df = pd.read_excel(io.BytesIO(content), engine="openpyxl", header=None)
            header_row = None
            for i, row in df.iterrows():
                vals = [str(v).lower() for v in row if pd.notna(v)]
                if any(re.search(r"txn.*date|tran.*date|date", v) for v in vals):
                    header_row = i
                    break
            if header_row is not None:
                df.columns = df.iloc[header_row]
                df = df.iloc[header_row + 1:].reset_index(drop=True)
                txns = _df_to_transactions(df)
                if txns:
                    logger.info("Parsed %d txns via openpyxl strategy", len(txns))
                    return txns
        except Exception as e:
            logger.warning("openpyxl parse failed: %s", e)

    # Strategy 3: Real Excel via xlrd (old .xls binary)
    if not fname_lower.endswith(".csv"):
        try:
            import xlrd
            wb = xlrd.open_workbook(file_contents=content)
            ws = wb.sheet_by_index(0)
            data = [ws.row_values(i) for i in range(ws.nrows)]
            if data:
                df = pd.DataFrame(data[1:], columns=data[0])
                txns = _df_to_transactions(df)
                if txns:
                    logger.info("Parsed %d txns via xlrd strategy", len(txns))
                    return txns
        except Exception as e:
            logger.warning("xlrd parse failed: %s", e)

    # Strategy 4: CSV (comma, semicolon, or tab)
    for sep in (",", ";", "	"):
        try:
            df = pd.read_csv(io.BytesIO(content), sep=sep)
            txns = _df_to_transactions(df)
            if txns:
                logger.info("Parsed %d txns via CSV sep=%r strategy", len(txns), sep)
                return txns
        except Exception:
            continue

    return []


def detect_bank(filename: str, first_text: str) -> str:
    """Detect bank from filename or text."""
    combined = (filename + " " + first_text).lower()
    if "sbi" in combined or "state bank" in combined:
        return "SBI"
    if "hdfc" in combined:
        return "HDFC"
    if "icici" in combined:
        return "ICICI"
    if "axis" in combined:
        return "Axis"
    if "kotak" in combined:
        return "Kotak"
    if "baroda" in combined or "bob" in combined:
        return "Bank of Baroda"
    return "Unknown"


# ═══════════════════════════════════════════════════════════════════════════
# AI CATEGORISATION
# ═══════════════════════════════════════════════════════════════════════════

def _rule_categorise(description: str, is_credit: bool) -> Dict:
    """Rule-based categorisation using CATEGORISATION_RULES."""
    desc_upper = description.upper()
    for rule in CATEGORISATION_RULES:
        for pat in rule["patterns"]:
            if re.search(pat, desc_upper, re.IGNORECASE):
                if is_credit and "credit_ac" in rule:
                    return {
                        "debit_account":  "1002",                 # Bank A/c
                        "credit_account": rule["credit_ac"],
                        "narration":      rule["narration"],
                        "confidence":     "rule",
                    }
                elif not is_credit and "debit_ac" in rule:
                    return {
                        "debit_account":  rule["debit_ac"],
                        "credit_account": "1002",                 # Bank A/c
                        "narration":      rule["narration"],
                        "confidence":     "rule",
                    }
    # Fallback
    if is_credit:
        return {"debit_account": "1002", "credit_account": "4009", "narration": "Miscellaneous Income", "confidence": "fallback"}
    else:
        return {"debit_account": "6015", "credit_account": "1002", "narration": "Miscellaneous Expense", "confidence": "fallback"}


async def ai_categorise_transaction(description: str, amount: float, is_credit: bool, org_id: str) -> Dict:
    """Try OpenAI/Gemini; fall back to rule-based on failure."""
    # ── Rule-based first (fast path) ─────────────────────────────────────
    rule_result = _rule_categorise(description, is_credit)
    if rule_result["confidence"] == "rule":
        return rule_result

    # ── AI path (OpenAI) ─────────────────────────────────────────────────
    try:
        import openai
        api_key = os.getenv("OPENAI_API_KEY")
        if not api_key:
            return rule_result

        account_list = "\n".join(
            f"{a['code']}: {a['name']} ({a['type']})"
            for a in DEFAULT_ACCOUNTS
        )
        direction = "CREDIT (money received into bank)" if is_credit else "DEBIT (money paid out of bank)"
        prompt = (
            f"You are an Indian CA. Categorise this bank transaction:\n"
            f"Description: {description}\nAmount: ₹{amount:,.2f}\nDirection: {direction}\n\n"
            f"Chart of Accounts:\n{account_list}\n\n"
            f"Return JSON only:\n"
            f"{{\"debit_account\":\"<code>\",\"credit_account\":\"<code>\",\"narration\":\"<short narration>\"}}"
        )

        client = openai.AsyncOpenAI(api_key=api_key)
        resp = await client.chat.completions.create(
            model="gpt-3.5-turbo",
            messages=[{"role": "user", "content": prompt}],
            max_tokens=200,
            temperature=0,
        )
        raw = resp.choices[0].message.content.strip()
        # Strip markdown fences if present
        raw = re.sub(r"```json|```", "", raw).strip()
        data = json.loads(raw)
        return {
            "debit_account":  data.get("debit_account", rule_result["debit_account"]),
            "credit_account": data.get("credit_account", rule_result["credit_account"]),
            "narration":      data.get("narration", rule_result["narration"]),
            "confidence":     "ai",
        }
    except Exception as e:
        logger.warning("AI categorisation failed (%s), using rule fallback", e)
        return rule_result


# ═══════════════════════════════════════════════════════════════════════════
# DOUBLE-ENTRY ENGINE
# ═══════════════════════════════════════════════════════════════════════════

async def _get_account(org_id: str, code: str) -> Optional[Dict]:
    return await db.accounts.find_one({"org_id": org_id, "code": code})


async def _get_or_create_account(org_id: str, code: str) -> Dict:
    acc = await _get_account(org_id, code)
    if acc:
        return acc
    # Find in defaults
    default = next((a for a in DEFAULT_ACCOUNTS if a["code"] == code), None)
    if not default:
        raise HTTPException(400, f"Account code {code} not found")
    new_acc = {**default, "org_id": org_id, "_id": str(uuid.uuid4()), "opening_balance": 0.0,
               "created_at": datetime.now(timezone.utc).isoformat()}
    await db.accounts.insert_one(new_acc)
    return new_acc


async def post_journal_entry(org_id: str, entry_data: dict, user_id: str) -> str:
    """Post a balanced journal entry and update account balances."""
    lines = entry_data["lines"]

    # Validate balance
    total_dr = sum(l["amount"] for l in lines if l["type"] == "Dr")
    total_cr = sum(l["amount"] for l in lines if l["type"] == "Cr")
    if abs(total_dr - total_cr) > 0.01:
        raise HTTPException(400, f"Journal entry not balanced: Dr={total_dr} Cr={total_cr}")

    entry_id = str(uuid.uuid4())
    entry = {
        "_id": entry_id,
        "org_id": org_id,
        "date": entry_data["date"],
        "narration": entry_data.get("narration", ""),
        "ref_no": entry_data.get("ref_no", ""),
        "bank_statement_id": entry_data.get("bank_statement_id"),
        "lines": lines,
        "total_amount": total_dr,
        "created_by": user_id,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "status": "posted",
    }
    await db.journal_entries.insert_one(entry)

    # Update running balances on ledger
    for line in lines:
        acc = await _get_or_create_account(org_id, line["account_code"])
        normal = acc.get("normal_balance", "Dr")
        amount = line["amount"]
        # Increase balance if same as normal, decrease if opposite
        if line["type"] == normal:
            delta = amount
        else:
            delta = -amount
        await db.accounts.update_one(
            {"org_id": org_id, "code": line["account_code"]},
            {"$inc": {"running_balance": delta}}
        )
    return entry_id


# ═══════════════════════════════════════════════════════════════════════════
# FINANCIAL STATEMENT BUILDERS
# ═══════════════════════════════════════════════════════════════════════════

async def _build_ledger_balances(org_id: str, from_date: str, to_date: str) -> Dict[str, float]:
    """Compute balance per account from journal entries within date range."""
    pipeline = [
        {"$match": {"org_id": org_id, "date": {"$gte": from_date, "$lte": to_date}}},
        {"$unwind": "$lines"},
        {"$group": {
            "_id": "$lines.account_code",
            "dr_total": {"$sum": {"$cond": [{"$eq": ["$lines.type", "Dr"]}, "$lines.amount", 0]}},
            "cr_total": {"$sum": {"$cond": [{"$eq": ["$lines.type", "Cr"]}, "$lines.amount", 0]}},
        }},
    ]
    rows = await db.journal_entries.aggregate(pipeline).to_list(None)
    balances: Dict[str, float] = {}

    # Fetch all accounts to know normal balance
    accs = {a["code"]: a async for a in db.accounts.find({"org_id": org_id})}

    for row in rows:
        code = row["_id"]
        dr, cr = row["dr_total"], row["cr_total"]
        acc = accs.get(code) or next((a for a in DEFAULT_ACCOUNTS if a["code"] == code), None)
        normal = acc.get("normal_balance", "Dr") if acc else "Dr"
        opening = float((accs.get(code) or {}).get("opening_balance", 0))
        if normal == "Dr":
            balances[code] = opening + dr - cr
        else:
            balances[code] = opening + cr - dr

    # Also include accounts with opening balances but no transactions
    for code, acc in accs.items():
        if code not in balances:
            ob = float(acc.get("opening_balance", 0))
            if ob != 0:
                balances[code] = ob
    return balances


async def build_trial_balance(org_id: str, from_date: str, to_date: str) -> Dict:
    balances = await _build_ledger_balances(org_id, from_date, to_date)
    accounts = {a["code"]: a async for a in db.accounts.find({"org_id": org_id})}
    # Supplement with defaults
    for d in DEFAULT_ACCOUNTS:
        if d["code"] not in accounts:
            accounts[d["code"]] = d

    rows = []
    for code, bal in balances.items():
        acc = accounts.get(code, {})
        normal = acc.get("normal_balance", "Dr")
        dr = bal if (normal == "Dr" and bal >= 0) or (normal == "Cr" and bal < 0) else 0
        cr = abs(bal) if dr == 0 else 0
        rows.append({
            "code": code,
            "name": acc.get("name", code),
            "type": acc.get("type", ""),
            "dr":   round(abs(dr), 2) if normal == "Dr" else 0,
            "cr":   round(abs(cr), 2) if normal == "Cr" else 0,
        })
    rows.sort(key=lambda x: x["code"])
    total_dr = sum(r["dr"] for r in rows)
    total_cr = sum(r["cr"] for r in rows)
    return {"rows": rows, "total_dr": round(total_dr, 2), "total_cr": round(total_cr, 2),
            "balanced": abs(total_dr - total_cr) < 1}


async def build_trading_account(org_id: str, from_date: str, to_date: str) -> Dict:
    balances = await _build_ledger_balances(org_id, from_date, to_date)
    accounts = {a["code"]: a async for a in db.accounts.find({"org_id": org_id})}
    for d in DEFAULT_ACCOUNTS:
        if d["code"] not in accounts:
            accounts[d["code"]] = d

    opening_stock  = balances.get("1030", 0)
    purchases      = balances.get("5001", 0) + balances.get("5002", 0) + balances.get("5003", 0)
    sales_revenue  = balances.get("4003", 0)
    closing_stock  = 0.0   # user must set; we default to 0

    gross_profit = sales_revenue + closing_stock - opening_stock - purchases

    return {
        "debit_side": [
            {"label": "Opening Stock",     "amount": round(opening_stock, 2)},
            {"label": "Purchases",         "amount": round(purchases, 2)},
            {"label": "Direct Expenses",   "amount": round(balances.get("5003", 0), 2)},
        ],
        "credit_side": [
            {"label": "Sales Revenue",     "amount": round(sales_revenue, 2)},
            {"label": "Closing Stock",     "amount": round(closing_stock, 2)},
        ],
        "gross_profit": round(gross_profit, 2),
        "total_debit":  round(opening_stock + purchases, 2),
        "total_credit": round(sales_revenue + closing_stock, 2),
    }


async def build_pl_account(org_id: str, from_date: str, to_date: str) -> Dict:
    balances = await _build_ledger_balances(org_id, from_date, to_date)
    accounts = {a["code"]: a async for a in db.accounts.find({"org_id": org_id})}
    for d in DEFAULT_ACCOUNTS:
        if d["code"] not in accounts:
            accounts[d["code"]] = d

    # Collect all revenue accounts
    income_items  = []
    expense_items = []
    for code, bal in balances.items():
        acc = accounts.get(code, {})
        if not acc:
            continue
        acc_type = acc.get("type", "")
        sub = acc.get("sub_type", "")
        if acc_type == "Revenue":
            income_items.append({"code": code, "name": acc.get("name", code), "amount": round(abs(bal), 2)})
        elif acc_type == "Expense" and sub == "Indirect Expense":
            expense_items.append({"code": code, "name": acc.get("name", code), "amount": round(abs(bal), 2)})

    # Trading gross profit passed into P&L
    trading = await build_trading_account(org_id, from_date, to_date)
    gross_profit = trading["gross_profit"]

    total_income  = sum(i["amount"] for i in income_items) + max(gross_profit, 0)
    total_expense = sum(e["amount"] for e in expense_items)
    net_profit    = total_income - total_expense

    income_items.sort(key=lambda x: x["code"])
    expense_items.sort(key=lambda x: x["code"])

    return {
        "income":       income_items,
        "expenses":     expense_items,
        "gross_profit": round(gross_profit, 2),
        "total_income": round(total_income, 2),
        "total_expense":round(total_expense, 2),
        "net_profit":   round(net_profit, 2),
        "from_date":    from_date,
        "to_date":      to_date,
    }


async def build_balance_sheet(org_id: str, as_on: str) -> Dict:
    balances = await _build_ledger_balances(org_id, "2000-01-01", as_on)
    accounts = {a["code"]: a async for a in db.accounts.find({"org_id": org_id})}
    for d in DEFAULT_ACCOUNTS:
        if d["code"] not in accounts:
            accounts[d["code"]] = d

    pl = await build_pl_account(org_id, "2000-01-01", as_on)
    retained_earnings = pl["net_profit"]

    assets      = {"Fixed Assets": [], "Current Assets": [], "Other Assets": []}
    liabilities = {"Capital & Reserves": [], "Long Term Liabilities": [], "Current Liabilities": []}

    for code, bal in balances.items():
        if abs(bal) < 0.01:
            continue
        acc = accounts.get(code, {})
        if not acc:
            continue
        acc_type = acc.get("type", "")
        sub      = acc.get("sub_type", "")
        name     = acc.get("name", code)
        entry    = {"code": code, "name": name, "amount": round(abs(bal), 2)}

        if acc_type == "Asset":
            if sub in ("Fixed Asset",):
                assets["Fixed Assets"].append(entry)
            elif sub in ("Bank", "Cash", "Receivable", "Current Asset", "Prepaid", "Tax Asset"):
                assets["Current Assets"].append(entry)
            else:
                assets["Other Assets"].append(entry)
        elif acc_type == "Liability":
            if sub in ("Long Term Liability",):
                liabilities["Long Term Liabilities"].append(entry)
            elif sub in ("Equity", "Retained Earnings"):
                liabilities["Capital & Reserves"].append(entry)
            else:
                liabilities["Current Liabilities"].append(entry)
        elif acc_type == "Capital":
            liabilities["Capital & Reserves"].append(entry)

    # Add retained earnings
    liabilities["Capital & Reserves"].append({
        "code": "RE", "name": "Net Profit (Current Year)", "amount": round(retained_earnings, 2)
    })

    total_assets = sum(e["amount"] for grp in assets.values() for e in grp)
    total_liab   = sum(e["amount"] for grp in liabilities.values() for e in grp)

    return {
        "assets":       assets,
        "liabilities":  liabilities,
        "total_assets": round(total_assets, 2),
        "total_liabilities": round(total_liab, 2),
        "as_on": as_on,
        "balanced": abs(total_assets - total_liab) < 100,
    }


# ═══════════════════════════════════════════════════════════════════════════
# API  ENDPOINTS
# ═══════════════════════════════════════════════════════════════════════════

# ─── Chart of Accounts ────────────────────────────────────────────────────

@router.get("/accounts")
async def list_accounts(current_user: User = Depends(get_current_user)):
    org_id = str(current_user.id)
    accs = await db.accounts.find({"org_id": org_id}).to_list(None)
    if not accs:
        # Seed defaults on first access
        docs = [{**a, "_id": str(uuid.uuid4()), "org_id": org_id,
                 "opening_balance": 0.0, "running_balance": 0.0,
                 "created_at": datetime.now(timezone.utc).isoformat()} for a in DEFAULT_ACCOUNTS]
        await db.accounts.insert_many(docs)
        accs = docs
    for a in accs:
        a.pop("_id", None)
    return accs


@router.post("/accounts", status_code=201)
async def create_account(payload: AccountCreate, current_user: User = Depends(get_current_user)):
    org_id = str(current_user.id)
    existing = await db.accounts.find_one({"org_id": org_id, "code": payload.code})
    if existing:
        raise HTTPException(400, "Account code already exists")
    doc = {**payload.dict(), "_id": str(uuid.uuid4()), "org_id": org_id,
           "running_balance": payload.opening_balance,
           "created_at": datetime.now(timezone.utc).isoformat()}
    await db.accounts.insert_one(doc)
    doc.pop("_id", None)
    return doc


@router.put("/accounts/{code}")
async def update_account(code: str, payload: dict, current_user: User = Depends(get_current_user)):
    org_id = str(current_user.id)
    payload.pop("org_id", None); payload.pop("code", None)
    r = await db.accounts.update_one({"org_id": org_id, "code": code}, {"$set": payload})
    if r.matched_count == 0:
        raise HTTPException(404, "Account not found")
    return {"message": "Updated"}


# ─── Journal Entries ──────────────────────────────────────────────────────

@router.get("/journal-entries")
async def list_journal_entries(
    from_date: Optional[str] = None,
    to_date: Optional[str] = None,
    account_code: Optional[str] = None,
    page: int = 1,
    limit: int = 50,
    current_user: User = Depends(get_current_user),
):
    org_id = str(current_user.id)
    query: Dict[str, Any] = {"org_id": org_id}
    if from_date or to_date:
        query["date"] = {}
        if from_date: query["date"]["$gte"] = from_date
        if to_date:   query["date"]["$lte"] = to_date
    if account_code:
        query["lines"] = {"$elemMatch": {"account_code": account_code}}

    total = await db.journal_entries.count_documents(query)
    entries = await db.journal_entries.find(query).sort("date", -1).skip((page-1)*limit).limit(limit).to_list(None)
    for e in entries:
        e.pop("_id", None)
    return {"entries": entries, "total": total, "page": page, "limit": limit}


@router.post("/journal-entries", status_code=201)
async def create_journal_entry(payload: JournalEntryCreate, current_user: User = Depends(get_current_user)):
    org_id  = str(current_user.id)
    user_id = str(current_user.id)
    lines = [l.dict() for l in payload.lines]
    entry_data = {**payload.dict(), "lines": lines}
    entry_id = await post_journal_entry(org_id, entry_data, user_id)
    return {"message": "Journal entry posted", "id": entry_id}


@router.delete("/journal-entries/{entry_id}")
async def delete_journal_entry(entry_id: str, current_user: User = Depends(get_current_user)):
    org_id = str(current_user.id)
    entry = await db.journal_entries.find_one({"_id": entry_id, "org_id": org_id})
    if not entry:
        raise HTTPException(404, "Entry not found")
    # Reverse balance updates
    accounts = {a["code"]: a async for a in db.accounts.find({"org_id": org_id})}
    for line in entry.get("lines", []):
        acc = accounts.get(line["account_code"])
        if acc:
            normal = acc.get("normal_balance", "Dr")
            amount = line["amount"]
            delta = amount if line["type"] == normal else -amount
            await db.accounts.update_one(
                {"org_id": org_id, "code": line["account_code"]},
                {"$inc": {"running_balance": -delta}}
            )
    await db.journal_entries.delete_one({"_id": entry_id, "org_id": org_id})
    return {"message": "Entry deleted"}


# ─── Bank Statement Upload & Processing ──────────────────────────────────

@router.post("/bank-statements/upload")
async def upload_bank_statement(
    file: UploadFile = File(...),
    bank_name: str = Query("SBI"),
    account_number: str = Query(""),
    bank_account_code: str = Query("1002"),
    financial_year: str = Query("2024-25"),
    current_user: User = Depends(get_current_user),
):
    org_id = str(current_user.id)
    content = await file.read()
    fname = file.filename or ""

    # ── Parse ──────────────────────────────────────────────────────────
    ext = fname.lower().split(".")[-1]
    if ext in ("xlsx", "xls", "csv"):
        raw_txns = parse_excel_statement(content, fname)
    else:
        # PDF
        detected_bank = detect_bank(fname, "")
        if detected_bank == "SBI":
            raw_txns = parse_sbi_pdf(content)
        else:
            raw_txns = parse_generic_pdf(content)
        if not raw_txns:
            raw_txns = parse_sbi_pdf(content)  # Try SBI as fallback

    if not raw_txns:
        raise HTTPException(422, "Could not extract transactions from the uploaded file. "
                                 "Please ensure it is a valid bank statement in PDF/Excel/CSV format.")

    # ── Save statement record ─────────────────────────────────────────
    stmt_id = str(uuid.uuid4())
    stmt_doc = {
        "_id": stmt_id,
        "org_id": org_id,
        "bank_name": bank_name,
        "account_number": account_number,
        "bank_account_code": bank_account_code,
        "financial_year": financial_year,
        "filename": fname,
        "transaction_count": len(raw_txns),
        "uploaded_by": str(current_user.id),
        "uploaded_at": datetime.now(timezone.utc).isoformat(),
        "status": "processing",
    }
    await db.bank_statements.insert_one(stmt_doc)

    # ── Categorise & create journal entries ──────────────────────────
    created = 0
    categorised_txns = []
    for txn in raw_txns:
        is_credit = txn["credit"] > 0
        amount    = txn["credit"] if is_credit else txn["debit"]
        cat       = await ai_categorise_transaction(txn["description"], amount, is_credit, org_id)

        # Override bank account with the one selected by user
        if is_credit:
            cat["debit_account"] = bank_account_code
        else:
            cat["credit_account"] = bank_account_code

        lines = [
            {"account_code": cat["debit_account"],  "type": "Dr", "amount": amount, "narration": cat["narration"]},
            {"account_code": cat["credit_account"], "type": "Cr", "amount": amount, "narration": cat["narration"]},
        ]
        entry_data = {
            "date":               txn["date"],
            "narration":          cat["narration"] + f" | {txn['description'][:80]}",
            "ref_no":             txn.get("ref_no", ""),
            "bank_statement_id":  stmt_id,
            "lines":              lines,
        }
        try:
            await post_journal_entry(org_id, entry_data, str(current_user.id))
            created += 1
        except Exception as e:
            logger.warning("Skip txn: %s", e)

        categorised_txns.append({
            **txn,
            "category": cat["narration"],
            "debit_account":  cat["debit_account"],
            "credit_account": cat["credit_account"],
            "confidence": cat.get("confidence", ""),
        })

    # ── Update statement status ───────────────────────────────────────
    await db.bank_statements.update_one(
        {"_id": stmt_id},
        {"$set": {"status": "done", "entries_created": created,
                  "transactions": categorised_txns}}
    )

    return {
        "statement_id":  stmt_id,
        "transactions_found": len(raw_txns),
        "entries_created":    created,
        "transactions":       categorised_txns,
    }


@router.get("/bank-statements")
async def list_bank_statements(current_user: User = Depends(get_current_user)):
    org_id = str(current_user.id)
    stmts = await db.bank_statements.find({"org_id": org_id}).sort("uploaded_at", -1).to_list(None)
    for s in stmts:
        s.pop("_id", None)
        s.pop("transactions", None)   # exclude raw transactions from list
    return stmts


@router.get("/bank-statements/{stmt_id}")
async def get_bank_statement(stmt_id: str, current_user: User = Depends(get_current_user)):
    org_id = str(current_user.id)
    stmt = await db.bank_statements.find_one({"_id": stmt_id, "org_id": org_id})
    if not stmt:
        raise HTTPException(404, "Statement not found")
    stmt.pop("_id", None)
    return stmt


@router.delete("/bank-statements/{stmt_id}")
async def delete_bank_statement(stmt_id: str, current_user: User = Depends(get_current_user)):
    org_id = str(current_user.id)
    await db.journal_entries.delete_many({"org_id": org_id, "bank_statement_id": stmt_id})
    await db.bank_statements.delete_one({"_id": stmt_id, "org_id": org_id})
    return {"message": "Statement and related journal entries deleted"}


# ─── Ledger ───────────────────────────────────────────────────────────────

@router.get("/ledger/{account_code}")
async def get_ledger(
    account_code: str,
    from_date: str = Query("2024-04-01"),
    to_date:   str = Query("2025-03-31"),
    current_user: User = Depends(get_current_user),
):
    org_id = str(current_user.id)
    acc = await db.accounts.find_one({"org_id": org_id, "code": account_code})
    if not acc:
        default = next((a for a in DEFAULT_ACCOUNTS if a["code"] == account_code), None)
        if not default:
            raise HTTPException(404, "Account not found")
        acc = default

    entries = await db.journal_entries.find(
        {"org_id": org_id, "date": {"$gte": from_date, "$lte": to_date},
         "lines": {"$elemMatch": {"account_code": account_code}}}
    ).sort("date", 1).to_list(None)

    normal = acc.get("normal_balance", "Dr")
    opening = float(acc.get("opening_balance", 0))

    rows = []
    running = opening
    for entry in entries:
        for line in entry.get("lines", []):
            if line["account_code"] != account_code:
                continue
            amount = line["amount"]
            if line["type"] == normal:
                running += amount
            else:
                running -= amount
            rows.append({
                "date":     entry["date"],
                "narration":entry.get("narration", ""),
                "ref_no":   entry.get("ref_no", ""),
                "dr":       round(amount, 2) if line["type"] == "Dr" else 0,
                "cr":       round(amount, 2) if line["type"] == "Cr" else 0,
                "balance":  round(running, 2),
            })

    return {
        "account": {"code": acc.get("code"), "name": acc.get("name"), "type": acc.get("type"),
                    "normal_balance": normal},
        "opening_balance": opening,
        "closing_balance": round(running, 2),
        "rows": rows,
    }


# ─── Financial Reports ────────────────────────────────────────────────────

@router.get("/reports/trial-balance")
async def trial_balance(
    from_date: str = Query("2024-04-01"),
    to_date:   str = Query("2025-03-31"),
    current_user: User = Depends(get_current_user),
):
    return await build_trial_balance(str(current_user.id), from_date, to_date)


@router.get("/reports/trading-account")
async def trading_account(
    from_date: str = Query("2024-04-01"),
    to_date:   str = Query("2025-03-31"),
    current_user: User = Depends(get_current_user),
):
    return await build_trading_account(str(current_user.id), from_date, to_date)


@router.get("/reports/profit-loss")
async def profit_loss(
    from_date: str = Query("2024-04-01"),
    to_date:   str = Query("2025-03-31"),
    current_user: User = Depends(get_current_user),
):
    return await build_pl_account(str(current_user.id), from_date, to_date)


@router.get("/reports/balance-sheet")
async def balance_sheet(
    as_on: str = Query("2025-03-31"),
    current_user: User = Depends(get_current_user),
):
    return await build_balance_sheet(str(current_user.id), as_on)


@router.get("/reports/summary")
async def accounting_summary(
    from_date: str = Query("2024-04-01"),
    to_date:   str = Query("2025-03-31"),
    current_user: User = Depends(get_current_user),
):
    org_id = str(current_user.id)
    total_entries  = await db.journal_entries.count_documents({"org_id": org_id})
    total_stmts    = await db.bank_statements.count_documents({"org_id": org_id})
    total_accounts = await db.accounts.count_documents({"org_id": org_id})
    pl  = await build_pl_account(org_id, from_date, to_date)
    bs  = await build_balance_sheet(org_id, to_date)

    return {
        "total_journal_entries": total_entries,
        "total_bank_statements": total_stmts,
        "total_accounts":        total_accounts,
        "net_profit":            pl["net_profit"],
        "total_income":          pl["total_income"],
        "total_expense":         pl["total_expense"],
        "total_assets":          bs["total_assets"],
        "total_liabilities":     bs["total_liabilities"],
    }


# ═══════════════════════════════════════════════════════════════════════════
# EXPORT ENDPOINTS  (Excel)
# ═══════════════════════════════════════════════════════════════════════════

def _make_workbook_bytes(sheets: Dict[str, list]) -> bytes:
    """Create an in-memory Excel workbook with one sheet per key."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        import subprocess, sys
        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    HEADER_FILL  = PatternFill("solid", fgColor="1F6FB2")
    HEADER_FONT  = Font(color="FFFFFF", bold=True, size=10)
    BOLD         = Font(bold=True)
    TOTAL_FILL   = PatternFill("solid", fgColor="E0F2FE")
    thin_border  = Border(bottom=Side(style="thin", color="CCCCCC"))
    center       = Alignment(horizontal="center")
    right_align  = Alignment(horizontal="right")

    for sheet_name, rows in sheets.items():
        ws = wb.create_sheet(title=sheet_name[:31])
        if not rows:
            continue
        # Header row
        headers = list(rows[0].keys())
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.fill  = HEADER_FILL
            cell.font  = HEADER_FONT
            cell.alignment = center
        # Data rows
        for row_idx, row in enumerate(rows[1:], 2):
            is_total = str(row.get(headers[0], "")).upper() in ("TOTAL", "NET PROFIT", "NET LOSS", "GROSS PROFIT")
            for col_idx, h in enumerate(headers, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=row.get(h, ""))
                cell.border = thin_border
                if is_total:
                    cell.font = BOLD
                    cell.fill = TOTAL_FILL
                if isinstance(row.get(h), float):
                    cell.number_format = '#,##0.00'
                    cell.alignment = right_align
        # Auto-fit columns
        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


@router.get("/reports/export/trial-balance")
async def export_trial_balance(
    from_date: str = Query("2024-04-01"),
    to_date:   str = Query("2025-03-31"),
    current_user: User = Depends(get_current_user),
):
    org_id = str(current_user.id)
    tb = await build_trial_balance(org_id, from_date, to_date)
    rows = [{"Code": r["code"], "Account Name": r["name"], "Type": r["type"],
             "Debit (₹)": r["dr"], "Credit (₹)": r["cr"]}
            for r in tb["rows"] if r["dr"] > 0 or r["cr"] > 0]
    rows.append({"Code": "", "Account Name": "TOTAL", "Type": "",
                 "Debit (₹)": tb["total_dr"], "Credit (₹)": tb["total_cr"]})
    xlsx = _make_workbook_bytes({"Trial Balance": rows})
    return StreamingResponse(io.BytesIO(xlsx),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=trial_balance_{from_date}_{to_date}.xlsx"})


@router.get("/reports/export/profit-loss")
async def export_pl(
    from_date: str = Query("2024-04-01"),
    to_date:   str = Query("2025-03-31"),
    current_user: User = Depends(get_current_user),
):
    org_id = str(current_user.id)
    pl = await build_pl_account(org_id, from_date, to_date)

    income_rows  = [{"Account": i["name"], "Code": i["code"], "Amount (₹)": i["amount"]} for i in pl["income"]]
    income_rows.append({"Account": "TOTAL INCOME", "Code": "", "Amount (₹)": pl["total_income"]})

    expense_rows = [{"Account": e["name"], "Code": e["code"], "Amount (₹)": e["amount"]} for e in pl["expenses"]]
    expense_rows.append({"Account": "TOTAL EXPENSES", "Code": "", "Amount (₹)": pl["total_expense"]})
    expense_rows.append({"Account": "NET PROFIT" if pl["net_profit"] >= 0 else "NET LOSS",
                          "Code": "", "Amount (₹)": pl["net_profit"]})

    xlsx = _make_workbook_bytes({"Income": income_rows, "Expenses": expense_rows})
    return StreamingResponse(io.BytesIO(xlsx),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=profit_loss_{from_date}_{to_date}.xlsx"})


@router.get("/reports/export/balance-sheet")
async def export_balance_sheet(
    as_on: str = Query("2025-03-31"),
    current_user: User = Depends(get_current_user),
):
    org_id = str(current_user.id)
    bs = await build_balance_sheet(org_id, as_on)

    asset_rows = []
    for grp, items in bs["assets"].items():
        for item in items:
            asset_rows.append({"Group": grp, "Account": item["name"], "Code": item["code"], "Amount (₹)": item["amount"]})
    asset_rows.append({"Group": "", "Account": "TOTAL ASSETS", "Code": "", "Amount (₹)": bs["total_assets"]})

    liab_rows = []
    for grp, items in bs["liabilities"].items():
        for item in items:
            liab_rows.append({"Group": grp, "Account": item["name"], "Code": item.get("code",""), "Amount (₹)": item["amount"]})
    liab_rows.append({"Group": "", "Account": "TOTAL LIABILITIES", "Code": "", "Amount (₹)": bs["total_liabilities"]})

    xlsx = _make_workbook_bytes({"Assets": asset_rows, "Liabilities & Capital": liab_rows})
    return StreamingResponse(io.BytesIO(xlsx),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=balance_sheet_{as_on}.xlsx"})


@router.get("/reports/export/ledger/{account_code}")
async def export_ledger(
    account_code: str,
    from_date: str = Query("2024-04-01"),
    to_date:   str = Query("2025-03-31"),
    current_user: User = Depends(get_current_user),
):
    org_id = str(current_user.id)
    ledger = await get_ledger.__wrapped__(account_code, from_date, to_date, current_user) \
        if hasattr(get_ledger, "__wrapped__") else None

    # Re-compute directly
    acc = await db.accounts.find_one({"org_id": org_id, "code": account_code})
    if not acc:
        default = next((a for a in DEFAULT_ACCOUNTS if a["code"] == account_code), None)
        if not default:
            raise HTTPException(404, "Account not found")
        acc = default

    entries = await db.journal_entries.find(
        {"org_id": org_id, "date": {"$gte": from_date, "$lte": to_date},
         "lines": {"$elemMatch": {"account_code": account_code}}}
    ).sort("date", 1).to_list(None)

    normal  = acc.get("normal_balance", "Dr")
    running = float(acc.get("opening_balance", 0))
    rows    = [{"Date": "Opening Balance", "Narration": "", "Ref No": "", "Dr (₹)": 0, "Cr (₹)": 0, "Balance (₹)": running}]

    for entry in entries:
        for line in entry.get("lines", []):
            if line["account_code"] != account_code:
                continue
            amount = line["amount"]
            running += amount if line["type"] == normal else -amount
            rows.append({
                "Date":        entry["date"],
                "Narration":   entry.get("narration", ""),
                "Ref No":      entry.get("ref_no", ""),
                "Dr (₹)":      round(amount, 2) if line["type"] == "Dr" else 0.0,
                "Cr (₹)":      round(amount, 2) if line["type"] == "Cr" else 0.0,
                "Balance (₹)": round(running, 2),
            })
    rows.append({"Date": "Closing Balance", "Narration": "", "Ref No": "",
                 "Dr (₹)": 0, "Cr (₹)": 0, "Balance (₹)": round(running, 2)})

    xlsx = _make_workbook_bytes({f"Ledger {account_code}": rows})
    return StreamingResponse(io.BytesIO(xlsx),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=ledger_{account_code}_{from_date}_{to_date}.xlsx"})


@router.get("/reports/export/journal-entries")
async def export_journal_entries(
    from_date: Optional[str] = None,
    to_date:   Optional[str] = None,
    current_user: User = Depends(get_current_user),
):
    org_id = str(current_user.id)
    query: Dict[str, Any] = {"org_id": org_id}
    if from_date or to_date:
        query["date"] = {}
        if from_date: query["date"]["$gte"] = from_date
        if to_date:   query["date"]["$lte"] = to_date

    entries = await db.journal_entries.find(query).sort("date", 1).to_list(None)
    accounts = {a["code"]: a["name"] async for a in db.accounts.find({"org_id": org_id})}

    rows = []
    for e in entries:
        for line in e.get("lines", []):
            rows.append({
                "Date":         e.get("date", ""),
                "Narration":    e.get("narration", ""),
                "Ref No":       e.get("ref_no", ""),
                "Account Code": line.get("account_code", ""),
                "Account Name": accounts.get(line.get("account_code",""), line.get("account_code","")),
                "Dr/Cr":        line.get("type", ""),
                "Amount (₹)":   float(line.get("amount", 0)),
            })

    xlsx = _make_workbook_bytes({"Journal Entries": rows})
    return StreamingResponse(io.BytesIO(xlsx),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=journal_entries.xlsx"})


# ═══════════════════════════════════════════════════════════════════════════
# RE-CATEGORISATION  (update a journal entry's account mapping)
# ═══════════════════════════════════════════════════════════════════════════

class RecategoriseRequest(BaseModel):
    entry_id: str
    new_debit_account:  str
    new_credit_account: str
    new_narration: Optional[str] = None


@router.post("/recategorise")
async def recategorise_entry(
    payload: RecategoriseRequest,
    current_user: User = Depends(get_current_user),
):
    """
    Update the account codes on a single journal entry (from bank statement import).
    Reverses old balance effects, applies new ones.
    """
    org_id = str(current_user.id)
    entry  = await db.journal_entries.find_one({"_id": payload.entry_id, "org_id": org_id})
    if not entry:
        raise HTTPException(404, "Entry not found")

    # Step 1: Reverse old balance effects
    accounts = {a["code"]: a async for a in db.accounts.find({"org_id": org_id})}
    for line in entry.get("lines", []):
        acc = accounts.get(line["account_code"])
        if acc:
            normal = acc.get("normal_balance", "Dr")
            delta  = line["amount"] if line["type"] == normal else -line["amount"]
            await db.accounts.update_one(
                {"org_id": org_id, "code": line["account_code"]},
                {"$inc": {"running_balance": -delta}}
            )

    # Step 2: Rebuild lines with new accounts
    amount    = entry["lines"][0]["amount"]  # both lines share same amount
    new_lines = [
        {"account_code": payload.new_debit_account,  "type": "Dr", "amount": amount,
         "narration": payload.new_narration or entry.get("narration","")},
        {"account_code": payload.new_credit_account, "type": "Cr", "amount": amount,
         "narration": payload.new_narration or entry.get("narration","")},
    ]
    updates = {"lines": new_lines}
    if payload.new_narration:
        updates["narration"] = payload.new_narration

    await db.journal_entries.update_one({"_id": payload.entry_id, "org_id": org_id}, {"$set": updates})

    # Step 3: Apply new balance effects
    for line in new_lines:
        acc = await _get_or_create_account(org_id, line["account_code"])
        normal = acc.get("normal_balance", "Dr")
        delta  = line["amount"] if line["type"] == normal else -line["amount"]
        await db.accounts.update_one(
            {"org_id": org_id, "code": line["account_code"]},
            {"$inc": {"running_balance": delta}}
        )

    return {"message": "Entry re-categorised successfully"}


# ═══════════════════════════════════════════════════════════════════════════
# OPENING BALANCES  (bulk update)
# ═══════════════════════════════════════════════════════════════════════════

class OpeningBalanceItem(BaseModel):
    code:            str
    opening_balance: float


@router.post("/accounts/opening-balances")
async def set_opening_balances(
    payload: List[OpeningBalanceItem],
    current_user: User = Depends(get_current_user),
):
    org_id = str(current_user.id)
    updated = 0
    for item in payload:
        r = await db.accounts.update_one(
            {"org_id": org_id, "code": item.code},
            {"$set": {"opening_balance": item.opening_balance, "running_balance": item.opening_balance}},
            upsert=False,
        )
        if r.matched_count:
            updated += 1
    return {"message": f"Updated opening balances for {updated} accounts"}
