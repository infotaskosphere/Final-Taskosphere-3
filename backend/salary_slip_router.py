# backend/salary_slip_router.py
# ----------------------------------------------------------------------------
# Salary Slip Generator — a Compliance module feature that lets the firm
# generate payslips for CLIENT COMPANIES' employees (as a service the firm
# performs for its clients), as distinct from the existing People Matrix →
# Payroll module which runs payroll for the firm's OWN staff.
#
# Mount this router from backend/server.py:
#
#     from backend.salary_slip_router import router as salary_slip_router, \
#         create_salary_slip_indexes
#     app.include_router(salary_slip_router)
#     # ...inside startup_event(): await create_salary_slip_indexes()
#
# All endpoints live under /api/compliance/salary-slips and are guarded by
# the "salary_slips" module permission (can_view_salary_slips /
# can_manage_salary_slips) via check_module_permission — see the matching
# entries added to backend/dependencies.py::MODULE_ACTION_MAP.
#
# Data model
# ----------
#   db.salary_manual_companies  — ad-hoc companies not in the Clients DB
#   db.salary_employees         — reusable employee master (linked to either
#                                  a Client record or a manual company)
#   db.salary_slips             — generated payslips (fully self-contained
#                                  snapshots, so a slip never changes even if
#                                  the employee/company record is edited later)
#
# A "company_key" is either "client:<client_id>" or "manual:<manual_company_id>".
# The frontend is expected to source the "client:*" options from the existing,
# already access-controlled GET /api/clients endpoint, and the "manual:*"
# options from this router's /manual-companies endpoints — this router does
# NOT re-derive client visibility rules itself, it only checks that a given
# client_id exists before snapshotting its details onto an employee/slip.
# ----------------------------------------------------------------------------

from __future__ import annotations

import uuid
import logging
from datetime import datetime, timezone
from io import BytesIO
from typing import Any, Dict, List, Optional

from fastapi import APIRouter, Depends, File, Form, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, EmailStr, Field

import pandas as pd

from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.lib import colors
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
    PageBreak, HRFlowable,
)

from backend.dependencies import db, get_current_user, check_module_permission
from backend.models import User

logger = logging.getLogger("salary_slip")
router = APIRouter(prefix="/api/compliance/salary-slips", tags=["Salary Slips"])

MONTH_NAMES = [
    "", "January", "February", "March", "April", "May", "June",
    "July", "August", "September", "October", "November", "December",
]


# ═════════════════════════════════════════════════════════════════════════
# HELPERS
# ═════════════════════════════════════════════════════════════════════════

def _now() -> str:
    return datetime.now(timezone.utc).isoformat()


def _new_id() -> str:
    return str(uuid.uuid4())


def _perm_ok(user: User, flag: str) -> bool:
    if user.role == "admin":
        return True
    perms = user.permissions if isinstance(user.permissions, dict) else (
        user.permissions.model_dump() if hasattr(user.permissions, "model_dump") else {}
    )
    return bool(perms.get(flag, False))


# ── Amount in words (Indian numbering: Crore / Lakh / Thousand) ────────────
# Deliberately self-contained (not imported from backend.invoicing) so this
# router has no dependency on the invoicing module.
_ONES = ["", "One", "Two", "Three", "Four", "Five", "Six", "Seven", "Eight", "Nine", "Ten",
         "Eleven", "Twelve", "Thirteen", "Fourteen", "Fifteen", "Sixteen", "Seventeen",
         "Eighteen", "Nineteen"]
_TENS = ["", "", "Twenty", "Thirty", "Forty", "Fifty", "Sixty", "Seventy", "Eighty", "Ninety"]


def amount_in_words(n: float) -> str:
    try:
        n = float(n or 0)
        rupees = int(n)
        paise = round((n - rupees) * 100)

        def _grp(num):
            if num == 0: return ""
            if num < 20: return _ONES[num] + " "
            if num < 100: return _TENS[num // 10] + (" " + _ONES[num % 10] if num % 10 else "") + " "
            return _ONES[num // 100] + " Hundred " + _grp(num % 100)

        def _convert(num):
            if num == 0: return "Zero "
            r = ""
            cr = num // 10_000_000; num %= 10_000_000
            lk = num // 100_000; num %= 100_000
            th = num // 1000; num %= 1000
            if cr: r += _grp(cr) + "Crore "
            if lk: r += _grp(lk) + "Lakh "
            if th: r += _grp(th) + "Thousand "
            r += _grp(num)
            return r

        r = _convert(rupees).strip()
        p = f" and {_convert(paise).strip()} Paise" if paise else ""
        return f"Rupees {r}{p} Only"
    except Exception:
        return f"Rupees {n:.2f} Only"


def inr(n: float) -> str:
    """Format a number with Indian digit grouping, e.g. 1234567.5 -> '12,34,567.50'."""
    try:
        n = float(n or 0)
    except Exception:
        n = 0.0
    neg = n < 0
    n = abs(n)
    whole = int(n)
    frac = round((n - whole) * 100)
    if frac >= 100:
        whole += 1
        frac -= 100
    s = str(whole)
    if len(s) > 3:
        last3, rest = s[-3:], s[:-3]
        groups = []
        while len(rest) > 2:
            groups.insert(0, rest[-2:])
            rest = rest[:-2]
        if rest:
            groups.insert(0, rest)
        s = ",".join(groups) + "," + last3
    return f"{'-' if neg else ''}Rs. {s}.{frac:02d}"


async def _resolve_company(company_key: str) -> Dict[str, Any]:
    """
    Resolve a "client:<id>" or "manual:<id>" key into a snapshot dict of
    company display fields. Raises 404 if the referenced record is missing.
    """
    if not company_key or ":" not in company_key:
        raise HTTPException(400, "Invalid company_key — expected 'client:<id>' or 'manual:<id>'")
    source, ref_id = company_key.split(":", 1)

    if source == "client":
        c = await db.clients.find_one({"id": ref_id}, {"_id": 0})
        if not c:
            raise HTTPException(404, "Linked client not found")
        return {
            "company_key":     company_key,
            "company_source":  "client",
            "company_ref_id":  ref_id,
            "company_name":    c.get("company_name") or "",
            "company_address": c.get("address") or c.get("gst_address") or "",
            "company_city":    c.get("city") or c.get("gst_city") or "",
            "company_state":   c.get("state") or c.get("gst_state") or "",
            "company_gstin":   c.get("gstin") or "",
            "company_pan":     c.get("pan") or "",
        }

    if source == "manual":
        c = await db.salary_manual_companies.find_one({"id": ref_id}, {"_id": 0})
        if not c:
            raise HTTPException(404, "Manual company not found")
        return {
            "company_key":     company_key,
            "company_source":  "manual",
            "company_ref_id":  ref_id,
            "company_name":    c.get("name") or "",
            "company_address": c.get("address") or "",
            "company_city":    c.get("city") or "",
            "company_state":   c.get("state") or "",
            "company_gstin":   c.get("gstin") or "",
            "company_pan":     c.get("pan") or "",
        }

    raise HTTPException(400, "company_key must start with 'client:' or 'manual:'")


def _line_items(items: Optional[List["SalaryLineItem"]]) -> List[Dict[str, Any]]:
    return [{"label": (it.label or "").strip(), "amount": round(float(it.amount or 0), 2)}
            for it in (items or []) if (it.label or "").strip()]


def _sum_items(items: List[Dict[str, Any]]) -> float:
    return round(sum(float(it.get("amount") or 0) for it in items), 2)


async def _next_slip_no(year: int) -> str:
    """Human-friendly sequential slip reference, e.g. PS-2026-0007.
    Uses count_documents (works against both the real MongoDB driver and the
    in-memory mock DB used in local/dev mode) rather than an atomic counter,
    which is an acceptable trade-off for a low-volume internal document.
    """
    count = await db.salary_slips.count_documents({"slip_year": year})
    return f"PS-{year}-{count + 1:04d}"


# ═════════════════════════════════════════════════════════════════════════
# PYDANTIC MODELS
# ═════════════════════════════════════════════════════════════════════════

class SalaryLineItem(BaseModel):
    label: str
    amount: float = 0


class ManualCompanyCreate(BaseModel):
    name: str = Field(..., min_length=2, max_length=255)
    address: Optional[str] = None
    city: Optional[str] = None
    state: Optional[str] = None
    gstin: Optional[str] = None
    pan: Optional[str] = None
    contact_email: Optional[EmailStr] = None
    contact_phone: Optional[str] = None


class ManualCompanyUpdate(BaseModel):
    name: Optional[str] = None
    address: Optional[str] = None
    city: Optional[str] = None
    state: Optional[str] = None
    gstin: Optional[str] = None
    pan: Optional[str] = None
    contact_email: Optional[EmailStr] = None
    contact_phone: Optional[str] = None


class EmployeeCreate(BaseModel):
    company_key: str
    name: str = Field(..., min_length=1, max_length=255)
    employee_code: Optional[str] = None
    designation: Optional[str] = None
    department: Optional[str] = None
    date_of_joining: Optional[str] = None
    dob: Optional[str] = None
    gender: Optional[str] = None
    pan: Optional[str] = None
    uan: Optional[str] = None
    pf_number: Optional[str] = None
    esic_number: Optional[str] = None
    bank_name: Optional[str] = None
    bank_account: Optional[str] = None
    ifsc: Optional[str] = None
    email: Optional[EmailStr] = None
    phone: Optional[str] = None
    status: str = "active"
    default_earnings: List[SalaryLineItem] = Field(default_factory=list)
    default_deductions: List[SalaryLineItem] = Field(default_factory=list)


class EmployeeUpdate(BaseModel):
    company_key: Optional[str] = None
    name: Optional[str] = None
    employee_code: Optional[str] = None
    designation: Optional[str] = None
    department: Optional[str] = None
    date_of_joining: Optional[str] = None
    dob: Optional[str] = None
    gender: Optional[str] = None
    pan: Optional[str] = None
    uan: Optional[str] = None
    pf_number: Optional[str] = None
    esic_number: Optional[str] = None
    bank_name: Optional[str] = None
    bank_account: Optional[str] = None
    ifsc: Optional[str] = None
    email: Optional[EmailStr] = None
    phone: Optional[str] = None
    status: Optional[str] = None
    default_earnings: Optional[List[SalaryLineItem]] = None
    default_deductions: Optional[List[SalaryLineItem]] = None


class SlipGenerate(BaseModel):
    employee_id: str
    month: int = Field(..., ge=1, le=12)
    year: int = Field(..., ge=2000, le=2100)
    pay_date: Optional[str] = None
    total_days: float = 30
    paid_days: float = 30
    lop_days: float = 0
    earnings: Optional[List[SalaryLineItem]] = None     # None -> use employee defaults
    deductions: Optional[List[SalaryLineItem]] = None
    template: str = "modern"                             # modern | classic
    notes: Optional[str] = None
    status: str = "final"                                 # draft | final


class SlipPeriod(BaseModel):
    month: int = Field(..., ge=1, le=12)
    year: int = Field(..., ge=2000, le=2100)


class BulkGenerateBody(BaseModel):
    company_key: str
    month: Optional[int] = Field(None, ge=1, le=12)        # single-period shorthand (back-compat)
    year: Optional[int] = Field(None, ge=2000, le=2100)
    periods: Optional[List[SlipPeriod]] = None             # multiple months/years in one request; wins over month/year when given
    employee_ids: Optional[List[str]] = None              # None -> all active employees
    pay_date: Optional[str] = None
    total_days: float = 30
    paid_days: float = 30
    lop_days: float = 0
    template: str = "modern"
    status: str = "final"


class SlipUpdate(BaseModel):
    pay_date: Optional[str] = None
    total_days: Optional[float] = None
    paid_days: Optional[float] = None
    lop_days: Optional[float] = None
    earnings: Optional[List[SalaryLineItem]] = None
    deductions: Optional[List[SalaryLineItem]] = None
    template: Optional[str] = None
    notes: Optional[str] = None
    status: Optional[str] = None


class DuplicateBody(BaseModel):
    month: int = Field(..., ge=1, le=12)
    year: int = Field(..., ge=2000, le=2100)


class BulkPdfBody(BaseModel):
    slip_ids: List[str]


# Common component presets the frontend can offer as one-click "add row"
# buttons — mirrors the fields seen on standard Indian payslips (including
# the firm's own internal Payroll module) without hard-coding a full
# statutory calculation engine, since accountants using this tool enter the
# already-computed figures themselves.
EARNING_PRESETS = ["Basic", "House Rent Allowance (HRA)", "Conveyance Allowance",
                    "Medical Allowance", "Special Allowance", "Children Education Allowance",
                    "Fixed Allowance", "Bonus", "Overtime", "Arrears"]
DEDUCTION_PRESETS = ["Provident Fund (EPF)", "Employee State Insurance (ESI)",
                      "Professional Tax", "TDS (Income Tax)", "Loan / Advance Recovery",
                      "Labour Welfare Fund (LWF)"]


# ═════════════════════════════════════════════════════════════════════════
# EXCEL IMPORT — bulk-add a company's employees from a spreadsheet
# ═════════════════════════════════════════════════════════════════════════
# The template exposes a fixed set of "plain" fields plus a handful of the
# most common earnings/deduction columns (mirrors the presets above). Any
# amount column left blank/zero is simply omitted from that employee's
# default_earnings/default_deductions — the accountant can still add more
# exotic components later from the employee's edit modal.

_IMPORT_PLAIN_FIELDS = [
    # (template header,               employee doc field)
    ("Name*",                         "name"),
    ("Employee Code",                 "employee_code"),
    ("Designation",                   "designation"),
    ("Department",                    "department"),
    ("Date of Joining (YYYY-MM-DD)",  "date_of_joining"),
    ("PAN",                           "pan"),
    ("UAN",                           "uan"),
    ("PF Number",                     "pf_number"),
    ("ESIC Number",                   "esic_number"),
    ("Bank Name",                     "bank_name"),
    ("Bank Account No.",              "bank_account"),
    ("IFSC",                          "ifsc"),
    ("Email",                         "email"),
    ("Phone",                         "phone"),
    ("Status (active/inactive)",      "status"),
]

_IMPORT_EARNING_COLS = [
    ("Basic",                        "Basic"),
    ("HRA",                          "House Rent Allowance (HRA)"),
    ("Conveyance Allowance",         "Conveyance Allowance"),
    ("Medical Allowance",            "Medical Allowance"),
    ("Special Allowance",            "Special Allowance"),
    ("Other Earnings",               "Other Earnings"),
]
_IMPORT_DEDUCTION_COLS = [
    ("PF (EPF)",                     "Provident Fund (EPF)"),
    ("ESI",                          "Employee State Insurance (ESI)"),
    ("Professional Tax",             "Professional Tax"),
    ("TDS",                          "TDS (Income Tax)"),
    ("Other Deductions",             "Other Deductions"),
]

_IMPORT_TEMPLATE_HEADERS = (
    [h for h, _ in _IMPORT_PLAIN_FIELDS]
    + [h for h, _ in _IMPORT_EARNING_COLS]
    + [h for h, _ in _IMPORT_DEDUCTION_COLS]
)

_IMPORT_EXAMPLE_ROW = [
    "Rahul Sharma", "EMP-001", "Accountant", "Finance", "2023-06-01",
    "ABCDE1234F", "100123456789", "PF/12345", "ESIC/98765",
    "HDFC Bank", "50100123456789", "HDFC0000123",
    "rahul@example.com", "9876543210", "active",
    25000, 10000, 1600, 1250, 0, 0,
    3000, 787, 200, 0, 0,
]


def _norm_header(h: str) -> str:
    """Loose header match: lowercase, strip, drop anything in parens/asterisks/punctuation."""
    import re
    h = str(h or "").lower()
    h = re.sub(r"\(.*?\)", "", h)          # drop "(YYYY-MM-DD)", "(EPF)" etc.
    h = re.sub(r"[^a-z0-9]+", " ", h)      # punctuation -> space
    return h.strip()


def _read_employee_file(contents: bytes, filename: str) -> pd.DataFrame:
    try:
        if (filename or "").lower().endswith(".csv"):
            return pd.read_csv(BytesIO(contents), dtype=str)
        return pd.read_excel(BytesIO(contents), dtype=str)
    except Exception as exc:
        raise HTTPException(400, f"Could not parse file: {exc}")


def _build_import_column_map(columns: List[str]) -> Dict[str, str]:
    """Maps each template header -> the actual column name found in the
    uploaded file (loose match), for every recognised header present."""
    norm_to_actual = {_norm_header(c): c for c in columns}
    mapping: Dict[str, str] = {}
    for header, _field in _IMPORT_PLAIN_FIELDS + _IMPORT_EARNING_COLS + _IMPORT_DEDUCTION_COLS:
        key = _norm_header(header)
        if key in norm_to_actual:
            mapping[header] = norm_to_actual[key]
    return mapping


def _num(val) -> float:
    try:
        if val is None or (isinstance(val, float) and pd.isna(val)):
            return 0.0
        s = str(val).replace(",", "").strip()
        if not s or s.lower() == "nan":
            return 0.0
        return round(float(s), 2)
    except (ValueError, TypeError):
        return 0.0


def _row_to_employee_fields(row: "pd.Series", col_map: Dict[str, str]):
    """Turns one spreadsheet row into (plain_fields, earnings, deductions)."""
    def get(header: str) -> str:
        actual = col_map.get(header)
        if not actual or actual not in row:
            return ""
        v = row[actual]
        if v is None or (isinstance(v, float) and pd.isna(v)):
            return ""
        return str(v).strip()

    fields: Dict[str, Any] = {}
    for header, field in _IMPORT_PLAIN_FIELDS:
        fields[field] = get(header)
    fields["status"] = fields.get("status", "").lower() if fields.get("status", "").lower() in ("active", "inactive") else "active"

    earnings = [{"label": label, "amount": _num(row.get(col_map[header]))}
                for header, label in _IMPORT_EARNING_COLS
                if header in col_map and _num(row.get(col_map[header])) > 0]
    deductions = [{"label": label, "amount": _num(row.get(col_map[header]))}
                  for header, label in _IMPORT_DEDUCTION_COLS
                  if header in col_map and _num(row.get(col_map[header])) > 0]

    return fields, earnings, deductions


def _build_import_template_bytes() -> bytes:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Employees"

    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=10)

    for col_idx, header in enumerate(_IMPORT_TEMPLATE_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = max(14, len(header) * 0.9)

    for col_idx, value in enumerate(_IMPORT_EXAMPLE_ROW, start=1):
        ws.cell(row=2, column=col_idx, value=value)

    ws.freeze_panes = "A2"

    notes = wb.create_sheet("Instructions")
    notes["A1"] = "How to use this template"
    notes["A1"].font = Font(bold=True, size=12)
    for i, line in enumerate([
        "Fill one row per employee on the Employees sheet. 'Name*' is the only required column.",
        "Leave any column blank if it doesn't apply — blank earning/deduction amounts (or 0) are simply skipped.",
        "'Status' accepts active or inactive (defaults to active if left blank).",
        "Re-uploading with the same Employee Code (or Name, if no code) updates that employee instead of creating a duplicate.",
        "Dates should be in YYYY-MM-DD format.",
    ], start=3):
        notes[f"A{i}"] = f"• {line}"
    notes.column_dimensions["A"].width = 100

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ═════════════════════════════════════════════════════════════════════════
# MANUAL (AD-HOC) COMPANIES — for companies not present in the Clients DB
# ═════════════════════════════════════════════════════════════════════════

@router.get("/manual-companies")
async def list_manual_companies(
    current_user: User = Depends(check_module_permission("salary_slips", "view")),
):
    items = await db.salary_manual_companies.find({}, {"_id": 0}).sort("name", 1).to_list(1000)
    return items


@router.post("/manual-companies")
async def create_manual_company(
    data: ManualCompanyCreate,
    current_user: User = Depends(check_module_permission("salary_slips", "create")),
):
    doc = {
        "id": _new_id(),
        **data.model_dump(),
        "created_by": current_user.id,
        "created_at": _now(),
        "updated_at": _now(),
    }
    await db.salary_manual_companies.insert_one({**doc, "_id": doc["id"]})
    doc.pop("_id", None)
    return doc


@router.patch("/manual-companies/{company_id}")
async def update_manual_company(
    company_id: str,
    data: ManualCompanyUpdate,
    current_user: User = Depends(check_module_permission("salary_slips", "edit")),
):
    existing = await db.salary_manual_companies.find_one({"id": company_id})
    if not existing:
        raise HTTPException(404, "Company not found")
    updates = {k: v for k, v in data.model_dump(exclude_unset=True).items()}
    updates["updated_at"] = _now()
    await db.salary_manual_companies.update_one({"id": company_id}, {"$set": updates})
    updated = await db.salary_manual_companies.find_one({"id": company_id}, {"_id": 0})
    return updated


@router.delete("/manual-companies/{company_id}")
async def delete_manual_company(
    company_id: str,
    current_user: User = Depends(check_module_permission("salary_slips", "delete")),
):
    existing = await db.salary_manual_companies.find_one({"id": company_id})
    if not existing:
        raise HTTPException(404, "Company not found")
    linked = await db.salary_employees.count_documents({"company_key": f"manual:{company_id}"})
    if linked:
        raise HTTPException(
            400,
            f"Cannot delete — {linked} employee(s) are linked to this company. "
            f"Remove or reassign them first.",
        )
    await db.salary_manual_companies.delete_one({"id": company_id})
    return {"deleted": True}


# ═════════════════════════════════════════════════════════════════════════
# EMPLOYEE MASTER
# ═════════════════════════════════════════════════════════════════════════

@router.get("/employees")
async def list_employees(
    company_key: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    search: Optional[str] = Query(None),
    current_user: User = Depends(check_module_permission("salary_slips", "view")),
):
    query: Dict[str, Any] = {}
    if company_key:
        query["company_key"] = company_key
    if status and status != "all":
        query["status"] = status
    if search and search.strip():
        query["name"] = {"$regex": search.strip(), "$options": "i"}
    items = await db.salary_employees.find(query, {"_id": 0}).sort("name", 1).to_list(2000)
    return items


@router.get("/employees/import-template")
async def download_employee_import_template(
    current_user: User = Depends(check_module_permission("salary_slips", "view")),
):
    # Registered ahead of GET /employees/{employee_id} — otherwise FastAPI
    # would match "import-template" as an employee_id path param.
    xlsx_bytes = _build_import_template_bytes()
    return StreamingResponse(
        BytesIO(xlsx_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=Employee_Import_Template.xlsx"},
    )


@router.get("/employees/{employee_id}")
async def get_employee(
    employee_id: str,
    current_user: User = Depends(check_module_permission("salary_slips", "view")),
):
    emp = await db.salary_employees.find_one({"id": employee_id}, {"_id": 0})
    if not emp:
        raise HTTPException(404, "Employee not found")
    return emp


@router.post("/employees")
async def create_employee(
    data: EmployeeCreate,
    current_user: User = Depends(check_module_permission("salary_slips", "create")),
):
    company = await _resolve_company(data.company_key)
    doc = {
        "id": _new_id(),
        **{k: v for k, v in data.model_dump().items() if k not in ("default_earnings", "default_deductions")},
        "default_earnings":   _line_items(data.default_earnings),
        "default_deductions": _line_items(data.default_deductions),
        "company_name":  company["company_name"],
        "created_by":    current_user.id,
        "created_at":    _now(),
        "updated_at":    _now(),
    }
    await db.salary_employees.insert_one({**doc, "_id": doc["id"]})
    doc.pop("_id", None)
    return doc


@router.patch("/employees/{employee_id}")
async def update_employee(
    employee_id: str,
    data: EmployeeUpdate,
    current_user: User = Depends(check_module_permission("salary_slips", "edit")),
):
    existing = await db.salary_employees.find_one({"id": employee_id})
    if not existing:
        raise HTTPException(404, "Employee not found")

    updates = data.model_dump(exclude_unset=True)
    if "default_earnings" in updates and updates["default_earnings"] is not None:
        updates["default_earnings"] = _line_items(data.default_earnings)
    if "default_deductions" in updates and updates["default_deductions"] is not None:
        updates["default_deductions"] = _line_items(data.default_deductions)
    if updates.get("company_key"):
        company = await _resolve_company(updates["company_key"])
        updates["company_name"] = company["company_name"]
    updates["updated_at"] = _now()

    await db.salary_employees.update_one({"id": employee_id}, {"$set": updates})
    updated = await db.salary_employees.find_one({"id": employee_id}, {"_id": 0})
    return updated


@router.delete("/employees/{employee_id}")
async def delete_employee(
    employee_id: str,
    current_user: User = Depends(check_module_permission("salary_slips", "delete")),
):
    existing = await db.salary_employees.find_one({"id": employee_id})
    if not existing:
        raise HTTPException(404, "Employee not found")
    await db.salary_employees.delete_one({"id": employee_id})
    # Note: previously generated slips are self-contained snapshots and are
    # intentionally left untouched so payslip history stays intact.
    return {"deleted": True}


# ── Excel import: preview / commit (template route is above, near GET /employees) ──

@router.post("/employees/import-preview")
async def preview_employee_import(
    company_key: str = Form(...),
    file: UploadFile = File(...),
    current_user: User = Depends(check_module_permission("salary_slips", "create")),
):
    """Parses the uploaded sheet and returns a row-by-row preview (name,
    code, gross earnings, will-create-vs-update) so the frontend can show a
    confirmation table before anything is written to the database."""
    await _resolve_company(company_key)  # 404s early if the company_key is bad

    contents = await file.read()
    df = _read_employee_file(contents, file.filename or "upload.xlsx")
    if df.empty:
        raise HTTPException(400, "The file has no data rows")

    df.columns = [str(c).strip() for c in df.columns]
    col_map = _build_import_column_map(list(df.columns))
    if "Name*" not in col_map:
        raise HTTPException(
            400,
            "Couldn't find a 'Name' column. Please use the downloaded template, "
            f"or make sure a column named 'Name' exists. Columns found: {list(df.columns)}",
        )

    existing = await db.salary_employees.find(
        {"company_key": company_key}, {"_id": 0, "id": 1, "name": 1, "employee_code": 1},
    ).to_list(2000)
    by_code = {e["employee_code"].strip().lower(): e for e in existing if e.get("employee_code")}
    by_name = {e["name"].strip().lower(): e for e in existing if e.get("name")}

    rows_preview = []
    for idx, row in df.iterrows():
        fields, earnings, deductions = _row_to_employee_fields(row, col_map)
        name = fields.get("name", "").strip()
        if not name:
            rows_preview.append({
                "row": int(idx) + 2, "name": "", "employee_code": fields.get("employee_code", ""),
                "status": "error", "reason": "Missing name — row will be skipped", "gross": 0,
            })
            continue

        code_key = (fields.get("employee_code") or "").strip().lower()
        match = by_code.get(code_key) if code_key else None
        if not match:
            match = by_name.get(name.lower())

        rows_preview.append({
            "row": int(idx) + 2,
            "name": name,
            "employee_code": fields.get("employee_code", ""),
            "designation": fields.get("designation", ""),
            "status": "update" if match else "new",
            "existing_id": match["id"] if match else None,
            "gross": _sum_items(earnings),
            "deductions_total": _sum_items(deductions),
        })

    error_count = sum(1 for r in rows_preview if r["status"] == "error")
    new_count = sum(1 for r in rows_preview if r["status"] == "new")
    update_count = sum(1 for r in rows_preview if r["status"] == "update")

    return {
        "rows": rows_preview,
        "total_rows": len(rows_preview),
        "new_count": new_count,
        "update_count": update_count,
        "error_count": error_count,
    }


@router.post("/employees/import")
async def import_employees(
    company_key: str = Form(...),
    file: UploadFile = File(...),
    skip_updates: bool = Form(False),
    current_user: User = Depends(check_module_permission("salary_slips", "create")),
):
    """Commits the bulk import: inserts new employees and — unless
    skip_updates is set — updates ones matched by Employee Code (or Name)."""
    company = await _resolve_company(company_key)

    contents = await file.read()
    df = _read_employee_file(contents, file.filename or "upload.xlsx")
    if df.empty:
        raise HTTPException(400, "The file has no data rows")

    df.columns = [str(c).strip() for c in df.columns]
    col_map = _build_import_column_map(list(df.columns))
    if "Name*" not in col_map:
        raise HTTPException(400, "Couldn't find a 'Name' column. Please use the downloaded template.")

    existing = await db.salary_employees.find(
        {"company_key": company_key}, {"_id": 0},
    ).to_list(2000)
    by_code = {e["employee_code"].strip().lower(): e for e in existing if e.get("employee_code")}
    by_name = {e["name"].strip().lower(): e for e in existing if e.get("name")}

    created, updated, skipped = [], [], []

    for idx, row in df.iterrows():
        fields, earnings, deductions = _row_to_employee_fields(row, col_map)
        name = fields.get("name", "").strip()
        if not name:
            skipped.append({"row": int(idx) + 2, "reason": "Missing name"})
            continue

        code_key = (fields.get("employee_code") or "").strip().lower()
        match = by_code.get(code_key) if code_key else None
        if not match:
            match = by_name.get(name.lower())

        if match:
            if skip_updates:
                skipped.append({"row": int(idx) + 2, "name": name, "reason": "Already exists — updates skipped"})
                continue
            # Blank cells in the sheet shouldn't blank out fields already saved
            # on the employee record — only overwrite with non-empty values.
            updates = {k: v for k, v in fields.items() if v not in ("", None)}
            updates["default_earnings"] = earnings
            updates["default_deductions"] = deductions
            updates["updated_at"] = _now()
            await db.salary_employees.update_one({"id": match["id"]}, {"$set": updates})
            updated.append({"row": int(idx) + 2, "id": match["id"], "name": name})
        else:
            doc = {
                "id": _new_id(),
                **fields,
                "company_key": company_key,
                "company_name": company["company_name"],
                "default_earnings": earnings,
                "default_deductions": deductions,
                "created_by": current_user.id,
                "created_at": _now(),
                "updated_at": _now(),
            }
            await db.salary_employees.insert_one({**doc, "_id": doc["id"]})
            created.append({"row": int(idx) + 2, "id": doc["id"], "name": name})
            # Guard against duplicate rows within the same file matching each other
            by_name[name.lower()] = doc
            if code_key:
                by_code[code_key] = doc

    return {
        "created_count": len(created),
        "updated_count": len(updated),
        "skipped_count": len(skipped),
        "created": created,
        "updated": updated,
        "skipped": skipped,
    }


# ═════════════════════════════════════════════════════════════════════════
# SALARY SLIPS — generate / list / update / delete / duplicate
# ═════════════════════════════════════════════════════════════════════════

def _compute_slip_doc(
    *, employee: dict, company: dict, month: int, year: int, pay_date: Optional[str],
    total_days: float, paid_days: float, lop_days: float,
    earnings: List[Dict[str, Any]], deductions: List[Dict[str, Any]],
    template: str, notes: Optional[str], status: str, created_by: str, slip_no: str,
) -> Dict[str, Any]:
    gross = _sum_items(earnings)
    total_ded = _sum_items(deductions)
    net_pay = round(gross - total_ded, 2)
    return {
        "id":               _new_id(),
        "slip_no":          slip_no,
        "slip_month":       month,
        "slip_year":        year,
        "period_label":     f"{MONTH_NAMES[month]} {year}",
        "pay_date":         pay_date,
        "total_days":       total_days,
        "paid_days":        paid_days,
        "lop_days":         lop_days,
        # employee snapshot
        "employee_id":      employee["id"],
        "employee_name":    employee.get("name", ""),
        "employee_code":    employee.get("employee_code") or "",
        "designation":      employee.get("designation") or "",
        "department":       employee.get("department") or "",
        "date_of_joining":  employee.get("date_of_joining") or "",
        "pan":              employee.get("pan") or "",
        "uan":              employee.get("uan") or "",
        "pf_number":        employee.get("pf_number") or "",
        "esic_number":      employee.get("esic_number") or "",
        "bank_name":        employee.get("bank_name") or "",
        "bank_account":     employee.get("bank_account") or "",
        "ifsc":             employee.get("ifsc") or "",
        # company snapshot
        "company_key":      company["company_key"],
        "company_name":     company["company_name"],
        "company_address":  company.get("company_address") or "",
        "company_city":     company.get("company_city") or "",
        "company_state":    company.get("company_state") or "",
        "company_gstin":    company.get("company_gstin") or "",
        "company_pan":      company.get("company_pan") or "",
        # money
        "earnings":         earnings,
        "deductions":       deductions,
        "gross_earnings":   gross,
        "total_deductions": total_ded,
        "net_pay":          net_pay,
        "net_pay_words":    amount_in_words(net_pay),
        "template":         template if template in ("modern", "classic") else "modern",
        "notes":            notes or "",
        "status":           status if status in ("draft", "final") else "final",
        "created_by":       created_by,
        "created_at":       _now(),
        "updated_at":       _now(),
    }


@router.post("/generate")
async def generate_slip(
    body: SlipGenerate,
    current_user: User = Depends(check_module_permission("salary_slips", "create")),
):
    employee = await db.salary_employees.find_one({"id": body.employee_id}, {"_id": 0})
    if not employee:
        raise HTTPException(404, "Employee not found")

    earnings = _line_items(body.earnings) if body.earnings is not None else \
        _line_items([SalaryLineItem(**it) for it in employee.get("default_earnings", [])])
    deductions = _line_items(body.deductions) if body.deductions is not None else \
        _line_items([SalaryLineItem(**it) for it in employee.get("default_deductions", [])])

    if not earnings:
        raise HTTPException(400, "Add at least one earnings component before generating a payslip")

    company = await _resolve_company(employee["company_key"])
    slip_no = await _next_slip_no(body.year)
    doc = _compute_slip_doc(
        employee=employee, company=company, month=body.month, year=body.year,
        pay_date=body.pay_date, total_days=body.total_days, paid_days=body.paid_days,
        lop_days=body.lop_days, earnings=earnings, deductions=deductions,
        template=body.template, notes=body.notes, status=body.status,
        created_by=current_user.id, slip_no=slip_no,
    )
    await db.salary_slips.insert_one({**doc, "_id": doc["id"]})
    doc.pop("_id", None)
    return doc


@router.post("/bulk-generate")
async def bulk_generate_slips(
    body: BulkGenerateBody,
    current_user: User = Depends(check_module_permission("salary_slips", "create")),
):
    # Accept either a single month/year (back-compat) or an explicit list of
    # periods — e.g. generating April, May and June payslips in one request.
    periods: List[SlipPeriod] = body.periods or (
        [SlipPeriod(month=body.month, year=body.year)] if body.month and body.year else []
    )
    if not periods:
        raise HTTPException(400, "Provide either month & year, or a list of periods")

    company = await _resolve_company(body.company_key)

    query: Dict[str, Any] = {"company_key": body.company_key}
    if body.employee_ids:
        query["id"] = {"$in": body.employee_ids}
    else:
        query["status"] = "active"
    employees = await db.salary_employees.find(query, {"_id": 0}).to_list(2000)

    # Pre-fetch existing slips for these employees across the requested years
    # so we can skip employee+period combos that already have a payslip,
    # rather than silently creating duplicates when a range overlaps months
    # that were already generated.
    emp_ids = [e["id"] for e in employees]
    years = sorted({p.year for p in periods})
    existing = await db.salary_slips.find(
        {"employee_id": {"$in": emp_ids}, "year": {"$in": years}},
        {"_id": 0, "employee_id": 1, "month": 1, "year": 1},
    ).to_list(20000) if emp_ids else []
    existing_keys = {(e["employee_id"], e["month"], e["year"]) for e in existing}

    generated: List[Dict[str, Any]] = []
    skipped: List[Dict[str, str]] = []

    for period in periods:
        for employee in employees:
            key = (employee["id"], period.month, period.year)
            if key in existing_keys:
                skipped.append({"employee_id": employee["id"], "name": employee.get("name", ""),
                                 "reason": f"Payslip already exists for {MONTH_NAMES[period.month]} {period.year}"})
                continue
            earnings = _line_items([SalaryLineItem(**it) for it in employee.get("default_earnings", [])])
            deductions = _line_items([SalaryLineItem(**it) for it in employee.get("default_deductions", [])])
            if not earnings:
                skipped.append({"employee_id": employee["id"], "name": employee.get("name", ""),
                                 "reason": "No default salary structure configured"})
                continue
            slip_no = await _next_slip_no(period.year)
            doc = _compute_slip_doc(
                employee=employee, company=company, month=period.month, year=period.year,
                pay_date=body.pay_date, total_days=body.total_days, paid_days=body.paid_days,
                lop_days=body.lop_days, earnings=earnings, deductions=deductions,
                template=body.template, notes=None, status=body.status,
                created_by=current_user.id, slip_no=slip_no,
            )
            await db.salary_slips.insert_one({**doc, "_id": doc["id"]})
            doc.pop("_id", None)
            generated.append(doc)
            existing_keys.add(key)  # guard against duplicate periods within the same request

    return {
        "generated_count": len(generated),
        "skipped_count":   len(skipped),
        "periods_count":   len(periods),
        "generated":       generated,
        "skipped":         skipped,
    }


@router.get("")
async def list_slips(
    company_key: Optional[str] = Query(None),
    employee_id: Optional[str] = Query(None),
    month: Optional[int] = Query(None),
    year: Optional[int] = Query(None),
    status: Optional[str] = Query(None),
    search: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200),
    current_user: User = Depends(check_module_permission("salary_slips", "view")),
):
    query: Dict[str, Any] = {}
    if company_key: query["company_key"] = company_key
    if employee_id: query["employee_id"] = employee_id
    if month: query["slip_month"] = month
    if year: query["slip_year"] = year
    if status and status != "all": query["status"] = status
    if search and search.strip():
        query["$or"] = [
            {"employee_name": {"$regex": search.strip(), "$options": "i"}},
            {"company_name":  {"$regex": search.strip(), "$options": "i"}},
            {"slip_no":       {"$regex": search.strip(), "$options": "i"}},
        ]

    total = await db.salary_slips.count_documents(query)
    skip = (page - 1) * page_size
    items = (
        await db.salary_slips.find(query, {"_id": 0})
        .sort("created_at", -1)
        .skip(skip)
        .limit(page_size)
        .to_list(page_size)
    )
    return {"items": items, "total": total, "page": page, "page_size": page_size}


@router.get("/dashboard-summary")
async def dashboard_summary(
    current_user: User = Depends(check_module_permission("salary_slips", "view")),
):
    now = datetime.now(timezone.utc)
    total_employees = await db.salary_employees.count_documents({"status": "active"})
    total_slips = await db.salary_slips.count_documents({})
    slips_this_month = await db.salary_slips.count_documents(
        {"slip_month": now.month, "slip_year": now.year}
    )
    manual_companies = await db.salary_manual_companies.count_documents({})
    # Avoid db.distinct() — not implemented by the in-memory mock DB used in
    # local/dev mode (no MONGO_URL set), so compute distinct keys in Python.
    emp_docs = await db.salary_employees.find({}, {"_id": 0, "company_key": 1}).to_list(5000)
    distinct_companies = len({d.get("company_key") for d in emp_docs if d.get("company_key")})
    return {
        "total_employees":   total_employees,
        "total_slips":       total_slips,
        "slips_this_month":  slips_this_month,
        "manual_companies":  manual_companies,
        "companies_covered": distinct_companies,
    }


@router.get("/{slip_id}")
async def get_slip(
    slip_id: str,
    current_user: User = Depends(check_module_permission("salary_slips", "view")),
):
    slip = await db.salary_slips.find_one({"id": slip_id}, {"_id": 0})
    if not slip:
        raise HTTPException(404, "Salary slip not found")
    return slip


@router.patch("/{slip_id}")
async def update_slip(
    slip_id: str,
    data: SlipUpdate,
    current_user: User = Depends(check_module_permission("salary_slips", "edit")),
):
    existing = await db.salary_slips.find_one({"id": slip_id})
    if not existing:
        raise HTTPException(404, "Salary slip not found")

    updates = data.model_dump(exclude_unset=True)
    if "earnings" in updates and updates["earnings"] is not None:
        updates["earnings"] = _line_items(data.earnings)
    if "deductions" in updates and updates["deductions"] is not None:
        updates["deductions"] = _line_items(data.deductions)

    earnings = updates.get("earnings", existing.get("earnings", []))
    deductions = updates.get("deductions", existing.get("deductions", []))
    if "earnings" in updates or "deductions" in updates:
        gross = _sum_items(earnings)
        total_ded = _sum_items(deductions)
        net_pay = round(gross - total_ded, 2)
        updates["gross_earnings"] = gross
        updates["total_deductions"] = total_ded
        updates["net_pay"] = net_pay
        updates["net_pay_words"] = amount_in_words(net_pay)

    updates["updated_at"] = _now()
    await db.salary_slips.update_one({"id": slip_id}, {"$set": updates})
    updated = await db.salary_slips.find_one({"id": slip_id}, {"_id": 0})
    return updated


@router.delete("/{slip_id}")
async def delete_slip(
    slip_id: str,
    current_user: User = Depends(check_module_permission("salary_slips", "delete")),
):
    existing = await db.salary_slips.find_one({"id": slip_id})
    if not existing:
        raise HTTPException(404, "Salary slip not found")
    await db.salary_slips.delete_one({"id": slip_id})
    return {"deleted": True}


@router.post("/{slip_id}/duplicate")
async def duplicate_slip(
    slip_id: str,
    body: DuplicateBody,
    current_user: User = Depends(check_module_permission("salary_slips", "create")),
):
    """Clone an existing slip's earnings/deductions/employee info into a new
    period — the common case of "same structure, next month"."""
    existing = await db.salary_slips.find_one({"id": slip_id}, {"_id": 0})
    if not existing:
        raise HTTPException(404, "Salary slip not found")

    slip_no = await _next_slip_no(body.year)
    new_doc = {
        **existing,
        "id": _new_id(),
        "slip_no": slip_no,
        "slip_month": body.month,
        "slip_year": body.year,
        "period_label": f"{MONTH_NAMES[body.month]} {body.year}",
        "pay_date": None,
        "status": "draft",
        "created_by": current_user.id,
        "created_at": _now(),
        "updated_at": _now(),
    }
    await db.salary_slips.insert_one({**new_doc, "_id": new_doc["id"]})
    new_doc.pop("_id", None)
    return new_doc


# ═════════════════════════════════════════════════════════════════════════
# PDF RENDERING
# ═════════════════════════════════════════════════════════════════════════

NAVY   = colors.HexColor("#0D3B66")
BLUE   = colors.HexColor("#1F6FB2")
LBLUE  = colors.HexColor("#EBF4FB")
BORDER = colors.HexColor("#D0DCE8")
TEXT   = colors.HexColor("#1A1A2E")
MUTED  = colors.HexColor("#5A6A7A")
SUBTLE = colors.HexColor("#F5F7FA")
WHITE  = colors.white
EMERALD    = colors.HexColor("#166534")
EMERALD_BG = colors.HexColor("#DCFCE7")
BLACK  = colors.HexColor("#111111")

PAGE_W, PAGE_H = A4
L_MARGIN = R_MARGIN = 16 * mm
CONTENT_W = PAGE_W - L_MARGIN - R_MARGIN


def _styles():
    return {
        "eyebrow": ParagraphStyle("eyebrow", fontName="Helvetica-Bold", fontSize=8.5,
                                   textColor=BLUE, leading=11, spaceAfter=1),
        "company": ParagraphStyle("company", fontName="Helvetica-Bold", fontSize=16,
                                   textColor=NAVY, leading=19),
        "addr":    ParagraphStyle("addr", fontName="Helvetica", fontSize=8.5,
                                   textColor=MUTED, leading=11),
        "h2":      ParagraphStyle("h2", fontName="Helvetica-Bold", fontSize=10,
                                   textColor=NAVY, leading=13, spaceAfter=3),
        "label":   ParagraphStyle("label", fontName="Helvetica", fontSize=8,
                                   textColor=MUTED, leading=10),
        "value":   ParagraphStyle("value", fontName="Helvetica-Bold", fontSize=8.5,
                                   textColor=TEXT, leading=11),
        "netlabel": ParagraphStyle("netlabel", fontName="Helvetica-Bold", fontSize=8.5,
                                    textColor=EMERALD, leading=11),
        "netvalue": ParagraphStyle("netvalue", fontName="Helvetica-Bold", fontSize=17,
                                    textColor=EMERALD, leading=20),
        "words":   ParagraphStyle("words", fontName="Helvetica-Oblique", fontSize=8.5,
                                   textColor=TEXT, leading=11),
        "footer":  ParagraphStyle("footer", fontName="Helvetica", fontSize=7.5,
                                   textColor=MUTED, leading=10, alignment=TA_CENTER),
        "classicTitle": ParagraphStyle("classicTitle", fontName="Helvetica-Bold", fontSize=18,
                                        textColor=WHITE, leading=22),
        "sigline": ParagraphStyle("sigline", fontName="Helvetica", fontSize=8,
                                   textColor=TEXT, leading=11),
    }


def _rule(color=BORDER, thickness=0.6):
    return HRFlowable(width="100%", thickness=thickness, color=color, spaceBefore=4, spaceAfter=6)


def _render_modern(slip: dict, st: dict) -> List[Any]:
    story: List[Any] = []

    # ── Header: company block (left) + payslip period + net pay card (right)
    header_tbl = Table(
        [[
            Paragraph(
                f"{slip['company_name']}<br/>"
                f"{', '.join(filter(None, [slip.get('company_address',''), slip.get('company_city',''), slip.get('company_state','')]))}",
                st["company"],
            ),
            Table(
                [[Paragraph("EMPLOYEE NET PAY", st["netlabel"])],
                 [Paragraph(inr(slip["net_pay"]), st["netvalue"])]],
                colWidths=[62 * mm],
                style=TableStyle([
                    ("BACKGROUND", (0, 0), (-1, -1), EMERALD_BG),
                    ("BOX", (0, 0), (-1, -1), 0.6, EMERALD),
                    ("LEFTPADDING", (0, 0), (-1, -1), 10),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 10),
                    ("TOPPADDING", (0, 0), (-1, 0), 8),
                    ("BOTTOMPADDING", (0, -1), (-1, -1), 8),
                    ("TOPPADDING", (0, 1), (-1, 1), 0),
                ]),
            ),
        ]],
        colWidths=[CONTENT_W - 62 * mm, 62 * mm],
    )
    header_tbl.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ALIGN", (1, 0), (1, 0), "RIGHT"),
    ]))
    story.append(header_tbl)
    story.append(Spacer(1, 4))
    story.append(Paragraph(f"PAYSLIP FOR {slip['period_label'].upper()}  ·  Ref: {slip['slip_no']}", st["eyebrow"]))
    story.append(_rule(BLUE, 1.1))

    # ── Employee summary grid ───────────────────────────────────────────
    def cell(label, value):
        return [Paragraph(label, st["label"]), Paragraph(value or "—", st["value"])]

    summary_rows = [
        [*cell("Employee Name", slip["employee_name"]), *cell("Employee Code", slip.get("employee_code"))],
        [*cell("Designation", slip.get("designation")), *cell("Department", slip.get("department"))],
        [*cell("Date of Joining", slip.get("date_of_joining")), *cell("PAN", slip.get("pan"))],
        [*cell("UAN", slip.get("uan")), *cell("Bank A/c No.", slip.get("bank_account"))],
        [*cell("Paid Days", f"{slip.get('paid_days')} / {slip.get('total_days')}"),
         *cell("LOP Days", str(slip.get("lop_days", 0)))],
    ]
    summary_tbl = Table(summary_rows, colWidths=[28 * mm, (CONTENT_W / 2) - 28 * mm] * 2)
    summary_tbl.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("LINEBELOW", (0, 0), (-1, -2), 0.4, BORDER),
        ("BACKGROUND", (0, 0), (-1, -1), SUBTLE),
    ]))
    story.append(summary_tbl)
    story.append(Spacer(1, 10))

    # ── Earnings / Deductions two-column table ──────────────────────────
    earnings = slip["earnings"] or []
    deductions = slip["deductions"] or []
    rows = [["EARNINGS", "AMOUNT", "DEDUCTIONS", "AMOUNT"]]
    for i in range(max(len(earnings), len(deductions))):
        e = earnings[i] if i < len(earnings) else None
        d = deductions[i] if i < len(deductions) else None
        rows.append([
            e["label"] if e else "", inr(e["amount"]) if e else "",
            d["label"] if d else "", inr(d["amount"]) if d else "",
        ])
    rows.append(["Gross Earnings", inr(slip["gross_earnings"]), "Total Deductions", inr(slip["total_deductions"])])

    ed_tbl = Table(rows, colWidths=[CONTENT_W * 0.32, CONTENT_W * 0.18, CONTENT_W * 0.32, CONTENT_W * 0.18])
    ed_tbl.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8.5),
        ("BACKGROUND", (0, 0), (-1, 0), NAVY),
        ("TEXTCOLOR", (0, 0), (-1, 0), WHITE),
        ("BACKGROUND", (0, -1), (-1, -1), LBLUE),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("ALIGN", (3, 0), (3, -1), "RIGHT"),
        ("GRID", (0, 0), (-1, -1), 0.4, BORDER),
        ("TOPPADDING", (0, 0), (-1, -1), 4.5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4.5),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
    ]))
    story.append(ed_tbl)
    story.append(Spacer(1, 8))
    story.append(Paragraph(f"Amount in words: {slip['net_pay_words']}", st["words"]))
    story.append(Spacer(1, 14))
    story.append(_rule())
    story.append(Paragraph(
        "This is a computer-generated payslip prepared on behalf of the employer and does not require a signature.",
        st["footer"],
    ))
    if slip.get("notes"):
        story.append(Spacer(1, 4))
        story.append(Paragraph(f"Note: {slip['notes']}", st["footer"]))
    return story


def _render_classic(slip: dict, st: dict) -> List[Any]:
    story: List[Any] = []

    title_tbl = Table(
        [[Paragraph("SALARY SLIP", st["classicTitle"]),
          Paragraph(f"Period<br/><b>{slip['period_label']}</b>", ParagraphStyle(
              "periodR", fontName="Helvetica", fontSize=9, textColor=WHITE, alignment=TA_RIGHT, leading=12))]],
        colWidths=[CONTENT_W * 0.7, CONTENT_W * 0.3],
    )
    title_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), BLACK),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 10),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
        ("LEFTPADDING", (0, 0), (-1, -1), 10),
        ("RIGHTPADDING", (0, 0), (-1, -1), 10),
    ]))
    story.append(title_tbl)
    story.append(Spacer(1, 8))
    story.append(Paragraph(
        f"{slip['company_name']} — "
        f"{', '.join(filter(None, [slip.get('company_address',''), slip.get('company_city','')]))}",
        st["addr"],
    ))
    story.append(Spacer(1, 6))

    info_rows = [
        ["Employee Name:", slip["employee_name"] or "—", "Designation:", slip.get("designation") or "—"],
        ["Employee Code:", slip.get("employee_code") or "—", "Month / Year:", slip["period_label"]],
        ["Bank Name:", slip.get("bank_name") or "—", "Cheque / A/c #:", slip.get("bank_account") or "—"],
    ]
    info_tbl = Table(info_rows, colWidths=[26 * mm, (CONTENT_W / 2) - 26 * mm] * 2)
    info_tbl.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTNAME", (2, 0), (2, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8.5),
        ("BOX", (0, 0), (-1, -1), 0.6, BLACK),
        ("INNERGRID", (0, 0), (-1, -1), 0.4, BORDER),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
    ]))
    story.append(info_tbl)
    story.append(Spacer(1, 0))

    earnings = slip["earnings"] or []
    deductions = slip["deductions"] or []
    rows = [["EARNINGS", "AMOUNT", "DEDUCTIONS", "AMOUNT"]]
    for i in range(max(len(earnings), len(deductions))):
        e = earnings[i] if i < len(earnings) else None
        d = deductions[i] if i < len(deductions) else None
        rows.append([
            e["label"] if e else "", inr(e["amount"]) if e else "",
            d["label"] if d else "", inr(d["amount"]) if d else "",
        ])
    rows.append(["Total Addition", inr(slip["gross_earnings"]), "Total Deduction", inr(slip["total_deductions"])])
    rows.append(["NET SALARY", inr(slip["net_pay"]), "", ""])

    ed_tbl = Table(rows, colWidths=[CONTENT_W * 0.32, CONTENT_W * 0.18, CONTENT_W * 0.32, CONTENT_W * 0.18])
    style = [
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2B2B2B")),
        ("TEXTCOLOR", (0, 0), (-1, 0), WHITE),
        ("FONTSIZE", (0, 0), (-1, -1), 8.5),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("ALIGN", (3, 0), (3, -1), "RIGHT"),
        ("GRID", (0, 0), (-1, -3), 0.4, BORDER),
        ("BOX", (0, 0), (-1, -1), 0.6, BLACK),
        ("TOPPADDING", (0, 0), (-1, -1), 4.5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4.5),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("BACKGROUND", (0, -2), (-1, -2), colors.HexColor("#E5E5E5")),
        ("FONTNAME", (0, -2), (-1, -2), "Helvetica-Bold"),
        ("SPAN", (2, -1), (3, -1)),
        ("BACKGROUND", (0, -1), (-1, -1), BLACK),
        ("TEXTCOLOR", (0, -1), (-1, -1), WHITE),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, -1), (-1, -1), 10),
        ("ALIGN", (0, -1), (0, -1), "LEFT"),
        ("ALIGN", (2, -1), (2, -1), "RIGHT"),
    ]
    ed_tbl.setStyle(TableStyle(style))
    story.append(ed_tbl)
    story.append(Spacer(1, 8))
    story.append(Paragraph(f"Amount in words: {slip['net_pay_words']}", st["words"]))
    story.append(Spacer(1, 26))

    sig_tbl = Table(
        [["Employee Signature", "Authorised Signatory"]],
        colWidths=[CONTENT_W / 2, CONTENT_W / 2],
    )
    sig_tbl.setStyle(TableStyle([
        ("LINEABOVE", (0, 0), (0, 0), 0.6, BLACK),
        ("LINEABOVE", (1, 0), (1, 0), 0.6, BLACK),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ALIGN", (1, 0), (1, 0), "RIGHT"),
    ]))
    story.append(sig_tbl)
    return story


def _render_slip_story(slip: dict, st: dict) -> List[Any]:
    return _render_modern(slip, st) if slip.get("template") != "classic" else _render_classic(slip, st)


def build_slip_pdf(slip: dict) -> bytes:
    buf = BytesIO()
    st = _styles()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=L_MARGIN, rightMargin=R_MARGIN,
        topMargin=14 * mm, bottomMargin=16 * mm,
        title=f"Payslip — {slip.get('employee_name','')} — {slip.get('period_label','')}",
    )
    doc.build(_render_slip_story(slip, st))
    return buf.getvalue()


def build_combined_slip_pdf(slips: List[dict]) -> bytes:
    buf = BytesIO()
    st = _styles()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=L_MARGIN, rightMargin=R_MARGIN,
        topMargin=14 * mm, bottomMargin=16 * mm,
        title="Payslips",
    )
    story: List[Any] = []
    for i, slip in enumerate(slips):
        if i > 0:
            story.append(PageBreak())
        story.extend(_render_slip_story(slip, st))
    doc.build(story)
    return buf.getvalue()


def _safe_filename(*parts: str) -> str:
    name = "_".join(p for p in parts if p)
    for ch in ("/", "\\", '"', "'"):
        name = name.replace(ch, "-")
    return name.replace(" ", "_")


@router.get("/{slip_id}/pdf")
async def download_slip_pdf(
    slip_id: str,
    current_user: User = Depends(check_module_permission("salary_slips", "view")),
):
    slip = await db.salary_slips.find_one({"id": slip_id}, {"_id": 0})
    if not slip:
        raise HTTPException(404, "Salary slip not found")
    pdf_bytes = build_slip_pdf(slip)
    filename = _safe_filename(slip.get("employee_name", "payslip"), slip.get("period_label", "")) + ".pdf"
    return StreamingResponse(
        iter([pdf_bytes]),
        media_type="application/pdf",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Content-Length": str(len(pdf_bytes)),
            "Cache-Control": "no-cache",
        },
    )


@router.post("/bulk-pdf")
async def download_bulk_pdf(
    body: BulkPdfBody,
    current_user: User = Depends(check_module_permission("salary_slips", "view")),
):
    if not body.slip_ids:
        raise HTTPException(400, "slip_ids must not be empty")
    slips = await db.salary_slips.find({"id": {"$in": body.slip_ids}}, {"_id": 0}).to_list(len(body.slip_ids))
    if not slips:
        raise HTTPException(404, "No matching salary slips found")
    # Preserve requested order
    order = {sid: i for i, sid in enumerate(body.slip_ids)}
    slips.sort(key=lambda s: order.get(s["id"], 0))

    pdf_bytes = build_combined_slip_pdf(slips)
    label = slips[0].get("period_label", "payslips")
    filename = _safe_filename("Payslips", label) + ".pdf"
    return StreamingResponse(
        iter([pdf_bytes]),
        media_type="application/pdf",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Content-Length": str(len(pdf_bytes)),
            "Cache-Control": "no-cache",
        },
    )


# ═════════════════════════════════════════════════════════════════════════
# MISC
# ═════════════════════════════════════════════════════════════════════════

@router.get("/meta/presets")
async def get_presets(
    current_user: User = Depends(check_module_permission("salary_slips", "view")),
):
    """Suggested earnings/deductions labels for the frontend's quick-add buttons."""
    return {"earning_presets": EARNING_PRESETS, "deduction_presets": DEDUCTION_PRESETS}


# ═════════════════════════════════════════════════════════════════════════
# INDEXES (call once at startup, idempotent)
# ═════════════════════════════════════════════════════════════════════════

async def create_salary_slip_indexes():
    try:
        await db.salary_manual_companies.create_index("id", unique=True, background=True)
        await db.salary_employees.create_index("id", unique=True, background=True)
        await db.salary_employees.create_index("company_key", background=True)
        await db.salary_employees.create_index("status", background=True)
        await db.salary_slips.create_index("id", unique=True, background=True)
        await db.salary_slips.create_index("employee_id", background=True)
        await db.salary_slips.create_index("company_key", background=True)
        await db.salary_slips.create_index([("slip_year", 1), ("slip_month", 1)], background=True)
        await db.salary_slips.create_index("created_at", background=True)
        logger.info("Salary slip indexes ensured")
    except Exception as exc:
        logger.warning("Salary slip index creation: %s", exc)
