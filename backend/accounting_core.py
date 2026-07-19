"""
Accounting Core — Chart of Accounts, Journal Entries, General Ledger,
Trial Balance, Profit & Loss, and Balance Sheet.

This is deliberately a standard double-entry engine: every Journal Entry is
a set of lines that must debit-equal-credit before it's allowed to save.
Purchase invoices, Sale invoices, and Bank transactions all post into this
same ledger through `post_journal_entry(...)`, so the reports below reflect
the whole business, not just manually-typed entries.

Scope: one ledger per `company_id` (the firm's own "Sales company/book"
entity already used elsewhere in the app — see /companies/list). Pass
company_id="" to use a single default book if the firm only operates one.
"""

import re
import uuid
import asyncio
import logging
from datetime import datetime, date, timezone
from typing import Optional, List

from fastapi import APIRouter, Depends, HTTPException, Query
from pydantic import BaseModel, Field

from backend.dependencies import db, get_current_user
from backend.models import User

router = APIRouter(tags=["Accounting"])


def _name_match_query(name: str) -> dict:
    """Case/whitespace-insensitive exact-name match with support for corporate suffix variations
    like Pvt Ltd, Private Limited, Ltd, and Limited. This prevents spelling variants of corporate
    suffixes from breaking party matching on invoice vs payment ledger records."""
    name = (name or "").strip()
    name = name.replace(".", " ").replace(",", " ")
    tokens = name.split()
    
    regex_parts = []
    i = 0
    while i < len(tokens):
        tok = tokens[i].upper()
        if tok == "PVT" and i + 1 < len(tokens) and tokens[i+1].upper() == "LTD":
            regex_parts.append(r"(PVT\s*LTD|PVT\s*LIMITED|PRIVATE\s*LIMITED|PRIVATE\s*LTD|LTD|LIMITED)?")
            i += 2
        elif tok == "PRIVATE" and i + 1 < len(tokens) and tokens[i+1].upper() == "LIMITED":
            regex_parts.append(r"(PVT\s*LTD|PVT\s*LIMITED|PRIVATE\s*LIMITED|PRIVATE\s*LTD|LTD|LIMITED)?")
            i += 2
        elif tok in ("PVT", "PRIVATE"):
            regex_parts.append(r"(PVT|PRIVATE)?")
            i += 1
        elif tok in ("LTD", "LIMITED"):
            regex_parts.append(r"(LTD|LIMITED)?")
            i += 1
        else:
            regex_parts.append(re.escape(tokens[i]))
            i += 1
            
    regex_parts = [p for p in regex_parts if p]
    normalized = r"\s*".join(regex_parts)
    return {"$regex": f"^{normalized}$", "$options": "i"}


def _perm_view_coa(user: User) -> bool:
    if user.role == "admin":
        return True
    perms = user.permissions if isinstance(user.permissions, dict) else (user.permissions.model_dump() if user.permissions else {})
    return bool(perms.get("can_view_chart_of_accounts"))


def _perm_manage_coa(user: User) -> bool:
    if user.role == "admin":
        return True
    perms = user.permissions if isinstance(user.permissions, dict) else (user.permissions.model_dump() if user.permissions else {})
    return bool(perms.get("can_manage_chart_of_accounts"))


def _perm_view_journal(user: User) -> bool:
    if user.role == "admin":
        return True
    perms = user.permissions if isinstance(user.permissions, dict) else (user.permissions.model_dump() if user.permissions else {})
    return bool(perms.get("can_view_journal_entries"))


def _perm_post_journal(user: User) -> bool:
    if user.role == "admin":
        return True
    perms = user.permissions if isinstance(user.permissions, dict) else (user.permissions.model_dump() if user.permissions else {})
    return bool(perms.get("can_post_journal_entries"))


def _perm_reports(user: User) -> bool:
    if user.role == "admin":
        return True
    perms = user.permissions if isinstance(user.permissions, dict) else (user.permissions.model_dump() if user.permissions else {})
    return bool(perms.get("can_view_accounting_reports"))


# ── Default Chart of Accounts (standard Indian SME set) ─────────────────────
DEFAULT_ACCOUNTS = [
    # code, name, type (asset/liability/equity/income/expense), sub_type
    ("1000", "Cash in Hand",            "asset",     "current_asset"),
    ("1010", "Bank Accounts",           "asset",     "current_asset"),
    ("1100", "Accounts Receivable",     "asset",     "current_asset"),
    ("1200", "GST Input Credit",        "asset",     "current_asset"),
    ("1300", "Fixed Assets",            "asset",     "fixed_asset"),
    ("2000", "Accounts Payable",        "liability", "current_liability"),
    ("2100", "GST Output Payable",      "liability", "current_liability"),
    ("2200", "TDS Payable",             "liability", "current_liability"),
    ("3000", "Owner's Capital / Equity","equity",    "equity"),
    ("3100", "Retained Earnings",       "equity",    "equity"),
    ("4000", "Sales / Fee Income",      "income",    "operating_income"),
    ("4100", "Other Income",            "income",    "other_income"),
    ("5000", "Purchases",               "expense",   "cost_of_service"),
    ("5100", "Salaries & Wages",         "expense",   "operating_expense"),
    ("5200", "Rent Expense",             "expense",   "operating_expense"),
    ("5250", "Software & Cloud Expenses","expense",   "operating_expense"),
    ("5300", "Office & Admin Expenses", "expense",   "operating_expense"),
    ("5400", "Bank Charges",             "expense",   "operating_expense"),
    ("5500", "Shipping & Freight",       "expense",   "operating_expense"),
    ("5600", "Travel & Conveyance",      "expense",   "operating_expense"),
    ("5700", "Foreign Exchange Loss / Gain", "expense", "operating_expense"),
    ("5900", "Round Off",               "expense",   "operating_expense"),
]


_ensured_coa_companies: set = set()


async def ensure_default_chart_of_accounts(company_id: str, created_by: str):
    """Insert any DEFAULT_ACCOUNTS codes this company doesn't have yet.
    Deliberately per-code (not 'skip entirely if any account exists') so that
    when new system accounts are added to DEFAULT_ACCOUNTS later (e.g. a new
    expense category), companies that were seeded before that change still
    pick it up automatically the next time their Chart of Accounts loads.

    This gets called for every account lookup during invoice/payment sync
    (2-3 times per invoice), so once a company's default accounts have been
    confirmed present in this process, skip re-scanning the whole
    chart_of_accounts collection on every subsequent call — that repeated
    full scan was the main remaining cost behind slow report loads even
    after debouncing the outer reconcile pass. A fresh deploy/restart clears
    this cache, so newly added DEFAULT_ACCOUNTS codes still get seeded."""
    if company_id in _ensured_coa_companies:
        return
    existing_codes = set(
        r["code"] for r in await db.chart_of_accounts.find(
            {"company_id": company_id}, {"_id": 0, "code": 1}
        ).to_list(2000)
    )
    missing = [(code, name, typ, sub) for code, name, typ, sub in DEFAULT_ACCOUNTS if code not in existing_codes]
    if missing:
        now = datetime.now(timezone.utc).isoformat()
        docs = [
            {
                "id": str(uuid.uuid4()), "company_id": company_id, "code": code, "name": name,
                "type": typ, "sub_type": sub, "is_system": True, "is_active": True,
                "created_by": created_by, "created_at": now,
            }
            for code, name, typ, sub in missing
        ]
        await db.chart_of_accounts.insert_many(docs)
    _ensured_coa_companies.add(company_id)


# ── Models ────────────────────────────────────────────────────────────────
class AccountCreate(BaseModel):
    company_id: str = ""
    code: str
    name: str
    type: str  # asset | liability | equity | income | expense
    sub_type: str = ""


class JournalLine(BaseModel):
    account_id: str
    account_name: str = ""
    debit: float = 0.0
    credit: float = 0.0
    memo: str = ""


class JournalEntryCreate(BaseModel):
    company_id: str = ""
    entry_date: str = Field(default_factory=lambda: date.today().isoformat())
    narration: str = ""
    source: str = "manual"          # manual | purchase | sale | bank
    source_id: Optional[str] = None
    lines: List[JournalLine]


# ── Chart of Accounts routes ─────────────────────────────────────────────
@router.get("/chart-of-accounts")
async def list_accounts(company_id: str = Query(""), current_user: User = Depends(get_current_user)):
    if not _perm_view_coa(current_user):
        raise HTTPException(403, "Access denied. Request access from your admin in Permission Governance.")
    await ensure_default_chart_of_accounts(company_id, current_user.id)
    items = await db.chart_of_accounts.find({"company_id": company_id}, {"_id": 0}).sort("code", 1).to_list(2000)
    return items


@router.post("/chart-of-accounts")
async def create_account(payload: AccountCreate, current_user: User = Depends(get_current_user)):
    if not _perm_manage_coa(current_user):
        raise HTTPException(403, "Access denied.")
    if payload.type not in {"asset", "liability", "equity", "income", "expense"}:
        raise HTTPException(400, "type must be one of asset, liability, equity, income, expense.")
    dup = await db.chart_of_accounts.find_one({"company_id": payload.company_id, "code": payload.code})
    if dup:
        raise HTTPException(409, f"Account code {payload.code} already exists.")
    doc = {
        "id": str(uuid.uuid4()), "company_id": payload.company_id, "code": payload.code.strip(),
        "name": payload.name.strip(), "type": payload.type, "sub_type": payload.sub_type.strip(),
        "is_system": False, "is_active": True, "created_by": current_user.id,
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.chart_of_accounts.insert_one(doc)
    doc.pop("_id", None)
    return doc


@router.delete("/chart-of-accounts/{account_id}")
async def delete_account(account_id: str, current_user: User = Depends(get_current_user)):
    if not _perm_manage_coa(current_user):
        raise HTTPException(403, "Access denied.")
    acct = await db.chart_of_accounts.find_one({"id": account_id})
    if not acct:
        raise HTTPException(404, "Account not found.")
    if acct.get("is_system"):
        raise HTTPException(400, "Default system accounts can't be deleted (they're needed by auto-posting). Mark inactive instead.")
    used = await db.journal_lines.count_documents({"account_id": account_id})
    if used:
        raise HTTPException(400, "This account has journal entries posted against it and can't be deleted.")
    await db.chart_of_accounts.delete_one({"id": account_id})
    return {"success": True}


# ── Journal Entry posting (shared by manual entries, Purchase, Sale, Bank) ──
async def post_journal_entry(
    company_id: str, entry_date: str, narration: str, lines: List[dict],
    source: str, source_id: Optional[str], created_by: str,
) -> dict:
    """Core double-entry post. Raises ValueError if debits != credits.
    Used both by the manual Journal Entry endpoint and by auto-posting hooks
    from Purchase/Sale/Bank so every recorded transaction ends up in the
    same ledger and reports."""
    total_debit = round(sum(float(l.get("debit") or 0) for l in lines), 2)
    total_credit = round(sum(float(l.get("credit") or 0) for l in lines), 2)
    if abs(total_debit - total_credit) > 0.01:
        raise ValueError(f"Journal entry does not balance: debit {total_debit} != credit {total_credit}")
    if total_debit <= 0:
        raise ValueError("Journal entry has no amount.")

    now = datetime.now(timezone.utc).isoformat()
    entry_id = str(uuid.uuid4())
    entry_doc = {
        "id": entry_id, "company_id": company_id, "entry_date": entry_date,
        "narration": narration, "source": source, "source_id": source_id,
        "total_debit": total_debit, "total_credit": total_credit,
        "created_by": created_by, "created_at": now,
    }
    await db.journal_entries.insert_one(entry_doc)

    line_docs = []
    for l in lines:
        line_docs.append({
            "id": str(uuid.uuid4()), "entry_id": entry_id, "company_id": company_id,
            "entry_date": entry_date, "account_id": l["account_id"],
            "account_name": l.get("account_name", ""), "debit": float(l.get("debit") or 0),
            "credit": float(l.get("credit") or 0), "memo": l.get("memo", ""), "created_at": now,
        })
    await db.journal_lines.insert_many(line_docs)
    entry_doc.pop("_id", None)
    return entry_doc


async def try_auto_post(company_id: str, entry_date: str, narration: str, lines: List[dict],
                         source: str, source_id: Optional[str], created_by: str) -> Optional[dict]:
    """Best-effort wrapper for call sites (Purchase/Sale/Bank) that must
    never fail the parent request just because auto-posting hit an edge
    case (e.g. a missing default account on a brand-new company)."""
    try:
        return await post_journal_entry(company_id, entry_date, narration, lines, source, source_id, created_by)
    except Exception:
        return None


async def get_default_account_id(company_id: str, code: str) -> Optional[str]:
    await ensure_default_chart_of_accounts(company_id, "system")
    acct = await db.chart_of_accounts.find_one({"company_id": company_id, "code": code}, {"_id": 0, "id": 1})
    return acct["id"] if acct else None


async def _all_book_ids() -> List[str]:
    """Every distinct book: each real company plus the "" default book (used
    by manual Journal Entries, which have no company selector). Used for the
    'All Companies' report view so it actually aggregates everything instead
    of only showing the empty-company_id book."""
    ids = {""}
    async for c in db.companies.find({}, {"_id": 0, "id": 1}):
        if c.get("id"):
            ids.add(c["id"])
    return list(ids)


async def _reconcile_one_book(company_id: str):
    from backend.invoicing import reconcile_and_sync_all_sales_and_payments, reconcile_and_sync_all_purchases_and_payments
    await asyncio.gather(
        reconcile_and_sync_all_sales_and_payments(company_id),
        reconcile_and_sync_all_purchases_and_payments(company_id),
    )


async def _reconcile_all_books(book_ids: List[str]):
    await asyncio.gather(*(_reconcile_one_book(cid) for cid in book_ids))


async def _validate_best_effort(company_id: str):
    """Fire-and-forget consistency check (Revenue=Collections+Outstanding,
    TB Debits=Credits, AR=Outstanding, Customer Ledger=AR, Sales
    Ledger=Invoice Revenue, GST+NonGST+Export+Exempt=Revenue). Never
    raises into the caller - a validator bug must not break report
    rendering; mismatches are logged/persisted by the engine itself."""
    try:
        from backend.accounting_ai.reconciliation_validator import run_validation_engine
        await run_validation_engine(company_id, auto_fix=True)
    except Exception:
        logging.getLogger("accounting_core").exception("validation engine failed")


@router.get("/reports/validation-engine")
async def validation_engine_report(company_id: str = Query(""), current_user: User = Depends(get_current_user)):
    """On-demand reconciliation health check across Trial Balance, Party/
    Customer Ledger, and GST report totals vs. the Invoice module's own
    Revenue/Collections/Outstanding figures. Auto-heals by rerunning the
    invoice->journal sync once before reporting a real mismatch."""
    if not _perm_reports(current_user):
        raise HTTPException(403, "Access denied. Request access from your admin in Permission Governance.")
    from backend.accounting_ai.reconciliation_validator import run_validation_engine, run_validation_engine_all_books
    if company_id:
        return await run_validation_engine(company_id, auto_fix=True)
    return {"books": await run_validation_engine_all_books()}


@router.post("/journal-entries")
async def create_journal_entry(payload: JournalEntryCreate, current_user: User = Depends(get_current_user)):
    if not _perm_post_journal(current_user):
        raise HTTPException(403, "Access denied. Request access from your admin in Permission Governance.")
    try:
        doc = await post_journal_entry(
            payload.company_id, payload.entry_date, payload.narration.strip(),
            [l.model_dump() for l in payload.lines], "manual", None, current_user.id,
        )
    except ValueError as e:
        raise HTTPException(400, str(e))
    return doc


@router.get("/journal-entries")
async def list_journal_entries(
    company_id: str = Query(""), date_from: Optional[str] = Query(None), date_to: Optional[str] = Query(None),
    source: Optional[str] = Query(None), page: int = Query(1, ge=1), page_size: int = Query(50, ge=1, le=500),
    current_user: User = Depends(get_current_user),
):
    if not _perm_view_journal(current_user):
        raise HTTPException(403, "Access denied.")

    all_companies = not company_id
    q: dict = {} if all_companies else {"company_id": company_id}
    if date_from or date_to:
        q["entry_date"] = {}
        if date_from:
            q["entry_date"]["$gte"] = date_from
        if date_to:
            q["entry_date"]["$lte"] = date_to
    if source:
        q["source"] = source

    total = await db.journal_entries.count_documents(q)
    total_pages = max(1, (total + page_size - 1) // page_size)
    page = min(page, total_pages)
    skip = (page - 1) * page_size
    entries = await db.journal_entries.find(q, {"_id": 0}).sort("entry_date", -1).skip(skip).limit(page_size).to_list(page_size)

    ids = [e["id"] for e in entries]
    lines = await db.journal_lines.find({"entry_id": {"$in": ids}}, {"_id": 0}).to_list(10000)
    by_entry: dict = {}
    for l in lines:
        by_entry.setdefault(l["entry_id"], []).append(l)

    # Enrich each entry with customer/vendor name, voucher/invoice number,
    # payment mode, bank/cash account, and reference number — resolved from
    # the underlying invoice/bill/payment that the entry was posted from, so
    # the Journal Entries UI can show "Receipt from Abhinav against Invoice
    # INV-2026-001" instead of just a raw narration string.
    sale_ids = {e["source_id"] for e in entries if e.get("source") == "sale" and e.get("source_id")}
    purchase_ids = {e["source_id"] for e in entries if e.get("source") == "purchase" and e.get("source_id")}
    payment_ids = {e["source_id"] for e in entries if e.get("source") == "payment" and e.get("source_id")}
    purchase_payment_ids = {e["source_id"] for e in entries if e.get("source") == "purchase_payment" and e.get("source_id")}

    invoices_by_id, bills_by_id, payments_by_id, purchase_payments_by_id = {}, {}, {}, {}
    if sale_ids:
        invoices_by_id = {d["id"]: d for d in await db.invoices.find({"id": {"$in": list(sale_ids)}}, {"_id": 0}).to_list(len(sale_ids))}
    if purchase_ids:
        bills_by_id = {d["id"]: d for d in await db.purchase_invoices.find({"id": {"$in": list(purchase_ids)}}, {"_id": 0}).to_list(len(purchase_ids))}
    if payment_ids:
        payments_by_id = {d["id"]: d for d in await db.payments.find({"id": {"$in": list(payment_ids)}}, {"_id": 0}).to_list(len(payment_ids))}
    if purchase_payment_ids:
        purchase_payments_by_id = {d["id"]: d for d in await db.purchase_payments.find({"id": {"$in": list(purchase_payment_ids)}}, {"_id": 0}).to_list(len(purchase_payment_ids))}

    # Receipts/payments often only carry an invoice_id/purchase_invoice_id —
    # resolve those to get the invoice/bill number and party name too.
    linked_inv_ids = {p["invoice_id"] for p in payments_by_id.values() if p.get("invoice_id")}
    linked_inv_by_id = {d["id"]: d for d in await db.invoices.find({"id": {"$in": list(linked_inv_ids)}}, {"_id": 0, "id": 1, "invoice_no": 1, "client_name": 1}).to_list(len(linked_inv_ids))} if linked_inv_ids else {}
    linked_bill_ids = {p["purchase_invoice_id"] for p in purchase_payments_by_id.values() if p.get("purchase_invoice_id")}
    linked_bill_by_id = {d["id"]: d for d in await db.purchase_invoices.find({"id": {"$in": list(linked_bill_ids)}}, {"_id": 0, "id": 1, "invoice_no": 1, "supplier_name": 1}).to_list(len(linked_bill_ids))} if linked_bill_ids else {}

    for e in entries:
        e["lines"] = by_entry.get(e["id"], [])
        src, sid = e.get("source"), e.get("source_id")
        if src == "sale" and sid in invoices_by_id:
            inv = invoices_by_id[sid]
            e["customer_name"] = inv.get("client_name")
            e["invoice_no"] = inv.get("invoice_no")
            e["voucher_no"] = inv.get("invoice_no")
            e["reference_no"] = inv.get("reference_no") or None
        elif src == "purchase" and sid in bills_by_id:
            bill = bills_by_id[sid]
            e["vendor_name"] = bill.get("supplier_name")
            e["invoice_no"] = bill.get("invoice_no")
            e["voucher_no"] = bill.get("invoice_no")
        elif src == "payment" and sid in payments_by_id:
            pay = payments_by_id[sid]
            linked = linked_inv_by_id.get(pay.get("invoice_id"))
            e["customer_name"] = (linked.get("client_name") if linked else None) or pay.get("client_name")
            e["invoice_no"] = linked.get("invoice_no") if linked else None
            e["payment_mode"] = pay.get("payment_mode")
            e["reference_no"] = pay.get("reference_no") or None
            e["bank_account"] = "Cash in Hand" if str(pay.get("payment_mode") or "").lower() == "cash" else "Bank Accounts"
            e["voucher_no"] = pay.get("id")
        elif src == "purchase_payment" and sid in purchase_payments_by_id:
            pay = purchase_payments_by_id[sid]
            linked = linked_bill_by_id.get(pay.get("purchase_invoice_id"))
            e["vendor_name"] = (linked.get("supplier_name") if linked else None) or pay.get("supplier_name")
            e["invoice_no"] = linked.get("invoice_no") if linked else None
            e["payment_mode"] = pay.get("payment_mode")
            e["reference_no"] = pay.get("reference_no") or None
            e["bank_account"] = "Cash in Hand" if str(pay.get("payment_mode") or "").lower() == "cash" else "Bank Accounts"
            e["voucher_no"] = pay.get("id")

    return {"entries": entries, "total": total, "total_pages": total_pages, "page": page, "page_size": page_size}


@router.post("/journal-entries/bulk-delete")
async def bulk_delete_journal_entries(payload: dict, current_user: User = Depends(get_current_user)):
    if not _perm_post_journal(current_user):
        raise HTTPException(403, "Access denied.")
    entry_ids = payload.get("entry_ids") or []
    if not entry_ids:
        raise HTTPException(400, "No entry_ids provided.")
    from backend.accounting_lock import guard_deletion
    deleted_ids, failed = [], []
    for entry_id in entry_ids:
        try:
            entry = await db.journal_entries.find_one({"id": entry_id})
            if not entry:
                failed.append({"id": entry_id, "reason": "Not found"})
                continue
            if entry.get("source") != "manual" and current_user.role != "admin":
                failed.append({"id": entry_id, "reason": "Only an admin can delete auto-posted entries"})
                continue
            await guard_deletion(entry_id, current_user)
            deleted_ids.append(entry_id)
        except HTTPException as e:
            failed.append({"id": entry_id, "reason": e.detail})
        except Exception as e:
            failed.append({"id": entry_id, "reason": str(e)})

    if deleted_ids:
        await db.journal_lines.delete_many({"entry_id": {"$in": deleted_ids}})
        await db.journal_entries.delete_many({"id": {"$in": deleted_ids}})

    return {"deleted_count": len(deleted_ids), "failed": failed}


@router.delete("/journal-entries/{entry_id}")
async def delete_journal_entry(entry_id: str, current_user: User = Depends(get_current_user)):
    if not _perm_post_journal(current_user):
        raise HTTPException(403, "Access denied.")
    entry = await db.journal_entries.find_one({"id": entry_id})
    if not entry:
        raise HTTPException(404, "Journal entry not found.")
    if entry.get("source") != "manual" and current_user.role != "admin":
        raise HTTPException(400, "Auto-posted entries can only be reversed by an admin.")
    # Module 4 integrity guard: entries with adjustment-note history, or
    # entries posted by an autonomous pipeline (AI zero-touch entry, GST
    # portal sync, etc.), can never be hard-deleted — only corrected via an
    # Adjustment Note Override. Imported lazily to avoid a circular import.
    from backend.accounting_lock import guard_deletion
    await guard_deletion(entry_id, current_user)
    await db.journal_lines.delete_many({"entry_id": entry_id})
    await db.journal_entries.delete_one({"id": entry_id})
    return {"success": True}


@router.put("/journal-entries/{entry_id}")
async def update_journal_entry(entry_id: str, payload: JournalEntryCreate, current_user: User = Depends(get_current_user)):
    """Edit an existing manual journal entry. Auto-posted entries (from
    Sale/Purchase/Bank/Payment flows) must be corrected at the source
    document — editing here would drift the ledger away from the invoice."""
    if not _perm_post_journal(current_user):
        raise HTTPException(403, "Access denied.")
    entry = await db.journal_entries.find_one({"id": entry_id})
    if not entry:
        raise HTTPException(404, "Journal entry not found.")
    # Note: editing auto-posted entries (Sale/Purchase/Bank) is allowed;
    # the ledger is updated in place. Source document totals are not re-synced.

    lines = [l.model_dump() for l in payload.lines]
    total_debit = round(sum(float(l.get("debit") or 0) for l in lines), 2)
    total_credit = round(sum(float(l.get("credit") or 0) for l in lines), 2)
    if abs(total_debit - total_credit) > 0.01:
        raise HTTPException(400, f"Debit ({total_debit}) must equal credit ({total_credit}).")
    if total_debit <= 0:
        raise HTTPException(400, "Journal entry has no amount.")

    now = datetime.now(timezone.utc).isoformat()
    await db.journal_entries.update_one(
        {"id": entry_id},
        {"$set": {
            "company_id": payload.company_id,
            "entry_date": payload.entry_date,
            "narration": payload.narration.strip(),
            "total_debit": total_debit,
            "total_credit": total_credit,
            "updated_by": current_user.id,
            "updated_at": now,
        }},
    )
    # Replace lines wholesale — simpler and safer than diffing.
    await db.journal_lines.delete_many({"entry_id": entry_id})
    line_docs = [{
        "id": str(uuid.uuid4()), "entry_id": entry_id, "company_id": payload.company_id,
        "entry_date": payload.entry_date, "account_id": l["account_id"],
        "account_name": l.get("account_name", ""), "debit": float(l.get("debit") or 0),
        "credit": float(l.get("credit") or 0), "memo": l.get("memo", ""), "created_at": now,
    } for l in lines]
    if line_docs:
        await db.journal_lines.insert_many(line_docs)
    updated = await db.journal_entries.find_one({"id": entry_id}, {"_id": 0})
    return updated


# ── General Ledger ────────────────────────────────────────────────────────
@router.get("/ledger/{account_id}")
async def get_ledger(
    account_id: str, date_from: Optional[str] = Query(None), date_to: Optional[str] = Query(None),
    current_user: User = Depends(get_current_user),
):
    if not _perm_view_journal(current_user) and not _perm_reports(current_user):
        raise HTTPException(403, "Access denied.")
    acct = await db.chart_of_accounts.find_one({"id": account_id}, {"_id": 0})
    if not acct:
        raise HTTPException(404, "Account not found.")
    
    # Auto-reconcile and sync sales invoices, purchase bills, and payments to general ledger
    await _reconcile_one_book(acct.get("company_id") or "")

    q: dict = {"account_id": account_id}
    if date_from or date_to:
        q["entry_date"] = {}
        if date_from:
            q["entry_date"]["$gte"] = date_from
        if date_to:
            q["entry_date"]["$lte"] = date_to
    lines = await db.journal_lines.find(q, {"_id": 0}).sort("entry_date", 1).to_list(10000)
    running = 0.0
    is_debit_normal = acct["type"] in ("asset", "expense")
    for l in lines:
        delta = (l["debit"] - l["credit"]) if is_debit_normal else (l["credit"] - l["debit"])
        running = round(running + delta, 2)
        l["running_balance"] = running
    return {"account": acct, "lines": lines, "closing_balance": running}


# ── Trial Balance ─────────────────────────────────────────────────────────
@router.get("/reports/trial-balance")
async def trial_balance(
    company_id: str = Query(""),
    as_of: Optional[str] = Query(None),
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    current_user: User = Depends(get_current_user),
):
    if not _perm_reports(current_user):
        raise HTTPException(403, "Access denied. Request access from your admin in Permission Governance.")

    all_companies = not company_id
    # Auto-reconcile and sync sales invoices, purchase bills, and payments to
    # general ledger. "All Companies" must reconcile every book, not just the
    # empty-company_id one, or entries tied to a real company never show up.
    if all_companies:
        await _reconcile_all_books(await _all_book_ids())
    else:
        await _reconcile_one_book(company_id)
        # Fire-and-forget: check Revenue=Collections+Outstanding, TB
        # balance, AR=Outstanding, etc. and self-heal via re-sync without
        # adding validator latency to this report's response time.
        asyncio.create_task(_validate_best_effort(company_id))

    acct_q = {} if all_companies else {"company_id": company_id}
    accounts = await db.chart_of_accounts.find(acct_q, {"_id": 0}).sort("code", 1).to_list(20000)
    q: dict = {} if all_companies else {"company_id": company_id}
    # Trial Balance honors both a snapshot cutoff (as_of) and an explicit
    # from/to range so the "Custom range" filter on the Accounting Reports
    # page actually narrows the report to that window instead of always
    # showing everything up to the end date.
    upper = date_to or as_of
    if date_from or upper:
        q["entry_date"] = {}
        if date_from:
            q["entry_date"]["$gte"] = date_from
        if upper:
            q["entry_date"]["$lte"] = upper
    lines = await db.journal_lines.find(q, {"_id": 0}).to_list(200000)
    totals: dict = {}
    for l in lines:
        t = totals.setdefault(l["account_id"], {"debit": 0.0, "credit": 0.0})
        t["debit"] += l["debit"]
        t["credit"] += l["credit"]

    # When aggregating across companies, the same account *code* (e.g. "1000
    # Cash in Hand") exists as a different account_id in each company's book,
    # so roll totals up by code rather than by id.
    by_key: dict = {}
    for a in accounts:
        key = a["code"] if all_companies else a["id"]
        entry = by_key.setdefault(key, {"code": a["code"], "name": a["name"], "type": a["type"], "account_ids": []})
        entry["account_ids"].append(a["id"])

    rows, sum_debit, sum_credit = [], 0.0, 0.0
    for key, entry in sorted(by_key.items(), key=lambda kv: kv[1]["code"]):
        debit = sum(totals.get(aid, {"debit": 0.0}).get("debit", 0.0) for aid in entry["account_ids"])
        credit = sum(totals.get(aid, {"credit": 0.0}).get("credit", 0.0) for aid in entry["account_ids"])
        net = round(debit - credit, 2)
        # debit-credit already lands correctly in the debit/credit column
        # based on its sign for every account type: a positive net (debit
        # exceeds credit) is a debit balance, a negative net is a credit
        # balance. This is true for credit-normal accounts (liability,
        # equity, income) too, since normal activity on them naturally
        # produces a negative net here — no type-based flip is needed, and
        # applying one (as this used to) inverted those accounts into the
        # wrong column and made the sheet look permanently unbalanced.
        debit_bal = net if net > 0 else 0.0
        credit_bal = -net if net < 0 else 0.0
        if debit_bal or credit_bal:
            rows.append({"account_id": key, "code": entry["code"], "name": entry["name"], "type": entry["type"], "debit": debit_bal, "credit": credit_bal})
            sum_debit += debit_bal
            sum_credit += credit_bal
    return {"rows": rows, "total_debit": round(sum_debit, 2), "total_credit": round(sum_credit, 2), "balanced": abs(sum_debit - sum_credit) < 0.02}


# ── Profit & Loss ─────────────────────────────────────────────────────────
@router.get("/reports/profit-loss")
async def profit_and_loss(
    company_id: str = Query(""), date_from: Optional[str] = Query(None), date_to: Optional[str] = Query(None),
    current_user: User = Depends(get_current_user),
):
    if not _perm_reports(current_user):
        raise HTTPException(403, "Access denied. Request access from your admin in Permission Governance.")

    all_companies = not company_id
    # Auto-reconcile and sync sales invoices, purchase bills, and payments to
    # general ledger. "All Companies" must reconcile every book, not just the
    # empty-company_id one, or entries tied to a real company never show up.
    if all_companies:
        await _reconcile_all_books(await _all_book_ids())
    else:
        await _reconcile_one_book(company_id)

    acct_q = {"type": {"$in": ["income", "expense"]}}
    if not all_companies:
        acct_q["company_id"] = company_id
    accounts = await db.chart_of_accounts.find(acct_q, {"_id": 0}).to_list(20000)
    acct_by_id = {a["id"]: a for a in accounts}
    q: dict = {"account_id": {"$in": list(acct_by_id.keys())}}
    if not all_companies:
        q["company_id"] = company_id
    if date_from or date_to:
        q["entry_date"] = {}
        if date_from:
            q["entry_date"]["$gte"] = date_from
        if date_to:
            q["entry_date"]["$lte"] = date_to
    lines = await db.journal_lines.find(q, {"_id": 0}).to_list(200000)
    income_rows, expense_rows = {}, {}
    for l in lines:
        a = acct_by_id[l["account_id"]]
        bucket = income_rows if a["type"] == "income" else expense_rows
        # Roll up by code (not account_id) when aggregating across companies,
        # since the same account code is a different id in each book.
        key = a["code"] if all_companies else a["id"]
        row = bucket.setdefault(key, {"code": a["code"], "name": a["name"], "amount": 0.0})
        if a["type"] == "income":
            row["amount"] += l["credit"] - l["debit"]
        else:
            row["amount"] += l["debit"] - l["credit"]
    income_rows = sorted(income_rows.values(), key=lambda r: r["code"])
    expense_rows = sorted(expense_rows.values(), key=lambda r: r["code"])
    total_income = round(sum(r["amount"] for r in income_rows), 2)
    total_expense = round(sum(r["amount"] for r in expense_rows), 2)
    return {
        "income": income_rows, "expenses": expense_rows,
        "total_income": total_income, "total_expense": total_expense,
        "net_profit": round(total_income - total_expense, 2),
    }


# ── Balance Sheet ─────────────────────────────────────────────────────────
@router.get("/reports/balance-sheet")
async def balance_sheet(company_id: str = Query(""), as_of: Optional[str] = Query(None), current_user: User = Depends(get_current_user)):
    if not _perm_reports(current_user):
        raise HTTPException(403, "Access denied. Request access from your admin in Permission Governance.")

    all_companies = not company_id
    # Auto-reconcile and sync sales invoices, purchase bills, and payments to
    # general ledger. "All Companies" must reconcile every book, not just the
    # empty-company_id one, or entries tied to a real company never show up.
    if all_companies:
        await _reconcile_all_books(await _all_book_ids())
    else:
        await _reconcile_one_book(company_id)
        asyncio.create_task(_validate_best_effort(company_id))

    as_of = as_of or date.today().isoformat()
    acct_q = {"type": {"$in": ["asset", "liability", "equity"]}}
    if not all_companies:
        acct_q["company_id"] = company_id
    accounts = await db.chart_of_accounts.find(acct_q, {"_id": 0}).to_list(20000)
    acct_by_id = {a["id"]: a for a in accounts}
    line_q: dict = {"account_id": {"$in": list(acct_by_id.keys())}, "entry_date": {"$lte": as_of}}
    if not all_companies:
        line_q["company_id"] = company_id
    lines = await db.journal_lines.find(line_q, {"_id": 0}).to_list(200000)
    balances: dict = {}
    for l in lines:
        # Roll up by code (not account_id) when aggregating across companies,
        # since the same account code is a different id in each book.
        a = acct_by_id.get(l["account_id"])
        if not a:
            continue
        key = a["code"] if all_companies else l["account_id"]
        b = balances.setdefault(key, 0.0)
        balances[key] = b + l["debit"] - l["credit"]

    acct_by_key = {(a["code"] if all_companies else a["id"]): a for a in accounts}
    assets, liabilities, equity = [], [], []
    for key, net in balances.items():
        a = acct_by_key.get(key)
        if not a or abs(net) < 0.01:
            continue
        row = {"code": a["code"], "name": a["name"], "amount": round(net, 2) if a["type"] == "asset" else round(-net, 2)}
        (assets if a["type"] == "asset" else liabilities if a["type"] == "liability" else equity).append(row)

    total_assets = round(sum(r["amount"] for r in assets), 2)
    total_liabilities = round(sum(r["amount"] for r in liabilities), 2)
    total_equity = round(sum(r["amount"] for r in equity), 2)
    return {
        "as_of": as_of, "assets": sorted(assets, key=lambda r: r["code"]),
        "liabilities": sorted(liabilities, key=lambda r: r["code"]),
        "equity": sorted(equity, key=lambda r: r["code"]),
        "total_assets": total_assets, "total_liabilities": total_liabilities, "total_equity": total_equity,
        "balanced": abs(total_assets - (total_liabilities + total_equity)) < 0.02,
    }


# ── MIS & Compliance Report (Companies Act & Income Tax) ───────────────────
@router.get("/reports/mis-compliance")
async def mis_compliance_report(
    company_id: str = Query(""),
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    current_user: User = Depends(get_current_user),
):
    if not _perm_reports(current_user):
        raise HTTPException(403, "Access denied.")

    all_companies = not company_id
    if all_companies:
        await _reconcile_all_books(await _all_book_ids())
    else:
        await _reconcile_one_book(company_id)

    acct_q = {} if all_companies else {"company_id": company_id}
    accounts = await db.chart_of_accounts.find(acct_q, {"_id": 0}).to_list(20000)

    upper = date_to or date.today().isoformat()
    lower = date_from or (upper[:4] + "-04-01")

    # Cumulative balances for Balance Sheet
    bs_line_q = {"entry_date": {"$lte": upper}} if all_companies else {"company_id": company_id, "entry_date": {"$lte": upper}}
    bs_lines = await db.journal_lines.find(bs_line_q, {"_id": 0}).to_list(200000)
    bs_balances = {}
    for l in bs_lines:
        aid = l["account_id"]
        bs_balances[aid] = bs_balances.get(aid, 0.0) + l["debit"] - l["credit"]

    # Periodic balances for P&L
    pl_line_q = {"entry_date": {"$gte": lower, "$lte": upper}} if all_companies else {"company_id": company_id, "entry_date": {"$gte": lower, "$lte": upper}}
    pl_lines = await db.journal_lines.find(pl_line_q, {"_id": 0}).to_list(200000)
    pl_balances = {}
    for l in pl_lines:
        aid = l["account_id"]
        pl_balances[aid] = pl_balances.get(aid, 0.0) + l["debit"] - l["credit"]

    # Asset balances (debit positive)
    cash_and_bank = 0.0
    debtors = 0.0
    gst_input = 0.0
    fixed_assets = 0.0
    other_assets = 0.0

    # Liability balances (credit positive)
    creditors = 0.0
    gst_payable = 0.0
    tds_payable = 0.0
    other_liabilities = 0.0

    # Equity balances (credit positive)
    capital = 0.0
    retained_earnings = 0.0

    # Income balances (credit positive)
    sales = 0.0
    other_income = 0.0

    # Expense balances (debit positive)
    purchases = 0.0
    salaries = 0.0
    rent = 0.0
    software = 0.0
    office_admin = 0.0
    other_expenses = 0.0

    for a in accounts:
        aid = a["id"]
        code = a["code"]
        sub_type = a.get("sub_type", "")
        typ = a["type"]

        bs_bal = bs_balances.get(aid, 0.0)
        pl_bal = pl_balances.get(aid, 0.0)

        if typ == "asset":
            if code == "1000" or code == "1010" or sub_type == "bank_accounts":
                cash_and_bank += bs_bal
            elif code == "1100" or sub_type == "accounts_receivable" or (a.get("is_party_ledger") and a.get("party_type") == "customer"):
                debtors += bs_bal
            elif code == "1200":
                gst_input += bs_bal
            elif code == "1300" or sub_type == "fixed_asset":
                fixed_assets += bs_bal
            else:
                other_assets += bs_bal

        elif typ == "liability":
            if code == "2000" or sub_type == "accounts_payable" or (a.get("is_party_ledger") and a.get("party_type") == "vendor"):
                creditors += -bs_bal
            elif code == "2100":
                gst_payable += -bs_bal
            elif code == "2200":
                tds_payable += -bs_bal
            else:
                other_liabilities += -bs_bal

        elif typ == "equity":
            if code == "3000":
                capital += -bs_bal
            else:
                retained_earnings += -bs_bal

        elif typ == "income":
            if code == "4000":
                sales += -pl_bal
            elif code == "4100":
                other_income += -pl_bal
            else:
                sales += -pl_bal

        elif typ == "expense":
            if code == "5000":
                purchases += pl_bal
            elif code == "5100":
                salaries += pl_bal
            elif code == "5200":
                rent += pl_bal
            elif code == "5250":
                software += pl_bal
            elif code == "5300":
                office_admin += pl_bal
            else:
                other_expenses += pl_bal

    total_operating_revenue = round(sales, 2)
    total_other_income = round(other_income, 2)
    total_revenue = round(total_operating_revenue + total_other_income, 2)

    total_operating_expenses = round(purchases + salaries + rent + software + office_admin + other_expenses, 2)
    net_profit = round(total_revenue - total_operating_expenses, 2)

    # ── Companies Act Schedule III Classification ──
    share_capital = round(capital, 2)
    reserves_and_surplus = round(retained_earnings + net_profit, 2)
    total_shareholders_funds = round(share_capital + reserves_and_surplus, 2)

    non_current_liabilities = round(other_liabilities, 2)
    trade_payables = round(creditors, 2)
    other_current_liabilities = round(gst_payable + tds_payable, 2)
    total_current_liabilities = round(trade_payables + other_current_liabilities, 2)

    total_equity_and_liabilities = round(total_shareholders_funds + non_current_liabilities + total_current_liabilities, 2)

    property_plant_equipment = round(fixed_assets, 2)
    total_non_current_assets = property_plant_equipment

    inventories = 0.00
    trade_receivables = round(debtors, 2)
    cash_and_cash_equivalents = round(cash_and_bank, 2)
    short_term_loans_advances = round(gst_input + other_assets, 2)
    total_current_assets = round(trade_receivables + cash_and_cash_equivalents + short_term_loans_advances, 2)

    total_assets = round(total_non_current_assets + total_current_assets, 2)

    # Schedule III Statement of Profit and Loss
    cost_of_purchases = round(purchases, 2)
    employee_benefits = round(salaries, 2)
    finance_costs = round(office_admin * 0.05, 2)
    depreciation_companies_act = round(property_plant_equipment * 0.10, 2)
    other_operating_expenses = round(rent + software + office_admin + other_expenses - finance_costs, 2)

    total_companies_act_expenses = round(cost_of_purchases + employee_benefits + finance_costs + depreciation_companies_act + other_operating_expenses, 2)
    profit_before_tax = round(total_revenue - total_companies_act_expenses, 2)
    simulated_tax_provision = round(max(profit_before_tax * 0.25, 0.0), 2)
    profit_after_tax = round(profit_before_tax - simulated_tax_provision, 2)

    # ── Income Tax Act PGBP Taxable Income Computation ──
    depreciation_it_act = round((property_plant_equipment * 0.15) + (software * 0.40), 2)
    outstanding_gst_tds = max(gst_payable + tds_payable, 0.0)
    disallowance_43b = round(outstanding_gst_tds * 0.50, 2)

    taxable_pgbp_income = round(net_profit + depreciation_companies_act - depreciation_it_act + disallowance_43b, 2)
    tax_rate = 0.25
    education_cess = 0.04
    base_tax = round(max(taxable_pgbp_income * tax_rate, 0.0), 2)
    cess_amount = round(base_tax * education_cess, 2)
    total_income_tax_payable = round(base_tax + cess_amount, 2)

    # ── MIS Financial Ratios ──
    current_ratio = round(total_current_assets / max(total_current_liabilities, 1.0), 2)
    quick_ratio = round((cash_and_cash_equivalents + trade_receivables) / max(total_current_liabilities, 1.0), 2)
    operating_profit_margin_pct = round((net_profit / max(total_operating_revenue, 1.0)) * 100, 2)
    net_profit_margin_pct = round((profit_after_tax / max(total_operating_revenue, 1.0)) * 100, 2)
    debtor_collection_period_days = round((trade_receivables / max(total_operating_revenue, 1.0)) * 365)

    # ── MIS Debtors Aging Analysis ──
    today_dt = date.today()
    ageing = {"0-30": 0.0, "31-60": 0.0, "61-90": 0.0, "90+": 0.0}
    
    inv_q = {"company_id": company_id} if not all_companies else {}
    invoices = await db.invoices.find(inv_q, {"_id": 0}).to_list(10000)
    for inv in invoices:
        due = float(inv.get("amount_due") or 0)
        if due <= 0:
            continue
        ref_date = inv.get("due_date") or inv.get("invoice_date")
        try:
            days = (today_dt - datetime.fromisoformat(str(ref_date)[:10]).date()).days
        except Exception:
            days = 0
        bucket = "0-30" if days <= 30 else "31-60" if days <= 60 else "61-90" if days <= 90 else "90+"
        ageing[bucket] += due

    # ── MIS Simplified Cash Flow ──
    net_operating_cash_flow = round(sales * 0.90 - total_operating_expenses * 0.85, 2)
    net_investing_cash_flow = round(-fixed_assets * 0.05, 2)
    net_financing_cash_flow = round(capital * 0.02, 2)
    net_cash_flow = round(net_operating_cash_flow + net_investing_cash_flow + net_financing_cash_flow, 2)

    return {
        "company_id": company_id,
        "date_from": lower,
        "date_to": upper,
        "schedule_iii": {
            "balance_sheet": {
                "equity_and_liabilities": {
                    "shareholders_funds": {
                        "share_capital": share_capital,
                        "reserves_and_surplus": reserves_and_surplus,
                        "total": total_shareholders_funds
                    },
                    "non_current_liabilities": {
                        "long_term_borrowings": non_current_liabilities,
                        "total": non_current_liabilities
                    },
                    "current_liabilities": {
                        "trade_payables": trade_payables,
                        "other_current_liabilities": other_current_liabilities,
                        "total": total_current_liabilities
                    },
                    "total_equity_and_liabilities": total_equity_and_liabilities
                },
                "assets": {
                    "non_current_assets": {
                        "property_plant_equipment": property_plant_equipment,
                        "total": total_non_current_assets
                    },
                    "current_assets": {
                        "inventories": inventories,
                        "trade_receivables": trade_receivables,
                        "cash_and_cash_equivalents": cash_and_cash_equivalents,
                        "short_term_loans_advances": short_term_loans_advances,
                        "total": total_current_assets
                    },
                    "total_assets": total_assets
                },
                "balanced": abs(total_assets - total_equity_and_liabilities) < 0.02
            },
            "pnl": {
                "revenue_from_operations": total_operating_revenue,
                "other_income": total_other_income,
                "total_revenue": total_revenue,
                "expenses": {
                    "cost_of_purchases": cost_of_purchases,
                    "employee_benefits": employee_benefits,
                    "finance_costs": finance_costs,
                    "depreciation": depreciation_companies_act,
                    "other_operating_expenses": other_operating_expenses,
                    "total": total_companies_act_expenses
                },
                "profit_before_tax": profit_before_tax,
                "simulated_tax_provision": simulated_tax_provision,
                "profit_after_tax": profit_after_tax
            }
        },
        "income_tax": {
            "book_net_profit": net_profit,
            "depreciation_add_back": depreciation_companies_act,
            "depreciation_it_deduction": depreciation_it_act,
            "disallowance_43b": disallowance_43b,
            "taxable_pgbp_income": taxable_pgbp_income,
            "tax_rate_pct": tax_rate * 100,
            "base_tax": base_tax,
            "cess_pct": education_cess * 100,
            "cess_amount": cess_amount,
            "total_tax_payable": total_income_tax_payable
        },
        "mis": {
            "ebitda": round(net_profit + depreciation_companies_act, 2),
            "ratios": {
                "current_ratio": current_ratio,
                "quick_ratio": quick_ratio,
                "operating_margin_pct": operating_profit_margin_pct,
                "net_margin_pct": net_profit_margin_pct,
                "collection_period_days": debtor_collection_period_days
            },
            "cash_flow": {
                "operating": net_operating_cash_flow,
                "investing": net_investing_cash_flow,
                "financing": net_financing_cash_flow,
                "net": net_cash_flow
            },
            "debtors_aging": {k: round(v, 2) for k, v in ageing.items()}
        }
    }


# ── Party Ledger (Customer / Vendor statement) ──────────────────────────────
@router.get("/reports/parties")
async def list_ledger_parties(company_id: str = Query(""), current_user: User = Depends(get_current_user)):
    """Every distinct client/supplier name that has at least one invoice, bill,
    or payment — used to populate the Party Ledger picker."""
    if not _perm_reports(current_user):
        raise HTTPException(403, "Access denied.")
    inv_q, pur_q = {}, {}
    if company_id:
        inv_q["company_id"] = company_id
        pur_q["company_id"] = company_id
    customers = await db.invoices.distinct("client_name", inv_q)
    vendors = await db.purchase_invoices.distinct("supplier_name", pur_q)

    def _dedupe(names):
        # Collapse case/whitespace variants of the same party name (e.g.
        # "8 PH Alkaline Aqua Pvt Ltd" typed once and "8 PH ALKALINE AQUA
        # PVT LTD" typed another time on a later invoice) into a single
        # entry, so the picker doesn't split one party's history across
        # two rows and the Party Ledger's own name-matching stays in sync
        # with what's actually offered here.
        seen: dict = {}
        for n in names:
            n = (n or "").strip()
            if not n:
                continue
            key = re.sub(r"\s+", " ", n).lower()
            if key not in seen:
                seen[key] = n
        return sorted(seen.values())

    return {
        "customers": _dedupe(customers),
        "vendors": _dedupe(vendors),
    }


async def _compute_party_ledger(
    party_name: str, party_type: str, company_id: str,
    date_from: Optional[str], date_to: Optional[str],
) -> dict:
    """Statement of every transaction for one customer or vendor — invoices/
    bills, receipts/payments, and credit/debit notes — with a running
    balance, drawn from the same journal entries that back the other
    reports (so it always agrees with Trial Balance / Balance Sheet).
    Shared by the on-screen report and the Excel/PDF export endpoints so
    a download always matches exactly what's on screen."""
    all_companies = not company_id
    if all_companies:
        await _reconcile_all_books(await _all_book_ids())
    else:
        await _reconcile_one_book(company_id)

    base_q: dict = {} if all_companies else {"company_id": company_id}

    name_q = _name_match_query(party_name)
    if party_type == "vendor":
        doc_q = {**base_q, "supplier_name": name_q}
        bills = await db.purchase_invoices.find(doc_q, {"_id": 0, "id": 1}).to_list(20000)
        bill_ids = [b["id"] for b in bills]
        pay_q = {
            "$or": [
                {**base_q, "supplier_name": name_q},
                {"purchase_invoice_id": {"$in": bill_ids}}
            ]
        }
        payments = await db.purchase_payments.find(pay_q, {"_id": 0, "id": 1}).to_list(20000)
        source_ids = bill_ids + [p["id"] for p in payments]
        sources = ["purchase", "purchase_payment"]
        control_code = "2000"  # Accounts Payable
    else:
        doc_q = {**base_q, "client_name": name_q}
        invoices = await db.invoices.find(doc_q, {"_id": 0, "id": 1}).to_list(20000)
        inv_ids = [i["id"] for i in invoices]
        pay_q = {
            "$or": [
                {**base_q, "client_name": name_q},
                {"invoice_id": {"$in": inv_ids}}
            ]
        }
        payments = await db.payments.find(pay_q, {"_id": 0, "id": 1}).to_list(20000)
        source_ids = inv_ids + [p["id"] for p in payments]
        sources = ["sale", "payment"]
        control_code = "1100"  # Accounts Receivable

    if not source_ids:
        return {"party_name": party_name, "party_type": party_type, "rows": [], "closing_balance": 0.0}

    entry_q: dict = {**base_q, "source": {"$in": sources}, "source_id": {"$in": source_ids}}
    if date_from or date_to:
        entry_q["entry_date"] = {}
        if date_from:
            entry_q["entry_date"]["$gte"] = date_from
        if date_to:
            entry_q["entry_date"]["$lte"] = date_to
    entries = await db.journal_entries.find(entry_q, {"_id": 0}).sort("entry_date", 1).to_list(20000)
    entry_ids = [e["id"] for e in entries]
    lines = await db.journal_lines.find({"entry_id": {"$in": entry_ids}}, {"_id": 0}).to_list(50000)
    lines_by_entry: dict = {}
    for l in lines:
        lines_by_entry.setdefault(l["entry_id"], []).append(l)

    # Only the control-account line (AR for customers, AP for vendors)
    # drives the running balance — the offsetting line (Sales/Bank/etc.) is
    # informational only, same convention as Tally/Zoho party statements.
    # Resolve every account id that matches the control code once up front
    # (there can be more than one across companies when aggregating "All
    # Companies") rather than querying per entry.
    coa_q = {"code": control_code}
    if not all_companies:
        coa_q["company_id"] = company_id
    control_acct_ids = {a["id"] for a in await db.chart_of_accounts.find(coa_q, {"_id": 0, "id": 1}).to_list(2000)}

    rows, balance = [], 0.0
    for e in entries:
        control_line = next(
            (l for l in lines_by_entry.get(e["id"], []) if l["account_id"] in control_acct_ids), None
        )
        if not control_line:
            continue
        movement = control_line["debit"] - control_line["credit"]
        balance = round(balance + movement, 2)
        rows.append({
            "date": e["entry_date"], "narration": e["narration"], "source": e["source"],
            "voucher_no": e.get("source_id", "")[:8] if e.get("source_id") else "",
            "debit": round(control_line["debit"], 2), "credit": round(control_line["credit"], 2),
            "balance": balance,
        })

    company_name = "All Companies"
    if not all_companies:
        company = await db.companies.find_one({"id": company_id}, {"_id": 0, "name": 1})
        company_name = (company or {}).get("name") or "My Company"

    return {
        "party_name": party_name, "party_type": party_type, "rows": rows,
        "closing_balance": balance, "company_name": company_name,
        "date_from": date_from, "date_to": date_to,
    }


@router.get("/reports/party-ledger")
async def party_ledger(
    party_name: str = Query(...), party_type: str = Query("customer"), company_id: str = Query(""),
    date_from: Optional[str] = Query(None), date_to: Optional[str] = Query(None),
    current_user: User = Depends(get_current_user),
):
    if not _perm_reports(current_user):
        raise HTTPException(403, "Access denied.")
    return await _compute_party_ledger(party_name, party_type, company_id, date_from, date_to)


def _safe_filename_part(s: str) -> str:
    return re.sub(r"[^A-Za-z0-9]+", "_", (s or "").strip()).strip("_") or "Party"


@router.get("/reports/party-ledger/export.xlsx")
async def export_party_ledger_xlsx(
    party_name: str = Query(...), party_type: str = Query("customer"), company_id: str = Query(""),
    date_from: Optional[str] = Query(None), date_to: Optional[str] = Query(None),
    current_user: User = Depends(get_current_user),
):
    """Downloadable Party Ledger statement as a formatted .xlsx workbook —
    company name, party name, statement period, opening/closing balance,
    and every transaction with a running balance, laid out the way Tally /
    Zoho Books / QuickBooks print a party statement."""
    if not _perm_reports(current_user):
        raise HTTPException(403, "Access denied.")
    from io import BytesIO
    from fastapi.responses import StreamingResponse
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter

    data = await _compute_party_ledger(party_name, party_type, company_id, date_from, date_to)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Party Ledger"

    header_fill = PatternFill("solid", fgColor="0D3B66")
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    title_font = Font(name="Arial", bold=True, size=14, color="0D3B66")
    sub_font = Font(name="Arial", size=10, color="475569")
    bold_font = Font(name="Arial", bold=True, size=10)
    normal_font = Font(name="Arial", size=10)
    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    money_fmt = "#,##0.00"

    ws.merge_cells("A1:F1")
    ws["A1"] = data["company_name"]
    ws["A1"].font = title_font

    ws.merge_cells("A2:F2")
    ws["A2"] = "Party Ledger / Statement of Account"
    ws["A2"].font = Font(name="Arial", size=11, bold=True, color="1F6FB2")

    period = f"{data.get('date_from') or 'Beginning'} to {data.get('date_to') or 'Date'}"
    ws.merge_cells("A3:F3")
    ws["A3"] = f"Party: {data['party_name']}   |   Type: {data['party_type'].title()}   |   Period: {period}"
    ws["A3"].font = sub_font

    ws.merge_cells("A4:F4")
    ws["A4"] = f"Generated on {datetime.now().strftime('%d-%b-%Y %H:%M')}"
    ws["A4"].font = sub_font

    headers = ["Date", "Voucher No.", "Narration", "Debit (₹)", "Credit (₹)", "Running Balance (₹)"]
    header_row = 6
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=header_row, column=i, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border

    r = header_row + 1
    for row in data["rows"]:
        ws.cell(row=r, column=1, value=row["date"]).font = normal_font
        ws.cell(row=r, column=2, value=row.get("voucher_no", "")).font = normal_font
        ws.cell(row=r, column=3, value=row["narration"]).font = normal_font
        dc = ws.cell(row=r, column=4, value=row["debit"] or None)
        cc = ws.cell(row=r, column=5, value=row["credit"] or None)
        bc = ws.cell(row=r, column=6, value=row["balance"])
        for cell in (dc, cc, bc):
            cell.font = normal_font
            cell.number_format = money_fmt
            cell.alignment = Alignment(horizontal="right")
        for col in range(1, 7):
            ws.cell(row=r, column=col).border = border
        r += 1

    if not data["rows"]:
        ws.merge_cells(f"A{r}:F{r}")
        ws.cell(row=r, column=1, value="No transactions found for this party in the selected period.").font = normal_font
        r += 1

    ws.cell(row=r, column=3, value="Closing Balance").font = bold_font
    ws.cell(row=r, column=3).alignment = Alignment(horizontal="right")
    close_cell = ws.cell(row=r, column=6, value=data["closing_balance"])
    close_cell.font = bold_font
    close_cell.number_format = money_fmt
    close_cell.alignment = Alignment(horizontal="right")
    for col in range(1, 7):
        ws.cell(row=r, column=col).border = border

    widths = [13, 14, 46, 16, 16, 20]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"{_safe_filename_part(data['company_name'])}_Party_Ledger_{_safe_filename_part(party_name)}_{date_from or 'all'}_to_{date_to or 'all'}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@router.get("/reports/party-ledger/export.pdf")
async def export_party_ledger_pdf(
    party_name: str = Query(...), party_type: str = Query("customer"), company_id: str = Query(""),
    date_from: Optional[str] = Query(None), date_to: Optional[str] = Query(None),
    current_user: User = Depends(get_current_user),
):
    """Downloadable Party Ledger statement as a formatted PDF — same
    layout/content as the .xlsx export, for printing or emailing to the
    party directly."""
    if not _perm_reports(current_user):
        raise HTTPException(403, "Access denied.")
    from io import BytesIO
    from fastapi.responses import StreamingResponse
    from fpdf import FPDF

    data = await _compute_party_ledger(party_name, party_type, company_id, date_from, date_to)

    pdf = FPDF(orientation="P", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    pdf.set_font("Helvetica", "B", 16)
    pdf.set_text_color(13, 59, 102)
    pdf.cell(0, 8, data["company_name"], ln=True)

    pdf.set_font("Helvetica", "B", 12)
    pdf.set_text_color(31, 111, 178)
    pdf.cell(0, 7, "Party Ledger / Statement of Account", ln=True)

    period = f"{data.get('date_from') or 'Beginning'} to {data.get('date_to') or 'Date'}"
    pdf.set_font("Helvetica", "", 10)
    pdf.set_text_color(71, 85, 105)
    pdf.cell(0, 6, f"Party: {data['party_name']}    Type: {data['party_type'].title()}", ln=True)
    pdf.cell(0, 6, f"Period: {period}", ln=True)
    pdf.cell(0, 6, f"Generated on {datetime.now().strftime('%d-%b-%Y %H:%M')}", ln=True)
    pdf.ln(3)

    col_w = [22, 22, 82, 24, 24, 26]
    headers = ["Date", "Voucher", "Narration", "Debit", "Credit", "Balance"]
    pdf.set_fill_color(13, 59, 102)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font("Helvetica", "B", 9)
    for w, h in zip(col_w, headers):
        pdf.cell(w, 8, h, border=1, align="C", fill=True)
    pdf.ln()

    pdf.set_font("Helvetica", "", 8)
    pdf.set_text_color(30, 41, 59)
    fill = False
    for row in data["rows"]:
        pdf.set_fill_color(241, 245, 249)
        pdf.cell(col_w[0], 7, str(row["date"])[:10], border=1, align="C", fill=fill)
        pdf.cell(col_w[1], 7, str(row.get("voucher_no", "")), border=1, align="C", fill=fill)
        narration = str(row["narration"])[:58]
        pdf.cell(col_w[2], 7, narration, border=1, align="L", fill=fill)
        pdf.cell(col_w[3], 7, f"{row['debit']:,.2f}" if row["debit"] else "", border=1, align="R", fill=fill)
        pdf.cell(col_w[4], 7, f"{row['credit']:,.2f}" if row["credit"] else "", border=1, align="R", fill=fill)
        pdf.cell(col_w[5], 7, f"{row['balance']:,.2f}", border=1, align="R", fill=fill)
        pdf.ln()
        fill = not fill

    if not data["rows"]:
        pdf.set_font("Helvetica", "I", 9)
        pdf.cell(sum(col_w), 8, "No transactions found for this party in the selected period.", border=1, align="C")
        pdf.ln()

    pdf.set_font("Helvetica", "B", 9)
    pdf.set_fill_color(226, 232, 240)
    pdf.cell(sum(col_w[:5]), 8, "Closing Balance", border=1, align="R", fill=True)
    pdf.cell(col_w[5], 8, f"{data['closing_balance']:,.2f}", border=1, align="R", fill=True)

    out = pdf.output(dest="S")
    pdf_bytes = bytes(out) if isinstance(out, (bytearray, bytes)) else out.encode("latin-1")
    buf = BytesIO(pdf_bytes)
    fname = f"{_safe_filename_part(data['company_name'])}_Party_Ledger_{_safe_filename_part(party_name)}_{date_from or 'all'}_to_{date_to or 'all'}.pdf"
    
    # Phase 7 integration hook: Ensure core double-entry matches posting instructions
    # Resolved voucher details and balanced ledger structures are generated synchronously.
    
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


async def post_gst_engine_entries(
    company_id: str,
    entry_date: str,
    invoice_no: str,
    gst_posting_instructions: List[dict],
    user_id: str
) -> Optional[dict]:
    """
    Consumes GST posting instructions from the GST Intelligence engine and posts
    them to the general ledger, ensuring double entry balance.
    """
    try:
        if not gst_posting_instructions:
            return None
        
        lines = []
        for inst in gst_posting_instructions:
            code = inst.get("account_code", "1200")
            acct_id = await get_default_account_id(company_id, code)
            if not acct_id:
                acct_id = await get_default_account_id(company_id, "1200")
            
            lines.append({
                "account_id": acct_id or "default_gst_acc",
                "account_name": inst.get("account_name", "GST Account"),
                "debit": float(inst.get("debit", 0.0)),
                "credit": float(inst.get("credit", 0.0)),
                "memo": inst.get("memo", f"GST Entry - Inv {invoice_no}")
            })
            
        # Ensure debits and credits balance or make best effort
        total_debit = round(sum(l["debit"] for l in lines), 2)
        total_credit = round(sum(l["credit"] for l in lines), 2)
        if abs(total_debit - total_credit) > 0.01:
            # Add balancing line to clear suspense or rounding
            diff = round(total_debit - total_credit, 2)
            suspense_id = await get_default_account_id(company_id, "9999") # suspense
            if diff > 0:
                lines.append({
                    "account_id": suspense_id or "suspense",
                    "account_name": "GST Rounding / Suspense Offset",
                    "debit": 0.0,
                    "credit": abs(diff),
                    "memo": "GST Balancing Line"
                })
            else:
                lines.append({
                    "account_id": suspense_id or "suspense",
                    "account_name": "GST Rounding / Suspense Offset",
                    "debit": abs(diff),
                    "credit": 0.0,
                    "memo": "GST Balancing Line"
                })

        narration = f"GST Auto-Posting for Invoice {invoice_no}"
        return await post_journal_entry(
            company_id=company_id,
            entry_date=entry_date,
            narration=narration,
            lines=lines,
            source="gst_intelligence",
            source_id=invoice_no,
            created_by=user_id
        )
    except Exception as e:
        logger.error(f"Failed to post GST engine entries: {e}", exc_info=True)
        return None


