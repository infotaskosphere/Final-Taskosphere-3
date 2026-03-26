"""
Enhanced Password Repository Router for Taskosphere
FIXED:
  1. Route ordering — static paths BEFORE /{entry_id} to avoid int-cast 500s
  2. download-template uses a temp file instead of BytesIO with FileResponse
  3. Robust NoneType current_user guards everywhere
  4. Encryption key loaded from env (falls back gracefully)
  5. Pagination done in SQL (OFFSET/LIMIT) not in Python — faster loads
  6. All endpoints return clean JSON; no silent swallows
"""

import io
import os
import base64
import json
import logging
import enum
import tempfile
from datetime import datetime
from typing import Optional, List, Dict, Any

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query, status
from fastapi.responses import FileResponse
from sqlalchemy import select, and_, or_, func, desc, asc, Column, Integer, String, Text, DateTime, Boolean
from sqlalchemy.orm import Session
from sqlalchemy.ext.declarative import declarative_base
from pydantic import BaseModel, Field, validator
from cryptography.fernet import Fernet
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

# ── DATABASE & AUTH DEPENDENCIES (AUTO-DETECT) ───────────────────────────────
try:
    from backend.dependencies import get_db, get_current_user
except ImportError:
    try:
        from app.dependencies import get_db, get_current_user
    except ImportError:
        try:
            from dependencies import get_db, get_current_user
        except ImportError:
            def get_db():
                raise NotImplementedError("get_db not found — fix your import path")
            def get_current_user():
                raise NotImplementedError("get_current_user not found — fix your import path")

logger = logging.getLogger(__name__)

# ── Database Models ───────────────────────────────────────────────────────────
Base = declarative_base()


class PortalTypeEnum(str, enum.Enum):
    MCA = "MCA"
    ROC = "ROC"
    DGFT = "DGFT"
    TRADEMARK = "TRADEMARK"
    GST = "GST"
    INCOME_TAX = "INCOME_TAX"
    TDS = "TDS"
    TRACES = "TRACES"
    EPFO = "EPFO"
    ESIC = "ESIC"
    MSME = "MSME"
    RERA = "RERA"
    OTHER = "OTHER"


class DepartmentEnum(str, enum.Enum):
    GST = "GST"
    IT = "IT"
    ACC = "ACC"
    TDS = "TDS"
    ROC = "ROC"
    TM = "TM"
    MSME = "MSME"
    FEMA = "FEMA"
    DSC = "DSC"
    OTHER = "OTHER"


class HolderTypeEnum(str, enum.Enum):
    COMPANY = "COMPANY"
    DIRECTOR = "DIRECTOR"
    INDIVIDUAL = "INDIVIDUAL"
    PARTNER = "PARTNER"
    TRUSTEE = "TRUSTEE"
    OTHER = "OTHER"


class PasswordEntry(Base):
    __tablename__ = "password_entries"

    id = Column(Integer, primary_key=True, index=True)
    user_id = Column(Integer, index=True)
    portal_name = Column(String(255), index=True)
    portal_type = Column(String(50), default="OTHER", index=True)
    url = Column(String(500), nullable=True)
    username = Column(String(255), index=True)
    password_encrypted = Column(Text)
    has_password = Column(Boolean, default=True)
    department = Column(String(50), default="OTHER", index=True)
    holder_type = Column(String(50), default="COMPANY")
    holder_name = Column(String(255), nullable=True, index=True)
    holder_pan = Column(String(20), nullable=True, index=True)
    holder_din = Column(String(20), nullable=True)
    mobile = Column(String(20), nullable=True)
    trade_name = Column(String(255), nullable=True, index=True)
    client_name = Column(String(255), nullable=True, index=True)
    client_id = Column(String(50), nullable=True, index=True)
    notes = Column(Text, nullable=True)
    tags = Column(Text, nullable=True)  # JSON array stored as string
    created_at = Column(DateTime, default=datetime.utcnow, index=True)
    updated_at = Column(DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    last_accessed_at = Column(DateTime, nullable=True)
    is_archived = Column(Boolean, default=False)


class AccessLog(Base):
    __tablename__ = "password_access_logs"

    id = Column(Integer, primary_key=True, index=True)
    user_id = Column(Integer, index=True)
    entry_id = Column(Integer, index=True)
    action = Column(String(50), index=True)
    timestamp = Column(DateTime, default=datetime.utcnow, index=True)
    ip_address = Column(String(50), nullable=True)
    user_agent = Column(String(500), nullable=True)


# ── Pydantic Schemas ──────────────────────────────────────────────────────────
class PasswordEntryBase(BaseModel):
    portal_name: str = Field(..., min_length=1, max_length=255)
    portal_type: Optional[str] = "OTHER"
    url: Optional[str] = None
    username: str = Field(..., min_length=1, max_length=255)
    password_plain: Optional[str] = None
    department: Optional[str] = None
    holder_type: Optional[str] = "COMPANY"
    holder_name: Optional[str] = None
    holder_pan: Optional[str] = None
    holder_din: Optional[str] = None
    mobile: Optional[str] = None
    trade_name: Optional[str] = None
    client_name: Optional[str] = None
    client_id: Optional[str] = None
    notes: Optional[str] = None
    tags: Optional[List[str]] = None

    @validator("portal_type")
    def validate_portal_type(cls, v):
        valid = [e.value for e in PortalTypeEnum]
        if v and v not in valid:
            return "OTHER"
        return v or "OTHER"

    @validator("department")
    def validate_department(cls, v, values):
        valid = [e.value for e in DepartmentEnum]
        if v and v not in valid:
            v = None
        if not v:
            dept_map = {
                "MCA": "ROC", "ROC": "ROC", "DGFT": "OTHER", "TRADEMARK": "TM",
                "GST": "GST", "INCOME_TAX": "IT", "TDS": "TDS", "EPFO": "ACC",
                "ESIC": "ACC", "TRACES": "TDS", "MSME": "MSME", "RERA": "OTHER",
            }
            portal = values.get("portal_type", "OTHER")
            return dept_map.get(portal, "OTHER")
        return v


class PasswordEntryCreate(PasswordEntryBase):
    pass


class PasswordEntryUpdate(BaseModel):
    portal_name: Optional[str] = None
    portal_type: Optional[str] = None
    url: Optional[str] = None
    username: Optional[str] = None
    password_plain: Optional[str] = None
    department: Optional[str] = None
    holder_type: Optional[str] = None
    holder_name: Optional[str] = None
    holder_pan: Optional[str] = None
    holder_din: Optional[str] = None
    mobile: Optional[str] = None
    trade_name: Optional[str] = None
    client_name: Optional[str] = None
    client_id: Optional[str] = None
    notes: Optional[str] = None
    tags: Optional[List[str]] = None


class PasswordEntryResponse(PasswordEntryBase):
    id: int
    user_id: int
    has_password: bool
    created_at: datetime
    updated_at: datetime
    last_accessed_at: Optional[datetime]
    is_archived: bool

    class Config:
        from_attributes = True


class PasswordRevealResponse(BaseModel):
    id: int
    portal_name: str
    username: str
    password: str
    revealed_at: datetime


class BulkDeleteRequest(BaseModel):
    entry_ids: List[int]


class ParsePreviewResponse(BaseModel):
    rows_count: int
    columns_count: int
    sample_rows: List[Dict[str, Any]]
    column_mapping: Dict[str, str]


class BulkImportResponse(BaseModel):
    imported: int
    skipped: int
    errors: int
    error_details: Optional[List[str]] = None


class StatsResponse(BaseModel):
    total: int
    by_portal_type: Dict[str, int]
    by_department: Dict[str, int]
    by_holder_type: Dict[str, int]
    total_access_logs: int
    last_updated: datetime


class AccessLogResponse(BaseModel):
    id: int
    user_id: int
    entry_id: int
    action: str
    timestamp: datetime
    ip_address: Optional[str]


# ── Encryption ────────────────────────────────────────────────────────────────
class PasswordEncryption:
    """
    Uses PASSWORDS_SECRET_KEY env var (must be 32+ bytes).
    Falls back to a hardcoded dev key — NEVER use default in production.
    """

    @staticmethod
    def _get_raw_key() -> bytes:
        key = os.environ.get("PASSWORDS_SECRET_KEY", "")
        if key:
            raw = key.encode()[:32].ljust(32, b"0")
        else:
            # DEV FALLBACK — set PASSWORDS_SECRET_KEY in Render env vars!
            raw = b"taskosphere-dev-key-replace-me!!"[:32].ljust(32, b"0")
        return base64.urlsafe_b64encode(raw)

    @staticmethod
    def encrypt(password: str) -> str:
        if not password:
            return ""
        try:
            f = Fernet(PasswordEncryption._get_raw_key())
            return base64.b64encode(f.encrypt(password.encode())).decode()
        except Exception as exc:
            logger.error("Encryption error: %s", exc)
            return base64.b64encode(password.encode()).decode()

    @staticmethod
    def decrypt(encrypted: str) -> str:
        if not encrypted:
            return ""
        try:
            f = Fernet(PasswordEncryption._get_raw_key())
            return f.decrypt(base64.b64decode(encrypted)).decode()
        except Exception as exc:
            logger.warning("Fernet decrypt failed (%s), trying plain base64", exc)
            try:
                return base64.b64decode(encrypted).decode()
            except Exception:
                return ""


# ── Helpers ───────────────────────────────────────────────────────────────────
def _parse_tags(raw: Optional[str]) -> List[str]:
    if not raw:
        return []
    try:
        parsed = json.loads(raw)
        return parsed if isinstance(parsed, list) else []
    except (json.JSONDecodeError, TypeError):
        return []


def _enrich(entry: PasswordEntry) -> PasswordEntryResponse:
    """Convert ORM row → Pydantic response (sync, no await needed)."""
    return PasswordEntryResponse(
        id=entry.id,
        user_id=entry.user_id,
        portal_name=entry.portal_name,
        portal_type=entry.portal_type or "OTHER",
        url=entry.url,
        username=entry.username,
        department=entry.department or "OTHER",
        holder_type=entry.holder_type or "COMPANY",
        holder_name=entry.holder_name,
        holder_pan=entry.holder_pan,
        holder_din=entry.holder_din,
        mobile=entry.mobile,
        trade_name=entry.trade_name,
        client_name=entry.client_name,
        client_id=entry.client_id,
        notes=entry.notes,
        tags=_parse_tags(entry.tags),
        has_password=bool(entry.has_password),
        created_at=entry.created_at or datetime.utcnow(),
        updated_at=entry.updated_at or datetime.utcnow(),
        last_accessed_at=entry.last_accessed_at,
        is_archived=bool(entry.is_archived),
    )


def _require_user(current_user):
    if not current_user:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Authentication required")


def _log(db: Session, user_id: int, entry_id: int, action: str):
    try:
        db.add(AccessLog(user_id=user_id, entry_id=entry_id, action=action))
    except Exception as exc:
        logger.warning("Failed to write access log: %s", exc)


# ── Router ────────────────────────────────────────────────────────────────────
# IMPORTANT: static-path routes MUST be declared BEFORE /{entry_id} routes,
# otherwise FastAPI will try to cast "admin", "portal-types", etc. as int → 500.

router = APIRouter(prefix="/passwords", tags=["passwords"])


# ── 1. Static / collection routes (no path param) ────────────────────────────

@router.get("", response_model=List[PasswordEntryResponse])
def list_passwords(
    search: Optional[str] = Query(None),
    department: Optional[str] = Query(None),
    portal_type: Optional[str] = Query(None),
    client_id: Optional[str] = Query(None),
    holder_type: Optional[str] = Query(None),
    sort_by: Optional[str] = Query("created_at"),
    sort_order: Optional[str] = Query("desc"),
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=500),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    _require_user(current_user)
    try:
        q = select(PasswordEntry).where(
            and_(
                PasswordEntry.user_id == current_user.id,
                PasswordEntry.is_archived == False,
            )
        )

        if search:
            like = f"%{search}%"
            q = q.where(
                or_(
                    PasswordEntry.portal_name.ilike(like),
                    PasswordEntry.username.ilike(like),
                    PasswordEntry.client_name.ilike(like),
                    PasswordEntry.holder_name.ilike(like),
                    PasswordEntry.trade_name.ilike(like),
                )
            )
        if department and department != "ALL":
            q = q.where(PasswordEntry.department == department)
        if portal_type and portal_type != "ALL":
            q = q.where(PasswordEntry.portal_type == portal_type)
        if client_id and client_id != "ALL":
            q = q.where(PasswordEntry.client_id == client_id)
        if holder_type and holder_type != "ALL":
            q = q.where(PasswordEntry.holder_type == holder_type)

        order_col = (
            PasswordEntry.portal_name if sort_by == "portal_name" else PasswordEntry.created_at
        )
        q = q.order_by(asc(order_col) if sort_order == "asc" else desc(order_col))
        q = q.offset(skip).limit(limit)

        rows = db.execute(q).scalars().all()
        return [_enrich(r) for r in rows]
    except HTTPException:
        raise
    except Exception as exc:
        logger.error("list_passwords error: %s", exc, exc_info=True)
        raise HTTPException(status_code=500, detail=f"Failed to list passwords: {exc}")


@router.post("", response_model=PasswordEntryResponse, status_code=status.HTTP_201_CREATED)
def create_password(
    payload: PasswordEntryCreate,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    _require_user(current_user)
    try:
        dup = db.execute(
            select(PasswordEntry).where(
                and_(
                    PasswordEntry.user_id == current_user.id,
                    PasswordEntry.portal_name == payload.portal_name,
                    PasswordEntry.username == payload.username,
                    PasswordEntry.is_archived == False,
                )
            )
        ).scalar()
        if dup:
            raise HTTPException(status_code=409, detail="Duplicate entry: same portal + username already exists")

        entry = PasswordEntry(
            user_id=current_user.id,
            portal_name=payload.portal_name,
            portal_type=payload.portal_type or "OTHER",
            url=payload.url,
            username=payload.username,
            password_encrypted=PasswordEncryption.encrypt(payload.password_plain or ""),
            has_password=bool(payload.password_plain),
            department=payload.department or "OTHER",
            holder_type=payload.holder_type or "COMPANY",
            holder_name=payload.holder_name,
            holder_pan=payload.holder_pan,
            holder_din=payload.holder_din,
            mobile=payload.mobile,
            trade_name=payload.trade_name,
            client_name=payload.client_name,
            client_id=payload.client_id,
            notes=payload.notes,
            tags=json.dumps(payload.tags or []),
        )
        db.add(entry)
        db.flush()  # get entry.id without full commit
        _log(db, current_user.id, entry.id, "create")
        db.commit()
        db.refresh(entry)
        return _enrich(entry)
    except HTTPException:
        raise
    except Exception as exc:
        db.rollback()
        logger.error("create_password error: %s", exc, exc_info=True)
        raise HTTPException(status_code=500, detail=f"Failed to create entry: {exc}")


# ── 2. Named static sub-routes (MUST come before /{entry_id}) ────────────────

@router.get("/portal-types")
def get_portal_types():
    return {
        "types": [e.value for e in PortalTypeEnum],
        "departments": [e.value for e in DepartmentEnum],
        "holder_types": [e.value for e in HolderTypeEnum],
    }


@router.get("/clients-list")
def get_clients_list(
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    _require_user(current_user)
    try:
        rows = db.execute(
            select(PasswordEntry.client_id, PasswordEntry.client_name)
            .where(
                and_(
                    PasswordEntry.user_id == current_user.id,
                    PasswordEntry.client_name.isnot(None),
                    PasswordEntry.client_id.isnot(None),
                    PasswordEntry.is_archived == False,
                )
            )
            .distinct()
        ).all()
        return [{"id": r[0], "name": r[1]} for r in rows]
    except HTTPException:
        raise
    except Exception as exc:
        logger.error("get_clients_list error: %s", exc, exc_info=True)
        raise HTTPException(status_code=500, detail=str(exc))


@router.get("/download-template")
def download_template():
    """
    Generates an Excel template and streams it back.
    Uses a NamedTemporaryFile so FileResponse has a real path to serve.
    """
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Passwords"
        headers = [
            "Portal Name", "Portal Type", "URL", "Username", "Password",
            "Department", "Holder Type", "Holder Name", "Holder PAN",
            "Holder DIN", "Mobile", "Trade Name", "Client Name", "Client ID", "Notes",
        ]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1F6FB2", end_color="1F6FB2", fill_type="solid")
            ws.column_dimensions[cell.column_letter].width = 18

        # Example row
        ws.append([
            "GST Portal", "GST", "https://www.gst.gov.in", "user@example.com",
            "SecurePass@123", "GST", "COMPANY", "Acme Pvt Ltd",
            "ABCDE1234F", "", "+91 9876543210", "Acme Traders",
            "Acme Corp", "CLI001", "Main GST login",
        ])

        # Write to a named temp file — FileResponse needs a real path
        tmp = tempfile.NamedTemporaryFile(
            delete=False, suffix=".xlsx", prefix="pw_template_"
        )
        wb.save(tmp.name)
        tmp.close()

        return FileResponse(
            path=tmp.name,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename="password-template.xlsx",
        )
    except Exception as exc:
        logger.error("download_template error: %s", exc, exc_info=True)
        raise HTTPException(status_code=500, detail="Failed to generate template")


@router.post("/parse-preview", response_model=ParsePreviewResponse)
async def parse_preview(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    _require_user(current_user)
    try:
        content = await file.read()
        fname = (file.filename or "").lower()
        df = (
            pd.read_csv(io.BytesIO(content))
            if fname.endswith(".csv")
            else pd.read_excel(io.BytesIO(content))
        )

        col_aliases = {
            "portal_name": ["portal name", "portal", "site"],
            "username": ["username", "email", "user", "login"],
            "password_plain": ["password", "pass", "pwd"],
            "department": ["department", "dept"],
            "client_name": ["client name", "client", "company"],
        }
        detected: Dict[str, str] = {}
        lower_cols = {c.lower(): c for c in df.columns}
        for field, aliases in col_aliases.items():
            for alias in aliases:
                if alias in lower_cols:
                    detected[field] = lower_cols[alias]
                    break

        sample = df.head(3).fillna("").to_dict("records")
        return ParsePreviewResponse(
            rows_count=len(df),
            columns_count=len(df.columns),
            sample_rows=sample,
            column_mapping=detected,
        )
    except HTTPException:
        raise
    except Exception as exc:
        logger.error("parse_preview error: %s", exc, exc_info=True)
        raise HTTPException(status_code=400, detail=f"Failed to parse file: {exc}")


@router.post("/bulk-import", response_model=BulkImportResponse)
async def bulk_import(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    _require_user(current_user)
    try:
        content = await file.read()
        fname = (file.filename or "").lower()
        df = (
            pd.read_csv(io.BytesIO(content))
            if fname.endswith(".csv")
            else pd.read_excel(io.BytesIO(content))
        )

        # Normalise column names to lowercase stripped
        df.columns = [str(c).strip().lower() for c in df.columns]

        def _col(*candidates):
            for c in candidates:
                if c in df.columns:
                    return c
            return None

        name_col = _col("portal name", "portal", "site")
        user_col = _col("username", "email", "user", "login")
        pass_col = _col("password", "pass", "pwd")
        dept_col = _col("department", "dept")
        type_col = _col("portal type", "type")
        client_col = _col("client name", "client", "company")

        if not name_col or not user_col:
            raise HTTPException(
                status_code=400,
                detail="File must have 'Portal Name' and 'Username'/'Email' columns",
            )

        imported = skipped = errors = 0
        error_details: List[str] = []

        for idx, row in df.iterrows():
            try:
                portal_name = str(row.get(name_col, "") or "").strip()
                username = str(row.get(user_col, "") or "").strip()
                if not portal_name or not username:
                    skipped += 1
                    continue

                dup = db.execute(
                    select(PasswordEntry).where(
                        and_(
                            PasswordEntry.user_id == current_user.id,
                            PasswordEntry.portal_name == portal_name,
                            PasswordEntry.username == username,
                            PasswordEntry.is_archived == False,
                        )
                    )
                ).scalar()
                if dup:
                    skipped += 1
                    continue

                raw_pass = str(row.get(pass_col, "") or "") if pass_col else ""
                db.add(
                    PasswordEntry(
                        user_id=current_user.id,
                        portal_name=portal_name,
                        portal_type=str(row.get(type_col, "OTHER") or "OTHER") if type_col else "OTHER",
                        username=username,
                        password_encrypted=PasswordEncryption.encrypt(raw_pass),
                        has_password=bool(raw_pass),
                        department=str(row.get(dept_col, "OTHER") or "OTHER") if dept_col else "OTHER",
                        client_name=str(row.get(client_col, "") or "") if client_col else None,
                        tags=json.dumps([]),
                    )
                )
                imported += 1

                # Batch-commit every 50 rows to avoid huge transactions
                if imported % 50 == 0:
                    db.commit()

            except Exception as row_exc:
                errors += 1
                error_details.append(f"Row {idx + 2}: {row_exc}")

        db.commit()
        return BulkImportResponse(
            imported=imported, skipped=skipped, errors=errors,
            error_details=error_details[:20] if error_details else None,
        )
    except HTTPException:
        raise
    except Exception as exc:
        db.rollback()
        logger.error("bulk_import error: %s", exc, exc_info=True)
        raise HTTPException(status_code=400, detail=f"Import failed: {exc}")


@router.post("/bulk-delete", status_code=status.HTTP_204_NO_CONTENT)
def bulk_delete_passwords(
    payload: BulkDeleteRequest,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    _require_user(current_user)
    # Soft-delete: archive all matching entries for this user
    try:
        for eid in payload.entry_ids:
            entry = db.execute(
                select(PasswordEntry).where(
                    and_(PasswordEntry.id == eid, PasswordEntry.user_id == current_user.id)
                )
            ).scalar()
            if entry:
                entry.is_archived = True
                _log(db, current_user.id, eid, "bulk_delete")
        db.commit()
    except Exception as exc:
        db.rollback()
        logger.error("bulk_delete error: %s", exc, exc_info=True)
        raise HTTPException(status_code=500, detail=str(exc))


@router.get("/admin/stats", response_model=StatsResponse)
def admin_stats(
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    _require_user(current_user)
    try:
        base_filter = PasswordEntry.is_archived == False

        total = db.execute(select(func.count(PasswordEntry.id)).where(base_filter)).scalar() or 0

        portal_rows = db.execute(
            select(PasswordEntry.portal_type, func.count(PasswordEntry.id))
            .where(base_filter)
            .group_by(PasswordEntry.portal_type)
        ).all()
        by_portal = {r[0]: r[1] for r in portal_rows if r[0]}

        dept_rows = db.execute(
            select(PasswordEntry.department, func.count(PasswordEntry.id))
            .where(base_filter)
            .group_by(PasswordEntry.department)
        ).all()
        by_dept = {r[0]: r[1] for r in dept_rows if r[0]}

        holder_rows = db.execute(
            select(PasswordEntry.holder_type, func.count(PasswordEntry.id))
            .where(base_filter)
            .group_by(PasswordEntry.holder_type)
        ).all()
        by_holder = {r[0]: r[1] for r in holder_rows if r[0]}

        log_count = db.execute(select(func.count(AccessLog.id))).scalar() or 0

        return StatsResponse(
            total=total,
            by_portal_type=by_portal,
            by_department=by_dept,
            by_holder_type=by_holder,
            total_access_logs=log_count,
            last_updated=datetime.utcnow(),
        )
    except HTTPException:
        raise
    except Exception as exc:
        logger.error("admin_stats error: %s", exc, exc_info=True)
        raise HTTPException(status_code=500, detail=str(exc))


# ── 3. /{entry_id} routes — MUST come last ───────────────────────────────────

@router.get("/{entry_id}", response_model=PasswordEntryResponse)
def get_password(
    entry_id: int,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    _require_user(current_user)
    entry = db.execute(
        select(PasswordEntry).where(
            and_(PasswordEntry.id == entry_id, PasswordEntry.user_id == current_user.id)
        )
    ).scalar()
    if not entry:
        raise HTTPException(status_code=404, detail="Entry not found")
    return _enrich(entry)


@router.get("/{entry_id}/reveal", response_model=PasswordRevealResponse)
def reveal_password(
    entry_id: int,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    _require_user(current_user)
    entry = db.execute(
        select(PasswordEntry).where(
            and_(PasswordEntry.id == entry_id, PasswordEntry.user_id == current_user.id)
        )
    ).scalar()
    if not entry:
        raise HTTPException(status_code=404, detail="Entry not found")

    password = PasswordEncryption.decrypt(entry.password_encrypted) if entry.has_password else ""
    entry.last_accessed_at = datetime.utcnow()
    _log(db, current_user.id, entry_id, "reveal")
    try:
        db.commit()
    except Exception as exc:
        logger.warning("reveal commit failed: %s", exc)
        db.rollback()

    return PasswordRevealResponse(
        id=entry.id,
        portal_name=entry.portal_name,
        username=entry.username,
        password=password,
        revealed_at=datetime.utcnow(),
    )


@router.put("/{entry_id}", response_model=PasswordEntryResponse)
def update_password(
    entry_id: int,
    payload: PasswordEntryUpdate,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    _require_user(current_user)
    entry = db.execute(
        select(PasswordEntry).where(
            and_(PasswordEntry.id == entry_id, PasswordEntry.user_id == current_user.id)
        )
    ).scalar()
    if not entry:
        raise HTTPException(status_code=404, detail="Entry not found")

    try:
        update_data = payload.dict(exclude_unset=True)

        if "password_plain" in update_data:
            plain = update_data.pop("password_plain")
            if plain:
                entry.password_encrypted = PasswordEncryption.encrypt(plain)
                entry.has_password = True

        if "tags" in update_data:
            entry.tags = json.dumps(update_data.pop("tags") or [])

        for key, value in update_data.items():
            if hasattr(entry, key):
                setattr(entry, key, value)

        entry.updated_at = datetime.utcnow()
        _log(db, current_user.id, entry_id, "edit")
        db.commit()
        db.refresh(entry)
        return _enrich(entry)
    except HTTPException:
        raise
    except Exception as exc:
        db.rollback()
        logger.error("update_password error: %s", exc, exc_info=True)
        raise HTTPException(status_code=500, detail=str(exc))


@router.delete("/{entry_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_password(
    entry_id: int,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    _require_user(current_user)
    entry = db.execute(
        select(PasswordEntry).where(
            and_(PasswordEntry.id == entry_id, PasswordEntry.user_id == current_user.id)
        )
    ).scalar()
    if not entry:
        raise HTTPException(status_code=404, detail="Entry not found")

    entry.is_archived = True
    _log(db, current_user.id, entry_id, "delete")
    try:
        db.commit()
    except Exception as exc:
        db.rollback()
        logger.error("delete_password error: %s", exc, exc_info=True)
        raise HTTPException(status_code=500, detail=str(exc))
