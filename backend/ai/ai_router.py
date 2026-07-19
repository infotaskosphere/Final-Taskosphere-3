import uuid
import logging
import hashlib
import json
from datetime import datetime, timezone
from typing import Callable, Any

from backend.dependencies import db
from backend.ai.fingerprint import generate_document_fingerprint
from backend.ai.ai_memory import save_ai_memory, find_memory_by_fingerprint, update_ai_memory

logger = logging.getLogger("ai_router")

def parse_analysis_text(text: str) -> dict:
    import re
    res = {
        "document_type": "UNKNOWN",
        "vendor_name": "Unknown Vendor",
        "vendor_gstin": "",
        "invoice_number": "",
        "invoice_date": "",
        "invoice_total": 0.0,
        "taxable_amount": 0.0,
        "gst_amount": 0.0,
        "confidence": 1.0
    }
    if not text:
        return res
        
    text_upper = text.upper()
    
    # 1. Document type
    if "PURCHASE" in text_upper:
        res["document_type"] = "PURCHASE"
    elif "SALE" in text_upper:
        res["document_type"] = "SALE"
        
    # 2. GSTIN search
    gstin_re = re.compile(r'\b\d{2}[A-Z]{5}\d{4}[A-Z][A-Z\d]Z[A-Z\d]\b', re.I)
    gstin_match = gstin_re.search(text)
    if gstin_match:
        res["vendor_gstin"] = gstin_match.group(0).upper()
        
    # 3. Invoice Number
    inv_re = re.compile(r'(?:invoice\s*(?:no|number|#)?|inv\s*(?:no|num|number|#)?)\s*[:\-=]?\s*([a-z0-9\-/]+)', re.I)
    inv_match = inv_re.search(text)
    if inv_match:
        res["invoice_number"] = inv_match.group(1).strip()
        
    # 4. Vendor Name
    vendor_re = re.compile(r'(?:vendor|seller|supplier|party\s*name)\s*[:\-=]?\s*([^\n\r]+)', re.I)
    vendor_match = vendor_re.search(text)
    if vendor_match:
        res["vendor_name"] = vendor_match.group(1).strip()[:100]
        
    # 5. Date
    date_re = re.compile(r'(?:date)\s*[:\-=]?\s*(\d{4}-\d{2}-\d{2})', re.I)
    date_match = date_re.search(text)
    if date_match:
        res["invoice_date"] = date_match.group(1)
        
    # 6. Total Invoice Value
    total_re = re.compile(r'(?:total|grand\s*total|amount|value)\s*[:\-=]?\s*(?:rs\.?|inr|₹|\$)?\s*([0-9,]+\.?[0-9]*)', re.I)
    total_match = total_re.search(text)
    if total_match:
        try:
            res["invoice_total"] = float(total_match.group(1).replace(",", ""))
        except ValueError:
            pass
            
    return res

async def classify_document(
    contents: bytes,
    filename: str,
    raw_ocr_text: str = "",
    file_hash: str = "",
    document_id: str = ""
) -> dict:
    """
    Module level classifier function that calls the document_classifier.py
    """
    from backend.ai.document_classifier import classify_document as run_classifier
    return await run_classifier(contents, filename, raw_ocr_text, file_hash, document_id)

async def find_template(contents: bytes, filename: str, ocr_text: str = "") -> Any:
    """
    Exposes find_matching_template from template_engine to find matching layout templates.
    """
    from backend.ai.template_engine import find_matching_template
    return await find_matching_template(contents, filename, ocr_text)

async def learn_template(
    contents: bytes,
    filename: str,
    doc_type: str,
    extracted_json: dict,
    ocr_text: str
) -> Any:
    """
    Exposes learn_template from template_engine to automatically learn templates.
    """
    from backend.ai.template_engine import learn_template as run_learn_template
    return await run_learn_template(contents, filename, doc_type, extracted_json, ocr_text)

async def process_document(
    contents: bytes,
    filename: str,
    current_user: Any,
    existing_reader_func: Callable
) -> dict:
    """
    Coordinates fingerprint generation, memory lookup, calling existing reader,
    and memory save.
    """
    logger.info("Document Received")
    
    # 1. Compute file hash
    file_hash = hashlib.sha256(contents).hexdigest()
    
    # 2. Extract initial raw text to help look up
    from backend.ai.ai_router import get_document_text_content
    raw_ocr_text = get_document_text_content(contents, filename)
    normalized_ocr = ""
    if raw_ocr_text.strip():
        import re
        normalized_ocr = re.sub(r'[^a-z0-9]', '', raw_ocr_text.lower().strip())[:1000]

    # Generate Fingerprint (Step 2 of Phase 2 Workflow)
    from backend.ai.fingerprint import generate_document_fingerprint
    fingerprint = generate_document_fingerprint(
        vendor_name="",
        vendor_gstin="",
        invoice_number="",
        document_type="",
        raw_ocr_text=raw_ocr_text
    )
    logger.info("Fingerprint Generated")

    # 3. Check AI Memory (Step 3 of Phase 2 Workflow)
    memory_record = None
    try:
        if file_hash:
            memory_record = await db.ai_document_memory.find_one({"file_hash": file_hash}, {"_id": 0})
        
        if not memory_record and normalized_ocr:
            memory_record = await db.ai_document_memory.find_one({"raw_ocr_text_normalized": normalized_ocr}, {"_id": 0})
    except Exception as e:
        logger.error(f"Error checking AI Memory lookup: {e}", exc_info=True)

    # 4. Classify Document (Step 4 of Phase 2 Workflow)
    classification_res = {"document_type": "Other", "confidence": 0.0, "reason": "Failed to classify"}
    try:
        classification_res = await classify_document(contents, filename, raw_ocr_text, file_hash)
    except Exception as exc:
        logger.error(f"Classifier failed: {exc}", exc_info=True)
    
    # 5. Return Classification (Step 5 of Phase 2 Workflow / Continue existing processing)
    if memory_record:
        logger.info("Memory Hit")
        res_json = memory_record["extracted_json"]
        if isinstance(res_json, dict):
            res_json["classification"] = classification_res
            res_json["document_type"] = classification_res["document_type"]
        return res_json
        
    logger.info("Memory Miss")

    from backend.ai.vendor_learning import apply_vendor_defaults, learn_vendor_profile
    from backend.ai.vendor_mapper import lookup_vendor_from_ocr

    # Phase 3: Template Search
    from backend.ai.template_engine import find_matching_template, extract_using_template, learn_template
    from backend.ai.template_storage import increment_usage
    
    matched_template = None
    try:
        matched_template = await find_matching_template(contents, filename, raw_ocr_text)
    except Exception as exc:
        logger.error(f"Template Search failed: {exc}", exc_info=True)
        
    if matched_template:
        logger.info("Template Match Found")
        # Extract fields using template and bypass Gemini/Groq
        try:
            result = extract_using_template(matched_template, raw_ocr_text)
            await increment_usage(matched_template["template_id"], success=True)
            
            # Enrich with classification
            if isinstance(result, dict):
                result["classification"] = classification_res
                result["document_type"] = classification_res["document_type"]
                result["template_matched_id"] = matched_template["template_id"]
                result["extraction_method"] = "template"
                
                # Apply Vendor Defaults & Learning
                vendor_name = result.get("vendor_or_customer_name") or ""
                vendor_gstin = result.get("tax_registration_number") or ""
                result = await apply_vendor_defaults(vendor_name, vendor_gstin, result)
                try:
                    await learn_vendor_profile(vendor_name, vendor_gstin, result)
                    logger.info("Vendor Defaults Applied and Profile Learned from Template")
                except Exception as l_err:
                    logger.error(f"Failed to learn vendor profile from template: {l_err}", exc_info=True)
                
            # Save this to memory for future fingerprint hits as well
            try:
                classified_doc_type = classification_res["document_type"]
                
                parsed = {
                    "vendor_name": result.get("vendor_or_customer_name") or "",
                    "vendor_gstin": result.get("tax_registration_number") or "",
                    "invoice_number": result.get("invoice_number") or "",
                    "invoice_date": result.get("invoice_date") or "",
                    "invoice_total": float(result.get("total_invoice_value") or 0) if result.get("total_invoice_value") else 0.0,
                    "taxable_amount": float(result.get("taxable_value") or 0) if result.get("taxable_value") else 0.0,
                    "gst_amount": float(result.get("total_tax") or 0) if result.get("total_tax") else 0.0
                }
                
                fingerprint = generate_document_fingerprint(
                    vendor_name=parsed["vendor_name"],
                    vendor_gstin=parsed["vendor_gstin"],
                    invoice_number=parsed["invoice_number"],
                    document_type=classified_doc_type,
                    raw_ocr_text=raw_ocr_text
                )
                
                now = datetime.now(timezone.utc).isoformat()
                new_record = {
                    "document_id": str(uuid.uuid4()),
                    "fingerprint": fingerprint,
                    "document_type": classified_doc_type,
                    "vendor_name": parsed["vendor_name"],
                    "vendor_gstin": parsed["vendor_gstin"],
                    "invoice_number": parsed["invoice_number"],
                    "invoice_date": parsed["invoice_date"],
                    "invoice_total": parsed["invoice_total"],
                    "taxable_amount": parsed["taxable_amount"],
                    "gst_amount": parsed["gst_amount"],
                    "raw_ocr_text": raw_ocr_text,
                    "extracted_json": result,
                    "ledger_mapping": {},
                    "journal_entry": {},
                    "processing_engine": "template",
                    "ai_confidence": classification_res["confidence"],
                    "processing_status": "extracted",
                    "created_at": now,
                    "updated_at": now,
                    "file_hash": file_hash,
                    "raw_ocr_text_normalized": normalized_ocr
                }
                await save_ai_memory(new_record)
                logger.info("Memory Saved (from Template)")
            except Exception as e:
                logger.error(f"Error saving template extraction to AI Memory: {e}", exc_info=True)
                
            return result
        except Exception as e:
            logger.error(f"Failed to extract using matched template: {e}", exc_info=True)
            logger.info("Template extraction failed, falling back to AI extraction")
    
    logger.info("Template Miss")

    # Vendor Lookup from OCR early before calling existing reader
    early_vendor = None
    try:
        early_vendor = await lookup_vendor_from_ocr(raw_ocr_text)
        if early_vendor:
            logger.info(f"Vendor Lookup Found Early from OCR: {early_vendor.get('vendor_name')}")
    except Exception as exc:
        logger.error(f"Early Vendor Lookup failed: {exc}", exc_info=True)
    
    # 4. Execute existing reader workflow
    logger.info("Gemini Called")
    result = await existing_reader_func()
    logger.info("Gemini Success")
    
    if isinstance(result, dict):
        result["classification"] = classification_res
        result["document_type"] = classification_res["document_type"]
        
        # Apply Vendor Defaults & Update Learning
        try:
            vendor_name = result.get("vendor_or_customer_name") or (early_vendor.get("vendor_name") if early_vendor else "")
            vendor_gstin = result.get("tax_registration_number") or (early_vendor.get("gstin") if early_vendor else "")
            
            # Apply defaults to result
            result = await apply_vendor_defaults(vendor_name, vendor_gstin, result)
            
            # Learn vendor profile from results
            await learn_vendor_profile(vendor_name, vendor_gstin, result)
            logger.info("Vendor Profile Updated from Gemini Extraction")
        except Exception as l_err:
            logger.error(f"Failed to apply/learn vendor profile: {l_err}", exc_info=True)

    # 5. Save to Memory
    try:
        analysis_text = result.get("analysis") if isinstance(result, dict) else ""
        parsed = parse_analysis_text(analysis_text)
        
        # Use classified document type as the single source of truth
        classified_doc_type = classification_res["document_type"]
        
        # Trigger template learning automatically for successful high confidence extractions
        try:
            extracted_baseline = {
                "vendor_or_customer_name": parsed["vendor_name"],
                "tax_registration_number": parsed["vendor_gstin"],
                "invoice_number": parsed["invoice_number"],
                "invoice_date": parsed["invoice_date"],
                "total_invoice_value": parsed["invoice_total"],
                "taxable_value": parsed["taxable_amount"],
                "total_tax": parsed["gst_amount"],
                "confidence": 1.0,
                "analysis": analysis_text
            }
            await learn_template(
                contents=contents,
                filename=filename,
                doc_type=classified_doc_type,
                extracted_json=extracted_baseline,
                ocr_text=raw_ocr_text
            )
            logger.info("Template Created")
        except Exception as temp_err:
            logger.error(f"Error learning template: {temp_err}", exc_info=True)
        
        fingerprint = generate_document_fingerprint(
            vendor_name=parsed["vendor_name"],
            vendor_gstin=parsed["vendor_gstin"],
            invoice_number=parsed["invoice_number"],
            document_type=classified_doc_type,
            raw_ocr_text=raw_ocr_text
        )
        logger.info("Fingerprint Generated")
        
        now = datetime.now(timezone.utc).isoformat()
        new_record = {
            "document_id": str(uuid.uuid4()),
            "fingerprint": fingerprint,
            "document_type": classified_doc_type,
            "vendor_name": parsed["vendor_name"],
            "vendor_gstin": parsed["vendor_gstin"],
            "invoice_number": parsed["invoice_number"],
            "invoice_date": parsed["invoice_date"],
            "invoice_total": parsed["invoice_total"],
            "taxable_amount": parsed["taxable_amount"],
            "gst_amount": parsed["gst_amount"],
            "raw_ocr_text": raw_ocr_text,
            "extracted_json": result,
            "ledger_mapping": {},
            "journal_entry": {},
            "processing_engine": "gemini",
            "ai_confidence": classification_res["confidence"],
            "processing_status": "extracted",
            "created_at": now,
            "updated_at": now,
            "file_hash": file_hash,
            "raw_ocr_text_normalized": normalized_ocr
        }
        
        await save_ai_memory(new_record)
        logger.info("Memory Saved")
    except Exception as e:
        logger.error(f"Error saving to AI Memory: {e}", exc_info=True)
        
    return result

def get_document_text_content(contents: bytes, filename: str) -> str:
    import io
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    text_content = ""
    try:
        if ext in ("xlsx", "xlsm"):
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
            lines = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                lines.append(f"\n--- Sheet: {sheet_name} ---")
                for row in ws.iter_rows(values_only=True):
                    row_text = "\t".join("" if v is None else str(v) for v in row)
                    if row_text.strip():
                        lines.append(row_text)
            text_content = "\n".join(lines)
        elif ext == "xls":
            import pandas as pd
            df = pd.read_excel(io.BytesIO(contents), engine="xlrd")
            text_content = df.to_string(index=False)
        elif ext == "csv":
            text_content = contents.decode("utf-8", errors="replace")
        elif ext == "pdf":
            import pdfplumber
            extracted_pages = []
            with pdfplumber.open(io.BytesIO(contents)) as pdf:
                for i, page in enumerate(pdf.pages[:30]):
                    text = page.extract_text()
                    if text and text.strip():
                        extracted_pages.append(f"--- Page {i+1} ---\n{text.strip()}")
            text_content = "\n\n".join(extracted_pages)
    except Exception:
        pass
    return text_content

async def update_accounting_memory(doc: dict, entry: dict, user_id: str):
    """
    Called after successful journal creation to update the AI Memory.
    """
    try:
        extracted = doc.get("extracted") or {}
        vendor_name = extracted.get("vendor_or_customer_name") or ""
        vendor_gstin = extracted.get("tax_registration_number") or ""
        invoice_number = extracted.get("invoice_number") or ""
        doc_type = extracted.get("document_type") or ""
        
        from backend.ai.fingerprint import generate_document_fingerprint
        from backend.ai.ai_memory import save_ai_memory, find_memory_by_fingerprint, update_ai_memory
        
        raw_ocr_text = doc.get("raw_ocr_text") or ""
        fingerprint = generate_document_fingerprint(
            vendor_name=vendor_name,
            vendor_gstin=vendor_gstin,
            invoice_number=invoice_number,
            document_type=doc_type,
            raw_ocr_text=raw_ocr_text
        )
        
        memory_record = await find_memory_by_fingerprint(fingerprint)
        
        now = datetime.now(timezone.utc).isoformat()
        ledger_mapping = {
            "lines": entry.get("lines") or doc.get("preview", {}).get("lines") or [],
            "rules_used": doc.get("company_match_reason") or ""
        }
        
        update_data = {
            "ledger_mapping": ledger_mapping,
            "journal_entry": entry,
            "posting_status": "posted",
            "posting_timestamp": now,
            "updated_at": now
        }
        
        if memory_record:
            await update_ai_memory(fingerprint, update_data)
            logger.info("Journal Linked")
        else:
            # Create a new memory record if it didn't exist
            new_record = {
                "document_id": doc.get("id") or str(uuid.uuid4()),
                "fingerprint": fingerprint,
                "document_type": doc_type,
                "vendor_name": vendor_name,
                "vendor_gstin": vendor_gstin,
                "invoice_number": invoice_number,
                "invoice_date": extracted.get("invoice_date") or "",
                "invoice_total": float(extracted.get("total_invoice_value") or 0),
                "taxable_amount": float(extracted.get("taxable_value") or 0),
                "gst_amount": float(extracted.get("total_tax") or 0),
                "raw_ocr_text": raw_ocr_text,
                "extracted_json": extracted,
                "ledger_mapping": ledger_mapping,
                "journal_entry": entry,
                "processing_engine": "gemini",
                "ai_confidence": float(extracted.get("confidence") or 1.0),
                "processing_status": "posted",
                "created_at": now,
                "updated_at": now,
                "file_hash": doc.get("file_hash") or ""
            }
            new_record["posting_status"] = "posted"
            new_record["posting_timestamp"] = now
            await save_ai_memory(new_record)
            logger.info("Journal Linked")
    except Exception as e:
        logger.error(f"Error in update_accounting_memory: {e}", exc_info=True)
